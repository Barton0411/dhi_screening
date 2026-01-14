#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import os
import threading
import time
import pandas as pd
from datetime import datetime, timedelta
from typing import List, Dict, Any
import logging
import subprocess
import atexit
import requests
import importlib.util

# è®¾ç½®logger
logger = logging.getLogger(__name__)

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QWidget, 
    QLabel, QPushButton, QFileDialog, QProgressBar, QTextEdit,
    QGroupBox, QFormLayout, QSpinBox, QDoubleSpinBox, QCheckBox,
    QComboBox, QDateEdit, QTableWidget, QTableWidgetItem, 
    QTabWidget, QMessageBox, QSplitter, QHeaderView, QListWidget,
    QListWidgetItem, QFrame, QScrollArea, QMenuBar, QMenu, 
    QDialog, QDialogButtonBox, QSlider, QGridLayout,
    QColorDialog, QInputDialog, QLineEdit, QStyle, QRadioButton,
    QSizePolicy, QProgressDialog
)
from PyQt6.QtCore import QThread, pyqtSignal, QDate, Qt, QTimer, QSettings, QPropertyAnimation, QEasingCurve, pyqtProperty, QDateTime
from PyQt6.QtGui import QIcon, QFont, QPixmap, QColor, QAction
import yaml

# å¯¼å…¥æˆ‘ä»¬çš„æ•°æ®å¤„ç†æ¨¡å—
from data_processor import DataProcessor
from models import FilterConfig

# å¯¼å…¥è®¤è¯æ¨¡å—
from auth_module import LoginDialog, show_login_dialog
from auth_module.simple_auth_service import SimpleAuthService

# å¯¼å…¥éšæ€§ä¹³æˆ¿ç‚ç›‘æµ‹æ¨¡å—
from mastitis_monitoring import MastitisMonitoringCalculator

# å¯¼å…¥è¿›åº¦æ¡ç®¡ç†å™¨
from progress_manager import SmoothProgressDialog, AsyncProgressManager

# å¯¼å…¥å›¾è¡¨æœ¬åœ°åŒ–
ChinesePlotWidget = None
try:
    # å°è¯•ç›´æ¥å¯¼å…¥
    from chart_localization import ChinesePlotWidget
    logger.info("æˆåŠŸå¯¼å…¥ ChinesePlotWidget")
except ImportError as e:
    logger.warning(f"æ— æ³•å¯¼å…¥ ChinesePlotWidget: {e}")
    # å°è¯•ä»å½“å‰ç›®å½•å¯¼å…¥
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        chart_path = os.path.join(base_dir, 'chart_localization.py')
        if os.path.exists(chart_path):
            spec = importlib.util.spec_from_file_location("chart_localization", chart_path)
            chart_module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(chart_module)
            ChinesePlotWidget = chart_module.ChinesePlotWidget
            logger.info(f"ä»æ–‡ä»¶è·¯å¾„å¯¼å…¥ ChinesePlotWidget: {chart_path}")
        else:
            logger.warning(f"æ‰¾ä¸åˆ° chart_localization.py: {chart_path}")
    except Exception as e2:
        logger.error(f"ä»æ–‡ä»¶å¯¼å…¥ ChinesePlotWidget å¤±è´¥: {e2}")
except Exception as e:
    logger.error(f"å¯¼å…¥ ChinesePlotWidget æ—¶å‡ºé”™: {e}")


class DisplaySettingsDialog(QDialog):
    """æ˜¾ç¤ºè®¾ç½®å¯¹è¯æ¡†"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("ç•Œé¢æ˜¾ç¤ºè®¾ç½®")
        self.setModal(True)
        self.resize(450, 600)
        
        # åŠ è½½å½“å‰è®¾ç½®
        self.settings = QSettings("DHI", "ProteinScreening")
        current_scale = self.settings.value("display_scale", 100, type=int)
        current_font_color = self.settings.value("font_color", "#000000", type=str)
        current_bg_color = self.settings.value("background_color", "#ffffff", type=str)
        current_font_family = self.settings.value("font_family", "Microsoft YaHei", type=str)
        current_font_size = self.settings.value("font_size", 12, type=int)
        current_font_bold = self.settings.value("font_bold", False, type=bool)
        current_font_italic = self.settings.value("font_italic", False, type=bool)
        current_font_underline = self.settings.value("font_underline", False, type=bool)
        current_use_system_theme = self.settings.value("use_system_theme", True, type=bool)
        
        self.init_ui(current_scale, current_font_color, current_bg_color, 
                    current_font_family, current_font_size, current_font_bold, 
                    current_font_italic, current_font_underline, current_use_system_theme)
    
    def init_ui(self, current_scale, current_font_color, current_bg_color,
                current_font_family, current_font_size, current_font_bold,
                current_font_italic, current_font_underline, current_use_system_theme):
        """åˆå§‹åŒ–ç•Œé¢"""
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # æ ‡é¢˜
        title_label = QLabel("ç•Œé¢æ˜¾ç¤ºè®¾ç½®")
        title_label.setStyleSheet("font-weight: bold; color: #333;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        # åˆ›å»ºæ»šåŠ¨åŒºåŸŸ
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        scroll_layout.setSpacing(15)
        
        # ç¼©æ”¾è®¾ç½®åˆ†ç»„
        scale_group = QGroupBox("æ˜¾ç¤ºç¼©æ”¾")
        scale_group.setStyleSheet("QGroupBox { font-weight: bold; color: #333; }")
        scale_layout = QVBoxLayout(scale_group)
        
        # ç¼©æ”¾æ»‘å—
        scale_container = QWidget()
        scale_container_layout = QHBoxLayout(scale_container)
        scale_container_layout.setContentsMargins(0, 0, 0, 0)
        
        scale_container_layout.addWidget(QLabel("50%"))
        
        self.scale_slider = QSlider(Qt.Orientation.Horizontal)
        self.scale_slider.setRange(50, 200)
        self.scale_slider.setValue(current_scale)
        self.scale_slider.setTickPosition(QSlider.TickPosition.TicksBelow)
        self.scale_slider.setTickInterval(25)
        self.scale_slider.valueChanged.connect(self.update_scale_label)
        scale_container_layout.addWidget(self.scale_slider)
        
        scale_container_layout.addWidget(QLabel("200%"))
        
        scale_layout.addWidget(scale_container)
        
        # å½“å‰ç¼©æ”¾æ˜¾ç¤º
        self.scale_label = QLabel(f"å½“å‰ç¼©æ”¾: {current_scale}%")
        self.scale_label.setStyleSheet("color: #666; font-size: 14px;")
        self.scale_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        scale_layout.addWidget(self.scale_label)
        
        scroll_layout.addWidget(scale_group)
        
        # å­—ä½“è®¾ç½®åˆ†ç»„
        font_group = QGroupBox("å­—ä½“è®¾ç½®")
        font_group.setStyleSheet("QGroupBox { font-weight: bold; color: #333; }")
        font_layout = QVBoxLayout(font_group)
        
        # å­—ä½“ç³»åˆ—è®¾ç½®
        font_family_container = QWidget()
        font_family_layout = QHBoxLayout(font_family_container)
        font_family_layout.setContentsMargins(0, 0, 0, 0)
        
        font_family_label = QLabel("å­—ä½“ç±»å‹:")
        font_family_label.setStyleSheet("color: #333; min-width: 80px;")
        font_family_layout.addWidget(font_family_label)
        
        self.font_family_combo = QComboBox()
        # æ·»åŠ å¸¸ç”¨ä¸­æ–‡å­—ä½“
        fonts = [
            "Microsoft YaHei", "SimHei", "SimSun", "KaiTi", "FangSong",
            "Arial", "Times New Roman", "Calibri", "Consolas", "Verdana"
        ]
        self.font_family_combo.addItems(fonts)
        self.font_family_combo.setCurrentText(current_font_family)
        self.font_family_combo.currentTextChanged.connect(self.update_preview)
        font_family_layout.addWidget(self.font_family_combo)
        
        font_layout.addWidget(font_family_container)
        
        # å­—ä½“å¤§å°è®¾ç½®
        font_size_container = QWidget()
        font_size_layout = QHBoxLayout(font_size_container)
        font_size_layout.setContentsMargins(0, 0, 0, 0)
        
        font_size_label = QLabel("å­—ä½“å¤§å°:")
        font_size_label.setStyleSheet("color: #333; min-width: 80px;")
        font_size_layout.addWidget(font_size_label)
        
        self.font_size_spin = QSpinBox()
        self.font_size_spin.setRange(8, 32)
        self.font_size_spin.setValue(current_font_size)
        self.font_size_spin.setSuffix(" px")
        self.font_size_spin.valueChanged.connect(self.update_preview)
        font_size_layout.addWidget(self.font_size_spin)
        
        font_size_layout.addStretch()
        
        font_layout.addWidget(font_size_container)
        
        # å­—ä½“æ ·å¼è®¾ç½®
        font_style_container = QWidget()
        font_style_layout = QHBoxLayout(font_style_container)
        font_style_layout.setContentsMargins(0, 0, 0, 0)
        
        font_style_label = QLabel("å­—ä½“æ ·å¼:")
        font_style_label.setStyleSheet("color: #333; min-width: 80px;")
        font_style_layout.addWidget(font_style_label)
        
        self.font_bold_cb = QCheckBox("åŠ ç²—")
        self.font_bold_cb.setChecked(current_font_bold)
        self.font_bold_cb.stateChanged.connect(self.update_preview)
        font_style_layout.addWidget(self.font_bold_cb)
        
        self.font_italic_cb = QCheckBox("æ–œä½“")
        self.font_italic_cb.setChecked(current_font_italic)
        self.font_italic_cb.stateChanged.connect(self.update_preview)
        font_style_layout.addWidget(self.font_italic_cb)
        
        self.font_underline_cb = QCheckBox("ä¸‹åˆ’çº¿")
        self.font_underline_cb.setChecked(current_font_underline)
        self.font_underline_cb.stateChanged.connect(self.update_preview)
        font_style_layout.addWidget(self.font_underline_cb)
        
        font_style_layout.addStretch()
        
        font_layout.addWidget(font_style_container)
        
        scroll_layout.addWidget(font_group)
        
        # é¢œè‰²è®¾ç½®åˆ†ç»„
        color_group = QGroupBox("é¢œè‰²è®¾ç½®")
        color_group.setStyleSheet("QGroupBox { font-weight: bold; color: #333; }")
        color_layout = QVBoxLayout(color_group)
        
        # ç³»ç»Ÿä¸»é¢˜è·Ÿéšé€‰é¡¹
        self.use_system_theme_cb = QCheckBox("è·Ÿéšç³»ç»Ÿä¸»é¢˜ï¼ˆæ·±è‰²/æµ…è‰²æ¨¡å¼ï¼‰")
        self.use_system_theme_cb.setChecked(current_use_system_theme)
        self.use_system_theme_cb.setToolTip("è‡ªåŠ¨é€‚é…ç³»ç»Ÿçš„æ·±è‰²æˆ–æµ…è‰²ä¸»é¢˜")
        self.use_system_theme_cb.stateChanged.connect(self.on_system_theme_toggled)
        color_layout.addWidget(self.use_system_theme_cb)
        
        # å­—ä½“é¢œè‰²è®¾ç½®
        font_color_container = QWidget()
        font_color_layout = QHBoxLayout(font_color_container)
        font_color_layout.setContentsMargins(0, 0, 0, 0)
        
        font_color_label = QLabel("å­—ä½“é¢œè‰²:")
        font_color_label.setStyleSheet("color: #333; min-width: 80px;")
        font_color_layout.addWidget(font_color_label)
        
        self.font_color_btn = QPushButton()
        self.font_color_btn.setFixedSize(80, 30)
        self.font_color_btn.clicked.connect(self.choose_font_color)
        self.current_font_color = current_font_color
        self.update_color_button(self.font_color_btn, self.current_font_color)
        font_color_layout.addWidget(self.font_color_btn)
        
        font_color_layout.addStretch()
        
        # é‡ç½®ä¸ºé»˜è®¤é»‘è‰²æŒ‰é’®
        reset_font_btn = QPushButton("é‡ç½®ä¸ºé»‘è‰²")
        reset_font_btn.setFixedSize(80, 30)
        reset_font_btn.clicked.connect(lambda: self.reset_color('font'))
        reset_font_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        font_color_layout.addWidget(reset_font_btn)
        
        color_layout.addWidget(font_color_container)
        
        # èƒŒæ™¯é¢œè‰²è®¾ç½®
        bg_color_container = QWidget()
        bg_color_layout = QHBoxLayout(bg_color_container)
        bg_color_layout.setContentsMargins(0, 0, 0, 0)
        
        bg_color_label = QLabel("èƒŒæ™¯é¢œè‰²:")
        bg_color_label.setStyleSheet("color: #333; min-width: 80px;")
        bg_color_layout.addWidget(bg_color_label)
        
        self.bg_color_btn = QPushButton()
        self.bg_color_btn.setFixedSize(80, 30)
        self.bg_color_btn.clicked.connect(self.choose_bg_color)
        self.current_bg_color = current_bg_color
        self.update_color_button(self.bg_color_btn, self.current_bg_color)
        bg_color_layout.addWidget(self.bg_color_btn)
        
        bg_color_layout.addStretch()
        
        # é‡ç½®ä¸ºé»˜è®¤ç™½è‰²æŒ‰é’®
        reset_bg_btn = QPushButton("é‡ç½®ä¸ºç™½è‰²")
        reset_bg_btn.setFixedSize(80, 30)
        reset_bg_btn.clicked.connect(lambda: self.reset_color('bg'))
        reset_bg_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        bg_color_layout.addWidget(reset_bg_btn)
        
        color_layout.addWidget(bg_color_container)
        
        scroll_layout.addWidget(color_group)
        
        # é¢„è§ˆåŒºåŸŸ
        preview_group = QGroupBox("é¢„è§ˆæ•ˆæœ")
        preview_group.setStyleSheet("QGroupBox { font-weight: bold; color: #333; }")
        preview_layout = QVBoxLayout(preview_group)
        
        self.preview_text = QLabel("è¿™æ˜¯å­—ä½“å’ŒèƒŒæ™¯è‰²çš„é¢„è§ˆæ•ˆæœ\næ”¯æŒä¸­æ–‡å’ŒEnglishæ··åˆæ˜¾ç¤º\næ•°å­—123456789")
        self.preview_text.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.preview_text.setMinimumHeight(80)
        self.preview_text.setWordWrap(True)
        self.update_preview()
        preview_layout.addWidget(self.preview_text)
        
        scroll_layout.addWidget(preview_group)
        
        scroll.setWidget(scroll_widget)
        layout.addWidget(scroll)
        
        # æŒ‰é’®åŒºåŸŸ
        button_layout = QHBoxLayout()
        
        # æ¢å¤é»˜è®¤æŒ‰é’®
        restore_btn = QPushButton("æ¢å¤é»˜è®¤")
        restore_btn.clicked.connect(self.restore_defaults)
        restore_btn.setStyleSheet("""
            QPushButton {
                background-color: #ffc107;
                color: #212529;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #e0a800;
            }
        """)
        button_layout.addWidget(restore_btn)
        
        button_layout.addStretch()
        
        # å–æ¶ˆå’Œç¡®å®šæŒ‰é’®
        cancel_btn = QPushButton("å–æ¶ˆ")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        button_layout.addWidget(cancel_btn)
        
        ok_btn = QPushButton("ç¡®å®š")
        ok_btn.clicked.connect(self.accept)
        ok_btn.setStyleSheet("""
            QPushButton {
                background-color: #007bff;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
        """)
        button_layout.addWidget(ok_btn)
        
        layout.addLayout(button_layout)
    
    def update_color_button(self, button, color):
        """æ›´æ–°é¢œè‰²æŒ‰é’®çš„æ˜¾ç¤º"""
        button.setText(color)
        button.setStyleSheet(f"""
            QPushButton {{
                background-color: {color};
                color: {'white' if self.is_dark_color(color) else 'black'};
                border: 2px solid #333;
                border-radius: 4px;
                font-weight: bold;
                font-size: 11px;
            }}
        """)
    
    def is_dark_color(self, hex_color):
        """åˆ¤æ–­é¢œè‰²æ˜¯å¦ä¸ºæ·±è‰²"""
        try:
            # ç§»é™¤#å·
            hex_color = hex_color.lstrip('#')
            # è½¬æ¢ä¸ºRGB
            r = int(hex_color[0:2], 16)
            g = int(hex_color[2:4], 16)
            b = int(hex_color[4:6], 16)
            # è®¡ç®—äº®åº¦
            brightness = (r * 0.299 + g * 0.587 + b * 0.114)
            return brightness < 128
        except:
            return False
    
    def choose_font_color(self):
        """é€‰æ‹©å­—ä½“é¢œè‰²"""
        color = QColorDialog.getColor(QColor(self.current_font_color), self, "é€‰æ‹©å­—ä½“é¢œè‰²")
        if color.isValid():
            color_hex = color.name()
            
            # é˜²å‘†æ£€æŸ¥ï¼šæ£€æµ‹äº®åº¦è¿‡é«˜çš„é¢œè‰²
            try:
                hex_color = color_hex.lstrip('#')
                r = int(hex_color[0:2], 16)
                g = int(hex_color[2:4], 16)
                b = int(hex_color[4:6], 16)
                brightness = (r * 0.299 + g * 0.587 + b * 0.114) / 255
                
                if brightness > 0.9:
                    reply = QMessageBox.question(
                        self,
                        "å­—ä½“é¢œè‰²è¿‡æµ…æé†’",
                        f"âš ï¸ æ‚¨é€‰æ‹©çš„é¢œè‰² {color_hex} è¿‡äºæµ…æ·¡ï¼ˆäº®åº¦{brightness:.1%}ï¼‰ï¼\n\n"
                        "åœ¨ç™½è‰²èƒŒæ™¯ä¸Šå¯èƒ½çœ‹ä¸æ¸…æ–‡å­—ã€‚\n\n"
                        "å»ºè®®é€‰æ‹©æ·±è‰²å­—ä½“ä»¥ç¡®ä¿è‰¯å¥½çš„å¯è¯»æ€§ã€‚\n\n"
                        "æ˜¯å¦ä»è¦ä½¿ç”¨è¿™ä¸ªé¢œè‰²ï¼Ÿ",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                        QMessageBox.StandardButton.No
                    )
                    
                    if reply == QMessageBox.StandardButton.No:
                        return  # å–æ¶ˆè®¾ç½®ï¼Œä¿æŒåŸé¢œè‰²
            except:
                pass  # å¦‚æœæ£€æŸ¥å¤±è´¥ï¼Œç»§ç»­ä½¿ç”¨ç”¨æˆ·é€‰æ‹©çš„é¢œè‰²
            
            self.current_font_color = color_hex
            self.update_color_button(self.font_color_btn, self.current_font_color)
            self.update_preview()
    
    def choose_bg_color(self):
        """é€‰æ‹©èƒŒæ™¯é¢œè‰²"""
        color = QColorDialog.getColor(QColor(self.current_bg_color), self, "é€‰æ‹©èƒŒæ™¯é¢œè‰²")
        if color.isValid():
            self.current_bg_color = color.name()
            self.update_color_button(self.bg_color_btn, self.current_bg_color)
            self.update_preview()
    
    def reset_color(self, color_type):
        """é‡ç½®é¢œè‰²ä¸ºé»˜è®¤å€¼"""
        if color_type == 'font':
            self.current_font_color = "#000000"  # é»‘è‰²
            self.update_color_button(self.font_color_btn, self.current_font_color)
        elif color_type == 'bg':
            self.current_bg_color = "#ffffff"  # ç™½è‰²
            self.update_color_button(self.bg_color_btn, self.current_bg_color)
        self.update_preview()
    
    def restore_defaults(self):
        """æ¢å¤æ‰€æœ‰é»˜è®¤è®¾ç½®"""
        self.scale_slider.setValue(100)
        self.current_font_color = "#000000"
        self.current_bg_color = "#ffffff"
        self.font_family_combo.setCurrentText("Microsoft YaHei")
        self.font_size_spin.setValue(12)
        self.font_bold_cb.setChecked(False)
        self.font_italic_cb.setChecked(False)
        self.font_underline_cb.setChecked(False)
        self.update_color_button(self.font_color_btn, self.current_font_color)
        self.update_color_button(self.bg_color_btn, self.current_bg_color)
        self.update_preview()
    
    def update_preview(self):
        """æ›´æ–°é¢„è§ˆæ•ˆæœ"""
        font_family = self.font_family_combo.currentText()
        font_size = self.font_size_spin.value()
        font_weight = "bold" if self.font_bold_cb.isChecked() else "normal"
        font_style = "italic" if self.font_italic_cb.isChecked() else "normal"
        text_decoration = "underline" if self.font_underline_cb.isChecked() else "none"
        
        self.preview_text.setStyleSheet(f"""
            color: {self.current_font_color};
            background-color: {self.current_bg_color};
            font-family: {font_family};
            font-size: {font_size}px;
            font-weight: {font_weight};
            font-style: {font_style};
            text-decoration: {text_decoration};
            padding: 15px;
            border: 1px solid #ccc;
            border-radius: 4px;
            line-height: 1.5;
        """)

    def update_scale_label(self, value):
        """æ›´æ–°ç¼©æ”¾æ ‡ç­¾"""
        self.scale_label.setText(f"å½“å‰ç¼©æ”¾: {value}%")

    def set_scale(self, value):
        """è®¾ç½®ç¼©æ”¾å€¼"""
        self.scale_slider.setValue(value)

    def get_scale(self):
        """è·å–ç¼©æ”¾å€¼"""
        return self.scale_slider.value()
    
    def get_font_color(self):
        """è·å–å­—ä½“é¢œè‰²"""
        return self.current_font_color
    
    def get_bg_color(self):
        """è·å–èƒŒæ™¯é¢œè‰²"""
        return self.current_bg_color
    
    def get_font_family(self):
        """è·å–å­—ä½“ç±»å‹"""
        return self.font_family_combo.currentText()
    
    def get_font_size(self):
        """è·å–å­—ä½“å¤§å°"""
        return self.font_size_spin.value()
    
    def get_font_bold(self):
        """è·å–å­—ä½“åŠ ç²—"""
        return self.font_bold_cb.isChecked()
    
    def get_font_italic(self):
        """è·å–å­—ä½“æ–œä½“"""
        return self.font_italic_cb.isChecked()
    
    def get_font_underline(self):
        """è·å–å­—ä½“ä¸‹åˆ’çº¿"""
        return self.font_underline_cb.isChecked()
    
    def get_use_system_theme(self):
        """è·å–ç³»ç»Ÿä¸»é¢˜è·Ÿéšè®¾ç½®"""
        return self.use_system_theme_cb.isChecked()
    
    def on_system_theme_toggled(self, checked):
        """ç³»ç»Ÿä¸»é¢˜é€‰é¡¹å˜åŒ–æ—¶è§¦å‘"""
        self.update_preview()

    def save_settings(self):
        """ä¿å­˜è®¾ç½®"""
        scale = self.get_scale()
        font_color = self.get_font_color()
        bg_color = self.get_bg_color()
        font_family = self.get_font_family()
        font_size = self.get_font_size()
        font_bold = self.get_font_bold()
        font_italic = self.get_font_italic()
        font_underline = self.get_font_underline()
        
        self.settings.setValue("display_scale", scale)
        self.settings.setValue("font_color", font_color)
        self.settings.setValue("background_color", bg_color)
        self.settings.setValue("font_family", font_family)
        self.settings.setValue("font_size", font_size)
        self.settings.setValue("font_bold", font_bold)
        self.settings.setValue("font_italic", font_italic)
        self.settings.setValue("font_underline", font_underline)
        self.settings.setValue("use_system_theme", self.get_use_system_theme())
        self.settings.sync()
        
        return scale, font_color, bg_color, font_family, font_size, font_bold, font_italic, font_underline


# FarmIdUnificationDialog class removed - no longer needed for single-farm uploads
    


# BatchFarmIdInputDialog class removed - no longer needed for single-farm uploads


class EnhancedProgressDialog(QProgressDialog):
    """å¢å¼ºç‰ˆè¿›åº¦æ¡å¯¹è¯æ¡†ï¼Œæ”¯æŒå¹³æ»‘åŠ¨ç”»å’Œå‰©ä½™æ—¶é—´ä¼°ç®—"""
    
    def __init__(self, title, cancel_text, min_val, max_val, parent=None):
        super().__init__(title, cancel_text, min_val, max_val, parent)
        self.setWindowTitle("å¤„ç†ä¸­")
        self.setWindowModality(Qt.WindowModality.WindowModal)
        self.setAutoClose(True)
        self.setAutoReset(True)
        
        # åˆå§‹åŒ–æ—¶é—´è·Ÿè¸ª
        self.start_time = QDateTime.currentDateTime()
        self.last_update_time = self.start_time
        self.progress_history = []  # å­˜å‚¨(æ—¶é—´, è¿›åº¦)å…ƒç»„ç”¨äºè®¡ç®—é€Ÿåº¦
        
        # åˆ›å»ºè‡ªå®šä¹‰è¿›åº¦æ¡
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setBar(self.progress_bar)
        
        # è®¾ç½®æ ·å¼
        self.setStyleSheet("""
            QProgressDialog {
                background-color: #f8f9fa;
                border: 2px solid #dee2e6;
                border-radius: 10px;
                padding: 20px;
            }
            QProgressBar {
                border: 2px solid #0d6efd;
                border-radius: 5px;
                text-align: center;
                font-weight: bold;
                background-color: #e9ecef;
                height: 25px;
            }
            QProgressBar::chunk {
                background: qlineargradient(x1: 0, y1: 0, x2: 1, y2: 0,
                    stop: 0 #0d6efd, stop: 0.5 #0a58ca, stop: 1 #0d6efd);
                border-radius: 3px;
            }
            QLabel {
                color: #212529;
                font-size: 14px;
                font-weight: 500;
                margin: 10px 0;
            }
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #bb2d3b;
            }
        """)
        
        # åˆ›å»ºåŠ¨ç”»
        self._animation_value = 0
        self.animation = QPropertyAnimation(self, b"animationValue")
        self.animation.setEasingCurve(QEasingCurve.Type.InOutCubic)
        self.animation.setDuration(500)  # 500msçš„å¹³æ»‘è¿‡æ¸¡
        
        # æ·»åŠ å‰©ä½™æ—¶é—´æ ‡ç­¾
        self.time_label = QLabel("è®¡ç®—å‰©ä½™æ—¶é—´...")
        self.setLabel(self.time_label)
        
        # åŠ¨ç”»è®¡æ—¶å™¨
        self.animation_timer = QTimer()
        self.animation_timer.timeout.connect(self.update_animation)
        self.animation_timer.start(30)  # 30msæ›´æ–°ä¸€æ¬¡ï¼Œç¡®ä¿æµç•…
        
        self.setMinimumWidth(400)
        self.setMinimumHeight(150)
    
    def get_animationValue(self):
        return self._animation_value
    
    def set_animationValue(self, value):
        self._animation_value = value
        self.setValue(int(value))
    
    animationValue = pyqtProperty(float, get_animationValue, set_animationValue)
    
    def set_smooth_value(self, value):
        """å¹³æ»‘åœ°è®¾ç½®è¿›åº¦å€¼"""
        # è®°å½•è¿›åº¦å†å²
        current_time = QDateTime.currentDateTime()
        self.progress_history.append((current_time, value))
        
        # åªä¿ç•™æœ€è¿‘10ä¸ªè®°å½•
        if len(self.progress_history) > 10:
            self.progress_history.pop(0)
        
        # æ›´æ–°å‰©ä½™æ—¶é—´ä¼°ç®—
        self.update_time_estimate(value)
        
        # å¯åŠ¨å¹³æ»‘åŠ¨ç”»
        self.animation.stop()
        self.animation.setStartValue(self.value())
        self.animation.setEndValue(value)
        self.animation.start()
    
    def update_time_estimate(self, current_value):
        """æ›´æ–°å‰©ä½™æ—¶é—´ä¼°ç®—"""
        if current_value <= 0 or current_value >= self.maximum():
            self.time_label.setText("å¤„ç†å®Œæˆ" if current_value >= self.maximum() else "å‡†å¤‡ä¸­...")
            return
        
        # è®¡ç®—å·²ç”¨æ—¶é—´
        elapsed_ms = self.start_time.msecsTo(QDateTime.currentDateTime())
        
        # å¦‚æœæœ‰è¶³å¤Ÿçš„å†å²æ•°æ®ï¼Œè®¡ç®—å¹³å‡é€Ÿåº¦
        if len(self.progress_history) >= 2:
            # ä½¿ç”¨æœ€è¿‘çš„æ•°æ®è®¡ç®—é€Ÿåº¦
            recent_time, recent_progress = self.progress_history[-1]
            old_time, old_progress = self.progress_history[0]
            
            time_diff_ms = old_time.msecsTo(recent_time)
            progress_diff = recent_progress - old_progress
            
            if progress_diff > 0 and time_diff_ms > 0:
                # è®¡ç®—é€Ÿåº¦ï¼ˆè¿›åº¦/æ¯«ç§’ï¼‰
                speed = progress_diff / time_diff_ms
                
                # è®¡ç®—å‰©ä½™è¿›åº¦
                remaining_progress = self.maximum() - current_value
                
                # ä¼°ç®—å‰©ä½™æ—¶é—´
                remaining_ms = int(remaining_progress / speed)
                
                # æ ¼å¼åŒ–æ—¶é—´æ˜¾ç¤º
                remaining_text = self.format_time(remaining_ms)
                elapsed_text = self.format_time(elapsed_ms)
                
                self.time_label.setText(
                    f"å·²ç”¨æ—¶é—´: {elapsed_text} | é¢„è®¡å‰©ä½™: {remaining_text}"
                )
            else:
                self.time_label.setText(f"å·²ç”¨æ—¶é—´: {self.format_time(elapsed_ms)}")
        else:
            self.time_label.setText(f"å·²ç”¨æ—¶é—´: {self.format_time(elapsed_ms)}")
    
    def format_time(self, milliseconds):
        """æ ¼å¼åŒ–æ—¶é—´æ˜¾ç¤º"""
        seconds = milliseconds // 1000
        
        if seconds < 60:
            return f"{seconds}ç§’"
        elif seconds < 3600:
            minutes = seconds // 60
            secs = seconds % 60
            return f"{minutes}åˆ†{secs}ç§’"
        else:
            hours = seconds // 3600
            minutes = (seconds % 3600) // 60
            return f"{hours}å°æ—¶{minutes}åˆ†"
    
    def update_animation(self):
        """æ›´æ–°åŠ¨ç”»æ•ˆæœ"""
        # å¯ä»¥åœ¨è¿™é‡Œæ·»åŠ é¢å¤–çš„è§†è§‰æ•ˆæœ
        pass
    
    def set_label_text(self, text):
        """è®¾ç½®æ ‡ç­¾æ–‡æœ¬åŒæ—¶ä¿ç•™æ—¶é—´ä¿¡æ¯"""
        current_time_info = self.time_label.text()
        if "å·²ç”¨æ—¶é—´" in current_time_info:
            self.time_label.setText(f"{text}\n{current_time_info}")
        else:
            self.time_label.setText(text)


class FileProcessThread(QThread):
    """æ–‡ä»¶å¤„ç†çº¿ç¨‹"""
    progress_updated = pyqtSignal(str, int)  # çŠ¶æ€ä¿¡æ¯, è¿›åº¦ç™¾åˆ†æ¯”
    file_processed = pyqtSignal(str, bool, str, dict)  # æ–‡ä»¶å, æˆåŠŸ, æ¶ˆæ¯, æ•°æ®ä¿¡æ¯
    processing_completed = pyqtSignal(dict)  # å®Œæˆä¿¡æ¯
    log_updated = pyqtSignal(str)  # å¤„ç†è¿‡ç¨‹æ—¥å¿—
    
    def __init__(self, file_paths, filenames, urea_tracker=None):
        super().__init__()
        self.file_paths = file_paths
        self.filenames = filenames
        self.processor = DataProcessor()
        self.urea_tracker = urea_tracker
        
    def run(self):
        """è¿è¡Œæ–‡ä»¶å¤„ç†"""
        try:
            from datetime import datetime
            
            total_files = len(self.filenames)
            self.log_updated.emit(f"ğŸ“‚ å¼€å§‹å¤„ç† {total_files} ä¸ªæ–‡ä»¶")
            self.log_updated.emit(f"â° å¤„ç†å¼€å§‹æ—¶é—´: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            self.progress_updated.emit("å¼€å§‹å¤„ç†æ–‡ä»¶...", 5)
            
            # é€ä¸ªå¤„ç†æ–‡ä»¶ï¼Œæä¾›æ›´è¯¦ç»†çš„è¿›åº¦
            success_files = []
            failed_files = []
            all_data = []
            farm_ids = set()
            
            for i, (file_path, filename) in enumerate(zip(self.file_paths, self.filenames)):
                base_progress = 10 + int((i / total_files) * 70)  # 10-80% for file processing
                
                self.log_updated.emit(f"\nğŸ“„ æ­£åœ¨å¤„ç†æ–‡ä»¶ {i+1}/{total_files}: {filename}")
                self.progress_updated.emit(f"å¤„ç†æ–‡ä»¶ {i+1}/{total_files}: {filename}", base_progress)
                
                # ç»†åˆ†è¿›åº¦ï¼šè¯»å–æ–‡ä»¶
                self.progress_updated.emit(f"æ­£åœ¨è¯»å–: {filename}", base_progress + 2)
                
                try:
                    # å¦‚æœæ–‡ä»¶å¾ˆå¤§ï¼Œæ·»åŠ æ›´å¤šè¿›åº¦æ›´æ–°
                    file_size = os.path.getsize(file_path) / (1024 * 1024)  # MB
                    if file_size > 10:
                        self.progress_updated.emit(f"æ­£åœ¨è§£æå¤§æ–‡ä»¶: {filename} ({file_size:.1f}MB)", base_progress + 5)
                    
                    success, message, df = self.processor.process_uploaded_file(file_path, filename)
                    
                    if success and df is not None:
                        # è·å–æ•°æ®ä¿¡æ¯
                        row_count = len(df)
                        date_range = self.processor.extract_date_range_from_data(df)
                        
                        # æå–ç‰›åœºç¼–å·
                        if 'farm_id' in df.columns:
                            file_farm_ids = df['farm_id'].dropna().unique()
                            farm_ids.update(file_farm_ids)
                            self.log_updated.emit(f"   âœ… æˆåŠŸ: {row_count}è¡Œæ•°æ®ï¼Œç‰›åœºç¼–å·: {list(file_farm_ids)}")
                        else:
                            self.log_updated.emit(f"   âœ… æˆåŠŸ: {row_count}è¡Œæ•°æ®ï¼Œç¼ºå°‘ç‰›åœºç¼–å·")
                        
                        if date_range:
                            self.log_updated.emit(f"   ğŸ“… æ•°æ®æ—¶é—´èŒƒå›´: {date_range}")
                        
                        success_files.append({
                            'filename': filename,
                            'message': message,
                            'row_count': row_count,
                            'date_range': date_range
                        })
                        
                        all_data.append({
                            'filename': filename,
                            'data': df
                        })
                        
                        # æ·»åŠ æ•°æ®åˆ°å°¿ç´ æ°®è¿½è¸ªå™¨
                        if self.urea_tracker and date_range:
                            # ä»æ—¥æœŸèŒƒå›´ä¸­æå–å¹´æœˆ
                            try:
                                # date_range['year_month_range'] æ ¼å¼å¦‚ "2024å¹´1æœˆ - 2024å¹´3æœˆ"
                                year_month_str = date_range.get('year_month_range', '')
                                if ' - ' in year_month_str:
                                    # å–ç¬¬ä¸€ä¸ªæœˆä»½ä½œä¸ºæ•°æ®çš„ä»£è¡¨æœˆä»½
                                    first_month = year_month_str.split(' - ')[0]
                                    # è½¬æ¢ä¸º YYYY-MM æ ¼å¼
                                    import re
                                    match = re.match(r'(\d{4})å¹´(\d{1,2})æœˆ', first_month)
                                    if match:
                                        year = match.group(1)
                                        month = match.group(2).zfill(2)
                                        date_str = f"{year}-{month}"
                                        self.urea_tracker.add_dhi_data(df, date_str)
                                        self.log_updated.emit(f"   ğŸ§ª å·²æ·»åŠ åˆ°å°¿ç´ æ°®è¿½è¸ª: {date_str}")
                                else:
                                    # å•æœˆæ•°æ®
                                    match = re.match(r'(\d{4})å¹´(\d{1,2})æœˆ', year_month_str)
                                    if match:
                                        year = match.group(1)
                                        month = match.group(2).zfill(2)
                                        date_str = f"{year}-{month}"
                                        self.urea_tracker.add_dhi_data(df, date_str)
                                        self.log_updated.emit(f"   ğŸ§ª å·²æ·»åŠ åˆ°å°¿ç´ æ°®è¿½è¸ª: {date_str}")
                            except Exception as e:
                                self.log_updated.emit(f"   âš ï¸ å°¿ç´ æ°®æ•°æ®æ·»åŠ å¤±è´¥: {str(e)}")
                    else:
                        self.log_updated.emit(f"   âŒ å¤±è´¥: {message}")
                        failed_files.append({
                            'filename': filename,
                            'error': message
                        })
                    
                    # å‘é€å•ä¸ªæ–‡ä»¶å¤„ç†ç»“æœ
                    file_info = {
                        'filename': filename,
                        'row_count': len(df) if df is not None else 0,
                        'date_range': date_range if success else None
                    }
                    
                    if success and df is not None and hasattr(df, 'attrs') and 'missing_farm_id_info' in df.attrs:
                        file_info['missing_farm_id_info'] = df.attrs['missing_farm_id_info']
                    
                    self.file_processed.emit(filename, success, message, file_info)
                    
                except Exception as e:
                    error_msg = f"å¤„ç†å¤±è´¥: {str(e)}"
                    self.log_updated.emit(f"   âŒ å¼‚å¸¸: {error_msg}")
                    failed_files.append({
                        'filename': filename,
                        'error': error_msg
                    })
                    self.file_processed.emit(filename, False, error_msg, {})
            
            self.progress_updated.emit("æ±‡æ€»å¤„ç†ç»“æœ...", 85)
                    
                    # æ”¶é›†ç¼ºå°‘ç®¡ç†å·çš„æ–‡ä»¶ä¿¡æ¯
            missing_farm_id_files = []
            for data_item in all_data:
                df = data_item['data']
                if hasattr(df, 'attrs') and 'missing_farm_id_info' in df.attrs:
                    missing_info = df.attrs['missing_farm_id_info']
                    source_info = self._get_source_info(data_item['filename'])
                    missing_farm_id_files.append({
                        'filename': data_item['filename'],
                        'missing_info': missing_info,
                        'source_info': source_info,
                        'data': df
                    })
            
            # æ±‡æ€»ç»“æœ
            results = {
                'success_files': success_files,
                'failed_files': failed_files,
                'all_data': all_data,
                'farm_ids': sorted(list(farm_ids)),
                'missing_farm_id_files': missing_farm_id_files
            }
            
            self.progress_updated.emit("æ–‡ä»¶å¤„ç†å®Œæˆ", 100)
            self.log_updated.emit(f"\nğŸ“Š å¤„ç†å®Œæˆç»Ÿè®¡:")
            self.log_updated.emit(f"   âœ… æˆåŠŸ: {len(success_files)} ä¸ªæ–‡ä»¶")
            self.log_updated.emit(f"   âŒ å¤±è´¥: {len(failed_files)} ä¸ªæ–‡ä»¶")
            self.log_updated.emit(f"   ğŸ¢ å‘ç°ç‰›åœº: {len(farm_ids)} ä¸ª")
            if missing_farm_id_files:
                self.log_updated.emit(f"   âš ï¸ ç¼ºå°‘ç®¡ç†å·: {len(missing_farm_id_files)} ä¸ªæ–‡ä»¶")
            self.log_updated.emit(f"â° å¤„ç†å®Œæˆæ—¶é—´: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            self.processing_completed.emit(results)
            
        except Exception as e:
            error_msg = f"å¤„ç†è¿‡ç¨‹å‡ºç°ä¸¥é‡é”™è¯¯: {str(e)}"
            self.log_updated.emit(f"\nâŒ {error_msg}")
            import traceback
            self.log_updated.emit(f"é”™è¯¯è¯¦æƒ…:\n{traceback.format_exc()}")
            
            # å‘é€é”™è¯¯ä¿¡å·ç»™æ‰€æœ‰æ–‡ä»¶
            for filename in self.filenames:
                self.file_processed.emit(filename, False, error_msg, {})

    def _get_source_info(self, filename):
        """è·å–æ–‡ä»¶æ¥æºä¿¡æ¯"""
        # æ£€æŸ¥æ˜¯å¦æ¥è‡ªå‹ç¼©åŒ…
        for file_path in self.file_paths:
            if filename in file_path:
                if '.zip' in file_path.lower():
                    # ä»å‹ç¼©åŒ…ä¸­æå–çš„æ–‡ä»¶
                    zip_name = file_path.split('/')[-1] if '/' in file_path else file_path.split('\\')[-1]
                    return f"å‹ç¼©åŒ…: {zip_name}"
                else:
                    # å•ç‹¬ä¸Šä¼ çš„æ–‡ä»¶
                    return "å•ç‹¬ä¸Šä¼ "
        
        return "å•ç‹¬ä¸Šä¼ "


class FilterThread(QThread):
    """ç­›é€‰å¤„ç†çº¿ç¨‹"""
    progress_updated = pyqtSignal(str, int)
    filtering_completed = pyqtSignal(bool, str, pd.DataFrame, dict)  # æ·»åŠ ç»Ÿè®¡ä¿¡æ¯å­—å…¸
    log_updated = pyqtSignal(str)  # ç­›é€‰è¿‡ç¨‹æ—¥å¿—
    
    def __init__(self, data_list, filters, selected_files, processor=None, urea_tracker=None):
        super().__init__()
        self.data_list = data_list
        self.filters = filters
        self.selected_files = selected_files
        self.processor = processor if processor else DataProcessor()
        self.urea_tracker = urea_tracker
        self._should_stop = False  # åœæ­¢æ ‡å¿—
    
    def stop(self):
        """åœæ­¢ç­›é€‰"""
        self._should_stop = True
        self.log_updated.emit("â¹ï¸ ç”¨æˆ·è¯·æ±‚åœæ­¢ç­›é€‰...")
    
    def request_cancel(self):
        """è¯·æ±‚å–æ¶ˆï¼ˆåˆ«åï¼‰"""
        self.stop()
    
    def should_stop(self):
        """æ£€æŸ¥æ˜¯å¦åº”è¯¥åœæ­¢"""
        return self._should_stop
    
    def run(self):
        """æ‰§è¡Œç­›é€‰"""
        try:
            from datetime import datetime, timedelta
            
            self.log_updated.emit(f"\nğŸ” å¼€å§‹ç­›é€‰æ•°æ®")
            self.log_updated.emit(f"â° ç­›é€‰å¼€å§‹æ—¶é—´: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            self.progress_updated.emit("å¼€å§‹ç­›é€‰...", 5)
            
            # ç»Ÿè®¡å¯ç”¨çš„ç­›é€‰é¡¹
            enabled_filters = []
            for filter_name, filter_config in self.filters.items():
                if filter_config.get('enabled', False) and filter_name not in ['parity', 'date_range']:
                    enabled_filters.append(filter_name)
            
            self.log_updated.emit(f"ğŸ“‹ å¯ç”¨çš„ç­›é€‰é¡¹: {enabled_filters if enabled_filters else 'ä»…åŸºç¡€ç­›é€‰'}")
            
            self.progress_updated.emit("ç»Ÿè®¡æ•°æ®è§„æ¨¡...", 10)
            
            # è®¡ç®—å…¨éƒ¨æ•°æ®çš„ç‰›å¤´æ•°
            all_cows = set()
            for item in self.data_list:
                df = item['data']
                if 'management_id' in df.columns:
                    cow_ids = df['management_id'].dropna().unique()
                    for cow_id in cow_ids:
                        all_cows.add(cow_id)
            
            self.log_updated.emit(f"ğŸ“Š å…¨éƒ¨æ•°æ®: {len(all_cows)} å¤´ç‰›")
            
            # è®¡ç®—ç­›é€‰èŒƒå›´çš„ç‰›å¤´æ•°
            range_cows = set()
            selected_data = [item for item in self.data_list if item['filename'] in self.selected_files]
            for item in selected_data:
                df = item['data']
                if 'management_id' in df.columns:
                    cow_ids = df['management_id'].dropna().unique()
                    for cow_id in cow_ids:
                        range_cows.add(cow_id)
            
            self.log_updated.emit(f"ğŸ“Š ç­›é€‰èŒƒå›´: {len(range_cows)} å¤´ç‰› (æ¥è‡ª{len(self.selected_files)}ä¸ªæ–‡ä»¶)")
            
            self.progress_updated.emit("åº”ç”¨ç­›é€‰æ¡ä»¶...", 25)
            
            # ä½¿ç”¨æ–°çš„å¤šç­›é€‰é¡¹é€»è¾‘
            self.log_updated.emit(f"ğŸ”§ å¼€å§‹åº”ç”¨å¤šç­›é€‰é¡¹é€»è¾‘...")
            
            def progress_callback(message, progress):
                """è¿›åº¦å›è°ƒå‡½æ•°"""
                self.progress_updated.emit(message, progress)
                self.log_updated.emit(f"   {message}")
            
            filtered_df = self.processor.apply_multi_filter_logic(
                self.data_list, self.filters, self.selected_files,
                progress_callback=progress_callback, should_stop=self.should_stop
            )
            
            # æ£€æŸ¥æ˜¯å¦è¢«åœæ­¢
            if self._should_stop:
                self.log_updated.emit("âŒ ç­›é€‰å·²è¢«ç”¨æˆ·å–æ¶ˆ")
                self.filtering_completed.emit(False, "ç­›é€‰å·²è¢«ç”¨æˆ·å–æ¶ˆ", pd.DataFrame(), {})
                return
            
            basic_filter_count = len(filtered_df)
            self.log_updated.emit(f"ğŸ“Š åŸºç¡€ç­›é€‰å: {basic_filter_count} æ¡è®°å½•")
            
            self.progress_updated.emit("ç”Ÿæˆæœˆåº¦æŠ¥å‘Š...", 50)
            
            # åŠ¨æ€æ„å»ºdisplay_fieldsï¼ŒåŒ…å«æ‰€æœ‰å¯ç”¨çš„ç­›é€‰é¡¹
            display_fields = ['management_id', 'parity']
            
            # æ·»åŠ å¯ç”¨çš„ç­›é€‰é¡¹åˆ°display_fields
            # å®šä¹‰æ‰€æœ‰æ”¯æŒçš„å­—æ®µ
            supported_fields = [
                'protein_pct', 'fat_pct', 'fat_protein_ratio', 'somatic_cell_count', 
                'somatic_cell_score', 'urea_nitrogen', 'lactose_pct', 'milk_loss',
                'milk_payment_diff', 'economic_loss', 'corrected_milk', 'persistency',
                'whi', 'fore_milk_yield', 'fore_somatic_cell_count', 'fore_somatic_cell_score',
                'fore_milk_loss', 'peak_milk_yield', 'peak_days', 'milk_305',
                'total_milk_yield', 'total_fat_pct', 'total_protein_pct', 'mature_equivalent'
            ]
            
            for filter_name in enabled_filters:
                if filter_name in supported_fields:
                    display_fields.append(filter_name)
            
            # ç¡®ä¿åŒ…å«å¿…è¦çš„å­—æ®µ - ä»»ä½•æ€§çŠ¶ç­›é€‰éƒ½éœ€è¦æ³Œä¹³å¤©æ•°å’Œäº§å¥¶é‡
            if 'lactation_days' not in display_fields:
                display_fields.append('lactation_days')
            
            if 'milk_yield' not in display_fields:
                display_fields.append('milk_yield')
            
            self.log_updated.emit(f"ğŸ“‹ ç”Ÿæˆæœˆåº¦æŠ¥å‘Šï¼ŒåŒ…å«å­—æ®µ: {display_fields}")
            
            # æ€»æ˜¯è·å–è®¡åˆ’è°ƒç¾¤æ—¥æœŸï¼Œç”¨äºè®¡ç®—æœªæ¥æ³Œä¹³å¤©æ•°ï¼ˆæ— è®ºæ˜¯å¦å¯ç”¨ç­›é€‰ï¼‰
            plan_date = None
            if 'future_lactation_days' in self.filters:
                plan_date = self.filters['future_lactation_days'].get('plan_date')
            
            # å¦‚æœæœªè®¾ç½®ï¼Œä½¿ç”¨é»˜è®¤å€¼ï¼ˆå½“å‰æ—¥æœŸ+30å¤©ï¼‰
            if not plan_date:
                default_plan_date = datetime.now() + timedelta(days=30)
                plan_date = default_plan_date.strftime('%Y-%m-%d')
            
            self.log_updated.emit(f"ğŸ“… è®¡åˆ’è°ƒç¾¤æ—¥æœŸ: {plan_date}")
            
            monthly_report = self.processor.create_monthly_report(filtered_df, display_fields, plan_date)
            
            self.log_updated.emit(f"ğŸ“Š æœˆåº¦æŠ¥å‘Šç”Ÿæˆ: {len(monthly_report)} æ¡è®°å½•")
            
                        # å¦‚æœå¯ç”¨äº†æœªæ¥æ³Œä¹³å¤©æ•°ç­›é€‰ï¼Œå¯¹æœˆåº¦æŠ¥å‘Šè¿›è¡Œæœ€åçš„ç­›é€‰
            if 'future_lactation_days' in self.filters and self.filters['future_lactation_days'].get('enabled', False):
                self.progress_updated.emit("åº”ç”¨æœªæ¥æ³Œä¹³å¤©æ•°ç­›é€‰...", 75)
                future_filter = self.filters['future_lactation_days']
                before_future_count = len(monthly_report)
                
                self.log_updated.emit(f"ğŸ”® åº”ç”¨æœªæ¥æ³Œä¹³å¤©æ•°ç­›é€‰: {future_filter['min']}-{future_filter['max']}å¤©")
                
                # å¯¹æœˆåº¦æŠ¥å‘Šè¿›è¡Œæœªæ¥æ³Œä¹³å¤©æ•°ç­›é€‰
                if not monthly_report.empty and 'æœªæ¥æ³Œä¹³å¤©æ•°(å¤©)' in monthly_report.columns:
                    monthly_report = self.processor.apply_numeric_filter(monthly_report, 'æœªæ¥æ³Œä¹³å¤©æ•°(å¤©)', future_filter)
                    
                after_future_count = len(monthly_report)
                self.log_updated.emit(f"ğŸ“Š æœªæ¥æ³Œä¹³å¤©æ•°ç­›é€‰å: {after_future_count} æ¡è®°å½• (ç­›é™¤{before_future_count - after_future_count}æ¡)")
            
            # åº”ç”¨åœ¨ç¾¤ç‰›ç­›é€‰ï¼ˆæœ€åä¸€æ­¥ï¼‰
            if self.processor.active_cattle_enabled:
                self.progress_updated.emit("åº”ç”¨åœ¨ç¾¤ç‰›ç­›é€‰...", 90)
                before_active_count = len(monthly_report)
                
                cattle_count = len(self.processor.active_cattle_list) if self.processor.active_cattle_list else 0
                self.log_updated.emit(f"ğŸ„ åº”ç”¨åœ¨ç¾¤ç‰›ç­›é€‰ (åœ¨ç¾¤ç‰›æ¸…å•: {cattle_count}å¤´)")
                monthly_report = self.processor.apply_active_cattle_filter(monthly_report)
                
                after_active_count = len(monthly_report)
                self.log_updated.emit(f"ğŸ“Š åœ¨ç¾¤ç‰›ç­›é€‰å: {after_active_count} æ¡è®°å½• (ç­›é™¤{before_active_count - after_active_count}æ¡)")
            
            # è®¡ç®—ç­›é€‰ç»“æœçš„ç‰›å¤´æ•°
            result_cows = set()
            if not monthly_report.empty and 'management_id' in monthly_report.columns:
                cow_ids = monthly_report['management_id'].dropna().unique()
                for cow_id in cow_ids:
                    result_cows.add(cow_id)
            
            # è®¡ç®—ç­›é€‰ç‡
            filter_rate = (len(result_cows) / len(all_cows) * 100) if len(all_cows) > 0 else 0
            
            # æ„å»ºç»Ÿè®¡ä¿¡æ¯
            stats = {
                'total_cows': len(all_cows),
                'range_cows': len(range_cows), 
                'result_cows': len(result_cows),
                'filter_rate': filter_rate,
                'active_cattle_enabled': self.processor.active_cattle_enabled,
                'active_cattle_count': len(self.processor.active_cattle_list) if self.processor.active_cattle_list else 0
            }
            
            # å°¿ç´ æ°®è¿½è¸ªåˆ†æ
            urea_tracking_config = self.filters.get('urea_tracking', {})
            if urea_tracking_config.get('enabled', False) and self.urea_tracker:
                self.progress_updated.emit("æ‰§è¡Œå°¿ç´ æ°®è¿½è¸ªåˆ†æ...", 95)
                self.log_updated.emit("\nğŸ§ª æ‰§è¡Œå°¿ç´ æ°®è¿½è¸ªåˆ†æ...")
                
                urea_results = self.urea_tracker.analyze(
                    selected_groups=urea_tracking_config['selected_groups'],
                    filter_outliers=urea_tracking_config['filter_outliers'],
                    min_value=urea_tracking_config['min_value'],
                    max_value=urea_tracking_config['max_value'],
                    min_sample_size=urea_tracking_config['min_sample_size']
                )
                
                if 'error' not in urea_results:
                    stats['urea_tracking'] = {
                        'results': urea_results,
                        'value_type': urea_tracking_config['value_type']
                    }
                    group_count = len(urea_results)
                    total_history_points = sum(len(group['history']) for group in urea_results.values())
                    self.log_updated.emit(f"   âœ… å°¿ç´ æ°®åˆ†æå®Œæˆ: {group_count}ä¸ªç»„ï¼Œ{total_history_points}ä¸ªå†å²æ•°æ®ç‚¹")
                else:
                    self.log_updated.emit(f"   âŒ å°¿ç´ æ°®åˆ†æå¤±è´¥: {urea_results['error']}")
            
            self.progress_updated.emit("ç­›é€‰å®Œæˆ", 100)
            
            self.log_updated.emit(f"\nâœ… ç­›é€‰å®Œæˆç»Ÿè®¡:")
            self.log_updated.emit(f"   ğŸ“Š æœ€ç»ˆç»“æœ: {len(result_cows)} å¤´ç‰›ï¼Œ{len(monthly_report)} æ¡è®°å½•")
            self.log_updated.emit(f"   ğŸ“ˆ ç­›é€‰ç‡: {filter_rate:.1f}%")
            if self.processor.active_cattle_enabled:
                self.log_updated.emit(f"   ğŸ„ å·²åº”ç”¨åœ¨ç¾¤ç‰›ç­›é€‰")
            self.log_updated.emit(f"â° ç­›é€‰å®Œæˆæ—¶é—´: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            success_msg = f"ç­›é€‰å®Œæˆï¼Œå…±ç­›é€‰å‡º {len(result_cows)} å¤´ç‰›ï¼Œ{len(monthly_report)} æ¡è®°å½•"
            if self.processor.active_cattle_enabled:
                success_msg += f"ï¼ˆå·²åº”ç”¨åœ¨ç¾¤ç‰›ç­›é€‰ï¼‰"
            
            self.filtering_completed.emit(True, success_msg, monthly_report, stats)
            
        except Exception as e:
            error_msg = f"ç­›é€‰å¤±è´¥: {str(e)}"
            self.log_updated.emit(f"\nâŒ {error_msg}")
            import traceback
            self.log_updated.emit(f"é”™è¯¯è¯¦æƒ…:\n{traceback.format_exc()}")
            self.filtering_completed.emit(False, error_msg, pd.DataFrame(), {})


class MainWindow(QMainWindow):
    """ä¸»çª—å£"""
    
    def __init__(self, username=None, auth_service=None):
        super().__init__()
        self.username = username or "æœªç™»å½•ç”¨æˆ·"
        self.auth_service = auth_service
        self.data_list = []  # å­˜å‚¨æ‰€æœ‰å¤„ç†è¿‡çš„æ•°æ®
        self.processor = DataProcessor()
        self.data_processor = self.processor  # ä¸ºæ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥åŠŸèƒ½æä¾›åˆ«å
        self.current_results = pd.DataFrame()  # å½“å‰ç­›é€‰ç»“æœ
        self.heartbeat_timer = None  # å¿ƒè·³å®šæ—¶å™¨
        
        # åˆå§‹åŒ–å°¿ç´ æ°®è¿½è¸ªå™¨
        from urea_tracker import UreaTracker
        self.urea_tracker = UreaTracker()
        self.urea_tracking_results = None
        
        # åŠ è½½æ˜¾ç¤ºè®¾ç½®
        self.settings = QSettings("DHI", "ProteinScreening")
        self.display_scale = self.settings.value("display_scale", 100, type=int)
        
        # é˜²å‘†è®¾è®¡ï¼šæ£€æŸ¥å­—ä½“é¢œè‰²æ˜¯å¦è¿‡æµ…ï¼Œè‡ªåŠ¨ä¿®æ­£
        raw_font_color = self.settings.value("font_color", "#000000", type=str)
        self.font_color = self.validate_and_fix_font_color(raw_font_color)
        
        self.background_color = self.settings.value("background_color", "#ffffff", type=str)
        self.font_family = self.settings.value("font_family", "Microsoft YaHei", type=str)
        self.font_size = self.settings.value("font_size", 12, type=int)
        self.font_bold = self.settings.value("font_bold", False, type=bool)
        self.font_italic = self.settings.value("font_italic", False, type=bool)
        self.font_underline = self.settings.value("font_underline", False, type=bool)
        
        # åˆå§‹åŒ–ç­›é€‰ç›¸å…³å˜é‡
        self.added_other_filters = {}  # å­˜å‚¨æ·»åŠ çš„å…¶ä»–ç­›é€‰é¡¹
        self.dhi_processed_ok = False  # åŸºç¡€æ•°æ®æ˜¯å¦å·²å¤„ç†å®Œæ¯•æ ‡å¿—
        
        self.init_ui()
        self.load_config()
        
        # å¯åŠ¨æ—¶æ£€æŸ¥æ˜¯å¦æœ‰æ˜¾ç¤ºé—®é¢˜ï¼ˆé˜²å‘†åŠŸèƒ½ï¼‰
        QTimer.singleShot(1000, self.check_display_issues_on_startup)
        
        # å¯åŠ¨å¿ƒè·³æœºåˆ¶
        if self.auth_service:
            self.start_heartbeat()
    
    def validate_and_fix_font_color(self, color_str: str) -> str:
        """é˜²å‘†è®¾è®¡ï¼šéªŒè¯å¹¶ä¿®æ­£å­—ä½“é¢œè‰²ï¼Œé˜²æ­¢è®¾ç½®è¿‡æµ…çš„é¢œè‰²å¯¼è‡´æ–‡å­—ä¸å¯è§"""
        try:
            # ç§»é™¤#å·
            hex_color = color_str.lstrip('#')
            if len(hex_color) != 6:
                return "#000000"  # æ— æ•ˆæ ¼å¼ï¼Œè¿”å›é»‘è‰²
            
            # è½¬æ¢ä¸ºRGB
            r = int(hex_color[0:2], 16)
            g = int(hex_color[2:4], 16) 
            b = int(hex_color[4:6], 16)
            
            # è®¡ç®—äº®åº¦ï¼ˆä½¿ç”¨ç›¸å¯¹äº®åº¦å…¬å¼ï¼‰
            brightness = (r * 0.299 + g * 0.587 + b * 0.114) / 255
            
            # å¦‚æœäº®åº¦è¿‡é«˜ï¼ˆ> 0.9ï¼‰ï¼Œå¼ºåˆ¶ä½¿ç”¨é»‘è‰²
            if brightness > 0.9:
                print(f"âš ï¸ é˜²å‘†æé†’ï¼šæ£€æµ‹åˆ°è¿‡æµ…çš„å­—ä½“é¢œè‰² {color_str}ï¼ˆäº®åº¦{brightness:.1%}ï¼‰ï¼Œå·²è‡ªåŠ¨ä¿®æ­£ä¸ºé»‘è‰²")
                # åŒæ—¶æ›´æ–°è®¾ç½®ä¸­çš„å€¼ï¼Œé¿å…ä¸‹æ¬¡å¯åŠ¨å†æ¬¡è§¦å‘
                self.settings.setValue("font_color", "#000000")
                return "#000000"
                
            # é¢œè‰²åˆé€‚ï¼Œè¿”å›åŸå€¼
            return color_str
            
        except Exception as e:
            print(f"âš ï¸ å­—ä½“é¢œè‰²éªŒè¯å‡ºé”™ {color_str}: {e}ï¼Œä½¿ç”¨é»˜è®¤é»‘è‰²")
            return "#000000"
    
    def get_safe_screen_info(self):
        """å®‰å…¨åœ°è·å–å±å¹•ä¿¡æ¯ - æ›´å‡†ç¡®çš„DPIé€‚é…"""
        screen = QApplication.primaryScreen()
        if screen is None:
            return {
                'width': 1920,
                'height': 1080,
                'dpi_ratio': 1.0,
                'logical_dpi': 96.0,
                'physical_dpi': 96.0,
                'scale_factor': 1.0
            }
        else:
            geometry = screen.availableGeometry()
            logical_dpi = screen.logicalDotsPerInch()
            physical_dpi = screen.physicalDotsPerInch()
            device_pixel_ratio = screen.devicePixelRatio()
            
            # è®¡ç®—ç³»ç»Ÿç¼©æ”¾æ¯”ä¾‹ - æ›´å‡†ç¡®çš„æ–¹æ³•
            system_scale_factor = logical_dpi / 96.0  # Windowsæ ‡å‡†DPI
            
            return {
                'width': geometry.width(),
                'height': geometry.height(),
                'dpi_ratio': device_pixel_ratio,
                'logical_dpi': logical_dpi,
                'physical_dpi': physical_dpi,
                'scale_factor': system_scale_factor
            }
    
    def safe_show_status_message(self, message: str):
        """å®‰å…¨åœ°æ˜¾ç¤ºçŠ¶æ€æ æ¶ˆæ¯"""
        status_bar = self.statusBar()
        if status_bar is not None:
            status_bar.showMessage(message)
    
    def get_dpi_scaled_size(self, base_size: int) -> int:
        """æ ¹æ®ç³»ç»ŸDPIè®¾ç½®è®¡ç®—é€‚é…åçš„å°ºå¯¸"""
        screen_info = self.get_safe_screen_info()
        
        # ä½¿ç”¨ç³»ç»Ÿç¼©æ”¾æ¯”ä¾‹å’Œç”¨æˆ·è‡ªå®šä¹‰ç¼©æ”¾çš„ç»„åˆ
        system_scale = screen_info['scale_factor']
        user_scale = self.display_scale / 100.0
        
        # æœ€ç»ˆç¼©æ”¾æ¯”ä¾‹ = ç³»ç»Ÿç¼©æ”¾ Ã— ç”¨æˆ·ç¼©æ”¾
        final_scale = system_scale * user_scale
        
        # åº”ç”¨ç¼©æ”¾å¹¶ç¡®ä¿æœ€å°å€¼
        scaled_size = max(int(base_size * final_scale), base_size // 2)
        
        return scaled_size
    
    def get_dpi_scaled_font_size(self, base_font_size: int) -> int:
        """æ ¹æ®ç³»ç»ŸDPIè®¾ç½®è®¡ç®—é€‚é…åçš„å­—ä½“å¤§å°"""
        return self.get_dpi_scaled_size(base_font_size)
    
    def detect_system_theme(self):
        """æ£€æµ‹ç³»ç»Ÿä¸»é¢˜ï¼ˆæ·±è‰²/æµ…è‰²æ¨¡å¼ï¼‰"""
        try:
            import platform
            system = platform.system()
            
            if system == "Darwin":  # macOS
                import subprocess
                result = subprocess.run(
                    ["defaults", "read", "-g", "AppleInterfaceStyle"], 
                    capture_output=True, text=True
                )
                return "dark" if result.stdout.strip() == "Dark" else "light"
            elif system == "Windows":
                try:
                    import winreg
                    registry = winreg.ConnectRegistry(None, winreg.HKEY_CURRENT_USER)
                    key = winreg.OpenKey(registry, r"Software\Microsoft\Windows\CurrentVersion\Themes\Personalize")
                    value = winreg.QueryValueEx(key, "AppsUseLightTheme")[0]
                    winreg.CloseKey(key)
                    return "light" if value else "dark"
                except:
                    return "light"
            else:  # Linuxç­‰
                return "light"
        except:
            return "light"
    
    def get_system_theme_colors(self):
        """æ ¹æ®ç³»ç»Ÿä¸»é¢˜è·å–é€‚å½“çš„é¢œè‰²"""
        theme = self.detect_system_theme()
        
        if theme == "dark":
            return {
                'background': '#2b2b2b',
                'surface': '#3c3c3c', 
                'text': '#ffffff',
                'text_secondary': '#b0b0b0',
                'border': '#555555',
                'accent': '#0084ff',
                'card_bg': '#404040',
                'input_bg': '#353535'
            }
        else:
            return {
                'background': '#f8f9fa',
                'surface': '#ffffff',
                'text': '#000000', 
                'text_secondary': '#6c757d',
                'border': '#dee2e6',
                'accent': '#007bff',
                'card_bg': '#ffffff',
                'input_bg': '#ffffff'
                         }
    
    def apply_consistent_styling(self):
        """åº”ç”¨ç»Ÿä¸€çš„å­—ä½“å¤§å°å’Œç³»ç»Ÿä¸»é¢˜è·Ÿéšæ ·å¼"""
        # æ£€æµ‹æ˜¯å¦åº”è¯¥ä½¿ç”¨ç³»ç»Ÿä¸»é¢˜
        use_system_theme = self.settings.value("use_system_theme", True, type=bool)
        
        if use_system_theme:
            # ä½¿ç”¨ç³»ç»Ÿä¸»é¢˜
            theme_colors = self.get_system_theme_colors()
            font_color = theme_colors['text']
            background_color = theme_colors['input_bg']
            card_bg = theme_colors['card_bg']
            border_color = theme_colors['border']
            accent_color = theme_colors['accent']
        else:
            # ä½¿ç”¨ç”¨æˆ·è‡ªå®šä¹‰é¢œè‰²
            font_color = self.font_color
            background_color = self.background_color
            card_bg = "#ffffff"
            border_color = "#dee2e6"
            accent_color = "#007bff"
        
        # ç»Ÿä¸€çš„åŸºç¡€å­—ä½“å¤§å° - æ‰€æœ‰æ§ä»¶ä½¿ç”¨ç›¸åŒå¤§å°
        base_font_size = self.get_dpi_scaled_font_size(self.font_size)
        
        # æ„å»ºå­—ä½“æ ·å¼å­—ç¬¦ä¸²
        font_weight = "bold" if self.font_bold else "normal"
        font_style = "italic" if self.font_italic else "normal"
        text_decoration = "underline" if self.font_underline else "none"
        
        # åº”ç”¨å…¨å±€æ ·å¼ - ç¡®ä¿æ‰€æœ‰æ§ä»¶å­—ä½“å¤§å°å®Œå…¨ç»Ÿä¸€
        self.setStyleSheet(f"""
            /* å…¨å±€å­—ä½“è®¾ç½® - å¼ºåˆ¶æ‰€æœ‰æ§ä»¶ä½¿ç”¨ç»Ÿä¸€å­—ä½“å¤§å° */
            * {{
                font-family: '{self.font_family}', 'Microsoft YaHei', 'SimHei', sans-serif !important;
                font-size: {base_font_size}px !important;
                font-weight: {font_weight} !important;
                font-style: {font_style} !important;
            }}
            
            QMainWindow {{
                background-color: {theme_colors.get('background', '#f8f9fa') if use_system_theme else '#f8f9fa'};
                color: {font_color};
            }}
            
            QWidget {{
                color: {font_color};
                background-color: transparent;
            }}
            
            /* è¾“å…¥æ§ä»¶ */
            QLineEdit, QSpinBox, QDoubleSpinBox, QComboBox, QDateEdit {{
                color: black;
                background-color: white;
                border: 1px solid {border_color};
                padding: 6px;
                border-radius: 4px;
                font-weight: bold;
            }}
            
            QLineEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus, 
            QComboBox:focus, QDateEdit:focus {{
                border: 2px solid {accent_color};
                color: black;
            }}
            
            /* æ–‡æœ¬æ˜¾ç¤ºæ§ä»¶ */
            QLabel {{
                color: black;
                background-color: white;
                text-decoration: {text_decoration};
                font-weight: bold;
            }}
            
            /* æŒ‰é’® */
            QPushButton {{
                color: {font_color};
                background-color: #e9ecef;
                border: 1px solid {border_color};
                padding: 8px 16px;
                border-radius: 4px;
            }}
            
            QPushButton:hover {{
                background-color: #dee2e6;
            }}
            
            QPushButton:pressed {{
                background-color: #d3d9df;
            }}
            
            QPushButton:disabled {{
                background-color: #f8f9fa;
                color: #6c757d;
            }}
            
            /* å¤é€‰æ¡†å’Œå•é€‰æŒ‰é’® */
            QCheckBox, QRadioButton {{
                color: black;
                background-color: white;
                spacing: 6px;
                font-weight: bold;
            }}
            
            /* å¤é€‰æ¡†å’Œå•é€‰æ¡†ä¸éœ€è¦é€‰ä¸­æ—¶çš„èƒŒæ™¯ */
            
            QCheckBox::indicator, QRadioButton::indicator {{
                width: {self.get_dpi_scaled_size(16)}px;
                height: {self.get_dpi_scaled_size(16)}px;
                background-color: white;
                border: 2px solid #666666;
            }}
            
            QCheckBox::indicator {{
                border-radius: 3px;
            }}
            
            QRadioButton::indicator {{
                border-radius: {self.get_dpi_scaled_size(8)}px;
            }}
            
            QCheckBox::indicator:checked {{
                background-color: {accent_color};
                border: 2px solid {accent_color};
                image: url(data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIxNiIgaGVpZ2h0PSIxNiIgdmlld0JveD0iMCAwIDI0IDI0IiBmaWxsPSJ3aGl0ZSI+PHBhdGggZD0iTTkgMTYuMTdMNC44MyAxMmwtMS40MiAxLjQxTDkgMTkgMjEgN2wtMS40MS0xLjQxeiIvPjwvc3ZnPg==);
            }}
            
            QRadioButton::indicator:checked {{
                background-color: {accent_color};
                border: 2px solid {accent_color};
            }}
            
            QCheckBox::indicator:hover, QRadioButton::indicator:hover {{
                border-color: {accent_color};
            }}
            
            /* è¡¨æ ¼ */
            QTableWidget {{
                color: black;  /* å¼ºåˆ¶ä½¿ç”¨é»‘è‰²å­—ä½“ */
                background-color: white;  /* å¼ºåˆ¶ä½¿ç”¨ç™½è‰²èƒŒæ™¯ */
                gridline-color: {border_color};
                selection-background-color: {accent_color}40;
            }}
            
            QTableWidget::item {{
                color: black;  /* å¼ºåˆ¶ä½¿ç”¨é»‘è‰²å­—ä½“ */
                background-color: white;  /* å¼ºåˆ¶ä½¿ç”¨ç™½è‰²èƒŒæ™¯ */
                padding: 4px;
                border: none;
            }}
            
            QHeaderView::section {{
                color: {font_color};
                background-color: {card_bg};
                border: 1px solid {border_color};
                padding: 6px;
                font-weight: bold;
            }}
            
            /* æ–‡æœ¬ç¼–è¾‘å™¨ */
            QTextEdit {{
                color: {font_color};
                background-color: {background_color};
                border: 1px solid {border_color};
                border-radius: 4px;
                padding: 8px;
            }}
            
            /* åˆ†ç»„æ¡† */
            QGroupBox {{
                color: {font_color};
                border: 1px solid {border_color};
                border-radius: 6px;
                margin-top: 10px;
                background-color: transparent;
            }}
            
            QGroupBox::title {{
                color: {font_color};
                subcontrol-origin: margin;
                left: 12px;
                padding: 0 8px 0 8px;
                background-color: {theme_colors.get('background', '#f8f9fa') if use_system_theme else '#f8f9fa'};
            }}
            
            /* æ ‡ç­¾é¡µ */
            QTabWidget::pane {{
                border: 1px solid {border_color};
                border-radius: 4px;
                background-color: {card_bg};
            }}
            
            QTabBar::tab {{
                color: {font_color};
                background-color: {card_bg};
                border: 1px solid {border_color};
                padding: 8px 16px;
                margin-right: 2px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
            }}
            
            QTabBar::tab:selected {{
                background-color: {card_bg};
                border-bottom: 2px solid {accent_color};
                color: {accent_color};
                font-weight: bold;
            }}
            
            QTabBar::tab:hover {{
                background-color: {border_color};
            }}
            
            /* çŠ¶æ€æ  */
            QStatusBar {{
                background-color: {theme_colors.get('background', '#f8f9fa') if use_system_theme else '#f8f9fa'};
                color: {font_color};
                border-top: 1px solid {border_color};
                padding: 4px;
            }}
            
            /* èœå•æ  */
            QMenuBar {{
                background-color: {theme_colors.get('background', '#f8f9fa') if use_system_theme else '#f8f9fa'};
                color: {font_color};
                border-bottom: 1px solid {border_color};
            }}
            
            QMenuBar::item {{
                background-color: transparent;
                padding: 4px 8px;
                border-radius: 4px;
            }}
            
            QMenuBar::item:selected {{
                background-color: {border_color};
            }}
            
            QMenu {{
                background-color: {card_bg};
                border: 1px solid {border_color};
                border-radius: 4px;
            }}
            
            QMenu::item {{
                padding: 6px 12px;
            }}
            
            QMenu::item:selected {{
                background-color: {border_color};
            }}
            
            /* è¿›åº¦æ¡ */
            QProgressBar {{
                background-color: {background_color};
                border: 1px solid {border_color};
                border-radius: 4px;
                color: {font_color};
                text-align: center;
            }}
            
            QProgressBar::chunk {{
                background-color: {accent_color};
                border-radius: 3px;
            }}
            
            /* å·¥å…·æç¤º */
            QToolTip {{
                background-color: {card_bg};
                color: {font_color};
                border: 1px solid {border_color};
                padding: 4px;
                border-radius: 4px;
            }}
        """)
    
    def force_uniform_font_on_all_widgets(self):
        """éå†æ‰€æœ‰æ§ä»¶ï¼Œå¼ºåˆ¶è®¾ç½®ç»Ÿä¸€çš„å­—ä½“å¤§å°ï¼ˆæœ€ç»ˆä¿é™©æªæ–½ï¼‰"""
        try:
            base_font_size = self.get_dpi_scaled_font_size(self.font_size)
            
            # åˆ›å»ºç»Ÿä¸€çš„å­—ä½“å¯¹è±¡
            uniform_font = QFont(self.font_family)
            uniform_font.setPointSize(base_font_size)
            uniform_font.setBold(self.font_bold)
            uniform_font.setItalic(self.font_italic)
            uniform_font.setUnderline(self.font_underline)
            
            def apply_font_to_widget(widget):
                if widget is None:
                    return
                
                try:
                    # è®¾ç½®å­—ä½“
                    widget.setFont(uniform_font)
                    
                    # é€’å½’å¤„ç†æ‰€æœ‰å­æ§ä»¶
                    for child in widget.findChildren(QWidget):
                        child.setFont(uniform_font)
                        
                except Exception as e:
                    # å¿½ç•¥æ— æ³•è®¾ç½®å­—ä½“çš„æ§ä»¶
                    pass
            
            # ä»ä¸»çª—å£å¼€å§‹éå†
            apply_font_to_widget(self)
            
            print(f"âœ… å­—ä½“ç»Ÿä¸€å®Œæˆï¼šæ‰€æœ‰æ§ä»¶å·²è®¾ç½®ä¸º {self.font_family} {base_font_size}px")
            
        except Exception as e:
            print(f"âš ï¸ å­—ä½“ç»Ÿä¸€è¿‡ç¨‹ä¸­å‡ºç°é”™è¯¯: {e}")

     
    def init_ui(self):
        """åˆå§‹åŒ–ç•Œé¢"""
        self.setWindowTitle("DHIç­›æŸ¥åŠ©æ‰‹ - ä¼Šåˆ©æ¶²å¥¶å¥¶ç§‘é™¢")
        
        # è®¾ç½®å…¨å±€æ ·å¼ - ç¡®ä¿æ‰€æœ‰å¤é€‰æ¡†ã€å•é€‰æ¡†ã€è¾“å…¥æ¡†æœ‰è‰¯å¥½çš„å¯¹æ¯”åº¦
        self.setStyleSheet("""
            QCheckBox {
                color: black;
                background-color: transparent;
            }
            QCheckBox::indicator {
                width: 16px;
                height: 16px;
                background-color: white;
                border: 2px solid #666;
                border-radius: 3px;
            }
            QCheckBox::indicator:checked {
                background-color: #007AFF;
                border-color: #007AFF;
                image: url(data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIxNiIgaGVpZ2h0PSIxNiIgdmlld0JveD0iMCAwIDI0IDI0IiBmaWxsPSJ3aGl0ZSI+PHBhdGggZD0iTTkgMTYuMTdMNC44MyAxMmwtMS40MiAxLjQxTDkgMTkgMjEgN2wtMS40MS0xLjQxeiIvPjwvc3ZnPg==);
            }
            /* å¤é€‰æ¡†ä¸éœ€è¦é€‰ä¸­æ—¶çš„èƒŒæ™¯ */
            QCheckBox::indicator:hover {
                border-color: #007AFF;
            }
            QRadioButton {
                color: black;
                background-color: transparent;
            }
            /* å•é€‰æ¡†ä¸éœ€è¦é€‰ä¸­æ—¶çš„èƒŒæ™¯ */
            QRadioButton::indicator {
                width: 16px;
                height: 16px;
                background-color: white;
                border: 2px solid #666;
                border-radius: 8px;
            }
            QRadioButton::indicator:checked {
                background-color: #007AFF;
                border-color: #007AFF;
            }
            /* ç§»é™¤afterä¼ªå…ƒç´ ï¼Œä½¿å•é€‰æ¡†å®Œå…¨å¡«å…… */
            QSpinBox, QDoubleSpinBox, QLineEdit {
                background-color: white;
                color: black;
                border: 1px solid #ccc;
                padding: 5px;
                border-radius: 3px;
            }
            QSpinBox:focus, QDoubleSpinBox:focus, QLineEdit:focus {
                border-color: #007AFF;
            }
            QGroupBox {
                color: black;
                font-weight: bold;
                border: 1px solid #ddd;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
                background-color: transparent;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
                background-color: white;
                color: black;
            }
            QLabel {
                color: black;
                background-color: transparent;
            }
            /* ç¡®ä¿æ ‡ç­¾é¡µå†…å®¹åŒºåŸŸèƒŒæ™¯ä¸ºç™½è‰² */
            QTabWidget::pane {
                background-color: white;
            }
            QWidget {
                background-color: white;
            }
            /* è¡¨æ ¼æ ·å¼ */
            QTableWidget {
                background-color: white;
                color: black;
                gridline-color: #ddd;
                selection-background-color: #007AFF;
                selection-color: white;
            }
            QTableWidget::item {
                color: black;
                background-color: white;
            }
            QTableWidget::item:selected {
                background-color: #007AFF;
                color: white;
            }
            QHeaderView::section {
                background-color: #f8f9fa;
                color: black;
                font-weight: bold;
                border: 1px solid #ddd;
                padding: 5px;
            }
            /* æŒ‰é’®æ ·å¼ - ç¡®ä¿è‰¯å¥½å¯¹æ¯”åº¦ */
            QPushButton {
                background-color: #007AFF;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #0051D5;
            }
            QPushButton:pressed {
                background-color: #0041AB;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
            /* ç‰¹æ®ŠæŒ‰é’®æ ·å¼ä¿æŒåŸæ · */
            QPushButton#enableCheckbox {
                background-color: transparent;
            }
        """)
        
        # åˆ›å»ºèœå•æ  - åªåˆ›å»ºä¸€æ¬¡
        self.create_menu_bar()
        
        # è‡ªé€‚åº”çª—å£å¤§å° - æ ¹æ®å±å¹•å°ºå¯¸è®¾ç½®
        screen_info = self.get_safe_screen_info()
        screen_width = screen_info['width']
        screen_height = screen_info['height']
        
        # è®¾ç½®çª—å£ä¸ºå±å¹•çš„80%ï¼Œä½†æœ‰æœ€å°å’Œæœ€å¤§é™åˆ¶
        window_width = min(max(int(screen_width * 0.8), 1200), 1800)
        window_height = min(max(int(screen_height * 0.8), 800), 1200)
        
        # çª—å£å±…ä¸­
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.setGeometry(x, y, window_width, window_height)
        
        # è®¾ç½®æœ€å°å°ºå¯¸
        self.setMinimumSize(1000, 700)
        
        # è®¾ç½®å›¾æ ‡
        try:
            if os.path.exists("whg3r-qi1nv-001.ico"):
                self.setWindowIcon(QIcon("whg3r-qi1nv-001.ico"))
        except:
            pass
        
        # åº”ç”¨ç³»ç»Ÿä¸»é¢˜è·Ÿéšå’Œç»Ÿä¸€å­—ä½“å¤§å°
        self.apply_consistent_styling()
        

        
        # åˆ›å»ºä¸­å¤®éƒ¨ä»¶
        central_widget = QWidget()
        central_widget.setStyleSheet(f"background-color: #f8f9fa;")  # ä¸»çª—å£ä¿æŒç°è‰²èƒŒæ™¯
        self.setCentralWidget(central_widget)
        
        # åˆ›å»ºä¸»å¸ƒå±€
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # æ·»åŠ é¡¶éƒ¨æ ‡é¢˜æ  - é«˜åº¦ç›¸å¯¹äºå±å¹•
        header = self.create_header()
        main_layout.addWidget(header)
        
        # æ·»åŠ æ­¥éª¤æŒ‡ç¤ºå™¨
        steps_widget = self.create_steps_indicator()
        main_layout.addWidget(steps_widget)
        
        # åˆ›å»ºå¯åˆ†å‰²çš„å†…å®¹åŒºåŸŸ
        content_splitter = QSplitter(Qt.Orientation.Horizontal)
        content_splitter.setContentsMargins(10, 10, 10, 10)
        
        # ===========================================
        # å·¦ä¾§æ§åˆ¶é¢æ¿ - å¹³è¡¡æ»šåŠ¨ç­–ç•¥
        # ===========================================
        # ç»è¿‡UIä¼˜åŒ–åçš„æœ€ç»ˆæ–¹æ¡ˆï¼šä¿ç•™æ»šåŠ¨åŠŸèƒ½ä½†è®¾ç½®åˆç†é«˜åº¦
        # 
        # è®¾è®¡è€ƒè™‘ï¼š
        # 1. ä¿ç•™QScrollArea - å½“å†…å®¹è¿‡å¤šæ—¶å¯ä»¥æ»šåŠ¨ï¼Œé¿å…å†…å®¹è¢«æˆªæ–­
        # 2. è®¾ç½®600pxæœ€å°é«˜åº¦ - ç¡®ä¿ä¸»è¦å†…å®¹åœ¨é¦–å±å¯è§
        # 3. ç»“åˆæ ‡ç­¾é¡µå†…çš„é¡¶éƒ¨å¯¹é½è®¾è®¡ï¼Œè¾¾åˆ°æœ€ä½³ç”¨æˆ·ä½“éªŒ
        left_scroll = QScrollArea()
        
        # æ»šåŠ¨åŒºåŸŸåŸºæœ¬è®¾ç½®
        left_scroll.setWidgetResizable(True)  # å†…å®¹è‡ªé€‚åº”æ»šåŠ¨åŒºåŸŸå¤§å°
        left_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)  # ç¦ç”¨æ¨ªå‘æ»šåŠ¨
        left_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)     # æŒ‰éœ€æ˜¾ç¤ºçºµå‘æ»šåŠ¨
        
        # å®½åº¦é™åˆ¶ - ç¡®ä¿åœ¨ä¸åŒå±å¹•ä¸Šçš„é€‚é…æ€§
        left_scroll.setMinimumWidth(400)  # å‡å°æœ€å°å®½åº¦åˆ°400px
        # ç§»é™¤æœ€å¤§å®½åº¦é™åˆ¶ï¼Œå…è®¸ç”¨æˆ·è‡ªç”±è°ƒæ•´
        # left_scroll.setMaximumWidth(800)  # æ³¨é‡Šæ‰æœ€å¤§å®½åº¦é™åˆ¶
        
        # ğŸ¯ é«˜åº¦ç­–ç•¥ - å¹³è¡¡å†…å®¹å¯è§æ€§å’Œç©ºé—´æ•ˆç‡çš„å…³é”®è®¾ç½®
        left_scroll.setMinimumHeight(400)  # é™ä½åˆ°400px - é¿å…æ–‡ä»¶ä¸Šä¼ åŒºåŸŸè¿‡åº¦æ‹‰ä¼¸
        
        left_panel = self.create_control_panel()
        left_scroll.setWidget(left_panel)
        content_splitter.addWidget(left_scroll)
        
        # å³ä¾§ç»“æœæ˜¾ç¤º
        right_panel = self.create_result_panel()
        right_panel.setMinimumWidth(200)  # å‡å°‘å³ä¾§æœ€å°å®½åº¦é™åˆ¶ï¼Œå…è®¸æ‹–æ‹½æ¡æ›´çµæ´»
        content_splitter.addWidget(right_panel)
        
        # è®¾ç½®åˆ†å‰²å™¨æ¯”ä¾‹å’Œçº¦æŸ - è°ƒæ•´ä¸º4:6æ¯”ä¾‹ï¼Œç»™å³ä¾§æ›´å¤šç©ºé—´
        left_width = max(450, int(window_width * 0.4))  # å·¦ä¾§å 40%ï¼Œè‡³å°‘450px
        right_width = window_width - left_width
        content_splitter.setSizes([left_width, right_width])
        content_splitter.setCollapsible(0, False)
        content_splitter.setCollapsible(1, False)
        
        # è®¾ç½®åˆ†å‰²å™¨æ ·å¼ - æ›´æ˜æ˜¾çš„æ‹–æ‹½æ‰‹æŸ„
        content_splitter.setStyleSheet("""
            QSplitter::handle {
                background-color: #ced4da;
                width: 5px;
                margin: 1px;
                border-radius: 2px;
            }
            QSplitter::handle:hover {
                background-color: #007bff;
                width: 7px;
            }
            QSplitter::handle:pressed {
                background-color: #0056b3;
            }
        """)
        
        main_layout.addWidget(content_splitter)
        
        # çŠ¶æ€æ 
        self.setup_status_bar()
        
        # å¼ºåˆ¶ç»Ÿä¸€æ‰€æœ‰æ§ä»¶çš„å­—ä½“å¤§å°ï¼ˆæœ€ç»ˆä¿é™©æªæ–½ï¼‰
        QTimer.singleShot(500, self.force_uniform_font_on_all_widgets)
    
    def create_menu_bar(self):
        """åˆ›å»ºèœå•æ """
        menubar = self.menuBar()
        if menubar is None:
            return
        
        # è®¾ç½®èœå•æ æ ·å¼ - ä½¿ç”¨ç»Ÿä¸€å­—ä½“å¤§å°
        menubar.setStyleSheet(f"""
            QMenuBar {{
                background-color: #f8f9fa;
                border-bottom: 1px solid #dee2e6;
                padding: 4px;
            }}
            QMenuBar::item {{
                background-color: transparent;
                padding: 4px 8px;
                border-radius: 4px;
            }}
            QMenuBar::item:selected {{
                background-color: #e9ecef;
            }}
            QMenu {{
                background-color: white;
                border: 1px solid #dee2e6;
                border-radius: 4px;
            }}
            QMenu::item {{
                padding: 6px 12px;
            }}
            QMenu::item:selected {{
                background-color: #e9ecef;
            }}
        """)
        
        # è®¾ç½®èœå•
        settings_menu = menubar.addMenu("è®¾ç½®")
        
        # æ˜¾ç¤ºè®¾ç½®
        display_action = QAction("ç•Œé¢æ˜¾ç¤ºè®¾ç½®...", self)
        display_action.setStatusTip("è°ƒæ•´ç•Œé¢æ˜¾ç¤ºæ¯”ä¾‹")
        display_action.triggered.connect(self.show_display_settings)
        settings_menu.addAction(display_action)
        
        # æ¢å¤é»˜è®¤æ˜¾ç¤ºï¼ˆé˜²å‘†åŠŸèƒ½ï¼‰
        reset_display_action = QAction("ğŸ”§ æ¢å¤é»˜è®¤æ˜¾ç¤º", self)
        reset_display_action.setStatusTip("ä¸€é”®æ¢å¤é»˜è®¤å­—ä½“é¢œè‰²å’Œæ˜¾ç¤ºè®¾ç½®ï¼ˆè§£å†³æ–‡å­—çœ‹ä¸è§é—®é¢˜ï¼‰")
        reset_display_action.triggered.connect(self.reset_display_settings_to_default)
        settings_menu.addAction(reset_display_action)
        
        settings_menu.addSeparator()
        
        # å…³äº
        about_action = QAction("å…³äº", self)
        about_action.setStatusTip("å…³äºDHIç­›æŸ¥åŠ©æ‰‹")
        about_action.triggered.connect(self.show_about)
        settings_menu.addAction(about_action)
        
        # è´¦å·èœå• - å·²ç§»é™¤ï¼ŒåŠŸèƒ½é›†æˆåˆ°å…¶ä»–ä½ç½®
        # account_menu = menubar.addMenu("è´¦å·")
        
        # # å½“å‰ç”¨æˆ·æ˜¾ç¤º
        # user_display = self.username
        # if self.auth_service and hasattr(self.auth_service, 'get_user_name'):
        #     full_name = self.auth_service.get_user_name()
        #     if full_name and full_name != self.username:
        #         user_display = f"{self.username} ({full_name})"
        
        # user_info_action = QAction(f"å½“å‰ç”¨æˆ·: {user_display}", self)
        # user_info_action.setEnabled(False)
        # account_menu.addAction(user_info_action)
        
        # account_menu.addSeparator()
        
        # # ä¿®æ”¹å¯†ç 
        # change_password_action = QAction("ä¿®æ”¹å¯†ç ...", self)
        # change_password_action.setStatusTip("ä¿®æ”¹å½“å‰è´¦å·å¯†ç ")
        # change_password_action.triggered.connect(self.show_change_password)
        # account_menu.addAction(change_password_action)
        
        # # é‚€è¯·ç ç®¡ç†
        # invite_code_action = QAction("é‚€è¯·ç ç®¡ç†...", self)
        # invite_code_action.setStatusTip("ç®¡ç†ç³»ç»Ÿé‚€è¯·ç ")
        # invite_code_action.triggered.connect(self.show_invite_code_management)
        # account_menu.addAction(invite_code_action)
        
        # account_menu.addSeparator()
        
        # # é€€å‡ºç™»å½•
        # logout_action = QAction("é€€å‡ºç™»å½•", self)
        # logout_action.setStatusTip("é€€å‡ºå½“å‰è´¦å·")
        # logout_action.triggered.connect(self.logout)
        # account_menu.addAction(logout_action)
    
    def show_display_settings(self):
        """æ˜¾ç¤ºç•Œé¢è®¾ç½®å¯¹è¯æ¡†"""
        dialog = DisplaySettingsDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_scale, new_font_color, new_bg_color, new_font_family, new_font_size, new_font_bold, new_font_italic, new_font_underline = dialog.save_settings()
            
            # æç¤ºç”¨æˆ·é‡å¯ç¨‹åº
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Icon.Information)
            msg.setWindowTitle("è®¾ç½®å·²ä¿å­˜")
            msg.setText(f"æ˜¾ç¤ºè®¾ç½®å·²æ›´æ–°")
            msg.setInformativeText(f"æ˜¾ç¤ºæ¯”ä¾‹: {new_scale}%\nå­—ä½“é¢œè‰²: {new_font_color}\nèƒŒæ™¯é¢œè‰²: {new_bg_color}\nå­—ä½“ç±»å‹: {new_font_family}\nå­—ä½“å¤§å°: {new_font_size}px\nå­—ä½“åŠ ç²—: {'æ˜¯' if new_font_bold else 'å¦'}\nå­—ä½“æ–œä½“: {'æ˜¯' if new_font_italic else 'å¦'}\nä¸‹åˆ’çº¿: {'æ˜¯' if new_font_underline else 'å¦'}\n\nå»ºè®®é‡å¯ç¨‹åºä»¥è·å¾—æœ€ä½³æ˜¾ç¤ºæ•ˆæœã€‚")
            
            restart_btn = msg.addButton("é‡å¯ç¨‹åº", QMessageBox.ButtonRole.AcceptRole)
            later_btn = msg.addButton("ç¨åé‡å¯", QMessageBox.ButtonRole.RejectRole)
            
            msg.exec()
            
            if msg.clickedButton() == restart_btn:
                self.restart_application()
            else:
                # å³æ—¶åº”ç”¨æ ·å¼æ›´æ–°ï¼Œæ— éœ€é‡å¯
                # é‡æ–°åŠ è½½è®¾ç½®
                self.display_scale = self.settings.value("display_scale", 100, type=int)
                self.font_color = self.settings.value("font_color", "#000000", type=str)
                self.background_color = self.settings.value("background_color", "#ffffff", type=str)
                self.font_family = self.settings.value("font_family", "Microsoft YaHei", type=str)
                self.font_size = self.settings.value("font_size", 12, type=int)
                self.font_bold = self.settings.value("font_bold", False, type=bool)
                self.font_italic = self.settings.value("font_italic", False, type=bool)
                self.font_underline = self.settings.value("font_underline", False, type=bool)
                
                # åº”ç”¨æ–°çš„æ ·å¼
                self.apply_consistent_styling()
                
                # å¼ºåˆ¶ç»Ÿä¸€æ‰€æœ‰æ§ä»¶çš„å­—ä½“å¤§å°
                QTimer.singleShot(100, self.force_uniform_font_on_all_widgets)
    
    def show_about(self):
        """æ˜¾ç¤ºå…³äºå¯¹è¯æ¡†"""
        QMessageBox.about(self, "å…³äºDHIç­›æŸ¥åŠ©æ‰‹",
                          "DHIç­›æŸ¥åŠ©æ‰‹ v3.0\n\n"
                         "ä¼Šåˆ©æ¶²å¥¶å¥¶ç§‘é™¢\n"
                         "ç”¨äºå¤„ç†DHIæŠ¥å‘Šæ•°æ®çš„ä¸“ä¸šåŠ©æ‰‹\n"
                         "æ”¯æŒæ‰¹é‡æ–‡ä»¶å¤„ç†å’Œå¤šç§ç­›é€‰æ¡ä»¶\n"
                         "ä¹³æˆ¿ç‚ç­›æŸ¥å’Œç›‘æµ‹åˆ†æåŠŸèƒ½\n\n"
                         "å¦‚æœ‰é—®é¢˜è¯·è”ç³»æŠ€æœ¯æ”¯æŒ")
    
    def restart_application(self):
        """é‡å¯åº”ç”¨ç¨‹åº"""
        import subprocess
        import sys
        
        QApplication.quit()
        subprocess.Popen([sys.executable] + sys.argv)
    
    def create_styled_message_box(self, icon_type, title, text, buttons=None, default_button=None):
        """åˆ›å»ºå¸¦æœ‰ç»Ÿä¸€æ ·å¼çš„æ¶ˆæ¯æ¡†"""
        msg = QMessageBox(self)
        msg.setIcon(icon_type)
        msg.setWindowTitle(title)
        msg.setText(text)
        
        # è®¾ç½®ç»Ÿä¸€çš„æ ·å¼ - ç¡®ä¿æ–‡å­—æ¸…æ™°å¯è§
        msg.setStyleSheet("""
            QMessageBox {
                background-color: white;
                color: black;
                font-family: 'Microsoft YaHei', 'SimHei', sans-serif;
                font-size: 12px;
                border: 1px solid #ddd;
                border-radius: 8px;
            }
            QMessageBox QLabel {
                color: black;
                background-color: white;
                font-weight: bold;
                padding: 10px;
                font-size: 13px;
            }
            QMessageBox QPushButton {
                background-color: #007bff;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 12px;
                min-width: 60px;
            }
            QMessageBox QPushButton:hover {
                background-color: #0056b3;
            }
            QMessageBox QPushButton:pressed {
                background-color: #004085;
            }
            QMessageBox QPushButton:default {
                background-color: #28a745;
            }
            QMessageBox QPushButton:default:hover {
                background-color: #1e7e34;
            }
        """)
        
        if buttons:
            msg.setStandardButtons(buttons)
        if default_button:
            msg.setDefaultButton(default_button)
            
        return msg
    
    def show_info(self, title, text):
        """æ˜¾ç¤ºä¿¡æ¯æç¤ºæ¡†"""
        msg = self.create_styled_message_box(QMessageBox.Icon.Information, title, text)
        return msg.exec()
    
    def show_warning(self, title, text):
        """æ˜¾ç¤ºè­¦å‘Šæç¤ºæ¡†"""
        msg = self.create_styled_message_box(QMessageBox.Icon.Warning, title, text)
        return msg.exec()
    
    def show_error(self, title, text):
        """æ˜¾ç¤ºé”™è¯¯æç¤ºæ¡†"""
        msg = self.create_styled_message_box(QMessageBox.Icon.Critical, title, text)
        return msg.exec()
    
    def show_question(self, title, text, buttons=None, default_button=None):
        """æ˜¾ç¤ºé—®é¢˜å¯¹è¯æ¡†"""
        if buttons is None:
            buttons = QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        if default_button is None:
            default_button = QMessageBox.StandardButton.No
            
        msg = self.create_styled_message_box(QMessageBox.Icon.Question, title, text, buttons, default_button)
        return msg.exec()
    
    def reset_display_settings_to_default(self):
        """é˜²å‘†åŠŸèƒ½ï¼šæ¢å¤é»˜è®¤æ˜¾ç¤ºè®¾ç½®"""
        reply = QMessageBox.question(
            self, 
            "æ¢å¤é»˜è®¤æ˜¾ç¤ºè®¾ç½®", 
            "ğŸ”§ è¿™å°†æ¢å¤æ‰€æœ‰æ˜¾ç¤ºè®¾ç½®ä¸ºé»˜è®¤å€¼ï¼š\n\n"
            "â€¢ å­—ä½“é¢œè‰²ï¼šé»‘è‰²\n"
            "â€¢ èƒŒæ™¯é¢œè‰²ï¼šç™½è‰²\n" 
            "â€¢ å­—ä½“ç±»å‹ï¼šMicrosoft YaHei\n"
            "â€¢ å­—ä½“å¤§å°ï¼š12px\n"
            "â€¢ æ˜¾ç¤ºæ¯”ä¾‹ï¼š100%\n"
            "â€¢ å…¶ä»–å­—ä½“æ ·å¼ï¼šå–æ¶ˆåŠ ç²—/æ–œä½“/ä¸‹åˆ’çº¿\n"
            "â€¢ è·Ÿéšç³»ç»Ÿä¸»é¢˜ï¼šå¯ç”¨\n\n"
            "ğŸ’¡ è¿™å¯ä»¥è§£å†³æ–‡å­—çœ‹ä¸è§æˆ–æ˜¾ç¤ºå¼‚å¸¸çš„é—®é¢˜ã€‚\n\n"
            "æ˜¯å¦ç¡®è®¤æ¢å¤é»˜è®¤è®¾ç½®ï¼Ÿ",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            # æ¢å¤æ‰€æœ‰é»˜è®¤å€¼
            self.settings.setValue("display_scale", 100)
            self.settings.setValue("font_color", "#000000")
            self.settings.setValue("background_color", "#ffffff")
            self.settings.setValue("font_family", "Microsoft YaHei")
            self.settings.setValue("font_size", 12)
            self.settings.setValue("font_bold", False)
            self.settings.setValue("font_italic", False)
            self.settings.setValue("font_underline", False)
            self.settings.setValue("use_system_theme", True)
            
            # ç«‹å³åº”ç”¨æ–°è®¾ç½®
            self.display_scale = 100
            self.font_color = "#000000"
            self.background_color = "#ffffff" 
            self.font_family = "Microsoft YaHei"
            self.font_size = 12
            self.font_bold = False
            self.font_italic = False
            self.font_underline = False
            
            # é‡æ–°åº”ç”¨æ ·å¼
            self.apply_consistent_styling()
            QTimer.singleShot(100, self.force_uniform_font_on_all_widgets)
            
            # æ˜¾ç¤ºæˆåŠŸæ¶ˆæ¯
            QMessageBox.information(
                self,
                "è®¾ç½®å·²æ¢å¤",
                "âœ… æ˜¾ç¤ºè®¾ç½®å·²æˆåŠŸæ¢å¤ä¸ºé»˜è®¤å€¼ï¼\n\n"
                "ğŸ¯ æ‰€æœ‰æ–‡å­—ç°åœ¨åº”è¯¥æ¸…æ™°å¯è§ã€‚\n"
                "ğŸ’¡ å¦‚æœä»æœ‰æ˜¾ç¤ºé—®é¢˜ï¼Œå»ºè®®é‡å¯ç¨‹åºè·å¾—æœ€ä½³æ•ˆæœã€‚",
                QMessageBox.StandardButton.Ok
            )
            
            print("âœ… ç”¨æˆ·æ‰‹åŠ¨æ¢å¤äº†é»˜è®¤æ˜¾ç¤ºè®¾ç½®")
    
    def check_display_issues_on_startup(self):
        """å¯åŠ¨æ—¶æ£€æŸ¥æ˜¾ç¤ºé—®é¢˜ï¼ˆé˜²å‘†åŠŸèƒ½ï¼‰"""
        try:
            # æ£€æŸ¥å­—ä½“é¢œè‰²æ˜¯å¦è¿‡æµ…
            hex_color = self.font_color.lstrip('#')
            if len(hex_color) == 6:
                r = int(hex_color[0:2], 16)
                g = int(hex_color[2:4], 16)
                b = int(hex_color[4:6], 16)
                brightness = (r * 0.299 + g * 0.587 + b * 0.114) / 255
                
                # å¦‚æœäº®åº¦è¿‡é«˜ä¸”èƒŒæ™¯ä¹Ÿæ˜¯ç™½è‰²ï¼Œæç¤ºç”¨æˆ·
                if brightness > 0.85 and self.background_color.lower() in ['#ffffff', '#fff', 'white']:
                    reply = QMessageBox.question(
                        self,
                        "æ˜¾ç¤ºé—®é¢˜æé†’",
                        f"ğŸ” æ£€æµ‹åˆ°å¯èƒ½çš„æ˜¾ç¤ºé—®é¢˜ï¼š\n\n"
                        f"å½“å‰å­—ä½“é¢œè‰²ï¼š{self.font_color} ï¼ˆäº®åº¦{brightness:.1%}ï¼‰\n"
                        f"å½“å‰èƒŒæ™¯é¢œè‰²ï¼š{self.background_color}\n\n"
                        f"âš ï¸ æµ…è‰²å­—ä½“åœ¨ç™½è‰²èƒŒæ™¯ä¸Šå¯èƒ½çœ‹ä¸æ¸…æ¥šï¼Œ\n"
                        f"å¦‚æœæ‚¨é‡åˆ°æ–‡å­—æ˜¾ç¤ºé—®é¢˜ï¼Œå»ºè®®æ¢å¤é»˜è®¤è®¾ç½®ã€‚\n\n"
                        f"ğŸ’¡ æ‚¨å¯ä»¥é€šè¿‡èœå•\"è®¾ç½® â†’ æ¢å¤é»˜è®¤æ˜¾ç¤º\"æ¥è§£å†³ã€‚\n\n"
                        f"æ˜¯å¦ç°åœ¨æ¢å¤é»˜è®¤æ˜¾ç¤ºè®¾ç½®ï¼Ÿ",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                        QMessageBox.StandardButton.No
                    )
                    
                    if reply == QMessageBox.StandardButton.Yes:
                        self.reset_display_settings_to_default()
                        
        except Exception as e:
            # é™é»˜å¤±è´¥ï¼Œä¸å½±å“ç¨‹åºå¯åŠ¨
            print(f"æ˜¾ç¤ºé—®é¢˜æ£€æŸ¥å¤±è´¥: {e}")
    
    def show_change_password(self):
        """æ˜¾ç¤ºä¿®æ”¹å¯†ç å¯¹è¯æ¡†"""
        from auth_module.change_password_dialog import ChangePasswordDialog
        dialog = ChangePasswordDialog(self, self.username)
        dialog.exec()
    
    def show_invite_code_management(self):
        """æ˜¾ç¤ºé‚€è¯·ç ç®¡ç†å¯¹è¯æ¡†"""
        from auth_module.invite_code_dialog import InviteCodeDialog
        dialog = InviteCodeDialog(self)
        dialog.exec()
    
    def logout(self):
        """é€€å‡ºç™»å½•"""
        reply = QMessageBox.question(
            self, 
            "é€€å‡ºç™»å½•", 
            "ç¡®å®šè¦é€€å‡ºå½“å‰è´¦å·å—ï¼Ÿ",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            # åœæ­¢å¿ƒè·³
            if self.heartbeat_timer:
                self.heartbeat_timer.stop()
            
            # è°ƒç”¨è®¤è¯æœåŠ¡çš„ç™»å‡ºæ–¹æ³•
            if self.auth_service:
                self.auth_service.logout()
            
            # å…³é—­ä¸»çª—å£
            self.close()
            
            # é‡å¯åº”ç”¨ç¨‹åºå›åˆ°ç™»å½•ç•Œé¢
            import subprocess
            subprocess.Popen([sys.executable] + sys.argv)
            QApplication.quit()
    
    def setup_status_bar(self):
        """è®¾ç½®çŠ¶æ€æ """
        status_bar = self.statusBar()
        if status_bar is not None:
            status_bar.showMessage("ä¼Šåˆ©æ¶²å¥¶å¥¶ç§‘é™¢ DHIç­›æŸ¥åŠ©æ‰‹ - å‡†å¤‡å°±ç»ª")
            status_bar.setStyleSheet(f"""
                QStatusBar {{
                    background-color: #e9ecef;
                    border-top: 1px solid #dee2e6;
                    padding: 8px 15px;
                }}
            """)
    
    def create_header(self):
        """åˆ›å»ºé¡¶éƒ¨æ ‡é¢˜æ """
        header = QWidget()
        
        # æ ¹æ®å±å¹•é«˜åº¦è®¾ç½®æ ‡é¢˜æ é«˜åº¦
        screen_info = self.get_safe_screen_info()
        screen_height = screen_info['height']
        header_height = max(min(int(screen_height * 0.04), 50), 30)
        header.setFixedHeight(header_height)
        
        header.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                    stop:0 #4285f4, stop:1 #1976d2);
                border: none;
            }
        """)
        
        layout = QHBoxLayout(header)
        # å‡å°‘ä¸Šä¸‹è¾¹è·ï¼Œä¿æŒå·¦å³è¾¹è·
        margin_h = max(int(header_height * 0.25), 15)
        margin_v = max(int(header_height * 0.15), 5)  # å‡å°‘å‚ç›´è¾¹è·
        layout.setContentsMargins(margin_h, margin_v, margin_h, margin_v)
        
        # å·¦ä¾§å›¾æ ‡å’Œæ ‡é¢˜
        title_layout = QHBoxLayout()
        
        # å›¾æ ‡
        icon_label = QLabel("ğŸ¥›")
        icon_label.setStyleSheet(f"background: transparent;")
        title_layout.addWidget(icon_label)
        
        # æ ‡é¢˜æ–‡å­—
        title_label = QLabel("DHIç­›æŸ¥åŠ©æ‰‹")
        title_label.setStyleSheet(f"""
            font-weight: bold;
            color: white;
            background: transparent;
            margin-left: 10px;
        """)
        title_layout.addWidget(title_label)
        
        layout.addLayout(title_layout)
        layout.addStretch()
        
        # ç”¨æˆ·ä¿¡æ¯åŒºåŸŸ
        user_layout = QHBoxLayout()
        user_layout.setSpacing(8)  # é€‚ä¸­çš„é—´è·
        
        # ç”¨æˆ·å›¾æ ‡å’Œåç§°
        user_icon = QLabel("ğŸ‘¤")
        user_icon.setStyleSheet("background: transparent; color: white;")
        user_layout.addWidget(user_icon)
        
        # è·å–ç”¨æˆ·å§“å
        user_display = self.username
        if self.auth_service and hasattr(self.auth_service, 'get_user_name'):
            full_name = self.auth_service.get_user_name()
            if full_name and full_name != self.username:
                user_display = f"{self.username} ({full_name})"
        
        user_label = QLabel(f"å½“å‰ç”¨æˆ·: {user_display}")
        user_label.setStyleSheet("""
            color: white;
            background: transparent;
            font-weight: bold;
        """)
        user_layout.addWidget(user_label)
        
        # æ³¨é”€æŒ‰é’®
        logout_btn = QPushButton("æ³¨é”€")
        logout_btn.setToolTip("é€€å‡ºç™»å½•")
        logout_btn.clicked.connect(self.logout)
        # ç¡®ä¿æŒ‰é’®æœ‰è¶³å¤Ÿçš„å®½åº¦ï¼Œé™ä½é«˜åº¦
        logout_btn.setMinimumWidth(70)
        logout_btn.setMaximumHeight(26)
        logout_btn.setStyleSheet("""
            QPushButton {
                background-color: rgba(255, 255, 255, 0.2);
                border: 1px solid rgba(255, 255, 255, 0.3);
                border-radius: 3px;
                color: white;
                padding: 3px 15px;
                font-weight: bold;
                font-size: 13px;
                min-width: 70px;
                max-height: 26px;
            }
            QPushButton:hover {
                background-color: rgba(255, 255, 255, 0.3);
                border: 1px solid rgba(255, 255, 255, 0.5);
            }
            QPushButton:pressed {
                background-color: rgba(255, 255, 255, 0.4);
            }
        """)
        user_layout.addWidget(logout_btn)
        
        layout.addLayout(user_layout)
        
        # å¿«é€Ÿè®¾ç½®æŒ‰é’®
        settings_btn = QPushButton("âš™ï¸")
        settings_btn.setToolTip("æ˜¾ç¤ºè®¾ç½®")
        settings_btn.clicked.connect(self.show_display_settings)
        btn_size = max(int(header_height * 0.6), 20)
        settings_btn.setFixedSize(btn_size, btn_size)
        settings_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: rgba(255, 255, 255, 0.2);
                border: 1px solid rgba(255, 255, 255, 0.3);
                border-radius: {btn_size // 2}px;
                color: white;
            }}
            QPushButton:hover {{
                background-color: rgba(255, 255, 255, 0.3);
            }}
            QPushButton:pressed {{
                background-color: rgba(255, 255, 255, 0.4);
            }}
        """)
        layout.addWidget(settings_btn)
        
        # å³ä¾§å‰¯æ ‡é¢˜
        subtitle_label = QLabel("ä¼Šåˆ©æ¶²å¥¶å¥¶ç§‘é™¢ | DHIç­›æŸ¥åŠ©æ‰‹")
        subtitle_label.setStyleSheet(f"""
            color: rgba(255, 255, 255, 0.8);
            background: transparent;
            margin-left: 10px;
        """)
        layout.addWidget(subtitle_label)
        
        return header
    
    def create_steps_indicator(self):
        """åˆ›å»ºæ­¥éª¤æŒ‡ç¤ºå™¨"""
        steps_widget = QWidget()
        
        # æ ¹æ®å±å¹•å¤§å°è®¾ç½®æ­¥éª¤æŒ‡ç¤ºå™¨é«˜åº¦
        screen_info = self.get_safe_screen_info()
        screen_width = screen_info['width']
        steps_height = max(min(int(screen_width * 0.05), 90), 70)
        steps_widget.setFixedHeight(steps_height)
        steps_widget.setStyleSheet("background-color: white; border-bottom: 1px solid #e0e0e0;")
        
        layout = QHBoxLayout(steps_widget)
        margin_h = max(int(screen_width * 0.03), 30)
        margin_v = max(int(steps_height * 0.2), 15)
        layout.setContentsMargins(margin_h, margin_v, margin_h, margin_v)
        
        # æ­¥éª¤æ•°æ®
        steps = [
            {"number": "1", "title": "ä¸Šä¼ æ–‡ä»¶", "status": "completed"},
            {"number": "2", "title": "è®¾ç½®ç­›é€‰æ¡ä»¶", "status": "current"},
            {"number": "3", "title": "æŸ¥çœ‹ç»“æœ", "status": "pending"}
        ]
        
        for i, step in enumerate(steps):
            if i > 0:
                # æ·»åŠ è¿æ¥çº¿
                line = QLabel()
                line.setFixedSize(80, 2)
                line.setStyleSheet("background-color: #e0e0e0; margin-top: 20px;")
                layout.addWidget(line)
            
            # æ­¥éª¤åœ†åœˆå’Œæ–‡å­— - æ°´å¹³å¸ƒå±€
            step_layout = QHBoxLayout()
            step_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            step_layout.setSpacing(10)
            
            # åœ†åœˆ
            circle = QLabel(step["number"])
            # è·å–DPIä¿¡æ¯
            screen_info = self.get_safe_screen_info()
            dpi_ratio = screen_info['dpi_ratio']
            circle_size = max(int(24 * dpi_ratio * 0.6), 20)
            circle.setFixedSize(circle_size, circle_size)
            circle.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            circle_font_size = max(int(12 * dpi_ratio * 0.6), 10)
            border_radius = circle_size // 2
            
            if step["status"] == "completed":
                circle.setStyleSheet(f"""
                    background-color: #28a745;
                    color: white;
                    border-radius: {border_radius}px;
                    font-weight: bold;
                    font-size: {circle_font_size}px;
                """)
            elif step["status"] == "current":
                circle.setStyleSheet(f"""
                    background-color: #007bff;
                    color: white;
                    border-radius: {border_radius}px;
                    font-weight: bold;
                    font-size: {circle_font_size}px;
                """)
            else:
                circle.setStyleSheet(f"""
                    background-color: #e0e0e0;
                    color: #666;
                    border-radius: {border_radius}px;
                    font-weight: bold;
                    font-size: {circle_font_size}px;
                """)
            
            step_layout.addWidget(circle)
            
            # æ ‡é¢˜
            title = QLabel(step["title"])
            title.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            step_title_font_size = max(int(13 * dpi_ratio * 0.7), 12)
            title.setStyleSheet(f"""
                font-size: {step_title_font_size}px;
                color: #333;
                margin-left: 5px;
            """)
            step_layout.addWidget(title)
            
            step_container = QWidget()
            step_container.setLayout(step_layout)
            layout.addWidget(step_container)
        
        layout.addStretch()
        return steps_widget
    
    def create_card_widget(self, title):
        """åˆ›å»ºå¡ç‰‡æ ·å¼çš„å®¹å™¨
        
        è¿™æ˜¯æ‰€æœ‰åŠŸèƒ½åŒºåŸŸçš„ç»Ÿä¸€å®¹å™¨ç»„ä»¶ï¼Œç»è¿‡UIä¼˜åŒ–åé‡‡ç”¨ç´§å‡‘è®¾è®¡ï¼š
        
        ğŸ¯ ä¼˜åŒ–é‡ç‚¹ï¼š
        1. æ ‡é¢˜æ padding: 10px â†’ 4px (å‡å°‘60%)
        2. æ ‡é¢˜æ margin: 10px â†’ 4px (å‡å°‘60%) 
        3. æ ‡é¢˜å­—ä½“: 16px â†’ 13px (å‡å°‘19%)
        4. ä¿æŒåœ†è§’å’Œè¾¹æ¡†æ ·å¼ä¸å˜
        
        ğŸ’¡ è®¾è®¡ç†å¿µï¼š
        - åœ¨ä¿æŒç¾è§‚çš„å‰æä¸‹æœ€å¤§åŒ–ç©ºé—´åˆ©ç”¨ç‡
        - æ‰€æœ‰å¡ç‰‡ä½¿ç”¨ç»Ÿä¸€çš„ç´§å‡‘æ ·å¼
        - ç¡®ä¿æ–‡å­—ä»ç„¶æ¸…æ™°å¯è¯»
        """
        # åˆ›å»ºå¡ç‰‡ä¸»å®¹å™¨
        card = QWidget()
        card.setStyleSheet("""
            QWidget {
                background-color: white;
                border: 1px solid #e0e0e0;
                border-radius: 8px;
                margin: 2px;
            }
        """)
        # è®¾ç½®å¡ç‰‡çš„å¤§å°ç­–ç•¥ - é˜²æ­¢è¿‡åº¦æ‹‰ä¼¸
        from PyQt6.QtWidgets import QSizePolicy
        card.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Maximum)
        
        # å¡ç‰‡ä¸»å¸ƒå±€ - é›¶è¾¹è·é›¶é—´è·ï¼Œæœ€å¤§åŒ–å†…å®¹ç©ºé—´
        main_layout = QVBoxLayout(card)
        main_layout.setContentsMargins(0, 0, 0, 0)  # å»é™¤æ‰€æœ‰å¤–è¾¹è·
        main_layout.setSpacing(0)  # å»é™¤æ‰€æœ‰é—´è·
        
        # ===========================================
        # ğŸ¯ æ ‡é¢˜æ ä¼˜åŒ– - ç©ºé—´æ•ˆç‡æå‡çš„å…³é”®
        # ===========================================
        title_widget = QWidget()
        title_widget.setStyleSheet("""
            QWidget {
                background-color: #f8f9fa;
                border-bottom: 1px solid #e0e0e0;
                border-radius: 8px 8px 0 0;
                padding: 2px 4px;
            }
        """)  # âœ… è¿›ä¸€æ­¥å‹ç¼©paddingåˆ°2px 4px
        
        title_layout = QHBoxLayout(title_widget)
        title_layout.setContentsMargins(2, 2, 2, 2)  # âœ… è¿›ä¸€æ­¥å‹ç¼©marginåˆ°2px
        
        # æ ‡é¢˜æ–‡å­— - å­—ä½“ä¼˜åŒ–
        title_label = QLabel(title)
        title_label.setStyleSheet("""
            font-size: 13px;
            font-weight: bold;
            color: #333;
            background: transparent;
        """)  # âœ… å­—ä½“ä»16pxå‹ç¼©åˆ°13pxï¼Œä¿æŒåŠ ç²—ç¡®ä¿å¯è¯»æ€§
        title_layout.addWidget(title_label)
        title_layout.addStretch()
        
        main_layout.addWidget(title_widget)
        
        # å†…å®¹åŒºåŸŸ
        content_widget = QWidget()
        content_widget.setStyleSheet("background: transparent; border: none;")
        main_layout.addWidget(content_widget)
        
        # å°†å†…å®¹åŒºåŸŸçš„å¼•ç”¨ä¿å­˜åˆ°å¡ç‰‡ä¸Šï¼Œé˜²æ­¢è¢«åƒåœ¾å›æ”¶
        setattr(card, 'content_widget', content_widget)
        
        return card
    
    def get_responsive_button_styles(self):
        """è·å–è‡ªé€‚åº”æŒ‰é’®æ ·å¼ - ä½¿ç”¨ç»Ÿä¸€å­—ä½“å¤§å°"""
        # ä½¿ç”¨ç»Ÿä¸€çš„DPIç¼©æ”¾æ–¹æ³•
        padding_v = self.get_dpi_scaled_size(8)
        padding_h = self.get_dpi_scaled_size(16)
        border_radius = self.get_dpi_scaled_size(5)
        min_height = self.get_dpi_scaled_size(32)  # ç»Ÿä¸€æŒ‰é’®æœ€å°é«˜åº¦
        
        return {
            'primary': f"""
                QPushButton {{
                    background-color: #007bff;
                    color: white;
                    border: none;
                    border-radius: {border_radius}px;
                    padding: {padding_v}px {padding_h}px;
                    font-weight: bold;
                    min-height: {min_height}px;
                }}
                QPushButton:hover {{
                    background-color: #0056b3;
                }}
                QPushButton:pressed {{
                    background-color: #004085;
                }}
                QPushButton:disabled {{
                    background-color: #6c757d;
                    color: #adb5bd;
                }}
            """,
            'success': f"""
                QPushButton {{
                    background-color: #28a745;
                    color: white;
                    border: none;
                    border-radius: {border_radius}px;
                    padding: {padding_v}px {padding_h}px;
                    font-weight: bold;
                    min-height: {min_height}px;
                }}
                QPushButton:hover {{
                    background-color: #218838;
                }}
                QPushButton:pressed {{
                    background-color: #1e7e34;
                }}
                QPushButton:disabled {{
                    background-color: #6c757d;
                    color: #adb5bd;
                }}
            """,
            'warning': f"""
                QPushButton {{
                    background-color: #ffc107;
                    color: #212529;
                    border: none;
                    border-radius: {border_radius}px;
                    padding: {padding_v}px {padding_h}px;
                    font-weight: bold;
                    min-height: {min_height}px;
                }}
                QPushButton:hover {{
                    background-color: #e0a800;
                }}
                QPushButton:pressed {{
                    background-color: #d39e00;
                }}
                QPushButton:disabled {{
                    background-color: #6c757d;
                    color: #adb5bd;
                }}
            """,
            'info': f"""
                QPushButton {{
                    background-color: #17a2b8;
                    color: white;
                    border: none;
                    border-radius: {border_radius}px;
                    padding: {padding_v}px {padding_h}px;
                    font-weight: bold;
                    min-height: {min_height}px;
                }}
                QPushButton:hover {{
                    background-color: #138496;
                }}
                QPushButton:pressed {{
                    background-color: #117a8b;
                }}
                QPushButton:disabled {{
                    background-color: #6c757d;
                    color: #adb5bd;
                }}
            """,
            'secondary': f"""
                QPushButton {{
                    background-color: #6c757d;
                    color: white;
                    border: none;
                    border-radius: {border_radius}px;
                    padding: {padding_v}px {padding_h}px;
                    font-weight: bold;
                    min-height: {min_height}px;
                }}
                QPushButton:hover {{
                    background-color: #5a6268;
                }}
                QPushButton:pressed {{
                    background-color: #545b62;
                }}
                QPushButton:disabled {{
                    background-color: #e9ecef;
                    color: #6c757d;
                }}
            """,
            'danger': f"""
                QPushButton {{
                    background-color: #dc3545;
                    color: white;
                    border: none;
                    border-radius: {border_radius}px;
                    padding: {padding_v}px {padding_h}px;
                    font-weight: bold;
                    min-height: {min_height}px;
                }}
                QPushButton:hover {{
                    background-color: #c82333;
                }}
                QPushButton:pressed {{
                    background-color: #bd2130;
                }}
                QPushButton:disabled {{
                    background-color: #e9ecef;
                    color: #6c757d;
                }}
            """
        }
    
    def get_responsive_form_styles(self):
        """è·å–è‡ªé€‚åº”è¡¨å•æ ·å¼ - ä½¿ç”¨ç»Ÿä¸€å­—ä½“å¤§å°"""
        # ä½¿ç”¨ç»Ÿä¸€çš„DPIç¼©æ”¾æ–¹æ³•
        padding_v = self.get_dpi_scaled_size(6)  # å‡å°‘å‚ç›´å†…è¾¹è·
        padding_h = self.get_dpi_scaled_size(12)  # ä¿æŒæ°´å¹³å†…è¾¹è·
        border_radius = self.get_dpi_scaled_size(4)
        min_height = self.get_dpi_scaled_size(28)  # ç»Ÿä¸€è¾“å…¥æ¡†é«˜åº¦ä¸º28px
        
        return f"""
            QComboBox, QSpinBox, QDoubleSpinBox, QDateEdit {{
                border: 2px solid #ced4da;
                border-radius: {border_radius}px;
                padding: {padding_v}px {padding_h}px;
                background-color: white;
                min-height: {min_height}px;
                color: #495057;
                selection-background-color: #007bff;
                selection-color: white;
            }}
            QComboBox:focus, QSpinBox:focus, QDoubleSpinBox:focus, QDateEdit:focus {{
                border-color: #007bff;
                outline: none;
                border-width: 2px;
            }}
            QComboBox:hover, QSpinBox:hover, QDoubleSpinBox:hover, QDateEdit:hover {{
                border-color: #adb5bd;
            }}
            QLabel {{
                font-weight: 500;
                color: #495057;
                padding: 2px 0px;
            }}
            
            /* ç¡®ä¿ä¸‹æ‹‰æ¡†å’Œæ•°å­—è¾“å…¥æ¡†çš„æŒ‰é’®æ˜¾ç¤ºæ­£å¸¸ */
            QComboBox::drop-down {{
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: {self.get_dpi_scaled_size(20)}px;
                border-left: 1px solid #ced4da;
                background-color: #f8f9fa;
                border-top-right-radius: {border_radius}px;
                border-bottom-right-radius: {border_radius}px;
            }}
            QComboBox::drop-down:hover {{
                background-color: #e9ecef;
            }}
            
            QSpinBox::up-button, QSpinBox::down-button,
            QDoubleSpinBox::up-button, QDoubleSpinBox::down-button {{
                background-color: #f8f9fa;
                border: 1px solid #ced4da;
                width: {self.get_dpi_scaled_size(16)}px;
                min-height: {self.get_dpi_scaled_size(14)}px;
            }}
            QSpinBox::up-button:hover, QSpinBox::down-button:hover,
            QDoubleSpinBox::up-button:hover, QDoubleSpinBox::down-button:hover {{
                background-color: #e9ecef;
            }}
            
            QDateEdit::up-button, QDateEdit::down-button {{
                background-color: #f8f9fa;
                border: 1px solid #ced4da;
                width: {self.get_dpi_scaled_size(16)}px;
                min-height: {self.get_dpi_scaled_size(14)}px;
            }}
            QDateEdit::drop-down {{
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: {self.get_dpi_scaled_size(20)}px;
                border-left: 1px solid #ced4da;
                background-color: #f8f9fa;
                border-top-right-radius: {border_radius}px;
                border-bottom-right-radius: {border_radius}px;
            }}
        """
    
    def create_control_panel(self):
        """åˆ›å»ºå·¦ä¾§æ§åˆ¶é¢æ¿ - é‡æ„ä¸º4ä¸ªåŠŸèƒ½æ ‡ç­¾é¡µ"""
        panel = QWidget()
        panel.setStyleSheet("""
            QWidget {
                background-color: white;
                border-radius: 8px;
            }
        """)
        layout = QVBoxLayout(panel)
        
        # è‡ªé€‚åº”è¾¹è·å’Œé—´è· - ä½¿ç”¨ä¼˜åŒ–çš„DPIé€‚é…
        margin = self.get_dpi_scaled_size(8)
        spacing = self.get_dpi_scaled_size(8)
        
        layout.setContentsMargins(margin, margin, margin, margin)
        layout.setSpacing(spacing)

        # åˆ›å»ºåŠŸèƒ½æ ‡ç­¾é¡µ
        self.function_tabs = QTabWidget()
        
        # è‡ªé€‚åº”æ ‡ç­¾é¡µæ ·å¼ - ä½¿ç”¨ä¼˜åŒ–çš„DPIé€‚é…
        tab_font_size = self.get_dpi_scaled_font_size(16)  # ä¸»æ ‡ç­¾é¡µç»Ÿä¸€ä¸º16px
        tab_padding_v = self.get_dpi_scaled_size(12)
        tab_padding_h = self.get_dpi_scaled_size(20)
        tab_border_radius = self.get_dpi_scaled_size(5)
        
        self.function_tabs.setStyleSheet(f"""
            QTabWidget::pane {{
                border: 1px solid #e0e0e0;
                border-radius: {tab_border_radius}px;
                background-color: white;
                margin-top: -1px;
            }}
            QTabBar::tab {{
                background-color: #f8f9fa;
                border: 1px solid #e0e0e0;
                padding: {tab_padding_v}px {tab_padding_h}px;
                margin-right: 2px;
                border-top-left-radius: {tab_border_radius}px;
                border-top-right-radius: {tab_border_radius}px;
                font-size: {tab_font_size}px;
                font-weight: 500;
                color: #495057;
                min-width: 120px;
                min-height: 35px;
            }}
            QTabBar::tab:selected {{
                background-color: white;
                border-bottom: 1px solid white;
                color: #007bff;
                font-weight: bold;
            }}
            QTabBar::tab:hover {{
                background-color: #e9ecef;
            }}
        """)
        
        # åˆ›å»º4ä¸ªåŠŸèƒ½æ ‡ç­¾é¡µ
        self.create_basic_data_tab()
        self.create_dhi_filter_tab()
        self.create_mastitis_screening_tab()
        
        # éšæ€§ä¹³æˆ¿ç‚æœˆåº¦ç›‘æµ‹æ ‡ç­¾é¡µ
        self.create_mastitis_monitoring_tab()
        
        layout.addWidget(self.function_tabs)
        
        return panel

    def create_basic_data_tab(self):
        """åˆ›å»ºåŸºç¡€æ•°æ®æ ‡ç­¾é¡µï¼šæ–‡ä»¶ä¸Šä¼  + åœ¨ç¾¤ç‰›æ–‡ä»¶ + åŸºç¡€ç­›é€‰æ¡ä»¶
        
        è¿™æ˜¯ä¸»è¦çš„æ•°æ®è¾“å…¥æ ‡ç­¾é¡µï¼ŒåŒ…å«ä¸‰ä¸ªæ ¸å¿ƒåŠŸèƒ½åŒºåŸŸï¼š
        1. ğŸ“ æ–‡ä»¶ä¸Šä¼  - DHI Excelæ–‡ä»¶é€‰æ‹©å’Œå¤„ç†
        2. ğŸ„ åœ¨ç¾¤ç‰›æ–‡ä»¶ - å¯é€‰çš„åœ¨ç¾¤ç‰›æ•°æ®æ–‡ä»¶
        3. ğŸ”§ åŸºç¡€ç­›é€‰æ¡ä»¶ - èƒæ¬¡å’Œæ—¥æœŸèŒƒå›´è®¾ç½®
        
        å¸ƒå±€ç­–ç•¥ï¼šé‡‡ç”¨é¡¶éƒ¨å¯¹é½è®¾è®¡ï¼Œæ‰€æœ‰å†…å®¹ç´§è´´ä¸Šæ–¹æ’åˆ—ï¼Œä¸‹æ–¹ç•™ç©º
        è¿™æ ·ç”¨æˆ·æ‰“å¼€æ ‡ç­¾é¡µå°±èƒ½ç«‹å³çœ‹åˆ°æ‰€æœ‰é‡è¦åŠŸèƒ½ï¼Œæ— éœ€æ»šåŠ¨
        """
        # åˆ›å»ºæ ‡ç­¾é¡µä¸»å®¹å™¨
        tab_widget = QWidget()
        tab_layout = QVBoxLayout(tab_widget)
        
        # è®¾ç½®å¸ƒå±€å‚æ•° - æåº¦ç´§å‡‘
        tab_layout.setSpacing(3)  # ç»„ä»¶é—´è·ï¼š6px â†’ 3pxï¼Œè¿›ä¸€æ­¥å‹ç¼©
        tab_layout.setContentsMargins(4, 4, 4, 4)  # å¤–è¾¹è·ï¼š8px â†’ 4pxï¼Œæœ€å°åŒ–ç©ºé—´æµªè´¹
        
        # è·å–è‡ªé€‚åº”æ ·å¼ - æ ¹æ®å±å¹•DPIè‡ªåŠ¨è°ƒæ•´
        button_styles = self.get_responsive_button_styles()  # æŒ‰é’®æ ·å¼å­—å…¸
        form_styles = self.get_responsive_form_styles()      # è¡¨å•æ§ä»¶æ ·å¼
        
        # å¡ç‰‡å†…è¾¹è·è®¾ç½® - å¹³è¡¡ç©ºé—´åˆ©ç”¨å’Œå†…å®¹å¯è§æ€§
        card_margin = self.get_dpi_scaled_size(8)  # 8pxè¾¹è·ï¼Œæ—¢ç´§å‡‘åˆä¸å‹ç¼©å†…å®¹
        
        # ===========================================
        # 1. ğŸ“ DHIæ•°æ®æ–‡ä»¶ - è¶…æç®€ç‰ˆæœ¬ï¼ˆæ— æ ‡é¢˜æ ï¼‰
        # ===========================================
        # ç›´æ¥åˆ›å»ºå®¹å™¨ï¼Œè·³è¿‡create_card_widgetä»¥èŠ‚çœæ ‡é¢˜æ ç©ºé—´
        upload_group = QWidget()
        upload_group.setStyleSheet("""
            QWidget {
                background-color: white;
                border: 1px solid #e0e0e0;
                border-radius: 4px;
            }
        """)
        # è®¾ç½®æ–‡ä»¶ä¸Šä¼ åŒºåŸŸçš„å¤§å°ç­–ç•¥ - é˜²æ­¢è¿‡åº¦æ‹‰ä¼¸
        from PyQt6.QtWidgets import QSizePolicy
        upload_group.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Maximum)
        upload_layout = QVBoxLayout(upload_group)
        upload_layout.setContentsMargins(4, 4, 4, 4)  # è¿›ä¸€æ­¥å‹ç¼©è¾¹è·
        upload_layout.setSpacing(1)  # è¿›ä¸€æ­¥å‹ç¼©é—´è·
        
        # åˆ›å»ºæ‹–æ”¾ä¸Šä¼ åŒºåŸŸ - å¤§å¹…å¢å¤§é«˜åº¦
        drop_area = QWidget()
        drop_area.setFixedHeight(self.get_dpi_scaled_size(150))  # å¢å¤§é«˜åº¦ä¸º150pxï¼ˆçº¦3å€ï¼‰ï¼Œæ›´æ˜“æ‹–æ‹½
        drop_area.setStyleSheet("""
            QWidget {
                border: 2px dashed #007bff;
                border-radius: 8px;
                background-color: #f8f9fa;
                margin: 2px;
            }
            QWidget:hover {
                background-color: #e9f4ff;
                border-color: #0056b3;
                border-width: 3px;
            }
        """)
        
        # æ‹–æ”¾åŒºåŸŸå¸ƒå±€ - ä¼˜åŒ–è®¾è®¡
        drop_layout = QHBoxLayout(drop_area)  # æ”¹ä¸ºæ°´å¹³å¸ƒå±€ï¼ŒèŠ‚çœå‚ç›´ç©ºé—´
        drop_layout.setContentsMargins(20, 15, 20, 15)  # è¿›ä¸€æ­¥å¢å¤§å†…è¾¹è·
        drop_layout.setSpacing(15)  # è¿›ä¸€æ­¥å¢å¤§é—´è·
        
        # ä¸Šä¼ å›¾æ ‡
        upload_icon = QLabel("ğŸ“¤")
        upload_icon.setStyleSheet("font-size: 32px; background: transparent; border: none;")
        drop_layout.addWidget(upload_icon)
        
        # æ–‡å­—ä¿¡æ¯ï¼ˆå‚ç›´å¸ƒå±€ï¼‰
        text_widget = QWidget()
        text_layout = QVBoxLayout(text_widget)
        text_layout.setContentsMargins(0, 0, 0, 0)
        text_layout.setSpacing(4)
        
        upload_text = QLabel("æ‹–æ‹½æˆ–ç‚¹å‡»é€‰æ‹©DHI Excelæ–‡ä»¶")
        upload_text.setStyleSheet("font-size: 16px; color: #6c757d; background: transparent; border: none; font-weight: 500;")
        text_layout.addWidget(upload_text)
        
        format_hint = QLabel("æ”¯æŒ .xlsx, .xls æ ¼å¼")
        format_hint.setStyleSheet("font-size: 13px; color: #9ca3af; background: transparent; border: none;")
        text_layout.addWidget(format_hint)
        
        drop_layout.addWidget(text_widget)
        drop_layout.addStretch()
        
        upload_layout.addWidget(drop_area)
        
        # é€‰æ‹©æ–‡ä»¶æŒ‰é’®ï¼ˆä½œä¸ºå¤‡é€‰æ–¹å¼ï¼‰
        self.upload_btn = QPushButton("ğŸ“‚ æµè§ˆæ–‡ä»¶")
        self.upload_btn.setStyleSheet(button_styles['secondary'])
        self.upload_btn.clicked.connect(self.select_files)
        # ç§»é™¤æœ€å¤§é«˜åº¦é™åˆ¶ï¼Œä½¿ç”¨æ ·å¼ä¸­çš„ç»Ÿä¸€é«˜åº¦
        upload_layout.addWidget(self.upload_btn)
        
        # å·²é€‰æ–‡ä»¶æ˜¾ç¤ºåŒºåŸŸ - å¤§å¹…å¢å¤§é«˜åº¦ï¼ˆä»…åœ¨æœ‰æ–‡ä»¶æ—¶æ˜¾ç¤ºï¼‰
        files_container = QWidget()
        self.files_layout = QVBoxLayout(files_container)
        self.files_layout.setContentsMargins(2, 2, 2, 2)  # é€‚å½“å¢åŠ è¾¹è·
        self.files_layout.setSpacing(2)  # é€‚å½“å¢åŠ é—´è·
        
        # æ–‡ä»¶åˆ—è¡¨å®¹å™¨ï¼ˆç”¨äºåŠ¨æ€æ·»åŠ æ–‡ä»¶æ ‡ç­¾ï¼‰
        self.file_list = QListWidget()  # ä¿æŒå…¼å®¹æ€§
        self.file_list.setVisible(False)  # éšè—ä¼ ç»Ÿåˆ—è¡¨
        
        # æ–‡ä»¶æ ‡ç­¾å®¹å™¨
        self.file_tags_widget = QWidget()
        self.file_tags_layout = QVBoxLayout(self.file_tags_widget)
        self.file_tags_layout.setContentsMargins(4, 4, 4, 4)  # è¿›ä¸€æ­¥å¢åŠ å†…è¾¹è·
        self.file_tags_layout.setSpacing(6)  # è¿›ä¸€æ­¥å¢åŠ é—´è·
        
        no_files_label = QLabel("å°šæœªé€‰æ‹©æ–‡ä»¶")
        no_files_label.setStyleSheet("color: #9ca3af; font-size: 11px; font-style: italic;")
        self.file_tags_layout.addWidget(no_files_label)
        self.file_tags_layout.addStretch()
        
        # ç”¨QScrollAreaåŒ…è£¹æ–‡ä»¶æ ‡ç­¾åŒºåŸŸï¼Œå›ºå®šä¸º12æ¡æ•°æ®çš„é«˜åº¦
        from PyQt6.QtWidgets import QScrollArea
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QScrollArea.Shape.NoFrame)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        
        # è®¡ç®—12æ¡æ•°æ®çš„é«˜åº¦ï¼šæ¯æ¡36px + é—´è·6px = 42pxï¼Œ12æ¡ = 504px
        single_file_height = self.get_dpi_scaled_size(36)  # å•ä¸ªæ–‡ä»¶æ ‡ç­¾é«˜åº¦
        spacing = self.get_dpi_scaled_size(6)  # é—´è·
        total_height = (single_file_height + spacing) * 12  # 12æ¡æ•°æ®çš„æ€»é«˜åº¦
        scroll_area.setFixedHeight(total_height)  # å›ºå®šé«˜åº¦ï¼Œè¶…å‡ºæ—¶æ˜¾ç¤ºæ»šåŠ¨æ¡
        
        scroll_area.setWidget(self.file_tags_widget)
        
        self.files_layout.addWidget(scroll_area)
        upload_layout.addWidget(files_container)
        
        # æ“ä½œæŒ‰é’®åŒºåŸŸ - æåº¦å‹ç¼©
        buttons_container = QWidget()
        buttons_layout = QHBoxLayout(buttons_container)
        buttons_layout.setContentsMargins(0, 1, 0, 0)  # è¿›ä¸€æ­¥å‹ç¼©åˆ°1px
        buttons_layout.setSpacing(4)  # è¿›ä¸€æ­¥å‹ç¼©åˆ°4px
        
        # å¤„ç†æŒ‰é’®
        self.process_btn = QPushButton("ğŸš€ å¼€å§‹å¤„ç†")
        self.process_btn.setStyleSheet(button_styles['primary'])
        self.process_btn.clicked.connect(self.process_files)
        self.process_btn.setEnabled(False)
        # ç§»é™¤æœ€å¤§é«˜åº¦é™åˆ¶ï¼Œä½¿ç”¨æ ·å¼ä¸­çš„ç»Ÿä¸€é«˜åº¦
        buttons_layout.addWidget(self.process_btn)
        
        # æ¸…ç©ºæŒ‰é’®
        clear_btn = QPushButton("ğŸ—‘ï¸ æ¸…ç©º")
        clear_btn.setStyleSheet(button_styles['danger'])
        clear_btn.clicked.connect(self.clear_files)
        # ç§»é™¤æœ€å¤§é«˜åº¦é™åˆ¶ï¼Œä½¿ç”¨æ ·å¼ä¸­çš„ç»Ÿä¸€é«˜åº¦
        buttons_layout.addWidget(clear_btn)
        
        buttons_layout.addStretch()
        upload_layout.addWidget(buttons_container)
        
        # è¿›åº¦æ¡ - ä¼˜åŒ–è®¾è®¡
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setMaximumHeight(self.get_dpi_scaled_size(6))  # ä¼˜åŒ–é«˜åº¦ä¸º6px
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: none;
                border-radius: 3px;
                background-color: #e9ecef;
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: #007bff;
                border-radius: 3px;
            }
        """)
        upload_layout.addWidget(self.progress_bar)
        
        # è¿›åº¦æ ‡ç­¾ï¼ˆéšè—ï¼‰
        self.progress_label = QLabel("")
        self.progress_label.setVisible(False)
        
        tab_layout.addWidget(upload_group)
        
        # 2. åœ¨ç¾¤ç‰›æ–‡ä»¶ä¸Šä¼ åŒºåŸŸ - ç®€æ´è®¾è®¡ï¼Œä¸ä¸»ä¸Šä¼ åŒºåŸŸä¿æŒä¸€è‡´
        active_cattle_group = self.create_card_widget("ğŸ„ åœ¨ç¾¤ç‰›æ–‡ä»¶")
        active_cattle_layout = QVBoxLayout(getattr(active_cattle_group, 'content_widget'))
        active_cattle_layout.setContentsMargins(0, 0, 0, 0)  # ç§»é™¤åŒé‡è¾¹è·
        active_cattle_layout.setSpacing(6)  # ä¸ä¸»ä¸Šä¼ åŒºåŸŸä¿æŒä¸€è‡´çš„é—´è·
        
        # åœ¨ç¾¤ç‰›æ–‡ä»¶é€‰æ‹©æŒ‰é’®
        self.active_cattle_btn = QPushButton("ğŸ“‹ é€‰æ‹©åœ¨ç¾¤ç‰›æ–‡ä»¶")
        self.active_cattle_btn.setStyleSheet(button_styles['secondary'])
        self.active_cattle_btn.clicked.connect(self.select_active_cattle_file)
        # ç§»é™¤æœ€å¤§é«˜åº¦é™åˆ¶ï¼Œä½¿ç”¨æ ·å¼ä¸­çš„ç»Ÿä¸€é«˜åº¦
        active_cattle_layout.addWidget(self.active_cattle_btn)
        
        # åœ¨ç¾¤ç‰›æ–‡ä»¶çŠ¶æ€æ ‡ç­¾ - éšè—çŠ¶æ€æ˜¾ç¤º
        self.active_cattle_label = QLabel("")
        self.active_cattle_label.setVisible(False)  # éšè—çŠ¶æ€æ˜¾ç¤º
        
        # æ¸…é™¤åœ¨ç¾¤ç‰›æŒ‰é’®
        self.clear_active_cattle_btn = QPushButton("ğŸ—‘ï¸ æ¸…é™¤åœ¨ç¾¤ç‰›")
        self.clear_active_cattle_btn.setStyleSheet(button_styles['danger'])
        self.clear_active_cattle_btn.clicked.connect(self.clear_active_cattle)
        self.clear_active_cattle_btn.setVisible(False)
        # ç§»é™¤æœ€å¤§é«˜åº¦é™åˆ¶ï¼Œä½¿ç”¨æ ·å¼ä¸­çš„ç»Ÿä¸€é«˜åº¦
        active_cattle_layout.addWidget(self.clear_active_cattle_btn)
        
        tab_layout.addWidget(active_cattle_group)
        
        # 3. åŸºç¡€ç­›é€‰æ¡ä»¶åŒºåŸŸ - ç®€æ´è®¾è®¡ï¼Œç»Ÿä¸€è¾¹è·ç­–ç•¥
        basic_filter_group = self.create_card_widget("ğŸ”§ åŸºç¡€ç­›é€‰æ¡ä»¶")
        basic_filter_layout = QFormLayout(getattr(basic_filter_group, 'content_widget'))
        basic_filter_layout.setContentsMargins(0, 4, 0, 4)  # åªä¿ç•™ä¸Šä¸‹4pxçš„ç»†å¾®è¾¹è·
        basic_filter_layout.setVerticalSpacing(8)  # é€‚ä¸­çš„è¡¨å•é¡¹é—´è·
        basic_filter_layout.setHorizontalSpacing(10)
        basic_filter_layout.setLabelAlignment(Qt.AlignmentFlag.AlignLeft)  # è®¾ç½®æ ‡ç­¾å·¦å¯¹é½
        basic_filter_layout.setFormAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)  # è®¾ç½®è¡¨å•å·¦å¯¹é½
        
        # èƒæ¬¡èŒƒå›´ç­›é€‰
        parity_layout = QHBoxLayout()
        parity_layout.setSpacing(4)  # å‡å°‘é—´è·
        self.parity_min = QSpinBox()
        self.parity_min.setRange(1, 99)
        self.parity_min.setValue(1)
        self.parity_min.setStyleSheet(form_styles)
        # ç§»é™¤æœ€å¤§é«˜åº¦é™åˆ¶ï¼Œä½¿ç”¨æ ·å¼ä¸­çš„ç»Ÿä¸€é«˜åº¦
        
        self.parity_max = QSpinBox()
        self.parity_max.setRange(1, 99)
        self.parity_max.setValue(99)
        self.parity_max.setStyleSheet(form_styles)
        # ç§»é™¤æœ€å¤§é«˜åº¦é™åˆ¶ï¼Œä½¿ç”¨æ ·å¼ä¸­çš„ç»Ÿä¸€é«˜åº¦
        
        parity_layout.addWidget(QLabel("ä»"))
        parity_layout.addWidget(self.parity_min)
        parity_layout.addWidget(QLabel("åˆ°"))
        parity_layout.addWidget(self.parity_max)
        parity_layout.addWidget(QLabel("èƒ"))
        parity_layout.addStretch()
        basic_filter_layout.addRow("èƒæ¬¡èŒƒå›´:", parity_layout)
        
        # æ—¥æœŸèŒƒå›´ç­›é€‰
        date_layout = QHBoxLayout()
        date_layout.setSpacing(4)  # å‡å°‘é—´è·
        self.date_start = QDateEdit()
        self.date_start.setCalendarPopup(True)
        self.date_start.setDate(QDate.currentDate().addMonths(-12))  # é»˜è®¤ä¸€å¹´å‰
        self.date_start.setStyleSheet(form_styles)
        # ç§»é™¤æœ€å¤§é«˜åº¦é™åˆ¶ï¼Œä½¿ç”¨æ ·å¼ä¸­çš„ç»Ÿä¸€é«˜åº¦
        
        self.date_end = QDateEdit()
        self.date_end.setCalendarPopup(True)
        self.date_end.setDate(QDate.currentDate())  # é»˜è®¤ä»Šå¤©
        self.date_end.setStyleSheet(form_styles)
        # ç§»é™¤æœ€å¤§é«˜åº¦é™åˆ¶ï¼Œä½¿ç”¨æ ·å¼ä¸­çš„ç»Ÿä¸€é«˜åº¦
        
        date_layout.addWidget(self.date_start)
        date_layout.addWidget(QLabel("è‡³"))
        date_layout.addWidget(self.date_end)
        date_layout.addStretch()
        basic_filter_layout.addRow("æ—¥æœŸèŒƒå›´:", date_layout)
        
        tab_layout.addWidget(basic_filter_group)
        
        # ===========================================
        # ğŸ¯ å…³é”®å¸ƒå±€ç­–ç•¥ï¼šé¡¶éƒ¨å¯¹é½è®¾è®¡
        # ===========================================
        # è¿™æ˜¯ç•Œé¢ä¼˜åŒ–çš„æ ¸å¿ƒï¼addStretch()è®©æ‰€æœ‰å†…å®¹ç´§è´´ä¸Šæ–¹æ’åˆ—
        # 
        # æ•ˆæœè¯´æ˜ï¼š
        # - æ–‡ä»¶ä¸Šä¼ ã€åœ¨ç¾¤ç‰›æ–‡ä»¶ã€åŸºç¡€ç­›é€‰æ¡ä»¶éƒ½é›†ä¸­åœ¨é¡¶éƒ¨
        # - ç”¨æˆ·æ‰“å¼€æ ‡ç­¾é¡µç«‹å³çœ‹åˆ°æ‰€æœ‰é‡è¦åŠŸèƒ½ï¼Œæ— éœ€æ»šåŠ¨
        # - ä¸‹æ–¹ç•™ç©ºä¸å½±å“ä½¿ç”¨ï¼Œç¬¦åˆç°ä»£ç•Œé¢è®¾è®¡ä¹ æƒ¯
        # - ç±»ä¼¼ç½‘é¡µè®¾è®¡çš„é¡¶éƒ¨å¯¹é½å¸ƒå±€
        #
        # å¯¹æ¯”å…¶ä»–æ ‡ç­¾é¡µï¼š
        # - åŸºç¡€æ•°æ®ã€éšæ€§ä¹³æˆ¿ç‚ç›‘æµ‹ï¼šä½¿ç”¨ addStretch() å®Œå…¨é¡¶éƒ¨å¯¹é½
        # - DHIç­›é€‰ã€æ…¢æ€§ä¹³æˆ¿ç‚ï¼šä½¿ç”¨ addStretch(1) é€‚åº¦åˆ†å¸ƒ
        tab_layout.addStretch()  # ğŸš€ é¡¶éƒ¨å¯¹é½çš„å…³é”®ä»£ç  - å†…å®¹é›†ä¸­ä¸Šæ–¹ï¼Œä¸‹æ–¹ç•™ç©º
        
        # æ·»åŠ åˆ°æ ‡ç­¾é¡µå®¹å™¨
        self.function_tabs.addTab(tab_widget, "ğŸ“Š åŸºç¡€æ•°æ®")

    def create_dhi_filter_tab(self):
        """åˆ›å»ºDHIåŸºç¡€ç­›é€‰æ ‡ç­¾é¡µï¼šè›‹ç™½ç­›é€‰ã€ä½“ç»†èƒæ•°ç­›é€‰ã€å…¶ä»–ç­›é€‰é¡¹ç›®ã€æœªæ¥æ³Œä¹³å¤©æ•°ç­›é€‰"""
        tab_widget = QWidget()
        tab_layout = QVBoxLayout(tab_widget)
        
        # è·å–è‡ªé€‚åº”æ ·å¼
        button_styles = self.get_responsive_button_styles()
        form_styles = self.get_responsive_form_styles()
        
        # ä½¿ç”¨åˆç†çš„è¾¹è·ï¼Œç¡®ä¿å†…å®¹å¯è§
        card_margin = self.get_dpi_scaled_size(10)
        
        # 1. è›‹ç™½ç‡ç­›é€‰
        protein_group = self.create_special_filter_group("ğŸ¥› è›‹ç™½ç‡ç­›é€‰", "protein")
        tab_layout.addWidget(protein_group)
        
        # 2. ä½“ç»†èƒæ•°ç­›é€‰
        somatic_group = self.create_special_filter_group("ğŸ”¬ ä½“ç»†èƒæ•°ç­›é€‰", "somatic")
        tab_layout.addWidget(somatic_group)
        
        # 3. å…¶ä»–ç­›é€‰é¡¹ç›®
        other_filters_group = self.create_card_widget("ğŸ“‹ å…¶ä»–ç­›é€‰é¡¹ç›®")
        other_filters_layout = QVBoxLayout(getattr(other_filters_group, 'content_widget'))
        other_filters_layout.setContentsMargins(card_margin, card_margin, card_margin, card_margin)
        
        # å¿«æ·é€‰æ‹©æŒ‰é’®
        buttons_layout = QHBoxLayout()
        
        add_common_btn = QPushButton("â• å¿«é€Ÿæ·»åŠ å¸¸ç”¨é¡¹")
        add_common_btn.setStyleSheet(button_styles['info'])
        add_common_btn.clicked.connect(self.quick_add_common_filters)
        buttons_layout.addWidget(add_common_btn)
        
        select_all_btn = QPushButton("âœ… å…¨é€‰")
        select_all_btn.setStyleSheet(button_styles['secondary'])
        select_all_btn.clicked.connect(self.select_all_filters)
        buttons_layout.addWidget(select_all_btn)
        
        clear_all_btn = QPushButton("âŒ å…¨æ¸…")
        clear_all_btn.setStyleSheet(button_styles['secondary'])
        clear_all_btn.clicked.connect(self.clear_all_filters)
        buttons_layout.addWidget(clear_all_btn)
        
        buttons_layout.addStretch()
        other_filters_layout.addLayout(buttons_layout)
        
        # åˆ›å»ºå¤šé€‰ç•Œé¢
        multi_select_widget = QWidget()
        multi_select_layout = QVBoxLayout(multi_select_widget)
        
        # æ ‡é¢˜å’Œä¸€é”®æ·»åŠ æŒ‰é’®
        header_layout = QHBoxLayout()
        select_label = QLabel("é€‰æ‹©è¦æ·»åŠ çš„ç­›é€‰é¡¹ç›®ï¼ˆå¯å¤šé€‰ï¼‰:")
        select_label.setStyleSheet("color: #495057; font-weight: bold; font-size: 13px;")
        header_layout.addWidget(select_label)
        header_layout.addStretch()
        
        apply_btn = QPushButton("åº”ç”¨é€‰ä¸­é¡¹ç›®")
        apply_btn.setStyleSheet(button_styles['success'])
        apply_btn.clicked.connect(self.apply_selected_filters)
        header_layout.addWidget(apply_btn)
        
        multi_select_layout.addLayout(header_layout)
        
        # ç­›é€‰é¡¹ç›®å¤é€‰æ¡†ç½‘æ ¼
        self.filter_checkboxes = {}
        filters_grid = QGridLayout()
        
        # ä»é…ç½®æ–‡ä»¶åŠ è½½å¯é€‰ç­›é€‰é¡¹
        available_filters = {
            'fat_pct': 'ä¹³è„‚ç‡(%)',
            'fat_protein_ratio': 'è„‚è›‹æ¯”',
            'somatic_cell_count': 'ä½“ç»†èƒæ•°(ä¸‡/ml)',
            'somatic_cell_score': 'ä½“ç»†èƒåˆ†',
            'urea_nitrogen': 'å°¿ç´ æ°®(mg/dl)',
            'lactose_pct': 'ä¹³ç³–ç‡',
            'milk_loss': 'å¥¶æŸå¤±(Kg)',
            'milk_payment_diff': 'å¥¶æ¬¾å·®',
            'economic_loss': 'ç»æµæŸå¤±',
            'corrected_milk': 'æ ¡æ­£å¥¶(Kg)',
            'persistency': 'æŒç»­åŠ›',
            'whi': 'WHI',
            'fore_milk_yield': 'å‰å¥¶é‡(Kg)',
            'fore_somatic_cell_count': 'å‰ä½“ç»†èƒ(ä¸‡/ml)',
            'fore_somatic_cell_score': 'å‰ä½“ç»†èƒåˆ†',
            'fore_milk_loss': 'å‰å¥¶æŸå¤±(Kg)',
            'peak_milk_yield': 'é«˜å³°å¥¶(Kg)',
            'peak_days': 'é«˜å³°æ—¥(å¤©)',
            'milk_305': '305å¥¶é‡(Kg)',
            'total_milk_yield': 'æ€»å¥¶é‡(Kg)',
            'total_fat_pct': 'æ€»ä¹³è„‚(%)',
            'total_protein_pct': 'æ€»è›‹ç™½(%)',
            'mature_equivalent': 'æˆå¹´å½“é‡(Kg)'
        }
        
        row = 0
        col = 0
        for filter_key, chinese_name in available_filters.items():
            checkbox = QCheckBox(chinese_name)
            checkbox.setStyleSheet(f"""
                QCheckBox {{
                    font-size: 12px;
                    color: #495057;
                    spacing: 6px;
                }}
                QCheckBox::indicator {{
                    width: 14px;
                    height: 14px;
                    border: 2px solid #ced4da;
                    border-radius: 3px;
                    background-color: white;
                }}
                QCheckBox::indicator:checked {{
                    background-color: #007bff;
                    border-color: #007bff;
                }}
            """)
            checkbox.toggled.connect(lambda checked, key=filter_key: self.on_filter_checkbox_toggled(key, checked))
            
            filters_grid.addWidget(checkbox, row, col)
            self.filter_checkboxes[filter_key] = checkbox
            
            col += 1
            if col >= 3:  # æ¯è¡Œ3ä¸ª
                col = 0
                row += 1
        
        multi_select_layout.addLayout(filters_grid)
        other_filters_layout.addWidget(multi_select_widget)
        
        # å·²æ·»åŠ çš„ç­›é€‰é¡¹å®¹å™¨
        added_label = QLabel("å·²æ·»åŠ çš„ç­›é€‰é¡¹:")
        added_label.setStyleSheet("color: #495057; font-weight: bold; font-size: 13px; margin-top: 10px;")
        other_filters_layout.addWidget(added_label)
        
        # åŠ¨æ€è°ƒæ•´çš„ç­›é€‰é¡¹å®¹å™¨ï¼ˆæ— æ»šåŠ¨æ¡ï¼‰
        self.filters_container = QWidget()
        self.filters_container.setMinimumWidth(580)  # è¿›ä¸€æ­¥å¢åŠ æœ€å°å®½åº¦
        self.filters_container.setMinimumHeight(200)  # è®¾ç½®æœ€å°é«˜åº¦ç¡®ä¿è¶³å¤Ÿæ˜¾ç¤ºç©ºé—´
        self.other_filters_layout = QVBoxLayout(self.filters_container)
        self.other_filters_layout.setContentsMargins(8, 8, 8, 8)
        self.other_filters_layout.setSpacing(8)  # å¢åŠ ç»„ä»¶é—´è·
        # å»æ‰addStretchï¼Œé¿å…å‹ç¼©ç­›é€‰é¡¹
        
        # ç›´æ¥æ·»åŠ å®¹å™¨ï¼Œä¸ä½¿ç”¨æ»šåŠ¨åŒºåŸŸ
        other_filters_layout.addWidget(self.filters_container)
        
        tab_layout.addWidget(other_filters_group)
        
        # 4. æœªæ¥æ³Œä¹³å¤©æ•°ç­›é€‰
        future_days_group = self.create_card_widget("ğŸ“… æœªæ¥æ³Œä¹³å¤©æ•°ç­›é€‰")
        future_days_layout = QVBoxLayout(getattr(future_days_group, 'content_widget'))
        future_days_layout.setContentsMargins(card_margin, card_margin, card_margin, card_margin)
        
        # å¯ç”¨å¼€å…³
        self.future_days_enabled = QCheckBox("å¯ç”¨æœªæ¥æ³Œä¹³å¤©æ•°ç­›é€‰")
        self.future_days_enabled.setChecked(False)
        self.future_days_enabled.setStyleSheet(f"""
            QCheckBox {{
                font-size: 13px;
                color: #495057;
                spacing: 8px;
                font-weight: bold;
            }}
            QCheckBox::indicator {{
                width: 16px;
                height: 16px;
                border: 2px solid #ced4da;
                border-radius: 3px;
                background-color: white;
            }}
            QCheckBox::indicator:checked {{
                background-color: #28a745;
                border-color: #28a745;
            }}
        """)
        future_days_layout.addWidget(self.future_days_enabled)
        
        # å¤©æ•°èŒƒå›´è®¾ç½®
        future_range_layout = QHBoxLayout()
        
        range_label = QLabel("æœªæ¥æ³Œä¹³å¤©æ•°èŒƒå›´:")
        range_label.setStyleSheet("color: #495057; font-weight: bold;")
        future_range_layout.addWidget(range_label)
        
        self.future_days_min = QSpinBox()
        self.future_days_min.setRange(1, 500)
        self.future_days_min.setValue(50)
        self.future_days_min.setStyleSheet(form_styles)
        future_range_layout.addWidget(self.future_days_min)
        
        dash_label4 = QLabel("â€”")
        dash_label4.setAlignment(Qt.AlignmentFlag.AlignCenter)
        dash_margin = self.get_dpi_scaled_size(8)
        dash_label4.setStyleSheet(f"color: #6c757d; margin: 0 {dash_margin}px;")
        future_range_layout.addWidget(dash_label4)
        
        self.future_days_max = QSpinBox()
        self.future_days_max.setRange(1, 500)
        self.future_days_max.setValue(350)
        self.future_days_max.setStyleSheet(form_styles)
        future_range_layout.addWidget(self.future_days_max)
        
        future_range_layout.addStretch()
        
        future_range_widget = QWidget()
        future_range_widget.setLayout(future_range_layout)
        future_days_layout.addWidget(future_range_widget)
        
        # è®¡åˆ’è°ƒç¾¤æ—¥æœŸé€‰æ‹©å™¨
        plan_date_widget = QWidget()
        plan_date_layout = QHBoxLayout()
        plan_date_layout.setContentsMargins(0, 0, 0, 0)
        
        plan_date_label = QLabel("è®¡åˆ’è°ƒç¾¤æ—¥æœŸ:")
        plan_date_label.setStyleSheet("font-weight: bold;")
        plan_date_layout.addWidget(plan_date_label)
        
        self.plan_date = QDateEdit()
        self.plan_date.setCalendarPopup(True)
        self.plan_date.setDate(QDate.currentDate().addDays(30))  # é»˜è®¤30å¤©å
        self.plan_date.setStyleSheet(form_styles)
        plan_date_layout.addWidget(self.plan_date)
        
        plan_date_layout.addStretch()
        plan_date_widget.setLayout(plan_date_layout)
        future_days_layout.addWidget(plan_date_widget)
        
        # æ§åˆ¶èŒƒå›´è®¾ç½®çš„å¯ç”¨çŠ¶æ€
        def toggle_future_days_range():
            enabled = self.future_days_enabled.isChecked()
            self.future_days_min.setEnabled(enabled)
            self.future_days_max.setEnabled(enabled)
            self.plan_date.setEnabled(enabled)
            dash_label4.setEnabled(enabled)
            plan_date_label.setEnabled(enabled)
        
        self.future_days_enabled.toggled.connect(toggle_future_days_range)
        toggle_future_days_range()
        
        tab_layout.addWidget(future_days_group)
        
        # 5. æ“ä½œæŒ‰é’®
        action_group = self.create_card_widget("ğŸš€ æ“ä½œ")
        action_layout = QVBoxLayout(getattr(action_group, 'content_widget'))
        action_layout.setContentsMargins(card_margin, card_margin, card_margin, card_margin)
        action_spacing = self.get_dpi_scaled_size(10)
        action_layout.setSpacing(action_spacing)
        
        # ç­›é€‰æŒ‰é’®å®¹å™¨
        filter_buttons_layout = QHBoxLayout()
        
        # ç­›é€‰æŒ‰é’®
        self.filter_btn = QPushButton("ğŸ” å¼€å§‹ç­›é€‰")
        self.filter_btn.setStyleSheet(button_styles['warning'])
        self.filter_btn.clicked.connect(self.start_filtering)
        self.filter_btn.setEnabled(False)
        filter_buttons_layout.addWidget(self.filter_btn)
        
        # å–æ¶ˆç­›é€‰æŒ‰é’®
        self.cancel_filter_btn = QPushButton("â¹ï¸ å–æ¶ˆç­›é€‰")
        self.cancel_filter_btn.setStyleSheet(button_styles['danger'])
        self.cancel_filter_btn.clicked.connect(self.cancel_filtering)
        self.cancel_filter_btn.setEnabled(False)
        self.cancel_filter_btn.setVisible(False)
        filter_buttons_layout.addWidget(self.cancel_filter_btn)
        
        filter_buttons_widget = QWidget()
        filter_buttons_widget.setLayout(filter_buttons_layout)
        action_layout.addWidget(filter_buttons_widget)
        
        # å¯¼å‡ºæŒ‰é’®
        self.export_btn = QPushButton("ğŸ“¥ å¯¼å‡ºExcel")
        self.export_btn.setStyleSheet(button_styles['info'])
        self.export_btn.clicked.connect(self.export_results)
        self.export_btn.setEnabled(False)
        action_layout.addWidget(self.export_btn)
        
        # ç­›é€‰è¿›åº¦
        self.filter_progress = QProgressBar()
        self.filter_progress.setVisible(False)
        
        progress_border_radius = self.get_dpi_scaled_size(4)
        progress_padding = self.get_dpi_scaled_size(2)
        progress_chunk_radius = self.get_dpi_scaled_size(3)
        min_height = self.get_dpi_scaled_size(20)
        
        self.filter_progress.setStyleSheet(f"""
            QProgressBar {{
                border: 1px solid #e0e0e0;
                border-radius: {progress_border_radius}px;
                text-align: center;
                padding: {progress_padding}px;
                background-color: #f8f9fa;
                min-height: {min_height}px;
            }}
            QProgressBar::chunk {{
                background-color: #ffc107;
                border-radius: {progress_chunk_radius}px;
            }}
        """)
        action_layout.addWidget(self.filter_progress)
        
        self.filter_label = QLabel("")
        self.filter_label.setStyleSheet(f"color: #495057; font-weight: 500;")
        action_layout.addWidget(self.filter_label)
        
        tab_layout.addWidget(action_group)
        
        # æ·»åŠ é€‚é‡å¼¹æ€§ç©ºé—´ï¼Œä¿æŒå¸ƒå±€å¹³è¡¡
        tab_layout.addStretch(1)  # æ¢å¤å°‘é‡å¼¹æ€§ç©ºé—´ï¼Œé¿å…å†…å®¹è¿‡åº¦å‹ç¼©
        
        self.function_tabs.addTab(tab_widget, "ğŸ”¬ DHIåŸºç¡€ç­›é€‰")

    def load_filter_config(self, filter_key):
        """ä»é…ç½®æ–‡ä»¶åŠ è½½ç­›é€‰é¡¹ç›®é…ç½®"""
        # å®šä¹‰è‹±æ–‡åˆ°ä¸­æ–‡çš„æ˜ å°„
        name_mapping = {
            'fat_pct': 'ä¹³è„‚ç‡(%)',
            'fat_protein_ratio': 'è„‚è›‹æ¯”',
            'somatic_cell_count': 'ä½“ç»†èƒæ•°(ä¸‡/ml)',
            'somatic_cell_score': 'ä½“ç»†èƒåˆ†',
            'urea_nitrogen': 'å°¿ç´ æ°®(mg/dl)',
            'lactose_pct': 'ä¹³ç³–ç‡',
            'milk_loss': 'å¥¶æŸå¤±(Kg)',
            'milk_payment_diff': 'å¥¶æ¬¾å·®',
            'economic_loss': 'ç»æµæŸå¤±',
            'corrected_milk': 'æ ¡æ­£å¥¶(Kg)',
            'persistency': 'æŒç»­åŠ›',
            'whi': 'WHI',
            'fore_milk_yield': 'å‰å¥¶é‡(Kg)',
            'fore_somatic_cell_count': 'å‰ä½“ç»†èƒ(ä¸‡/ml)',
            'fore_somatic_cell_score': 'å‰ä½“ç»†èƒåˆ†',
            'fore_milk_loss': 'å‰å¥¶æŸå¤±(Kg)',
            'peak_milk_yield': 'é«˜å³°å¥¶(Kg)',
            'peak_days': 'é«˜å³°æ—¥(å¤©)',
            'milk_305': '305å¥¶é‡(Kg)',
            'total_milk_yield': 'æ€»å¥¶é‡(Kg)',
            'total_fat_pct': 'æ€»ä¹³è„‚(%)',
            'total_protein_pct': 'æ€»è›‹ç™½(%)',
            'mature_equivalent': 'æˆå¹´å½“é‡(Kg)'
        }
        
        try:
            import yaml
            with open('rules.yaml', 'r', encoding='utf-8') as f:
                config = yaml.safe_load(f)
                
            optional_filters = config.get('optional_filters', {})
            if filter_key in optional_filters:
                filter_config = optional_filters[filter_key]
                # å¦‚æœé…ç½®ä¸­æ²¡æœ‰ä¸­æ–‡åç§°ï¼Œä½¿ç”¨æ˜ å°„è¡¨
                if 'chinese_name' not in filter_config:
                    filter_config['chinese_name'] = name_mapping.get(filter_key, filter_key)
                return filter_config
            else:
                # è¿”å›é»˜è®¤é…ç½®ï¼Œä½¿ç”¨æ˜ å°„è¡¨ä¸­çš„ä¸­æ–‡åç§°
                return {
                    'chinese_name': name_mapping.get(filter_key, filter_key),
                    'min': 0.0,
                    'max': 100.0,
                    'min_match_months': 3,
                    'treat_empty_as_match': False
                }
        except Exception as e:
            print(f"åŠ è½½ç­›é€‰é…ç½®å¤±è´¥: {e}")
            # å‡ºé”™æ—¶è¿”å›é»˜è®¤é…ç½®ï¼Œä½¿ç”¨æ˜ å°„è¡¨ä¸­çš„ä¸­æ–‡åç§°
            return {
                'chinese_name': name_mapping.get(filter_key, filter_key),
                'min': 0.0,
                'max': 100.0,
                'min_match_months': 3,
                'treat_empty_as_match': False
            }
    
    def on_system_type_changed(self, system_type: str):
        """ç³»ç»Ÿç±»å‹æ”¹å˜æ—¶çš„å¤„ç†å‡½æ•°"""
        try:
            print(f"ç³»ç»Ÿç±»å‹å·²åˆ‡æ¢åˆ°: {system_type}")
            
            # è¿™é‡Œå¯ä»¥æ ¹æ®ç³»ç»Ÿç±»å‹æ‰§è¡Œç›¸åº”çš„é€»è¾‘
            # ä¾‹å¦‚ï¼šæ›´æ–°ç•Œé¢æ˜¾ç¤ºã€é‡ç½®æŸäº›è®¾ç½®ç­‰
            
            # æ›´æ–°çŠ¶æ€æ æ¶ˆæ¯
            self.safe_show_status_message(f"å·²åˆ‡æ¢åˆ°{system_type}ç³»ç»Ÿ")
            
            # å¦‚æœéœ€è¦ï¼Œå¯ä»¥åœ¨è¿™é‡Œæ·»åŠ æ›´å¤šçš„ç³»ç»Ÿç±»å‹åˆ‡æ¢é€»è¾‘
            # æ¯”å¦‚ï¼š
            # - æ›´æ–°é»˜è®¤çš„ç­›é€‰æ¡ä»¶
            # - é‡ç½®æŸäº›ç•Œé¢å…ƒç´ 
            # - åŠ è½½ç³»ç»Ÿç‰¹å®šçš„é…ç½®
            
        except Exception as e:
            print(f"ç³»ç»Ÿç±»å‹åˆ‡æ¢å¤„ç†å‡ºé”™: {e}")
            # ä¸æŠ›å‡ºå¼‚å¸¸ï¼Œé¿å…å½±å“ç¨‹åºè¿è¡Œ

    def create_mastitis_screening_tab(self):
        """åˆ›å»ºç‰§åœºæ…¢æ€§ä¹³æˆ¿ç‚æ„ŸæŸ“ç‰›ç­›æŸ¥å¤„ç½®æ ‡ç­¾é¡µ"""
        tab_widget = QWidget()
        tab_layout = QVBoxLayout(tab_widget)
        
        # è·å–è‡ªé€‚åº”æ ·å¼
        button_styles = self.get_responsive_button_styles()
        form_styles = self.get_responsive_form_styles()
        card_margin = self.get_dpi_scaled_size(12)
        
        # 1. ç³»ç»Ÿé€‰æ‹©åŒºåŸŸ
        system_group = self.create_card_widget("ğŸ¥ é€‰æ‹©æ•°æ®ç®¡ç†ç³»ç»Ÿ")
        system_layout = QVBoxLayout(getattr(system_group, 'content_widget'))
        system_layout.setContentsMargins(card_margin, card_margin, card_margin, card_margin)
        
        # ç³»ç»Ÿç±»å‹é€‰æ‹©
        system_selection_layout = QHBoxLayout()
        
        self.mastitis_system_group = QWidget()
        system_radio_layout = QHBoxLayout(self.mastitis_system_group)
        system_radio_layout.setContentsMargins(0, 0, 0, 0)
        
        self.yiqiniu_radio = QCheckBox("ä¼Šèµ·ç‰›ç³»ç»Ÿ")
        self.yiqiniu_radio.setStyleSheet("color: black; background-color: white; font-weight: bold;")
        self.huimuyun_radio = QCheckBox("æ…§ç‰§äº‘ç³»ç»Ÿ")
        self.huimuyun_radio.setStyleSheet("color: black; background-color: white; font-weight: bold;")
        self.custom_radio = QCheckBox("å…¶ä»–ç³»ç»Ÿ")
        self.custom_radio.setStyleSheet("color: black; background-color: white; font-weight: bold;")
        
        # è®¾ç½®ä¸ºå•é€‰æ¨¡å¼
        self.yiqiniu_radio.toggled.connect(lambda checked: self.on_mastitis_system_selected('yiqiniu', checked))
        self.huimuyun_radio.toggled.connect(lambda checked: self.on_mastitis_system_selected('huimuyun', checked))
        self.custom_radio.toggled.connect(lambda checked: self.on_mastitis_system_selected('custom', checked))
        
        system_radio_layout.addWidget(self.yiqiniu_radio)
        system_radio_layout.addWidget(self.huimuyun_radio)
        system_radio_layout.addWidget(self.custom_radio)
        system_radio_layout.addStretch()
        
        system_selection_layout.addWidget(self.mastitis_system_group)
        system_layout.addLayout(system_selection_layout)
        
        tab_layout.addWidget(system_group)
        
        # 2. æ–‡ä»¶ä¸Šä¼ åŒºåŸŸ - åŠ¨æ€æ˜¾ç¤º
        self.mastitis_upload_group = self.create_card_widget("ğŸ“ ä¸Šä¼ ç›¸å…³æ•°æ®æ–‡ä»¶")
        self.mastitis_upload_layout = QVBoxLayout(getattr(self.mastitis_upload_group, 'content_widget'))
        self.mastitis_upload_layout.setContentsMargins(card_margin, card_margin, card_margin, card_margin)
        
        # åˆå§‹åŒ–æ–‡ä»¶ä¸Šä¼ æ§ä»¶
        self.mastitis_file_uploads = {}
        
        tab_layout.addWidget(self.mastitis_upload_group)
        self.mastitis_upload_group.setVisible(False)  # é»˜è®¤éšè—
        
        # 3. æ…¢æ€§æ„ŸæŸ“ç‰›è¯†åˆ«é…ç½®
        chronic_group = self.create_card_widget("ğŸ”¬ æ…¢æ€§æ„ŸæŸ“ç‰›è¯†åˆ«æ ‡å‡†")
        chronic_layout = QFormLayout(getattr(chronic_group, 'content_widget'))
        chronic_layout.setContentsMargins(card_margin, card_margin, card_margin, card_margin)
        chronic_layout.setLabelAlignment(Qt.AlignmentFlag.AlignLeft)  # è®¾ç½®æ ‡ç­¾å·¦å¯¹é½
        chronic_layout.setFormAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)  # è®¾ç½®è¡¨å•å·¦å¯¹é½
        
        # æœˆä»½é€‰æ‹©è®¾ç½®
        self.chronic_months_widget = QWidget()
        chronic_months_layout = QGridLayout(self.chronic_months_widget)
        chronic_months_layout.setContentsMargins(0, 0, 0, 0)
        
        # åˆå§‹åŒ–æœˆä»½å¤é€‰æ¡†å­—å…¸
        self.chronic_month_checkboxes = {}
        
        # é»˜è®¤æ˜¾ç¤ºæç¤ºä¿¡æ¯
        no_data_label = QLabel("è¯·å…ˆä¸Šä¼ DHIæ•°æ®ä»¥é€‰æ‹©æœˆä»½")
        no_data_label.setStyleSheet("color: #495057; font-style: italic; font-weight: 500;")
        chronic_months_layout.addWidget(no_data_label, 0, 0, 1, 3)
        
        chronic_months_label = QLabel("é€‰æ‹©æ£€æŸ¥æœˆä»½:")
        chronic_months_label.setStyleSheet("color: black; background-color: white; font-weight: bold;")
        chronic_layout.addRow(chronic_months_label, self.chronic_months_widget)
        
        # ä½“ç»†èƒæ•°é˜ˆå€¼è®¾ç½®
        scc_threshold_layout = QHBoxLayout()
        scc_threshold_combo = QComboBox()
        scc_threshold_combo.addItems(["<", "<=", "=", ">=", ">"])
        scc_threshold_combo.setCurrentText(">=")
        scc_threshold_combo.setStyleSheet(form_styles)
        scc_threshold_combo.setFixedWidth(70)
        
        self.scc_threshold_spin = QDoubleSpinBox()
        self.scc_threshold_spin.setRange(0.1, 100.0)
        self.scc_threshold_spin.setValue(20.0)
        self.scc_threshold_spin.setSuffix("ä¸‡/ml")
        self.scc_threshold_spin.setDecimals(1)
        self.scc_threshold_spin.setStyleSheet(form_styles)
        
        scc_threshold_layout.addWidget(scc_threshold_combo)
        scc_threshold_layout.addWidget(self.scc_threshold_spin)
        scc_threshold_layout.addStretch()
        
        scc_threshold_widget = QWidget()
        scc_threshold_widget.setLayout(scc_threshold_layout)
        scc_label = QLabel("ä½“ç»†èƒæ•°:")
        scc_label.setStyleSheet("color: black; background-color: white; font-weight: bold;")
        chronic_layout.addRow(scc_label, scc_threshold_widget)
        
        # å­˜å‚¨é˜ˆå€¼æ¯”è¾ƒç¬¦å·
        self.scc_threshold_combo = scc_threshold_combo
        
        tab_layout.addWidget(chronic_group)
        
        # 4. å¤„ç½®åŠæ³•é…ç½®
        treatment_group = self.create_card_widget("âš•ï¸ å¤„ç½®åŠæ³•é…ç½®")
        treatment_layout = QVBoxLayout(getattr(treatment_group, 'content_widget'))
        treatment_layout.setContentsMargins(card_margin, card_margin, card_margin, card_margin)
        
        # åˆ›å»º5ç§å¤„ç½®åŠæ³•çš„é…ç½®ç•Œé¢
        self.treatment_configs = {}
        
        treatment_methods = [
            ('cull', 'æ·˜æ±°', 'ğŸ—‘ï¸'),
            ('isolate', 'ç¦é…éš”ç¦»', 'ğŸš«'),
            ('blind_quarter', 'çä¹³åŒº', 'ğŸ‘ï¸'),
            ('early_dry', 'æå‰å¹²å¥¶', 'â°'),
            ('treatment', 'æ²»ç–—', 'ğŸ’Š')
        ]
        
        for method_key, method_name, icon in treatment_methods:
            method_widget = self.create_treatment_config_widget(method_key, method_name, icon, form_styles)
            self.treatment_configs[method_key] = method_widget
            treatment_layout.addWidget(method_widget)
        
        tab_layout.addWidget(treatment_group)
        
        # 5. æ“ä½œæŒ‰é’®åŒºåŸŸ
        action_group = self.create_card_widget("ğŸš€ æ‰§è¡Œç­›æŸ¥")
        action_layout = QVBoxLayout(getattr(action_group, 'content_widget'))
        action_layout.setContentsMargins(card_margin, card_margin, card_margin, card_margin)
        
        buttons_layout = QHBoxLayout()
        
        self.mastitis_screen_btn = QPushButton("ğŸ” å¼€å§‹æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥")
        self.mastitis_screen_btn.setStyleSheet(button_styles['primary'])
        self.mastitis_screen_btn.clicked.connect(self.start_mastitis_screening)
        self.mastitis_screen_btn.setEnabled(False)
        buttons_layout.addWidget(self.mastitis_screen_btn)
        
        self.mastitis_export_btn = QPushButton("ğŸ“¤ å¯¼å‡ºç­›æŸ¥ç»“æœ")
        self.mastitis_export_btn.setStyleSheet(button_styles['success'])
        self.mastitis_export_btn.clicked.connect(self.export_mastitis_results)
        self.mastitis_export_btn.setEnabled(False)
        buttons_layout.addWidget(self.mastitis_export_btn)
        
        buttons_layout.addStretch()
        action_layout.addLayout(buttons_layout)
        
        # è¿›åº¦æ˜¾ç¤º
        progress_widget = QWidget()
        progress_layout = QVBoxLayout(progress_widget)
        progress_layout.setContentsMargins(0, 5, 0, 5)
        
        # è¿›åº¦æ¡
        self.mastitis_progress = QProgressBar()
        self.mastitis_progress.setRange(0, 100)
        self.mastitis_progress.setValue(0)
        self.mastitis_progress.setVisible(False)
        self.mastitis_progress.setStyleSheet("""
            QProgressBar {
                border: 1px solid #ccc;
                border-radius: 5px;
                text-align: center;
                padding: 1px;
                background-color: #f0f0f0;
                height: 25px;
                font-size: 12px;
                font-weight: bold;
            }
            QProgressBar::chunk {
                background-color: #4CAF50;
                border-radius: 4px;
            }
        """)
        progress_layout.addWidget(self.mastitis_progress)
        
        # è¿›åº¦çŠ¶æ€æ ‡ç­¾
        self.progress_status_label = QLabel("")
        self.progress_status_label.setStyleSheet("font-size: 12px; color: #495057; margin-top: 2px; font-weight: 500;")
        self.progress_status_label.setVisible(False)
        self.progress_status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        progress_layout.addWidget(self.progress_status_label)
        
        action_layout.addWidget(progress_widget)
        
        self.mastitis_status_label = QLabel("è¯·é€‰æ‹©æ•°æ®ç®¡ç†ç³»ç»Ÿå¹¶ä¸Šä¼ ç›¸å…³æ–‡ä»¶")
        self.mastitis_status_label.setStyleSheet("color: #495057; font-size: 14px; padding: 10px; font-weight: 500;")
        action_layout.addWidget(self.mastitis_status_label)
        
        tab_layout.addWidget(action_group)
        
        # æ·»åŠ é€‚é‡å¼¹æ€§ç©ºé—´ï¼Œä¿æŒå¸ƒå±€å¹³è¡¡
        tab_layout.addStretch(1)  # æ¢å¤å°‘é‡å¼¹æ€§ç©ºé—´ï¼Œé¿å…å†…å®¹è¿‡åº¦å‹ç¼©
        
        # åˆå§‹åŒ–å˜é‡
        self.current_mastitis_system = None
        self.mastitis_screening_results = None
        
        self.function_tabs.addTab(tab_widget, "ğŸ¥ æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥")

    def create_mastitis_monitoring_tab(self):
        """åˆ›å»ºéšæ€§ä¹³æˆ¿ç‚æœˆåº¦ç›‘æµ‹æ ‡ç­¾é¡µ"""
        try:
            import pyqtgraph as pg
        except ImportError:
            # å¦‚æœPyQtGraphæœªå®‰è£…ï¼Œæ˜¾ç¤ºé”™è¯¯ä¿¡æ¯
            tab_widget = QWidget()
            tab_layout = QVBoxLayout(tab_widget)
            error_label = QLabel("ç¼ºå°‘PyQtGraphä¾èµ–ï¼Œè¯·å®‰è£…: pip install pyqtgraph")
            error_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            error_label.setStyleSheet("color: #dc3545; padding: 20px;")
            tab_layout.addWidget(error_label)
            # æ·»åŠ å¼¹æ€§ç©ºé—´ï¼Œè®©å†…å®¹ç´§è´´ä¸Šæ–¹
            tab_layout.addStretch()  # å†…å®¹é›†ä¸­åœ¨ä¸Šæ–¹æ˜¾ç¤ºï¼Œä¸‹æ–¹ç•™ç©º
            self.function_tabs.addTab(tab_widget, "ğŸ‘ï¸ éšæ€§ä¹³æˆ¿ç‚ç›‘æµ‹")
            return

        tab_widget = QWidget()
        tab_layout = QVBoxLayout(tab_widget)
        tab_layout.setSpacing(10)
        tab_layout.setContentsMargins(15, 15, 15, 15)
        
        # åˆ›å»ºæ ‡é¢˜
        title_label = QLabel("éšæ€§ä¹³æˆ¿ç‚æœˆåº¦ç›‘æµ‹")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: black;
                background-color: white;
                padding: 10px 0px;
                border-bottom: 2px solid #3498db;
                margin-bottom: 15px;
            }
        """)
        tab_layout.addWidget(title_label)
        
        # 1. é¡¶éƒ¨é…ç½®åŒºåŸŸï¼ˆç®€åŒ–ä¸ºä¸€è¡Œï¼‰
        config_card = self.create_card_widget("ğŸ› ï¸ ç›‘æµ‹é…ç½®")
        config_layout = QHBoxLayout()
        config_layout.setSpacing(15)
        
        # ä½“ç»†èƒé˜ˆå€¼è®¾ç½®
        threshold_label = QLabel("ä½“ç»†èƒé˜ˆå€¼:")
        threshold_label.setStyleSheet("font-weight: bold; color: black; background-color: white;")
        
        self.monitoring_scc_threshold = QDoubleSpinBox()
        self.monitoring_scc_threshold.setRange(1.0, 100.0)
        self.monitoring_scc_threshold.setValue(20.0)
        self.monitoring_scc_threshold.setSuffix(" ä¸‡/ml")
        self.monitoring_scc_threshold.setMaximumWidth(150)
        self.monitoring_scc_threshold.setStyleSheet("""
            QDoubleSpinBox {
                border: 2px solid #ddd;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
                color: black;
                background-color: white;
                font-weight: bold;
            }
            QDoubleSpinBox:focus {
                border-color: #3498db;
                color: black;
            }
        """)
        
        # æŒ‰é’®ç»„
        button_styles = self.get_responsive_button_styles()
        
        self.start_monitoring_btn = QPushButton("ğŸ” å¼€å§‹åˆ†æ")
        self.start_monitoring_btn.setStyleSheet(button_styles['primary'])
        self.start_monitoring_btn.clicked.connect(self.start_mastitis_monitoring)
        self.start_monitoring_btn.setMaximumWidth(120)
        
        self.export_monitoring_btn = QPushButton("ğŸ“¤ å¯¼å‡ºExcel")
        self.export_monitoring_btn.setStyleSheet(button_styles['success'])
        self.export_monitoring_btn.clicked.connect(self.export_monitoring_results)
        self.export_monitoring_btn.setEnabled(False)
        self.export_monitoring_btn.setMaximumWidth(120)
        
        # çŠ¶æ€æ˜¾ç¤º
        self.monitoring_status_label = QLabel("è¯·å…ˆä¸Šä¼ ç‰›ç¾¤åŸºç¡€ä¿¡æ¯ï¼Œç„¶åä¸Šä¼ DHIæ•°æ®è¿›è¡Œåˆ†æ")
        self.monitoring_status_label.setStyleSheet("""
            QLabel {
                color: black;
                font-size: 12px;
                padding: 8px 12px;
                background-color: #f8f9fa;
                border-radius: 4px;
                border-left: 4px solid #ffc107;
                font-weight: bold;
            }
        """)
        

        
        # æ•°æ®çŠ¶æ€æ˜¾ç¤º
        self.monitoring_data_status = QLabel()
        self.monitoring_data_status.setStyleSheet("""
            QLabel {
                color: black;
                font-size: 11px;
                padding: 8px;
                background-color: #f8f9fa;
                border-radius: 4px;
                border-left: 4px solid #17a2b8;
                font-weight: bold;
            }
        """)
        self.update_monitoring_data_status()
        
        # æ·»åŠ åˆ°é…ç½®å¸ƒå±€
        config_layout.addWidget(threshold_label)
        config_layout.addWidget(self.monitoring_scc_threshold)
        config_layout.addWidget(self.start_monitoring_btn)
        config_layout.addWidget(self.export_monitoring_btn)
        config_layout.addStretch()
        
        config_card.layout().addLayout(config_layout)
        config_card.layout().addWidget(self.monitoring_data_status)
        config_card.layout().addWidget(self.monitoring_status_label)
        tab_layout.addWidget(config_card)
        
        # æ·»åŠ å¼¹æ€§ç©ºé—´ï¼Œè®©å†…å®¹ç´§è´´ä¸Šæ–¹
        tab_layout.addStretch()  # å†…å®¹é›†ä¸­åœ¨ä¸Šæ–¹æ˜¾ç¤ºï¼Œä¸‹æ–¹ç•™ç©º
        
        # åˆå§‹åŒ–ç›‘æµ‹è®¡ç®—å™¨
        self.mastitis_monitoring_calculator = None
        self.mastitis_monitoring_results = None
        
        self.function_tabs.addTab(tab_widget, "ğŸ‘ï¸ éšæ€§ä¹³æˆ¿ç‚ç›‘æµ‹")
        
        # åˆ›å»ºå°¿ç´ æ°®è¿½è¸ªæ ‡ç­¾é¡µ
        self.create_urea_tracking_tab()
    
    def create_urea_tracking_tab(self):
        """åˆ›å»ºå°¿ç´ æ°®è¿½è¸ªæ ‡ç­¾é¡µ"""
        tab_widget = QWidget()
        tab_layout = QVBoxLayout(tab_widget)
        tab_layout.setContentsMargins(20, 20, 20, 20)
        tab_layout.setSpacing(15)
        
        # åŠŸèƒ½è¯´æ˜æŒ‰é’®
        help_btn = QPushButton("â“ è¯´æ˜")
        help_btn.setObjectName("helpButton")
        help_btn.clicked.connect(self.show_urea_tracking_help)
        help_btn.setMaximumWidth(120)
        # å¼ºåˆ¶è®¾ç½®æŒ‰é’®æ ·å¼ï¼Œç¡®ä¿æ–‡å­—æ˜¯é»‘è‰²
        help_btn.setStyleSheet("""
            QPushButton {
                background-color: #f8f9fa;
                color: black;
                border: 1px solid #dee2e6;
                padding: 6px 12px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #e9ecef;
                border-color: #adb5bd;
            }
            QPushButton:pressed {
                background-color: #dee2e6;
            }
        """)
        tab_layout.addWidget(help_btn, alignment=Qt.AlignmentFlag.AlignRight)
        
        # ç§»é™¤å¯ç”¨å¼€å…³ï¼Œå› ä¸ºä¸éœ€è¦
        
        # åˆ†æè®¾ç½®ç»„
        analysis_group = QGroupBox("åˆ†æè®¾ç½®")
        analysis_layout = QVBoxLayout(analysis_group)
        
        self.urea_arithmetic_radio = QRadioButton("ç®—æœ¯å¹³å‡å€¼")
        self.urea_weighted_radio = QRadioButton("åŠ æƒå¹³å‡å€¼")
        self.urea_both_radio = QRadioButton("åŒæ—¶æ˜¾ç¤ºä¸¤è€…ï¼ˆé»˜è®¤ï¼‰")
        self.urea_both_radio.setChecked(True)
        
        analysis_layout.addWidget(self.urea_arithmetic_radio)
        analysis_layout.addWidget(self.urea_weighted_radio)
        analysis_layout.addWidget(self.urea_both_radio)
        
        tab_layout.addWidget(analysis_group)
        
        # æ•°æ®ç­›é€‰ç»„
        filter_group = QGroupBox("æ•°æ®ç­›é€‰")
        filter_layout = QGridLayout(filter_group)
        
        self.urea_filter_outliers = QCheckBox("ç­›é€‰å¼‚å¸¸å€¼")
        self.urea_filter_outliers.toggled.connect(self.on_urea_filter_toggled)
        filter_layout.addWidget(self.urea_filter_outliers, 0, 0, 1, 2)
        
        filter_layout.addWidget(QLabel("å°¿ç´ æ°®èŒƒå›´ï¼š"), 1, 0, 1, 2)
        
        filter_layout.addWidget(QLabel("æœ€å°å€¼ï¼š"), 2, 0)
        self.urea_min_value = QSpinBox()
        self.urea_min_value.setRange(0, 50)
        self.urea_min_value.setValue(5)
        self.urea_min_value.setSuffix(" mg/dl")
        self.urea_min_value.setEnabled(False)
        filter_layout.addWidget(self.urea_min_value, 2, 1)
        
        filter_layout.addWidget(QLabel("æœ€å¤§å€¼ï¼š"), 3, 0)
        self.urea_max_value = QSpinBox()
        self.urea_max_value.setRange(0, 100)
        self.urea_max_value.setValue(30)
        self.urea_max_value.setSuffix(" mg/dl")
        self.urea_max_value.setEnabled(False)
        filter_layout.addWidget(self.urea_max_value, 3, 1)
        
        tab_layout.addWidget(filter_group)
        
        # æ³Œä¹³å¤©æ•°åˆ†ç»„
        lactation_group = QGroupBox("æ³Œä¹³å¤©æ•°åˆ†ç»„")
        lactation_layout = QVBoxLayout(lactation_group)
        
        # åˆ›å»ºåˆ†ç»„å¤é€‰æ¡†
        self.urea_lactation_groups = {}
        groups = [
            "1-30å¤©", "31-60å¤©", "61-90å¤©", "91-120å¤©",
            "121-150å¤©", "151-180å¤©", "181-210å¤©", "211-240å¤©",
            "241-270å¤©", "271-300å¤©", "301-330å¤©", "331å¤©ä»¥ä¸Š"
        ]
        
        # ä½¿ç”¨ç½‘æ ¼å¸ƒå±€æ˜¾ç¤ºåˆ†ç»„
        groups_widget = QWidget()
        groups_grid = QGridLayout(groups_widget)
        groups_grid.setSpacing(10)
        
        for i, group in enumerate(groups):
            checkbox = QCheckBox(group)
            checkbox.setChecked(True)  # é»˜è®¤å…¨é€‰
            self.urea_lactation_groups[group] = checkbox
            groups_grid.addWidget(checkbox, i // 3, i % 3)
        
        lactation_layout.addWidget(groups_widget)
        
        # å…¨é€‰/æ¸…é™¤æŒ‰é’®
        btn_layout = QHBoxLayout()
        select_all_btn = QPushButton("å…¨é€‰")
        select_all_btn.clicked.connect(lambda: self.toggle_urea_groups(True))
        clear_all_btn = QPushButton("æ¸…é™¤")
        clear_all_btn.clicked.connect(lambda: self.toggle_urea_groups(False))
        btn_layout.addWidget(select_all_btn)
        btn_layout.addWidget(clear_all_btn)
        btn_layout.addStretch()
        lactation_layout.addLayout(btn_layout)
        
        tab_layout.addWidget(lactation_group)
        
        # æœ€å°‘æ ·æœ¬æ•°è®¾ç½®
        sample_layout = QHBoxLayout()
        sample_layout.addWidget(QLabel("æœ€å°‘æ ·æœ¬æ•°ï¼š"))
        self.urea_min_sample = QSpinBox()
        self.urea_min_sample.setRange(1, 50)
        self.urea_min_sample.setValue(5)
        self.urea_min_sample.setSuffix("å¤´")
        sample_layout.addWidget(self.urea_min_sample)
        sample_layout.addWidget(QLabel("ï¼ˆå°‘äºæ­¤æ•°é‡çš„ç»„ä¸æ˜¾ç¤ºï¼‰"))
        sample_layout.addStretch()
        tab_layout.addLayout(sample_layout)
        
        # åˆ†ææŒ‰é’®
        btn_layout = QHBoxLayout()
        
        # å¼€å§‹åˆ†ææŒ‰é’®
        self.urea_analyze_btn = QPushButton("å¼€å§‹åˆ†æ")
        self.urea_analyze_btn.setEnabled(False)  # é»˜è®¤ç¦ç”¨ï¼Œç›´åˆ°æœ‰æ•°æ®
        self.urea_analyze_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
                border-radius: 5px;
                min-height: 35px;
            }
            QPushButton:hover {
                background-color: #218838;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)
        self.urea_analyze_btn.clicked.connect(self.perform_urea_tracking_analysis)
        btn_layout.addWidget(self.urea_analyze_btn)
        
        # å¯¼å‡ºExcelæŒ‰é’®  
        self.urea_export_btn = QPushButton("å¯¼å‡ºExcel")
        self.urea_export_btn.setEnabled(False)  # é»˜è®¤ç¦ç”¨ï¼Œç›´åˆ°æœ‰ç»“æœ
        self.urea_export_btn.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
                border-radius: 5px;
                min-height: 35px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)
        self.urea_export_btn.clicked.connect(self.export_urea_tracking_results)
        btn_layout.addWidget(self.urea_export_btn)
        
        btn_layout.addStretch()
        tab_layout.addLayout(btn_layout)
        
        # æ·»åŠ å¼¹æ€§ç©ºé—´
        tab_layout.addStretch()
        
        self.function_tabs.addTab(tab_widget, "ğŸ§ª å°¿ç´ æ°®è¿½è¸ª")
    
    def show_urea_tracking_help(self):
        """æ˜¾ç¤ºå°¿ç´ æ°®è¿½è¸ªåŠŸèƒ½è¯´æ˜"""
        help_text = """
        <h2>å°¿ç´ æ°®è¿½è¸ªåˆ†æè¯´æ˜</h2>
        
        <h3>åŠŸèƒ½æ¦‚è¿°</h3>
        <p>æœ¬åŠŸèƒ½é€šè¿‡è¿½è¸ªä¸åŒæ³Œä¹³é˜¶æ®µç‰›ç¾¤çš„å°¿ç´ æ°®æ°´å¹³å˜åŒ–ï¼Œå¸®åŠ©æ‚¨ä¼˜åŒ–æ—¥ç²®é…æ–¹å’Œè›‹ç™½è´¨åˆ©ç”¨æ•ˆç‡ã€‚</p>
        
        <h3>åˆ†æé€»è¾‘</h3>
        <ol>
            <li><b>åˆ†ç»„ä¾æ®</b>ï¼šåŸºäºæœ€æ–°ä¸€æ¬¡DHIæ•°æ®ä¸­æ¯å¤´ç‰›çš„æ³Œä¹³å¤©æ•°è¿›è¡Œåˆ†ç»„<br>
                ä¾‹å¦‚ï¼š2024å¹´4æœˆæ•°æ®ä¸­ï¼Œæ³Œä¹³å¤©æ•°75å¤©çš„ç‰›è¢«åˆ†åˆ°"2024å¹´4æœˆ 61-90å¤©"ç»„</li>
            <li><b>å†å²è¿½è¸ª</b>ï¼š
                <ul>
                    <li>ç¡®å®šå½“å‰ç»„å†…çš„ç‰›åªï¼ˆå¦‚50å¤´ï¼‰</li>
                    <li>åœ¨å†å²DHIæ•°æ®ä¸­æŸ¥æ‰¾è¿™äº›ç‰›çš„å°¿ç´ æ°®è®°å½•</li>
                    <li>å¦‚æŸæœˆåªæœ‰30å¤´æœ‰æ•°æ®ï¼Œåˆ™è®¡ç®—è¿™30å¤´çš„å¹³å‡å€¼</li>
                </ul>
            </li>
            <li><b>è®¡ç®—æ–¹æ³•</b>ï¼š
                <ul>
                    <li>ç®—æœ¯å¹³å‡ï¼šæ‰€æœ‰ç‰›åªå°¿ç´ æ°®å€¼çš„ç®€å•å¹³å‡</li>
                    <li>åŠ æƒå¹³å‡ï¼šè€ƒè™‘äº§å¥¶é‡çš„åŠ æƒå¹³å‡<br>
                        å…¬å¼ï¼šÎ£(å°¿ç´ æ°®Ã—äº§å¥¶é‡) / Î£(äº§å¥¶é‡)</li>
                </ul>
            </li>
        </ol>
        
        <h3>ä½¿ç”¨è¯´æ˜</h3>
        <ol>
            <li>ç¡®ä¿å·²åœ¨"åŸºç¡€æ•°æ®"ä¸­ä¸Šä¼ å¤šä¸ªæœˆä»½çš„DHIæ•°æ®</li>
            <li>é€‰æ‹©è¦åˆ†æçš„æ³Œä¹³å¤©æ•°ç»„</li>
            <li>è®¾ç½®å¼‚å¸¸å€¼ç­›é€‰èŒƒå›´ï¼ˆå¯é€‰ï¼‰</li>
            <li>ç‚¹å‡»ä¸»ç•Œé¢çš„"å¼€å§‹ç­›é€‰"æŒ‰é’®</li>
            <li>åœ¨"ç­›é€‰ç»“æœ-å°¿ç´ æ°®è¿½è¸ª"æŸ¥çœ‹ç»“æœ</li>
        </ol>
        
        <h3>ç»“æœè§£è¯»</h3>
        <ul>
            <li><b>ä¸Šå‡è¶‹åŠ¿</b>ï¼šå¯èƒ½è¡¨ç¤ºæ—¥ç²®è›‹ç™½è¿‡é‡æˆ–ç˜¤èƒƒèƒ½é‡ä¸è¶³</li>
            <li><b>ä¸‹é™è¶‹åŠ¿</b>ï¼šæ—¥ç²®è°ƒæ•´æœ‰æ•ˆï¼Œè›‹ç™½åˆ©ç”¨ç‡æé«˜</li>
            <li><b>æ­£å¸¸èŒƒå›´</b>ï¼šä¸€èˆ¬ä¸º12-18 mg/dl</li>
        </ul>
        
        <h3>æ³¨æ„äº‹é¡¹</h3>
        <ul>
            <li>åˆ†ç»„åŸºäºæœ€æ–°æ•°æ®ï¼Œå†å²æ•°æ®ä»…ç”¨äºè¶‹åŠ¿åˆ†æ</li>
            <li>æ ·æœ¬æ•°è¿‡å°‘çš„ç»„å¯èƒ½å¯¼è‡´æ•°æ®æ³¢åŠ¨è¾ƒå¤§</li>
        </ul>
        """
        
        dialog = QDialog(self)
        dialog.setWindowTitle("å°¿ç´ æ°®è¿½è¸ªåŠŸèƒ½è¯´æ˜")
        dialog.setMinimumSize(800, 600)
        
        layout = QVBoxLayout(dialog)
        
        # ä½¿ç”¨ QTextEdit æ˜¾ç¤º HTML
        text_edit = QTextEdit()
        text_edit.setReadOnly(True)
        text_edit.setHtml(help_text)
        layout.addWidget(text_edit)
        
        # å…³é—­æŒ‰é’®
        close_btn = QPushButton("å…³é—­")
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn, alignment=Qt.AlignmentFlag.AlignCenter)
        
        dialog.exec()
    
    def on_urea_filter_toggled(self, checked):
        """åˆ‡æ¢å¼‚å¸¸å€¼ç­›é€‰æ—¶å¯ç”¨/ç¦ç”¨è¾“å…¥æ¡†"""
        self.urea_min_value.setEnabled(checked)
        self.urea_max_value.setEnabled(checked)
    
    def toggle_urea_groups(self, select_all):
        """å…¨é€‰æˆ–æ¸…é™¤å°¿ç´ æ°®åˆ†ç»„"""
        for checkbox in self.urea_lactation_groups.values():
            checkbox.setChecked(select_all)
    
    def perform_urea_tracking_analysis(self):
        """æ‰§è¡Œå°¿ç´ æ°®è¿½è¸ªåˆ†æ"""
        # ç§»é™¤å¯ç”¨æ£€æŸ¥ï¼Œç›´æ¥è¿›è¡Œåˆ†æ
        
        if not self.urea_tracker.dhi_data_dict:
            QMessageBox.warning(self, "è­¦å‘Š", "æ²¡æœ‰å¯ç”¨çš„DHIæ•°æ®ï¼Œè¯·å…ˆä¸Šä¼ DHIæ–‡ä»¶")
            return
        
        # æ”¶é›†é€‰ä¸­çš„æ³Œä¹³å¤©æ•°ç»„
        selected_groups = []
        for group_name, checkbox in self.urea_lactation_groups.items():
            if checkbox.isChecked():
                selected_groups.append(group_name)
        
        if not selected_groups:
            QMessageBox.warning(self, "è­¦å‘Š", "è¯·è‡³å°‘é€‰æ‹©ä¸€ä¸ªæ³Œä¹³å¤©æ•°ç»„")
            return
        
        # è·å–æ˜¾ç¤ºæ¨¡å¼
        if self.urea_arithmetic_radio.isChecked():
            value_type = 'arithmetic'
        elif self.urea_weighted_radio.isChecked():
            value_type = 'weighted'
        else:
            value_type = 'both'
        
        try:
            # å®šä¹‰å¼‚æ­¥åˆ†æä»»åŠ¡
            def analyze_task(progress_callback=None, status_callback=None, check_cancelled=None):
                if status_callback:
                    status_callback("æ­£åœ¨å‡†å¤‡åˆ†æ...")
                
                # æ‰§è¡Œåˆ†æ
                results = self.urea_tracker.analyze(
                    selected_groups=selected_groups,
                    filter_outliers=self.urea_filter_outliers.isChecked(),
                    min_value=self.urea_min_value.value() if self.urea_filter_outliers.isChecked() else 5.0,
                    max_value=self.urea_max_value.value() if self.urea_filter_outliers.isChecked() else 30.0,
                    min_sample_size=self.urea_min_sample.value()
                )
                
                if progress_callback:
                    progress_callback(100)
                    
                return results
            
            # ä½¿ç”¨å¼‚æ­¥è¿›åº¦ç®¡ç†å™¨æ‰§è¡Œåˆ†æ
            progress_manager = AsyncProgressManager(self)
            results = progress_manager.execute_with_progress(
                analyze_task,
                title="å°¿ç´ æ°®è¿½è¸ªåˆ†æ",
                cancel_text="å–æ¶ˆ",
                total_steps=100
            )
            
            if 'error' in results:
                QMessageBox.warning(self, "è­¦å‘Š", results['error'])
                return
            
            if not results:
                QMessageBox.information(self, "æç¤º", "æ²¡æœ‰ç¬¦åˆæ¡ä»¶çš„æ•°æ®")
                return
            
            # ä¿å­˜ç»“æœ
            self.urea_tracking_results = {
                'results': results,
                'value_type': value_type
            }
            
            # åœ¨ç»“æœæ ‡ç­¾é¡µä¸­æ·»åŠ å°¿ç´ æ°®è¿½è¸ªæ ‡ç­¾
            self.add_urea_tracking_tab()
            
            # å¯ç”¨å¯¼å‡ºæŒ‰é’®
            self.urea_export_btn.setEnabled(True)
            
            # æ˜¾ç¤ºæˆåŠŸæ¶ˆæ¯
            group_count = len(results)
            total_history_points = sum(len(group['history']) for group in results.values())
            QMessageBox.information(
                self, 
                "åˆ†æå®Œæˆ", 
                f"å°¿ç´ æ°®è¿½è¸ªåˆ†æå®Œæˆï¼\n\n"
                f"åˆ†æäº† {group_count} ä¸ªæ³Œä¹³å¤©æ•°ç»„\n"
                f"å…± {total_history_points} ä¸ªå†å²æ•°æ®ç‚¹\n\n"
                f"è¯·æŸ¥çœ‹å³ä¾§'ç­›é€‰ç»“æœ'ä¸­çš„'å°¿ç´ æ°®è¿½è¸ª'æ ‡ç­¾é¡µ"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "é”™è¯¯", f"åˆ†æå¤±è´¥: {str(e)}")
    
    def create_urea_tracking_result_tab(self):
        """åˆ›å»ºå°¿ç´ æ°®è¿½è¸ªç»“æœæ ‡ç­¾é¡µï¼ˆåˆå§‹ä¸ºç©ºï¼‰"""
        tab_widget = QWidget()
        tab_layout = QVBoxLayout(tab_widget)
        
        # åˆ›å»ºä¸€ä¸ªå ä½æ ‡ç­¾
        placeholder = QLabel("å°¿ç´ æ°®è¿½è¸ªåˆ†æç»“æœå°†åœ¨è¿™é‡Œæ˜¾ç¤º\n\nè¯·å…ˆä¸Šä¼ DHIæ•°æ®ï¼Œç„¶ååœ¨å·¦ä¾§'å°¿ç´ æ°®è¿½è¸ª'æ ‡ç­¾é¡µä¸­ç‚¹å‡»'å¼€å§‹åˆ†æ'")
        placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        placeholder.setStyleSheet("""
            QLabel {
                color: #666;
                font-size: 14px;
                padding: 50px;
            }
        """)
        tab_layout.addWidget(placeholder)
        
        # ä¿å­˜å ä½ç¬¦å¼•ç”¨ï¼Œä»¥ä¾¿åç»­æ›¿æ¢
        self.urea_placeholder = placeholder
        self.urea_result_tab_widget = tab_widget
        self.urea_result_tab_layout = tab_layout
        
        self.result_sub_tabs.addTab(tab_widget, "ğŸ§ª å°¿ç´ æ°®è¿½è¸ª")
    
    def add_urea_tracking_tab(self):
        """æ›´æ–°å°¿ç´ æ°®è¿½è¸ªç»“æœæ ‡ç­¾é¡µçš„å†…å®¹"""
        if not self.urea_tracking_results or not hasattr(self, 'urea_result_tab_layout'):
            return
        
        # æ¸…ç©ºç°æœ‰å†…å®¹
        while self.urea_result_tab_layout.count():
            child = self.urea_result_tab_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
        
        # åˆ›å»ºåˆ†å‰²å™¨
        splitter = QSplitter(Qt.Orientation.Vertical)
        
        # ä¸Šéƒ¨åˆ†ï¼šå›¾è¡¨æ˜¾ç¤º
        chart_widget = QWidget()
        chart_layout = QVBoxLayout(chart_widget)
        
        # å›¾è¡¨æ ‡é¢˜
        chart_title = QLabel("å°¿ç´ æ°®å†å²è¶‹åŠ¿å›¾")
        chart_title.setStyleSheet("font-size: 16px; font-weight: bold; padding: 10px;")
        chart_layout.addWidget(chart_title)
        
        # åˆ›å»ºpyqtgraphå›¾è¡¨
        try:
            import pyqtgraph as pg
            
            # åˆ›å»ºå›¾è¡¨æ§ä»¶ï¼ˆä½¿ç”¨ä¸­æ–‡åŒ–çš„å›¾è¡¨éƒ¨ä»¶ï¼‰
            if ChinesePlotWidget:
                self.urea_chart = ChinesePlotWidget()
            else:
                self.urea_chart = pg.PlotWidget()
            
            self.urea_chart.setLabel('left', 'å°¿ç´ æ°® (mg/dl)', color='black')
            self.urea_chart.setLabel('bottom', 'æœˆä»½', color='black')
            self.urea_chart.showGrid(x=True, y=True, alpha=0.3)
            self.urea_chart.setMinimumHeight(250)
            
            # è®¾ç½®å›¾è¡¨èƒŒæ™¯è‰²å’Œè½´é¢œè‰²
            self.urea_chart.setBackground('white')
            axis_pen = pg.mkPen(color='black', width=1)
            self.urea_chart.getAxis('left').setPen(axis_pen)
            self.urea_chart.getAxis('left').setTextPen('black')
            self.urea_chart.getAxis('bottom').setPen(axis_pen)
            self.urea_chart.getAxis('bottom').setTextPen('black')
            
            # è®¾ç½®å›¾ä¾‹æ ·å¼
            legend = self.urea_chart.addLegend()
            legend.setBrush(pg.mkBrush(255, 255, 255, 200))  # ç™½è‰²åŠé€æ˜èƒŒæ™¯
            legend.setPen(pg.mkPen(0, 0, 0))  # é»‘è‰²è¾¹æ¡†
            
            # ç»˜åˆ¶æ•°æ®
            self.plot_urea_tracking_data()
            
            chart_layout.addWidget(self.urea_chart)
        except ImportError:
            # å¦‚æœæ²¡æœ‰å®‰è£…pyqtgraphï¼Œæ˜¾ç¤ºæç¤ºä¿¡æ¯
            chart_placeholder = QTextEdit()
            chart_placeholder.setReadOnly(True)
            chart_placeholder.setPlainText("å›¾è¡¨åŠŸèƒ½éœ€è¦å®‰è£…pyqtgraphåº“\n\nè¯·è¿è¡Œ: pip install pyqtgraph")
            chart_layout.addWidget(chart_placeholder)
        
        splitter.addWidget(chart_widget)
        
        # ä¸‹éƒ¨åˆ†ï¼šæ•°æ®è¡¨æ ¼
        table_widget = QWidget()
        table_layout = QVBoxLayout(table_widget)
        
        # è¡¨æ ¼æ ‡é¢˜
        table_title = QLabel("å°¿ç´ æ°®åˆ†ææ•°æ®è¡¨")
        table_title.setStyleSheet("font-size: 16px; font-weight: bold; padding: 10px;")
        table_layout.addWidget(table_title)
        
        # åˆ›å»ºæ•°æ®è¡¨æ ¼
        data_table = QTableWidget()
        
        # è®¾ç½®è¡¨æ ¼å“åº”å¼ç‰¹æ€§
        data_table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        data_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        data_table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        
        # ä»ç»“æœä¸­æ„å»ºè¡¨æ ¼æ•°æ®
        results = self.urea_tracking_results['results']
        value_type = self.urea_tracking_results['value_type']
        
        # è·å–æ±‡æ€»DataFrame
        summary_df = self.urea_tracker.get_summary_dataframe(results)
        
        if not summary_df.empty:
            # è®¾ç½®è¡¨æ ¼
            data_table.setRowCount(len(summary_df))
            data_table.setColumnCount(len(summary_df.columns))
            data_table.setHorizontalHeaderLabels(summary_df.columns.tolist())
            
            # å¡«å……æ•°æ®
            for row_idx in range(len(summary_df)):
                for col_idx in range(len(summary_df.columns)):
                    value = summary_df.iloc[row_idx, col_idx]
                    item = QTableWidgetItem(str(value))
                    data_table.setItem(row_idx, col_idx, item)
            
            # è®¾ç½®åˆ—å®½è°ƒæ•´æ¨¡å¼
            header = data_table.horizontalHeader()
            # è®¾ç½®å‰å‡ åˆ—ä¸ºå›ºå®šå®½åº¦
            header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # æ³Œä¹³å¤©æ•°ç»„
            # å…¶ä»–åˆ—è‡ªåŠ¨æ‹‰ä¼¸
            for i in range(1, data_table.columnCount()):
                header.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
            
            # è®¾ç½®æœ€åä¸€åˆ—æ‹‰ä¼¸å¡«å……å‰©ä½™ç©ºé—´
            header.setStretchLastSection(True)
            
            # åˆå§‹è°ƒæ•´åˆ—å®½
            data_table.resizeColumnsToContents()
            
            # è®¾ç½®æœ€å°åˆ—å®½
            for i in range(data_table.columnCount()):
                if data_table.columnWidth(i) < 80:
                    data_table.setColumnWidth(i, 80)
        
        table_layout.addWidget(data_table)
        
        splitter.addWidget(table_widget)
        
        # è®¾ç½®åˆ†å‰²æ¯”ä¾‹
        splitter.setSizes([300, 400])
        
        self.urea_result_tab_layout.addWidget(splitter)
        
        # åˆ‡æ¢åˆ°å°¿ç´ æ°®è¿½è¸ªæ ‡ç­¾é¡µ
        for i in range(self.result_sub_tabs.count()):
            if self.result_sub_tabs.tabText(i) == "ğŸ§ª å°¿ç´ æ°®è¿½è¸ª":
                self.result_sub_tabs.setCurrentIndex(i)
                break
        
        # åŒæ—¶åˆ‡æ¢ä¸»æ ‡ç­¾é¡µåˆ°ç­›é€‰ç»“æœ
        for i in range(self.tab_widget.count()):
            if self.tab_widget.tabText(i) == "ğŸ“Š ç­›é€‰ç»“æœ":
                self.tab_widget.setCurrentIndex(i)
                break
    
    def plot_urea_tracking_data(self):
        """ç»˜åˆ¶å°¿ç´ æ°®è¿½è¸ªæ•°æ®å›¾è¡¨"""
        if not self.urea_tracking_results or not hasattr(self, 'urea_chart'):
            return
        
        try:
            import pyqtgraph as pg
            
            # æ¸…ç©ºç°æœ‰å›¾è¡¨
            self.urea_chart.clear()
            
            # è·å–æ•°æ®
            results = self.urea_tracking_results['results']
            value_type = self.urea_tracking_results['value_type']
            
            # è·å–å›¾è¡¨æ•°æ®
            chart_data = self.urea_tracker.get_chart_data(results, value_type)
            
            if not chart_data['dates']:
                return
            
            # å‡†å¤‡Xè½´æ•°æ®ï¼ˆæœˆä»½ç´¢å¼•ï¼‰
            x_values = list(range(len(chart_data['dates'])))
            
            # é¢œè‰²åˆ—è¡¨
            colors = [
                (255, 0, 0),      # çº¢è‰²
                (0, 255, 0),      # ç»¿è‰²
                (0, 0, 255),      # è“è‰²
                (255, 255, 0),    # é»„è‰²
                (255, 0, 255),    # å“çº¢
                (0, 255, 255),    # é’è‰²
                (255, 128, 0),    # æ©™è‰²
                (128, 0, 255),    # ç´«è‰²
                (0, 128, 255),    # å¤©è“
                (255, 0, 128),    # ç«çº¢
                (128, 255, 0),    # é»„ç»¿
                (128, 128, 255),  # æ·¡ç´«
            ]
            
            # ç»˜åˆ¶æ¯ä¸ªç³»åˆ—çš„æ•°æ®
            for i, series in enumerate(chart_data['series']):
                # è¿‡æ»¤æ‰Noneå€¼
                valid_points = []
                valid_x = []
                for j, value in enumerate(series['data']):
                    if value is not None:
                        valid_points.append(value)
                        valid_x.append(x_values[j])
                
                if valid_points:
                    color = colors[i % len(colors)]
                    pen = pg.mkPen(color=color, width=2)
                    
                    # ç»˜åˆ¶çº¿æ¡
                    self.urea_chart.plot(
                        valid_x, valid_points,
                        pen=pen,
                        symbol='o',
                        symbolSize=8,
                        symbolBrush=color,
                        name=series['name']
                    )
            
            # è®¾ç½®Xè½´æ ‡ç­¾
            x_axis = self.urea_chart.getAxis('bottom')
            x_axis.setTicks([[(i, date) for i, date in enumerate(chart_data['dates'])]])
            
            # è®¾ç½®Yè½´èŒƒå›´
            self.urea_chart.setYRange(0, 35)  # å°¿ç´ æ°®çš„åˆç†èŒƒå›´
            
        except Exception as e:
            logger.error(f"ç»˜åˆ¶å°¿ç´ æ°®å›¾è¡¨å¤±è´¥: {e}")
            print(f"ç»˜åˆ¶å°¿ç´ æ°®å›¾è¡¨å¤±è´¥: {e}")
    
    def export_urea_tracking_results(self):
        """å¯¼å‡ºå°¿ç´ æ°®è¿½è¸ªç»“æœåˆ°Excel"""
        if not self.urea_tracking_results:
            QMessageBox.warning(self, "è­¦å‘Š", "æ²¡æœ‰å¯å¯¼å‡ºçš„å°¿ç´ æ°®è¿½è¸ªç»“æœ")
            return
        
        try:
            # é€‰æ‹©ä¿å­˜æ–‡ä»¶
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "ä¿å­˜å°¿ç´ æ°®è¿½è¸ªç»“æœ",
                f"å°¿ç´ æ°®è¿½è¸ªåˆ†æ_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excelæ–‡ä»¶ (*.xlsx)"
            )
            
            if not filename:
                return
            
            # åˆ›å»ºExcelå†™å…¥å™¨
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # 1. æ±‡æ€»è¡¨
                results = self.urea_tracking_results['results']
                summary_df = self.urea_tracker.get_summary_dataframe(results)
                if not summary_df.empty:
                    summary_df.to_excel(writer, sheet_name='æ±‡æ€»æ•°æ®', index=False)
                
                # 2. è¯¦ç»†ç‰›åªæ¸…å•
                detail_df = self.urea_tracker.get_detail_dataframe(results)
                if not detail_df.empty:
                    detail_df.to_excel(writer, sheet_name='è¯¦ç»†ç‰›åªæ¸…å•', index=False)
                
                # 3. åˆ†æè¯´æ˜
                info_data = {
                    'é¡¹ç›®': ['åˆ†ææ—¶é—´', 'æœ€æ–°æ•°æ®æœˆä»½', 'åˆ†æç»„æ•°', 'æ€»è®°å½•æ•°'],
                    'å€¼': [
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        self.urea_tracker.latest_date or 'æ— ',
                        len(results),
                        len(detail_df) if not detail_df.empty else 0
                    ]
                }
                info_df = pd.DataFrame(info_data)
                info_df.to_excel(writer, sheet_name='åˆ†æä¿¡æ¯', index=False)
            
            QMessageBox.information(self, "æˆåŠŸ", f"å°¿ç´ æ°®è¿½è¸ªç»“æœå·²å¯¼å‡ºåˆ°:\n{filename}")
            
        except Exception as e:
            QMessageBox.critical(self, "é”™è¯¯", f"å¯¼å‡ºå¤±è´¥: {str(e)}")
    
    def update_monitoring_data_status(self):
        """æ›´æ–°éšæ€§ä¹³æˆ¿ç‚ç›‘æµ‹çš„æ•°æ®çŠ¶æ€æ˜¾ç¤º - å–æ¶ˆæ‰€æœ‰çŠ¶æ€æ˜¾ç¤º"""
        # æ¸…ç©ºçŠ¶æ€æ˜¾ç¤º
        self.monitoring_data_status.setText("")
    
    def get_mastitis_monitoring_formula_html(self):
        """è·å–éšæ€§ä¹³æˆ¿ç‚ç›‘æµ‹å…¬å¼è¯´æ˜HTML"""
        return """
        <div style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <h3 style="color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 5px;">ğŸ“Š éšæ€§ä¹³æˆ¿ç‚æœˆåº¦ç›‘æµ‹æŒ‡æ ‡è®¡ç®—å…¬å¼</h3>
            
            <h4 style="color: #e67e22;">1. å½“æœˆæµè¡Œç‡ (%)</h4>
            <p><strong>å…¬å¼:</strong> ä½“ç»†èƒæ•°(ä¸‡/ml) > é˜ˆå€¼çš„ç‰›å¤´æ•° Ã· å½“æœˆå‚æµ‹ç‰›å¤´æ•° Ã— 100</p>
            <p><strong>æ•°æ®æ¥æº:</strong> DHIæŠ¥å‘Š - ä½“ç»†èƒè®¡æ•°å­—æ®µ</p>
            
            <h4 style="color: #e67e22;">2. æ–°å‘æ„ŸæŸ“ç‡ (%)</h4>
            <p><strong>å…¬å¼:</strong> (å½“æœˆSCC>é˜ˆå€¼ ä¸” ä¸ŠæœˆSCCâ‰¤é˜ˆå€¼çš„ç‰›å¤´æ•°) Ã· (ä¸ŠæœˆSCCâ‰¤é˜ˆå€¼çš„ç‰›å¤´æ•°) Ã— 100</p>
            <p><strong>æ•°æ®æ¥æº:</strong> è¿ç»­ä¸¤ä¸ªæœˆDHIæŠ¥å‘Šå¯¹æ¯”</p>
            <p><strong>æ³¨æ„:</strong> éœ€è¦è‡³å°‘2ä¸ªæœˆçš„æ•°æ®ï¼ŒåŸºäºç®¡ç†å·åŒ¹é…é‡å ç‰›åª</p>
            
            <h4 style="color: #e67e22;">3. æ…¢æ€§æ„ŸæŸ“ç‡ (%)</h4>
            <p><strong>å…¬å¼:</strong> (å½“æœˆSCC>é˜ˆå€¼ ä¸” ä¸ŠæœˆSCC>é˜ˆå€¼çš„ç‰›å¤´æ•°) Ã· (ä¸ŠæœˆSCC>é˜ˆå€¼çš„ç‰›å¤´æ•°) Ã— 100</p>
            <p><strong>æ•°æ®æ¥æº:</strong> è¿ç»­ä¸¤ä¸ªæœˆDHIæŠ¥å‘Šå¯¹æ¯”</p>
            
            <h4 style="color: #e67e22;">4. æ…¢æ€§æ„ŸæŸ“ç‰›å æ¯” (%)</h4>
            <p><strong>å…¬å¼:</strong> (å½“æœˆSCC>é˜ˆå€¼ ä¸” ä¸ŠæœˆSCC>é˜ˆå€¼çš„ç‰›å¤´æ•°) Ã· (å½“æœˆå‚æµ‹ç‰›å¤´æ•°) Ã— 100</p>
            <p><strong>æ•°æ®æ¥æº:</strong> è¿ç»­ä¸¤ä¸ªæœˆDHIæŠ¥å‘Šå¯¹æ¯”</p>
            
            <h4 style="color: #e67e22;">5. å¤´èƒ/ç»äº§é¦–æµ‹æµè¡Œç‡ (%)</h4>
            <p><strong>å…¬å¼:</strong> (èƒæ¬¡=1/èƒæ¬¡>1 ä¸” DIM5-35å¤© ä¸” SCC>é˜ˆå€¼çš„ç‰›å¤´æ•°) Ã· (ç›¸åº”èƒæ¬¡ä¸”DIM5-35å¤©çš„å‚æµ‹ç‰›å¤´æ•°) Ã— 100</p>
            <p><strong>æ•°æ®æ¥æº:</strong> DHIæŠ¥å‘Š - èƒæ¬¡ã€æ³Œä¹³å¤©æ•°ã€ä½“ç»†èƒè®¡æ•°å­—æ®µ</p>
            
            <h4 style="color: #e67e22;">6. å¹²å¥¶å‰æµè¡Œç‡ (%)</h4>
            <p><strong>å…¬å¼:</strong> (åœ¨èƒå¤©æ•°>180å¤© ä¸” SCC>é˜ˆå€¼çš„ç‰›å¤´æ•°) Ã· (åœ¨èƒå¤©æ•°>180å¤©çš„å‚æµ‹ç‰›å¤´æ•°) Ã— 100</p>
            <p><strong>æ•°æ®æ¥æº:</strong> DHIæŠ¥å‘Š + ç‰›ç¾¤åŸºç¡€ä¿¡æ¯ (ç®¡ç†å·ä¸è€³å·åŒ¹é…)</p>
            <p><strong>æ•°æ®è¦æ±‚:</strong></p>
            <ul>
                <li>å¿…é¡»å…ˆåœ¨"æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥"ä¸­ä¸Šä¼ ç‰›ç¾¤åŸºç¡€ä¿¡æ¯</li>
                <li>ç‰›ç¾¤åŸºç¡€ä¿¡æ¯éœ€åŒ…å«è€³å·å’Œåœ¨èƒå¤©æ•°å­—æ®µ</li>
                <li>DHIæ•°æ®çš„ç®¡ç†å·éœ€è¦èƒ½ä¸ç‰›ç¾¤åŸºç¡€ä¿¡æ¯çš„è€³å·åŒ¹é…</li>
                <li>ç³»ç»Ÿä¼šè‡ªåŠ¨å»é™¤å‰å¯¼0è¿›è¡ŒåŒ¹é…</li>
            </ul>
            <p><strong>å¸¸è§é—®é¢˜:</strong></p>
            <ul>
                <li>å¦‚æœæ˜¾ç¤º"æ•°æ®æ— æ³•åŒ¹é…"ï¼Œé€šå¸¸æ˜¯DHIæ•°æ®ä¸ç‰›ç¾¤ä¿¡æ¯æ¥è‡ªä¸åŒæ—¶é—´ç‚¹</li>
                <li>å¦‚æœæ˜¾ç¤º"æ— åœ¨èƒå¤©æ•°æ•°æ®"ï¼Œè¯´æ˜åŒ¹é…çš„ç‰›åªå½“æ—¶å¤„äºç©ºæ€€çŠ¶æ€</li>
                <li>å¦‚æœæ˜¾ç¤º"æ— ç¬¦åˆæ¡ä»¶ç‰›åª"ï¼Œè¯´æ˜å½“å‰æ²¡æœ‰åœ¨èƒå¤©æ•°>180å¤©çš„ç‰›</li>
            </ul>
            <p><strong>æ•°æ®è¦æ±‚:</strong></p>
            <ul>
                <li>å¿…é¡»å…ˆåœ¨"æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥"ä¸­ä¸Šä¼ ç‰›ç¾¤åŸºç¡€ä¿¡æ¯</li>
                <li>ç‰›ç¾¤åŸºç¡€ä¿¡æ¯éœ€åŒ…å«è€³å·å’Œåœ¨èƒå¤©æ•°å­—æ®µ</li>
                <li>DHIæ•°æ®çš„ç®¡ç†å·éœ€è¦èƒ½ä¸ç‰›ç¾¤åŸºç¡€ä¿¡æ¯çš„è€³å·åŒ¹é…</li>
                <li>ç³»ç»Ÿä¼šè‡ªåŠ¨å»é™¤å‰å¯¼0è¿›è¡ŒåŒ¹é…</li>
            </ul>
            <p><strong>å¸¸è§é—®é¢˜:</strong></p>
            <ul>
                <li>å¦‚æœæ˜¾ç¤º"æ•°æ®æ— æ³•åŒ¹é…"ï¼Œé€šå¸¸æ˜¯DHIæ•°æ®ä¸ç‰›ç¾¤ä¿¡æ¯æ¥è‡ªä¸åŒæ—¶é—´ç‚¹</li>
                <li>å¦‚æœæ˜¾ç¤º"æ— åœ¨èƒå¤©æ•°æ•°æ®"ï¼Œè¯´æ˜åŒ¹é…çš„ç‰›åªå½“æ—¶å¤„äºç©ºæ€€çŠ¶æ€</li>
                <li>å¦‚æœæ˜¾ç¤º"æ— ç¬¦åˆæ¡ä»¶ç‰›åª"ï¼Œè¯´æ˜å½“å‰æ²¡æœ‰åœ¨èƒå¤©æ•°>180å¤©çš„ç‰›</li>
            </ul>
            
            <h4 style="color: #27ae60;">âš ï¸ é‡è¦è¯´æ˜</h4>
            <ul>
                <li><strong>ä½“ç»†èƒé˜ˆå€¼:</strong> é»˜è®¤20ä¸‡/mlï¼Œå¯åœ¨ç•Œé¢ä¸Šæ–¹è°ƒæ•´</li>
                <li><strong>æ•°æ®åŒ¹é…:</strong> åŸºäºç®¡ç†å·æ ‡å‡†åŒ–åŒ¹é…ï¼Œè‡ªåŠ¨å»é™¤å‰å¯¼0</li>
                <li><strong>æœˆä»½è¿ç»­æ€§:</strong> ç³»ç»Ÿä¼šæ£€æŸ¥å¹¶æç¤ºæœˆä»½ç¼ºå¤±æƒ…å†µ</li>
                <li><strong>ç»Ÿè®¡æ„ä¹‰:</strong> é‡å ç‰›åª<20å¤´æ—¶ä¼šæ˜¾ç¤ºç»Ÿè®¡è­¦å‘Š</li>
                <li><strong>è®¡ç®—é€æ˜åº¦:</strong> è¡¨æ ¼ä¸­æ˜¾ç¤ºæ¯ä¸ªæŒ‡æ ‡çš„è¯¦ç»†è®¡ç®—è¿‡ç¨‹</li>
            </ul>
        </div>
        """
    
    def toggle_widget_visibility(self, widget):
        """åˆ‡æ¢æ§ä»¶æ˜¾ç¤º/éšè—çŠ¶æ€"""
        if widget.isVisible():
            widget.hide()
        else:
            widget.show()
    
    def create_result_panel(self):
        """åˆ›å»ºå³ä¾§ç»“æœé¢æ¿"""
        panel = QWidget()
        panel.setStyleSheet("""
            QWidget {
                background-color: white;
                border-radius: 8px;
            }
        """)
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(0, 0, 0, 0)
        
        # åˆ›å»ºæ ‡ç­¾é¡µ
        self.tab_widget = QTabWidget()
        
        # è‡ªé€‚åº”æ ‡ç­¾é¡µæ ·å¼
        screen = QApplication.primaryScreen()
        dpi_ratio = screen.devicePixelRatio()
        
        tab_font_size = self.get_dpi_scaled_font_size(16)  # ä¸»æ ‡ç­¾é¡µç»Ÿä¸€ä¸º16px
        tab_padding_v = max(int(10 * dpi_ratio * 0.6), 8)
        tab_padding_h = max(int(14 * dpi_ratio * 0.6), 10)
        tab_border_radius = max(int(5 * dpi_ratio * 0.6), 4)
        tab_min_width = 120  # ä¸å·¦ä¾§æ ‡ç­¾å®½åº¦ç»Ÿä¸€
        
        self.tab_widget.setStyleSheet(f"""
            QTabWidget::pane {{
                border: 1px solid #e0e0e0;
                border-radius: {tab_border_radius}px;
                background-color: white;
                margin-top: -1px;
            }}
            QTabBar::tab {{
                background-color: #f8f9fa;
                border: 1px solid #e0e0e0;
                padding: {tab_padding_v}px {tab_padding_h}px;
                margin-right: 2px;
                border-top-left-radius: {tab_border_radius}px;
                border-top-right-radius: {tab_border_radius}px;
                font-size: {tab_font_size}px;
                font-weight: 500;
                color: #495057;
                min-width: {tab_min_width}px;
                min-height: 35px;
            }}
            QTabBar::tab:selected {{
                background-color: white;
                border-bottom: 1px solid white;
                color: #007bff;
                font-weight: bold;
            }}
            QTabBar::tab:hover {{
                background-color: #e9ecef;
            }}
        """)
        layout.addWidget(self.tab_widget)
        
        # æ–‡ä»¶ä¿¡æ¯æ ‡ç­¾é¡µ
        self.file_info_widget = QTextEdit()
        self.file_info_widget.setReadOnly(True)
        info_font_size = 13  # ç»Ÿä¸€å­—ä½“å¤§å°ä¸º13px
        info_padding = max(int(12 * dpi_ratio * 0.6), 10)
        self.file_info_widget.setStyleSheet(f"""
            QTextEdit {{
                border: none;
                background-color: #f8f9fa;
                color: black;  /* å¼ºåˆ¶ä½¿ç”¨é»‘è‰²å­—ä½“ */
                font-family: 'Consolas', 'Monaco', monospace;
                font-size: {info_font_size}px;
                padding: {info_padding}px;
                line-height: 1.5;
            }}
        """)
        self.tab_widget.addTab(self.file_info_widget, "ğŸ“ æ–‡ä»¶ä¿¡æ¯")
        
        # å¤„ç†è¿‡ç¨‹æ ‡ç­¾é¡µ
        self.process_log_widget = QTextEdit()
        self.process_log_widget.setReadOnly(True)
        self.process_log_widget.setStyleSheet(f"""
            QTextEdit {{
                border: none;
                background-color: #f8f9fa;
                font-family: 'Consolas', 'Monaco', monospace;
                font-size: {info_font_size}px;
                padding: {info_padding}px;
                line-height: 1.5;
                color: black;  /* å¼ºåˆ¶ä½¿ç”¨é»‘è‰²å­—ä½“ */
            }}
        """)
        # è®¾ç½®åˆå§‹å†…å®¹
        self.process_log_widget.setText("ğŸ”„ å¤„ç†è¿‡ç¨‹æ—¥å¿—\n\nç­‰å¾…å¼€å§‹å¤„ç†æ–‡ä»¶...\n")
        self.tab_widget.addTab(self.process_log_widget, "ğŸ”„ å¤„ç†è¿‡ç¨‹")
        
        # ç­›é€‰ç»“æœæ ‡ç­¾é¡µ
        self.result_table = QTableWidget()
        table_font_size = max(int(14 * dpi_ratio * 0.8), 12)
        table_padding = max(int(8 * dpi_ratio * 0.6), 6)
        header_padding = max(int(10 * dpi_ratio * 0.6), 8)
        self.result_table.setStyleSheet(f"""
            QTableWidget {{
                border: none;
                background-color: white;
                gridline-color: #e0e0e0;
                font-size: {table_font_size}px;
            }}
            QTableWidget::item {{
                padding: {table_padding}px;
                border-bottom: 1px solid #f0f0f0;
            }}
            QTableWidget::item:selected {{
                background-color: #e3f2fd;
                color: #1976d2;
            }}
            QHeaderView::section {{
                background-color: #f8f9fa;
                padding: {header_padding}px;
                border: none;
                border-bottom: 2px solid #e0e0e0;
                font-weight: bold;
                color: #495057;
            }}
        """)
        # åˆ›å»ºç­›é€‰ç»“æœçš„æ¬¡çº§æ ‡ç­¾é¡µç»“æ„
        self.result_widget = self.create_result_sub_tabs()
        self.tab_widget.addTab(self.result_widget, "ğŸ“Š ç­›é€‰ç»“æœ")
        
        # ç­›é€‰åˆ†ææ ‡ç­¾é¡µï¼ˆåˆå¹¶ç»Ÿè®¡ä¿¡æ¯ï¼‰
        self.analysis_widget = self.create_analysis_panel()
        self.tab_widget.addTab(self.analysis_widget, "ğŸ¯ ç­›é€‰åˆ†æ")
        
        return panel
    
    def create_result_sub_tabs(self):
        """åˆ›å»ºç­›é€‰ç»“æœçš„æ¬¡çº§æ ‡ç­¾é¡µ"""
        result_widget = QWidget()
        layout = QVBoxLayout(result_widget)
        layout.setContentsMargins(5, 5, 5, 5)
        
        # åˆ›å»ºæ¬¡çº§æ ‡ç­¾é¡µå®¹å™¨
        self.result_sub_tabs = QTabWidget()
        # è·å–DPIç¼©æ”¾æ¯”ä¾‹
        screen = QApplication.primaryScreen()
        dpi_ratio = screen.devicePixelRatio()
        sub_tab_min_width = max(int(100 * dpi_ratio * 0.6), 80)  # æ¬¡çº§æ ‡ç­¾é¡µæœ€å°å®½åº¦
        
        self.result_sub_tabs.setStyleSheet(f"""
            QTabWidget::pane {{
                border: 1px solid #c0c0c0;
                border-radius: 4px;
                background-color: white;
            }}
            QTabBar::tab {{
                background-color: #f0f0f0;
                border: 1px solid #c0c0c0;
                padding: 10px 16px;
                margin-right: 2px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                font-size: 13px;
                min-width: 100px;
                min-height: 30px;
            }}
            QTabBar::tab:selected {{
                background-color: white;
                border-bottom-color: white;
                font-weight: bold;
            }}
            QTabBar::tab:hover {{
                background-color: #e9ecef;
            }}
        """)
        
        # æ¬¡çº§æ ‡ç­¾é¡µ1: DHIåŸºç¡€ç­›é€‰ç»“æœ (ä¿ç•™åŸæœ‰çš„ç»“æœè¡¨æ ¼)
        self.result_table = QTableWidget()
        self.result_table.setStyleSheet("""
            QTableWidget {
                border: none;
                background-color: white;
                alternate-background-color: #f8f9fa;
                gridline-color: #dee2e6;
                selection-background-color: #007bff;
                selection-color: white;
            }
            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #dee2e6;
            }
            QTableWidget::item:selected {
                background-color: #007bff;
                color: white;
            }
            QHeaderView::section {
                background-color: #f8f9fa;
                padding: 10px;
                border: none;
                border-bottom: 2px solid #e0e0e0;
                font-weight: bold;
                color: #495057;
            }
        """)
        self.result_sub_tabs.addTab(self.result_table, "ğŸ“Š DHIåŸºç¡€ç­›é€‰")
        
        # æ¬¡çº§æ ‡ç­¾é¡µ2: æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥ç»“æœ
        self.create_mastitis_screening_result_tab()
        
        # æ¬¡çº§æ ‡ç­¾é¡µ3: éšæ€§ä¹³æˆ¿ç‚ç›‘æµ‹
        self.create_mastitis_monitoring_result_tab()
        
        # æ¬¡çº§æ ‡ç­¾é¡µ4: å°¿ç´ æ°®è¿½è¸ª
        self.create_urea_tracking_result_tab()
        
        layout.addWidget(self.result_sub_tabs)
        return result_widget
    
    def create_mastitis_screening_result_tab(self):
        """åˆ›å»ºæ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥ç»“æœæ ‡ç­¾é¡µ"""
        tab_widget = QWidget()
        tab_layout = QVBoxLayout(tab_widget)
        tab_layout.setSpacing(0)
        tab_layout.setContentsMargins(0, 0, 0, 0)
        
        # ç›´æ¥åˆ›å»ºè¡¨æ ¼ï¼Œä¸æ·»åŠ ä»»ä½•å…¶ä»–ç»„ä»¶
        self.mastitis_screening_table = QTableWidget()
        self.mastitis_screening_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #ddd;
                background-color: white;
                gridline-color: #e0e0e0;
            }
            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #e0e0e0;
            }
            QTableWidget::item:selected {
                background-color: #ffeaa7;
                color: #2d3436;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                padding: 10px;
                border: 1px solid #ddd;
                font-weight: bold;
            }
        """)
        
        # æ·»åŠ ç©ºçŠ¶æ€æç¤º
        self.mastitis_screening_table.setRowCount(1)
        self.mastitis_screening_table.setColumnCount(1)
        self.mastitis_screening_table.setHorizontalHeaderLabels(["çŠ¶æ€"])
        
        empty_item = QTableWidgetItem("æš‚æ— ç­›æŸ¥ç»“æœï¼Œè¯·åœ¨å·¦ä¾§'æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥'åŠŸèƒ½ä¸­è¿›è¡Œç­›æŸ¥")
        empty_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.mastitis_screening_table.setItem(0, 0, empty_item)
        
        # ç›´æ¥æ·»åŠ è¡¨æ ¼åˆ°å¸ƒå±€ï¼Œä¸ä½¿ç”¨å¡ç‰‡å®¹å™¨
        tab_layout.addWidget(self.mastitis_screening_table)
        
        self.result_sub_tabs.addTab(tab_widget, "ğŸ¥ æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥")
    
    def create_mastitis_monitoring_result_tab(self):
        """åˆ›å»ºéšæ€§ä¹³æˆ¿ç‚ç›‘æµ‹ç»“æœæ ‡ç­¾é¡µ"""
        # æ£€æŸ¥PyQtGraphä¾èµ–
        try:
            import pyqtgraph as pg
        except ImportError:
            tab_widget = QWidget()
            tab_layout = QVBoxLayout(tab_widget)
            error_label = QLabel("ç¼ºå°‘PyQtGraphä¾èµ–ï¼Œè¯·å®‰è£…: pip install pyqtgraph")
            error_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            error_label.setStyleSheet("color: #dc3545; padding: 20px;")
            tab_layout.addWidget(error_label)
            tab_layout.addStretch()
            self.result_sub_tabs.addTab(tab_widget, "ğŸ‘ï¸ éšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹")
            return

        tab_widget = QWidget()
        tab_layout = QVBoxLayout(tab_widget)
        tab_layout.setSpacing(5)
        tab_layout.setContentsMargins(5, 5, 5, 5)
        
        # æ·»åŠ è¯´æ˜ä¿¡æ¯
        info_label = QLabel("ğŸ’¡ è¯·åœ¨å·¦ä¾§ã€éšå½¢ä¹³æˆ¿ç‚æœˆåº¦ç›‘æµ‹ã€‘æ ‡ç­¾é¡µä¸­è¿›è¡Œé…ç½®å’Œåˆ†æï¼Œç»“æœå°†åœ¨æ­¤å¤„æ˜¾ç¤º")
        info_label.setStyleSheet("""
            QLabel {
                color: black;
                font-size: 12px;
                padding: 8px 12px;
                background-color: #f8f9fa;
                border-radius: 4px;
                border-left: 4px solid #17a2b8;
                margin-bottom: 10px;
                font-weight: bold;
            }
        """)
        tab_layout.addWidget(info_label)
        
        # åˆ›å»ºä¸»è¦å†…å®¹åŒºåŸŸ - æ°´å¹³åˆ†å‰²ï¼šå·¦ä¾§è¡¨æ ¼ã€å³ä¾§å›¾è¡¨å’Œå…¬å¼è¯´æ˜
        main_splitter = QSplitter(Qt.Orientation.Horizontal)
        main_splitter.setSizes([600, 400])  # å·¦ä¾§è¡¨æ ¼600ï¼Œå³ä¾§400
        
        # å·¦ä¾§ï¼šç›‘æµ‹ç»“æœè¡¨æ ¼
        table_widget = QWidget()
        table_layout = QVBoxLayout(table_widget)
        table_layout.setContentsMargins(0, 0, 5, 0)
        
        table_title = QLabel("ğŸ“Š ç›‘æµ‹ç»“æœ")
        table_title.setStyleSheet("font-weight: bold; font-size: 14px; color: black; background-color: white; margin-bottom: 8px;")
        table_layout.addWidget(table_title)
        
        self.mastitis_monitoring_table = QTableWidget()
        self.mastitis_monitoring_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #ddd;
                background-color: white;
                gridline-color: #e0e0e0;
                alternate-background-color: #f8f9fa;
                color: black;
            }
            QTableWidget::item {
                padding: 6px;
                border-bottom: 1px solid #e0e0e0;
                color: black;
                background-color: white;
                font-weight: bold;
            }
            QTableWidget::item:selected {
                background-color: #e3f2fd;
                color: black;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                padding: 8px;
                border: 1px solid #ddd;
                font-weight: bold;
                font-size: 11px;
                color: black;
            }
        """)
        table_layout.addWidget(self.mastitis_monitoring_table)
        
        # å³ä¾§ï¼šå›¾è¡¨å’Œå…¬å¼è¯´æ˜çš„å‚ç›´åˆ†å‰²
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(3, 0, 0, 0)
        right_layout.setSpacing(5)
        
        # ä¸Šéƒ¨ï¼šè¶‹åŠ¿å›¾è¡¨
        if ChinesePlotWidget:
            self.mastitis_monitoring_plot = ChinesePlotWidget()
        else:
            self.mastitis_monitoring_plot = pg.PlotWidget()
        self.mastitis_monitoring_plot.setLabel('left', 'ç™¾åˆ†æ¯” (%)')
        self.mastitis_monitoring_plot.setLabel('bottom', 'æœˆä»½')
        self.mastitis_monitoring_plot.showGrid(x=True, y=True, alpha=0.3)
        self.mastitis_monitoring_plot.setBackground('white')
        self.mastitis_monitoring_plot.addLegend()
        self.mastitis_monitoring_plot.setMinimumHeight(250)
        
        right_layout.addWidget(self.mastitis_monitoring_plot)
        
        # ä¸‹éƒ¨ï¼šå¯æŠ˜å çš„å…¬å¼è¯´æ˜
        formula_container = QWidget()
        formula_layout = QVBoxLayout(formula_container)
        formula_layout.setContentsMargins(0, 5, 0, 0)
        
        # å…¬å¼æ ‡é¢˜å’ŒæŠ˜å æŒ‰é’®
        formula_header = QHBoxLayout()
        formula_title = QLabel("ğŸ“– å…¬å¼è¯´æ˜")
        formula_title.setStyleSheet("font-weight: bold; font-size: 13px; color: #2c3e50;")
        
        self.formula_toggle_btn = QPushButton("â–¼ å±•å¼€")
        self.formula_toggle_btn.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: none;
                color: #3498db;
                font-size: 11px;
                font-weight: bold;
                padding: 2px 8px;
            }
            QPushButton:hover {
                color: #2980b9;
                background-color: #ecf0f1;
                border-radius: 3px;
            }
        """)
        self.formula_toggle_btn.clicked.connect(self.toggle_monitoring_formula_visibility)
        
        formula_header.addWidget(formula_title)
        formula_header.addStretch()
        formula_header.addWidget(self.formula_toggle_btn)
        formula_layout.addLayout(formula_header)
        
        # å…¬å¼å†…å®¹ï¼ˆåˆå§‹éšè—ï¼‰
        self.monitoring_formula_widget = QTextEdit()
        self.monitoring_formula_widget.setReadOnly(True)
        self.monitoring_formula_widget.setMaximumHeight(180)
        self.monitoring_formula_widget.setHtml(self.get_mastitis_monitoring_formula_html())
        self.monitoring_formula_widget.setStyleSheet("""
            QTextEdit {
                border: 1px solid #ddd;
                background-color: #f8f9fa;
                font-family: 'Arial', sans-serif;
                font-size: 11px;
                padding: 8px;
            }
        """)
        self.monitoring_formula_widget.setVisible(False)  # åˆå§‹éšè—
        formula_layout.addWidget(self.monitoring_formula_widget)
        
        right_layout.addWidget(formula_container)
        
        # æ·»åŠ åˆ°ä¸»åˆ†å‰²å™¨
        main_splitter.addWidget(table_widget)
        main_splitter.addWidget(right_widget)
        main_splitter.setSizes([350, 450])  # è¡¨æ ¼é€‚ä¸­ï¼Œå³ä¾§è¾ƒå®½
        
        tab_layout.addWidget(main_splitter)
        
        # åˆå§‹åŒ–å˜é‡
        self.mastitis_monitoring_calculator = None
        self.mastitis_monitoring_results = None
        
        self.result_sub_tabs.addTab(tab_widget, "ğŸ‘ï¸ éšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹")
    
    def create_analysis_panel(self):
        """åˆ›å»ºç­›é€‰åˆ†æé¢æ¿ï¼ˆåŒ…å«ç»Ÿè®¡ä¿¡æ¯ï¼‰"""
        panel = QWidget()
        main_layout = QVBoxLayout(panel)
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # åˆ›å»ºå‚ç›´åˆ†å‰²å™¨
        splitter = QSplitter(Qt.Orientation.Vertical)
        
        # ä¸ŠåŠéƒ¨åˆ†ï¼šç»Ÿè®¡å¡ç‰‡
        top_widget = QWidget()
        top_layout = QVBoxLayout(top_widget)
        
        # æ ‡é¢˜
        title_label = QLabel("ç­›é€‰ç»“æœåˆ†æ")
        title_font = QFont()
        title_font.setPointSize(max(int(14 * 0.6), 12))
        title_font.setBold(True)
        title_label.setFont(title_font)
        top_layout.addWidget(title_label)
        
        # ç»Ÿè®¡å¡ç‰‡å®¹å™¨
        cards_widget = QWidget()
        cards_layout = QHBoxLayout(cards_widget)
        
        # åˆ›å»º4ä¸ªç»Ÿè®¡å¡ç‰‡
        self.total_data_card = self.create_stat_card("å…¨éƒ¨æ•°æ®", "0", "æ‰€æœ‰ä¸Šä¼ æ–‡ä»¶çš„ç‰›å¤´æ•°", "#3498db")
        self.range_data_card = self.create_stat_card("ç­›é€‰èŒƒå›´", "0", "é€‰ä¸­æ–‡ä»¶çš„ç‰›å¤´æ•°", "#17a2b8")
        self.result_data_card = self.create_stat_card("ç­›é€‰ç»“æœ", "0", "ç¬¦åˆæ¡ä»¶çš„ç‰›å¤´æ•°", "#28a745")
        self.rate_data_card = self.create_stat_card("ç­›é€‰ç‡", "0%", "ç¬¦åˆæ¡ä»¶å å…¨éƒ¨æ•°æ®æ¯”ä¾‹", "#ffc107")
        
        cards_layout.addWidget(self.total_data_card)
        cards_layout.addWidget(self.range_data_card)
        cards_layout.addWidget(self.result_data_card)
        cards_layout.addWidget(self.rate_data_card)
        
        top_layout.addWidget(cards_widget)
        top_layout.addStretch()
        
        # ä¸‹åŠéƒ¨åˆ†ï¼šç»Ÿè®¡ä¿¡æ¯é€‰é¡¹å¡
        bottom_widget = QWidget()
        bottom_layout = QVBoxLayout(bottom_widget)
        bottom_layout.setContentsMargins(5, 5, 5, 5)
        
        # åˆ›å»ºç»Ÿè®¡ä¿¡æ¯æ ‡ç­¾é¡µ
        self.stats_tab_widget = QTabWidget()
        
        # è·å–å±å¹•DPIä¿¡æ¯ç”¨äºæ ·å¼
        screen = QApplication.primaryScreen()
        dpi_ratio = screen.devicePixelRatio()
        
        tab_font_size = 13  # ä¸å…¶ä»–æ¬¡çº§æ ‡ç­¾é¡µä¸€è‡´
        tab_padding_v = 10
        tab_padding_h = 16
        tab_border_radius = 4
        tab_min_width = 100  # ä¸å…¶ä»–æ¬¡çº§æ ‡ç­¾é¡µä¸€è‡´
        
        self.stats_tab_widget.setStyleSheet(f"""
            QTabWidget::pane {{
                border: 1px solid #bee5eb;
                border-radius: {tab_border_radius}px;
                background-color: white;
                margin-top: -1px;
            }}
            QTabBar::tab {{
                background-color: #d1ecf1;
                border: 1px solid #bee5eb;
                padding: {tab_padding_v}px {tab_padding_h}px;
                margin-right: 2px;
                border-top-left-radius: {tab_border_radius}px;
                border-top-right-radius: {tab_border_radius}px;
                font-size: {tab_font_size}px;
                font-weight: 500;
                color: #0c5460;
                min-width: {tab_min_width}px;
                min-height: 30px;
            }}
            QTabBar::tab:selected {{
                background-color: white;
                border-bottom: 1px solid white;
                color: #007bff;
                font-weight: bold;
            }}
            QTabBar::tab:hover {{
                background-color: #b8daff;
            }}
        """)
        
        # åˆ›å»ºå„ä¸ªç»Ÿè®¡é€‰é¡¹å¡
        self.create_statistics_tabs()
        
        bottom_layout.addWidget(self.stats_tab_widget)
        
        # æ·»åŠ åˆ°åˆ†å‰²å™¨
        splitter.addWidget(top_widget)
        splitter.addWidget(bottom_widget)
        
        # è®¾ç½®åˆ†å‰²å™¨åˆå§‹å¤§å°æ¯”ä¾‹ (å¡ç‰‡åŒºåŸŸ:ç»Ÿè®¡ä¿¡æ¯åŒºåŸŸ = 1:2)
        splitter.setSizes([150, 300])
        splitter.setCollapsible(0, False)  # å¡ç‰‡åŒºåŸŸä¸å¯å®Œå…¨æŠ˜å 
        splitter.setCollapsible(1, False)  # ç»Ÿè®¡ä¿¡æ¯åŒºåŸŸä¸å¯å®Œå…¨æŠ˜å 
        
        # è®¾ç½®åˆ†å‰²å™¨æ ·å¼
        splitter.setStyleSheet("""
            QSplitter::handle {
                background-color: #dee2e6;
                border: 1px solid #ced4da;
                margin: 2px;
            }
            QSplitter::handle:horizontal {
                width: 3px;
            }
            QSplitter::handle:vertical {
                height: 3px;
            }
            QSplitter::handle:hover {
                background-color: #007bff;
            }
        """)
        
        main_layout.addWidget(splitter)
        
        return panel
    
    def create_statistics_tabs(self, enabled_traits=None):
        """åˆ›å»ºç»Ÿè®¡ä¿¡æ¯çš„å„ä¸ªé€‰é¡¹å¡
        
        Args:
            enabled_traits: å¯ç”¨çš„æ€§çŠ¶åˆ—è¡¨ï¼Œå¦‚æœä¸ºNoneåˆ™åˆ›å»ºé»˜è®¤é€‰é¡¹å¡
        """
        # æ¸…ç©ºç°æœ‰é€‰é¡¹å¡
        self.stats_tab_widget.clear()
        
        # è·å–å±å¹•DPIä¿¡æ¯
        screen = QApplication.primaryScreen()
        dpi_ratio = screen.devicePixelRatio()
        stats_font_size = 13  # ç»Ÿä¸€å­—ä½“å¤§å°ä¸º13px
        stats_padding = max(int(8 * dpi_ratio * 0.6), 6)
        
        # é€šç”¨æ–‡æœ¬æ¡†æ ·å¼
        text_style = f"""
            QTextEdit {{
                border: 1px solid #bee5eb;
                border-radius: 4px;
                background-color: white;
                color: black;  /* å¼ºåˆ¶ä½¿ç”¨é»‘è‰²å­—ä½“ */
                font-family: 'Consolas', 'Monaco', monospace;
                font-size: {stats_font_size}px;
                padding: {stats_padding}px;
                line-height: 1.4;
            }}
        """
        
        # 1. æ€»ä½“ç»Ÿè®¡æ ‡ç­¾é¡µï¼ˆå§‹ç»ˆæ˜¾ç¤ºï¼‰
        self.overall_stats_widget = QTextEdit()
        self.overall_stats_widget.setReadOnly(True)
        self.overall_stats_widget.setStyleSheet(text_style)
        self.overall_stats_widget.setText("ğŸ“Š æ€»ä½“ç»Ÿè®¡ä¿¡æ¯\n\nè¯·å…ˆå¤„ç†æ–‡ä»¶å¹¶è¿›è¡Œç­›é€‰ï¼Œç„¶åæŸ¥çœ‹ç»Ÿè®¡ç»“æœã€‚")
        self.stats_tab_widget.addTab(self.overall_stats_widget, "ğŸ“Š æ€»ä½“ç»Ÿè®¡")
        
        # èƒæ¬¡åˆ†æï¼ˆå§‹ç»ˆæ˜¾ç¤ºï¼‰
        self.parity_stats_widget = QTextEdit()
        self.parity_stats_widget.setReadOnly(True)
        self.parity_stats_widget.setStyleSheet(text_style)
        self.parity_stats_widget.setText("ğŸ„ èƒæ¬¡åˆ†æ\n\nç­‰å¾…ç­›é€‰ç»“æœ...")
        self.stats_tab_widget.addTab(self.parity_stats_widget, "ğŸ„ èƒæ¬¡åˆ†æ")
        
        # åŠ¨æ€åˆ›å»ºæ€§çŠ¶é€‰é¡¹å¡å­—å…¸ï¼Œç”¨äºå­˜å‚¨å„ç§æ€§çŠ¶çš„åˆ†æwidget
        self.trait_stats_widgets = {}
        
        # å¦‚æœæä¾›äº†å¯ç”¨çš„æ€§çŠ¶åˆ—è¡¨ï¼Œåˆ™åŠ¨æ€åˆ›å»ºå¯¹åº”é€‰é¡¹å¡
        if enabled_traits:
            # å®šä¹‰æ€§çŠ¶çš„å›¾æ ‡å’Œä¸­æ–‡åç§°æ˜ å°„
            trait_config = {
                'protein_pct': {'icon': 'ğŸ¥›', 'name': 'è›‹ç™½ç‡åˆ†æ'},
                'somatic_cell_count': {'icon': 'ğŸ”¬', 'name': 'ä½“ç»†èƒæ•°åˆ†æ'},
                'fat_pct': {'icon': 'ğŸ§ˆ', 'name': 'ä¹³è„‚ç‡åˆ†æ'},
                'lactose_pct': {'icon': 'ğŸ¬', 'name': 'ä¹³ç³–ç‡åˆ†æ'},
                'milk_yield': {'icon': 'ğŸ¥›', 'name': 'äº§å¥¶é‡åˆ†æ'},
                'lactation_days': {'icon': 'ğŸ“…', 'name': 'æ³Œä¹³å¤©æ•°åˆ†æ'},
                'solids_pct': {'icon': 'ğŸ§ª', 'name': 'å›ºå½¢ç‰©åˆ†æ'},
                'fat_protein_ratio': {'icon': 'âš–ï¸', 'name': 'è„‚è›‹æ¯”åˆ†æ'},
                'urea_nitrogen': {'icon': 'ğŸ§¬', 'name': 'å°¿ç´ æ°®åˆ†æ'},
                'total_fat_pct': {'icon': 'ğŸ§ˆ', 'name': 'æ€»ä¹³è„‚åˆ†æ'},
                'total_protein_pct': {'icon': 'ğŸ¥›', 'name': 'æ€»è›‹ç™½åˆ†æ'},
                'mature_equivalent': {'icon': 'ğŸ„', 'name': 'æˆå¹´å½“é‡åˆ†æ'},
                'somatic_cell_score': {'icon': 'ğŸ”¬', 'name': 'ä½“ç»†èƒåˆ†åˆ†æ'},
                'freezing_point': {'icon': 'â„ï¸', 'name': 'å†°ç‚¹åˆ†æ'},
                'total_bacterial_count': {'icon': 'ğŸ¦ ', 'name': 'ç»†èŒæ€»æ•°åˆ†æ'},
                'dry_matter_intake': {'icon': 'ğŸŒ¾', 'name': 'å¹²ç‰©è´¨é‡‡é£Ÿé‡åˆ†æ'},
                'net_energy_lactation': {'icon': 'âš¡', 'name': 'æ³Œä¹³å‡€èƒ½åˆ†æ'},
                'metabolizable_protein': {'icon': 'ğŸ§¬', 'name': 'å¯ä»£è°¢è›‹ç™½åˆ†æ'},
                'crude_protein': {'icon': 'ğŸ«˜', 'name': 'ç²—è›‹ç™½åˆ†æ'},
                'neutral_detergent_fiber': {'icon': 'ğŸŒ¾', 'name': 'ä¸­æ€§æ´—æ¶¤çº¤ç»´åˆ†æ'},
                'acid_detergent_fiber': {'icon': 'ğŸŒ¾', 'name': 'é…¸æ€§æ´—æ¶¤çº¤ç»´åˆ†æ'},
                'starch': {'icon': 'ğŸŒ½', 'name': 'æ·€ç²‰åˆ†æ'},
                'ether_extract': {'icon': 'ğŸ§ª', 'name': 'é†šæå–ç‰©åˆ†æ'},
                'ash': {'icon': 'ğŸ”¥', 'name': 'ç°åˆ†åˆ†æ'},
                'calcium': {'icon': 'ğŸ¦´', 'name': 'é’™åˆ†æ'},
                'phosphorus': {'icon': 'âš—ï¸', 'name': 'ç£·åˆ†æ'},
                'magnesium': {'icon': 'ğŸ§ª', 'name': 'é•åˆ†æ'},
                'sodium': {'icon': 'ğŸ§‚', 'name': 'é’ åˆ†æ'},
                'potassium': {'icon': 'ğŸŒ', 'name': 'é’¾åˆ†æ'},
                'sulfur': {'icon': 'ğŸ’›', 'name': 'ç¡«åˆ†æ'}
            }
            
            # ä¸ºæ¯ä¸ªå¯ç”¨çš„æ€§çŠ¶åˆ›å»ºé€‰é¡¹å¡
            for trait in enabled_traits:
                if trait in trait_config:
                    config = trait_config[trait]
                    widget = QTextEdit()
                    widget.setReadOnly(True)
                    widget.setStyleSheet(text_style)
                    widget.setText(f"{config['icon']} {config['name']}\n\nç­‰å¾…ç­›é€‰ç»“æœ...")
                    
                    self.trait_stats_widgets[trait] = widget
                    self.stats_tab_widget.addTab(widget, f"{config['icon']} {config['name']}")
        
        else:
            # é»˜è®¤é€‰é¡¹å¡ï¼ˆå…¼å®¹æ—§ç‰ˆæœ¬ï¼‰
            # è›‹ç™½ç‡åˆ†ææ ‡ç­¾é¡µ
            self.protein_stats_widget = QTextEdit()
            self.protein_stats_widget.setReadOnly(True)
            self.protein_stats_widget.setStyleSheet(text_style)
            self.protein_stats_widget.setText("ğŸ¥› è›‹ç™½ç‡åˆ†æ\n\nç­‰å¾…ç­›é€‰ç»“æœ...")
            self.stats_tab_widget.addTab(self.protein_stats_widget, "ğŸ¥› è›‹ç™½ç‡åˆ†æ")
            self.trait_stats_widgets['protein_pct'] = self.protein_stats_widget
            
            # ä½“ç»†èƒæ•°åˆ†ææ ‡ç­¾é¡µ
            self.somatic_stats_widget = QTextEdit()
            self.somatic_stats_widget.setReadOnly(True)
            self.somatic_stats_widget.setStyleSheet(text_style)
            self.somatic_stats_widget.setText("ğŸ”¬ ä½“ç»†èƒæ•°åˆ†æ\n\nç­‰å¾…ç­›é€‰ç»“æœ...")
            self.stats_tab_widget.addTab(self.somatic_stats_widget, "ğŸ”¬ ä½“ç»†èƒæ•°åˆ†æ")
            self.trait_stats_widgets['somatic_cell_count'] = self.somatic_stats_widget
            
            # å…¶ä»–æ€§çŠ¶åˆ†ææ ‡ç­¾é¡µ
            self.other_traits_stats_widget = QTextEdit()
            self.other_traits_stats_widget.setReadOnly(True)
            self.other_traits_stats_widget.setStyleSheet(text_style)
            self.other_traits_stats_widget.setText("ğŸ“ˆ å…¶ä»–æ€§çŠ¶åˆ†æ\n\nç­‰å¾…ç­›é€‰ç»“æœ...")
            self.stats_tab_widget.addTab(self.other_traits_stats_widget, "ğŸ“ˆ å…¶ä»–æ€§çŠ¶")
        
        # è®¡ç®—æ–¹æ³•è¯´æ˜æ ‡ç­¾é¡µï¼ˆå§‹ç»ˆæ˜¾ç¤ºï¼‰
        method_widget = QTextEdit()
        method_widget.setReadOnly(True)
        method_widget.setStyleSheet(text_style)
        method_text = """ğŸ“‹ è®¡ç®—æ–¹æ³•è¯´æ˜

ğŸ”¢ åŸºç¡€ç»Ÿè®¡ï¼š
â€¢ ç‰›å¤´æ•°ç»Ÿè®¡åŸºäº'ç‰›åœºç¼–å·+ç®¡ç†å·'çš„å”¯ä¸€ç»„åˆ
â€¢ ç­›é€‰ç‡ = ç¬¦åˆæ¡ä»¶çš„ç‰›å¤´æ•° Ã· å…¨éƒ¨æ•°æ®çš„ç‰›å¤´æ•° Ã— 100%

ğŸ¥› è›‹ç™½ç‡è®¡ç®—ï¼š
â€¢ å•å¤´ç‰›å¹³å‡è›‹ç™½ç‡ = Î£(å•æ¬¡è›‹ç™½ç‡Ã—å•æ¬¡äº§å¥¶é‡) / Î£(å•æ¬¡äº§å¥¶é‡)
â€¢ å½“æœˆæ‰€æœ‰ç‰›å¹³å‡è›‹ç™½ç‡ = Î£(å„ç‰›å½“æœˆè›‹ç™½ç‡Ã—å„ç‰›å½“æœˆäº§å¥¶é‡) / Î£(å„ç‰›å½“æœˆäº§å¥¶é‡)
â€¢ æ€»ä½“å¹³å‡è›‹ç™½ç‡ = Î£(æ‰€æœ‰è®°å½•è›‹ç™½ç‡Ã—æ‰€æœ‰è®°å½•äº§å¥¶é‡) / Î£(æ‰€æœ‰è®°å½•äº§å¥¶é‡)
â€¢ é‡‡ç”¨äº§å¥¶é‡åŠ æƒå¹³å‡çš„æ–¹å¼è®¡ç®—å¹³å‡è›‹ç™½ç‡ï¼Œè€Œéç®€å•ç®—æœ¯å¹³å‡

ğŸ”¬ ä½“ç»†èƒæ•°è®¡ç®—ï¼š
â€¢ ä½“ç»†èƒæ•°ä»¥ä¸‡/mlä¸ºå•ä½
â€¢ æœˆåº¦å¹³å‡å€¼é‡‡ç”¨ç®—æœ¯å¹³å‡
â€¢ ç­›é€‰æ—¶æ”¯æŒç©ºå€¼åˆ¤æ–­å¤„ç†

ğŸ„ èƒæ¬¡åˆ†æï¼š
â€¢ èƒæ¬¡ä¸ºæœ€åä¸€æ¬¡å–æ ·æ—¶çš„èƒæ¬¡
â€¢ ç¾¤ä½“å¹³å‡èƒæ¬¡é‡‡ç”¨ç®—æœ¯å¹³å‡
â€¢ æ”¯æŒæŒ‰èƒæ¬¡èŒƒå›´ç­›é€‰

ğŸ“Š æœˆåº¦åˆ†æï¼š
â€¢ æŒ‰æ—¶é—´é¡ºåºå±•ç¤ºå„æœˆä»½æ•°æ®ç»Ÿè®¡
â€¢ æ˜¾ç¤ºæ•°æ®èŒƒå›´å’Œå¹³å‡å€¼
â€¢ æ”¯æŒå¤šæ€§çŠ¶åŒæ­¥åˆ†æ
"""
        method_widget.setText(method_text)
        self.stats_tab_widget.addTab(method_widget, "ğŸ“‹ è®¡ç®—æ–¹æ³•")
        
        # ä¿ç•™åŸå§‹çš„stats_widgetå¼•ç”¨ä»¥ä¾¿å…¼å®¹ç°æœ‰ä»£ç 
        self.stats_widget = self.overall_stats_widget
    
    def create_stat_card(self, title, value, description, color):
        """åˆ›å»ºç»Ÿè®¡å¡ç‰‡"""
        card = QWidget()
        card.setStyleSheet(f"""
            QWidget {{
                background-color: white;
                border: 1px solid #dee2e6;
                border-radius: 8px;
                padding: 15px;
                margin: 5px;
            }}
        """)
        
        layout = QVBoxLayout(card)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # æ ‡é¢˜
        title_label = QLabel(title)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        card_title_font_size = max(int(11 * 0.6), 10)
        title_label.setStyleSheet(f"color: black; background-color: white; font-size: {card_title_font_size}px; margin-bottom: 5px; font-weight: bold;")
        layout.addWidget(title_label)
        
        # æ•°å€¼
        value_label = QLabel(value)
        value_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        card_value_font_size = max(int(28 * 0.6), 20)
        value_label.setStyleSheet(f"color: {color}; font-size: {card_value_font_size}px; font-weight: bold; margin: 10px 0;")
        layout.addWidget(value_label)
        
        # æè¿°
        desc_label = QLabel(description)
        desc_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        card_desc_font_size = max(int(10 * 0.6), 9)
        desc_label.setStyleSheet(f"color: #495057; font-size: {card_desc_font_size}px; font-weight: 500;")
        desc_label.setWordWrap(True)
        layout.addWidget(desc_label)
        
        # ä¿å­˜value_labelå¼•ç”¨ä»¥ä¾¿åç»­æ›´æ–°
        setattr(card, 'value_label', value_label)
        
        return card
    
    def load_config(self):
        """åŠ è½½é…ç½®"""
        try:
            with open("config.yaml", "r", encoding="utf-8") as f:
                self.config = yaml.safe_load(f)
        except:
            self.config = {}
    
    def select_files(self):
        """é€‰æ‹©æ–‡ä»¶"""
        files, _ = QFileDialog.getOpenFileNames(
            self, 
            "é€‰æ‹©DHIæŠ¥å‘Šæ–‡ä»¶",
            "",
            "æ”¯æŒçš„æ–‡ä»¶ (*.zip *.xlsx *.xls);;ZIPæ–‡ä»¶ (*.zip);;Excelæ–‡ä»¶ (*.xlsx *.xls)"
        )
        
        if files:
            # æ›´æ–°ä¼ ç»Ÿæ–‡ä»¶åˆ—è¡¨ï¼ˆä¿æŒå…¼å®¹æ€§ï¼‰
            self.file_list.clear()
            self.selected_files = files
            
            for file in files:
                filename = os.path.basename(file)
                item = QListWidgetItem(filename)
                self.file_list.addItem(item)
            
            # æ›´æ–°æ–°çš„æ–‡ä»¶æ ‡ç­¾æ˜¾ç¤º
            self.update_file_tags_display(files)
            
            self.process_btn.setEnabled(True)
            self.safe_show_status_message(f"å·²é€‰æ‹© {len(files)} ä¸ªæ–‡ä»¶")
    
    def update_file_tags_display(self, files):
        """æ›´æ–°æ–‡ä»¶æ ‡ç­¾æ˜¾ç¤º"""
        if not hasattr(self, 'file_tags_layout'):
            return
        
        # æ¸…é™¤æ‰€æœ‰ç°æœ‰çš„æ–‡ä»¶æ ‡ç­¾
        for i in reversed(range(self.file_tags_layout.count())):
            item = self.file_tags_layout.itemAt(i)
            if item:
                widget = item.widget()
                if widget:
                    widget.setParent(None)
        
        # ä¸ºæ¯ä¸ªæ–‡ä»¶åˆ›å»ºæ ‡ç­¾
        for file_path in files:
            filename = os.path.basename(file_path)
            file_tag = self.create_file_tag(filename)
            self.file_tags_layout.addWidget(file_tag)
        
        # æ·»åŠ å¼¹æ€§ç©ºé—´
        self.file_tags_layout.addStretch()
    
    def create_file_tag(self, filename):
        """åˆ›å»ºæ–‡ä»¶æ ‡ç­¾"""
        tag_widget = QWidget()
        tag_widget.setMaximumHeight(self.get_dpi_scaled_size(36))  # è¿›ä¸€æ­¥å¢å¤§é«˜åº¦ä¸º36pxï¼Œæ›´æ˜“é˜…è¯»
        tag_widget.setStyleSheet("""
            QWidget {
                background-color: #e9f4ff;
                border: 1px solid #007bff;
                border-radius: 12px;
                margin: 2px;
            }
        """)
        
        tag_layout = QHBoxLayout(tag_widget)
        tag_layout.setContentsMargins(12, 6, 12, 6)  # è¿›ä¸€æ­¥å¢å¤§å†…è¾¹è·
        tag_layout.setSpacing(8)  # è¿›ä¸€æ­¥å¢å¤§é—´è·
        
        # æ–‡ä»¶å›¾æ ‡
        file_icon = QLabel("ğŸ“„")
        file_icon.setStyleSheet("background: transparent; border: none; font-size: 14px;")
        tag_layout.addWidget(file_icon)
        
        # æ–‡ä»¶å - æ”¯æŒæ–‡æœ¬æ¢è¡Œå’Œå®Œæ•´æ˜¾ç¤º
        file_label = QLabel(filename)
        file_label.setStyleSheet("background: transparent; border: none; font-size: 13px; color: #0056b3;")
        file_label.setWordWrap(True)  # å…è®¸æ–‡æœ¬æ¢è¡Œ
        file_label.setMinimumWidth(self.get_dpi_scaled_size(200))  # è®¾ç½®æœ€å°å®½åº¦ç¡®ä¿æ–‡æœ¬æœ‰è¶³å¤Ÿç©ºé—´
        file_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)  # å…è®¸æ°´å¹³æ‰©å±•
        tag_layout.addWidget(file_label)
        
        tag_layout.addStretch()
        
        return tag_widget
    
    def clear_files(self):
        """æ¸…ç©ºå·²é€‰æ‹©çš„æ–‡ä»¶"""
        if hasattr(self, 'file_list'):
            self.file_list.clear()
        
        # æ¸…ç©ºæ–‡ä»¶æ ‡ç­¾æ˜¾ç¤º
        if hasattr(self, 'file_tags_layout'):
            # æ¸…é™¤æ‰€æœ‰ç°æœ‰çš„æ–‡ä»¶æ ‡ç­¾
            for i in reversed(range(self.file_tags_layout.count())):
                item = self.file_tags_layout.itemAt(i)
                if item:
                    widget = item.widget()
                    if widget:
                        widget.setParent(None)
            
            # æ¢å¤"å°šæœªé€‰æ‹©æ–‡ä»¶"æç¤º
            no_files_label = QLabel("å°šæœªé€‰æ‹©æ–‡ä»¶")
            no_files_label.setStyleSheet("color: #9ca3af; font-size: 11px; font-style: italic;")
            self.file_tags_layout.addWidget(no_files_label)
            self.file_tags_layout.addStretch()
        
        # é‡ç½®çŠ¶æ€
        self.selected_files = []
        self.process_btn.setEnabled(False)
        self.progress_bar.setVisible(False)
        self.safe_show_status_message("å·²æ¸…ç©ºæ‰€æœ‰æ–‡ä»¶")
    
    def process_files(self):
        """å¤„ç†æ–‡ä»¶"""
        if not hasattr(self, 'selected_files') or not self.selected_files:
            return
        
        self.process_btn.setEnabled(False)
        
        # åˆ›å»ºå¢å¼ºè¿›åº¦æ¡å¯¹è¯æ¡†ï¼ˆä½¿ç”¨æ–°çš„æµç•…è¿›åº¦æ¡ï¼‰
        self.file_progress_dialog = SmoothProgressDialog(
            "æ­£åœ¨å¤„ç†æ–‡ä»¶...",
            "å–æ¶ˆ",
            0, 100,
            self
        )
        self.file_progress_dialog.setWindowTitle("æ–‡ä»¶å¤„ç†")
        self.file_progress_dialog.canceled.connect(self.cancel_file_processing)
        self.file_progress_dialog.show()
        
        # å¯åŠ¨å¤„ç†çº¿ç¨‹
        filenames = [os.path.basename(f) for f in self.selected_files]
        self.process_thread = FileProcessThread(self.selected_files, filenames, self.urea_tracker)
        self.process_thread.progress_updated.connect(self.update_file_progress)
        self.process_thread.file_processed.connect(self.file_processed)
        self.process_thread.processing_completed.connect(self.processing_completed)
        self.process_thread.log_updated.connect(self.update_process_log)
        self.process_thread.start()
        
        # åˆ‡æ¢åˆ°å¤„ç†è¿‡ç¨‹æ ‡ç­¾é¡µ
        self.tab_widget.setCurrentWidget(self.process_log_widget)
    
    def update_file_progress(self, status, progress):
        """æ›´æ–°æ–‡ä»¶å¤„ç†è¿›åº¦"""
        if hasattr(self, 'file_progress_dialog'):
            self.file_progress_dialog.setValue(progress)
            self.file_progress_dialog.setLabelText(status)
        self.statusBar().showMessage(status)
    
    def cancel_file_processing(self):
        """å–æ¶ˆæ–‡ä»¶å¤„ç†"""
        if hasattr(self, 'process_thread') and self.process_thread.isRunning():
            self.process_thread.terminate()
            self.process_thread.wait()
            self.process_btn.setEnabled(True)
            self.statusBar().showMessage("æ–‡ä»¶å¤„ç†å·²å–æ¶ˆ")
    
    def update_progress(self, status, progress):
        """æ›´æ–°è¿›åº¦ï¼ˆå…¼å®¹æ—§ä»£ç ï¼‰"""
        # ä¸æ›´æ–°progress_labelï¼Œåªæ›´æ–°è¿›åº¦æ¡å’ŒçŠ¶æ€æ 
        self.progress_bar.setValue(progress)
        self.statusBar().showMessage(status)
    
    def update_process_log(self, log_message):
        """æ›´æ–°å¤„ç†è¿‡ç¨‹æ—¥å¿—"""
        # åœ¨æ—¥å¿—æœ«å°¾æ·»åŠ æ–°æ¶ˆæ¯
        self.process_log_widget.append(log_message)
        # æ»šåŠ¨åˆ°åº•éƒ¨
        cursor = self.process_log_widget.textCursor()
        cursor.movePosition(cursor.MoveOperation.End)
        self.process_log_widget.setTextCursor(cursor)
    
    def file_processed(self, filename, success, message, file_info):
        """å•ä¸ªæ–‡ä»¶å¤„ç†å®Œæˆ"""
        if success:
            info = f"âœ… {filename}: {message}"
        else:
            info = f"âŒ {filename}: {message}"
        
        self.file_info_widget.append(info)
        self.statusBar().showMessage(f"å·²å¤„ç†: {filename}")
        
        # ç§»é™¤å•ç‹¬å¤„ç†ç¼ºå°‘ç‰›åœºç¼–å·çš„é€»è¾‘ï¼Œæ”¹ä¸ºåœ¨æ‰¹é‡å¤„ç†å®Œæˆæ—¶ç»Ÿä¸€å¤„ç†
    
    # handle_missing_farm_id and add_farm_id_to_data methods removed - no longer needed for single-farm uploads
    
    def processing_completed(self, results):
        """æ‰€æœ‰æ–‡ä»¶å¤„ç†å®Œæˆ"""
        # å…³é—­è¿›åº¦æ¡å¯¹è¯æ¡†
        if hasattr(self, 'file_progress_dialog'):
            self.file_progress_dialog.close()
        
        self.progress_bar.setVisible(False)
        self.process_btn.setEnabled(True)
        
        # ç›´æ¥å®Œæˆå¤„ç†ï¼Œæ— éœ€æ£€æŸ¥ç‰§åœºç¼–å·
        self.complete_processing(results)
    
    # handle_batch_missing_farm_id method removed - no longer needed for single-farm uploads
    
    # check_and_handle_farm_id_consistency method removed - no longer needed for single-farm uploads
    
    def complete_processing(self, results):
        """å®Œæˆå¤„ç†æµç¨‹"""
        # ä¿å­˜æ•°æ®
        self.data_list = results['all_data']
        
        # å¤„ç†æˆåŠŸåè®¾ç½®æ ‡å¿—
        self.dhi_processed_ok = True if self.data_list else False
        
        # è®¡ç®—æ€»ç‰›å¤´æ•°å¹¶æ›´æ–°åˆ†æé¢æ¿
        total_cows = set()
        all_data_combined = []
        
        for item in self.data_list:
            df = item['data']
            if df is not None and not df.empty:
                all_data_combined.append(df)
                if 'management_id' in df.columns:
                    cow_ids = df['management_id'].dropna().unique()
                    for cow_id in cow_ids:
                        total_cows.add(cow_id)
        
        # åˆå¹¶æ‰€æœ‰æ•°æ®ç”¨äºåˆ†æ
        if all_data_combined:
            combined_df = pd.concat(all_data_combined, ignore_index=True)
            self.update_filter_ranges(combined_df)
        
        # æ›´æ–°å…¨éƒ¨æ•°æ®ç»Ÿè®¡
        getattr(self.total_data_card, 'value_label').setText(str(len(total_cows)))
        
        # ç‰›åœºç¼–å·é€‰æ‹©å™¨å·²ç§»é™¤ - å•ç‰›åœºä¸Šä¼ ä¸å†éœ€è¦
        
        # æ£€æµ‹é‡å¤æ–‡ä»¶å¹¶åœ¨æ–‡ä»¶ä¿¡æ¯æ¡†æ˜¾ç¤º
        self.detect_and_display_duplicates()
        
        # æ˜¾ç¤ºå¤„ç†ç»“æœ
        success_count = len(results.get('success_files', []))
        failed_count = len(results.get('failed_files', []))
        summary = f"\nğŸ“Š å¤„ç†å®Œæˆï¼\n"
        summary += f"æˆåŠŸ: {success_count} ä¸ªæ–‡ä»¶\n"
        summary += f"å¤±è´¥: {failed_count} ä¸ªæ–‡ä»¶\n\n"
        
        self.file_info_widget.append(summary)
        
        # å¯ç”¨ç­›é€‰æŒ‰é’®ï¼ˆå¦‚æœå­˜åœ¨ï¼‰
        if success_count > 0 and hasattr(self, 'filter_btn'):
            self.filter_btn.setEnabled(True)
        
        status_msg = f"å¤„ç†å®Œæˆï¼šæˆåŠŸ {success_count} ä¸ªï¼Œå¤±è´¥ {failed_count} ä¸ª"
        self.statusBar().showMessage(status_msg)
        
        # æå–å¹¶æ›´æ–°æ…¢æ€§æ„ŸæŸ“ç‰›è¯†åˆ«çš„æœˆä»½é€‰æ‹©ï¼ˆå¦‚æœæœ‰DHIæ•°æ®ï¼‰
        dhi_months = set()
        for item in self.data_list:
            df = item['data']
            if 'sample_date' in df.columns and 'somatic_cell_count' in df.columns:
                # ä»æœ‰ä½“ç»†èƒæ•°æ®çš„æ–‡ä»¶ä¸­æå–æœˆä»½
                dates = pd.to_datetime(df['sample_date'], errors='coerce').dropna()
                months = dates.dt.strftime('%Yå¹´%mæœˆ').unique()
                dhi_months.update(months)
        
        if dhi_months:
            sorted_months = sorted(list(dhi_months))
            print(f"ä»DHIæ•°æ®ä¸­æå–åˆ°æœˆä»½: {sorted_months}")
            self.update_chronic_months_options(sorted_months)
        else:
            print("æœªæ‰¾åˆ°åŒ…å«ä½“ç»†èƒæ•°æ®çš„DHIæ–‡ä»¶ï¼Œæ— æ³•æ›´æ–°æœˆä»½é€‰æ‹©")
        
        # å¦‚æœå°¿ç´ æ°®è¿½è¸ªå™¨æœ‰æ•°æ®ï¼Œå¯ç”¨åˆ†ææŒ‰é’®
        if hasattr(self, 'urea_analyze_btn') and self.urea_tracker.dhi_data_dict:
            self.urea_analyze_btn.setEnabled(True)
    
    def detect_and_display_duplicates(self):
        """æ£€æµ‹é‡å¤æ–‡ä»¶å¹¶åœ¨æ–‡ä»¶ä¿¡æ¯æ¡†ä¸­æ˜¾ç¤ºè¯¦ç»†ä¿¡æ¯"""
        if not self.data_list or len(self.data_list) < 2:
            return
        
        try:
            # ä½¿ç”¨å¤„ç†å™¨çš„é‡å¤æ£€æµ‹åŠŸèƒ½
            duplicate_result = self.processor.detect_duplicate_data(self.data_list)
            
            if duplicate_result['has_duplicates']:
                duplicate_count = duplicate_result['duplicate_files_count']
                group_count = len(duplicate_result['duplicate_groups'])
                
                self.file_info_widget.append(f"\nâš ï¸ é‡å¤æ–‡ä»¶æ£€æµ‹ç»“æœ:")
                self.file_info_widget.append(f"å‘ç° {group_count} ç»„é‡å¤æ–‡ä»¶ï¼Œå…±æ¶‰åŠ {duplicate_count} ä¸ªæ–‡ä»¶\n")
                
                for i, group in enumerate(duplicate_result['duplicate_groups'], 1):
                    self.file_info_widget.append(f"ğŸ“‹ é‡å¤ç»„ {i}:")
                    
                    for j, file_info in enumerate(group):
                        filename = file_info['filename']
                        data = file_info['data']
                        
                        # è·å–æ–‡ä»¶çš„æœˆä»½ä¿¡æ¯
                        months_info = self._extract_file_months_info(data)
                        similarity_score = file_info.get('similarity_score', 'N/A')
                        
                        if j == 0:
                            # ç¬¬ä¸€ä¸ªæ–‡ä»¶ä½œä¸ºåŸºå‡†
                            self.file_info_widget.append(f"  ğŸ“„ {filename}")
                            self.file_info_widget.append(f"     ğŸ“… æ•°æ®æœˆä»½: {months_info}")
                        else:
                            # åç»­æ–‡ä»¶æ˜¾ç¤ºä¸åŸºå‡†çš„ç›¸ä¼¼åº¦
                            self.file_info_widget.append(f"  ğŸ“„ {filename} (ç›¸ä¼¼åº¦: {similarity_score:.1%})")
                            self.file_info_widget.append(f"     ğŸ“… æ•°æ®æœˆä»½: {months_info}")
                    
                    self.file_info_widget.append("")  # ç©ºè¡Œåˆ†éš”ä¸åŒç»„
                
                self.file_info_widget.append("ğŸ’¡ å»ºè®®: æ£€æŸ¥è¿™äº›é‡å¤æ–‡ä»¶çš„å†…å®¹ï¼Œç¡®è®¤æ˜¯å¦éœ€è¦ä¿ç•™æ‰€æœ‰æ–‡ä»¶ã€‚\n")
                
        except Exception as e:
            print(f"é‡å¤æ–‡ä»¶æ£€æµ‹æ—¶å‡ºé”™: {e}")
            # ä¸å½±å“ä¸»æµç¨‹ï¼Œåªè®°å½•é”™è¯¯
            self.file_info_widget.append(f"\nâš ï¸ é‡å¤æ–‡ä»¶æ£€æµ‹æ—¶å‡ºç°é”™è¯¯: {str(e)}\n")
    
    def _extract_file_months_info(self, df):
        """ä»æ•°æ®æ¡†ä¸­æå–æœˆä»½ä¿¡æ¯"""
        try:
            if 'sample_date' not in df.columns:
                return "æœªçŸ¥æœˆä»½"
            
            # è½¬æ¢æ—¥æœŸå¹¶æå–æœˆä»½
            dates = pd.to_datetime(df['sample_date'], errors='coerce').dropna()
            if dates.empty:
                return "æ— æœ‰æ•ˆæ—¥æœŸ"
            
            # è·å–å¹´æœˆä¿¡æ¯
            year_months = dates.dt.strftime('%Yå¹´%mæœˆ').unique()
            
            if len(year_months) == 1:
                return year_months[0]
            elif len(year_months) <= 3:
                return "ã€".join(sorted(year_months))
            else:
                sorted_months = sorted(year_months)
                return f"{sorted_months[0]}ï½{sorted_months[-1]} (å…±{len(year_months)}ä¸ªæœˆ)"
                
        except Exception as e:
            print(f"æå–æœˆä»½ä¿¡æ¯æ—¶å‡ºé”™: {e}")
            return "æœˆä»½ä¿¡æ¯æå–å¤±è´¥"
    
    def update_filter_ranges(self, df):
        """æ ¹æ®æ•°æ®æ›´æ–°ç­›é€‰æ¡ä»¶çš„èŒƒå›´å’Œé»˜è®¤å€¼"""
        try:
            # ä½¿ç”¨æ–°çš„æ•°æ®èŒƒå›´è®¡ç®—åŠŸèƒ½
            data_ranges = self.processor.get_data_ranges(self.data_list)
            self.current_data_ranges = data_ranges  # ä¿å­˜æ•°æ®èŒƒå›´ä¾›åç»­ä½¿ç”¨
            
            # è®¡ç®—æœˆæ•°ä¸Šé™
            max_months = data_ranges.get('months', {}).get('max', 12)
            
            print(f"æ•°æ®èŒƒå›´è®¡ç®—å®Œæˆï¼š")
            print(f"  æœˆæ•°èŒƒå›´: 0 - {max_months}ä¸ªæœˆ")
            for field, range_info in data_ranges.items():
                if field != 'months' and isinstance(range_info, dict):
                    print(f"  {field}: {range_info.get('description', 'æœªçŸ¥')}")
            
            # æ›´æ–°èƒæ¬¡èŒƒå›´
            if 'parity' in data_ranges:
                parity_range = data_ranges['parity']
                min_parity = int(parity_range['min'])
                max_parity = int(parity_range['max'])
                
                # æ›´æ–°èƒæ¬¡æ§ä»¶ï¼ˆå¦‚æœå­˜åœ¨ï¼‰
                if hasattr(self, 'parity_min') and hasattr(self, 'parity_max'):
                    # æ›´æ–°èŒƒå›´ - ä½¿ç”¨å®é™…æ•°æ®èŒƒå›´ä½†ä¸è®¾ç½®ä¸Šé™
                    self.parity_min.setRange(min_parity, 99)
                    self.parity_max.setRange(min_parity, 99)
                    
                    # è®¾ç½®é»˜è®¤å€¼ä¸ºæ•°æ®èŒƒå›´
                    self.parity_min.setValue(min_parity)
                    self.parity_max.setValue(max_parity)
            
                print(f"  èƒæ¬¡æ§ä»¶æ›´æ–°: {min_parity}-{max_parity}èƒ")
            
            # æ›´æ–°è›‹ç™½ç‡ç­›é€‰æ§ä»¶ï¼ˆå¦‚æœå­˜åœ¨ï¼‰
            if hasattr(self, 'protein_min') and 'protein_pct' in data_ranges:
                protein_range = data_ranges['protein_pct']
                
                # è®¾ç½®èŒƒå›´ - ç§»é™¤äººä¸ºä¸Šé™é™åˆ¶
                self.protein_min.setRange(0.0, 999999.99)
                self.protein_max.setRange(0.0, 999999.99)
                
                # ä½¿ç”¨å®é™…æ•°æ®çš„æœ€å°å€¼å’Œæœ€å¤§å€¼
                actual_min = protein_range['min']
                actual_max = protein_range['max']
                self.protein_min.setValue(actual_min)
                self.protein_max.setValue(actual_max)
                
                # æ›´æ–°æœ€å°‘ç¬¦åˆæœˆæ•°èŒƒå›´
                if hasattr(self, 'protein_months'):
                    self.protein_months.setRange(0, max_months)
                    self.protein_months.setValue(min(3, max_months // 2) if max_months > 0 else 1)
                
                print(f"  è›‹ç™½ç‡æ§ä»¶æ›´æ–°: {actual_min}-{actual_max}% (å®é™…æ•°æ®èŒƒå›´)")
            
            # æ›´æ–°ä½“ç»†èƒæ•°ç­›é€‰æ§ä»¶ï¼ˆå¦‚æœå­˜åœ¨ï¼‰
            if hasattr(self, 'somatic_min') and 'somatic_cell_count' in data_ranges:
                somatic_range = data_ranges['somatic_cell_count']
                
                # è®¾ç½®èŒƒå›´ - ç§»é™¤äººä¸ºä¸Šé™é™åˆ¶
                self.somatic_min.setRange(0.0, 999999.99)
                self.somatic_max.setRange(0.0, 999999.99)
                
                # ä½¿ç”¨å®é™…æ•°æ®çš„æœ€å°å€¼å’Œæœ€å¤§å€¼
                actual_min = somatic_range['min']
                actual_max = somatic_range['max']
                self.somatic_min.setValue(actual_min)
                self.somatic_max.setValue(actual_max)
                
                # æ›´æ–°æœ€å°‘ç¬¦åˆæœˆæ•°èŒƒå›´
                if hasattr(self, 'somatic_months'):
                    self.somatic_months.setRange(0, max_months)
                    self.somatic_months.setValue(min(3, max_months // 2) if max_months > 0 else 1)
                
                print(f"  ä½“ç»†èƒæ•°æ§ä»¶æ›´æ–°: {actual_min}-{actual_max}ä¸‡/ml (å®é™…æ•°æ®èŒƒå›´)")
            
            # æ›´æ–°å…¶ä»–ç­›é€‰é¡¹æ§ä»¶
            if hasattr(self, 'added_other_filters'):
                for filter_key, widget in self.added_other_filters.items():
                    if filter_key in data_ranges:
                        field_range = data_ranges[filter_key]
                        
                        # ä½¿ç”¨å®é™…æ•°æ®çš„æœ€å°å€¼å’Œæœ€å¤§å€¼
                        actual_min = field_range['min']
                        actual_max = field_range['max']
                        
                        # æ›´æ–°æ•°å€¼èŒƒå›´ - ç§»é™¤äººä¸ºä¸Šé™é™åˆ¶
                        widget.range_min.setRange(-999999.99, 999999.99)
                        widget.range_max.setRange(-999999.99, 999999.99)
                        widget.range_min.setValue(actual_min)
                        widget.range_max.setValue(actual_max)
                        
                        # æ›´æ–°æœ€å°‘ç¬¦åˆæœˆæ•°èŒƒå›´
                        widget.months_spinbox.setRange(0, max_months)
                        widget.months_spinbox.setValue(min(3, max_months // 2) if max_months > 0 else 1)
                        
                        print(f"  {filter_key}æ§ä»¶æ›´æ–°: {actual_min}-{actual_max} (å®é™…æ•°æ®èŒƒå›´)")
            
            # æ›´æ–°æ—¥æœŸèŒƒå›´
            if 'sample_date' in df.columns:
                date_data = pd.to_datetime(df['sample_date'], errors='coerce').dropna()
                if not date_data.empty:
                    min_date = date_data.min().date()
                    max_date = date_data.max().date()
                    
                    # æ›´æ–°æ—¥æœŸé€‰æ‹©å™¨ï¼ˆå¦‚æœå­˜åœ¨ï¼‰
                    if hasattr(self, 'date_start') and hasattr(self, 'date_end'):
                        self.date_start.setDate(QDate(min_date))
                        self.date_end.setDate(QDate(max_date))
            
                    print(f"  æ—¥æœŸèŒƒå›´æ›´æ–°: {min_date} åˆ° {max_date}")
            
            # æ›´æ–°æœªæ¥æ³Œä¹³å¤©æ•°çš„é»˜è®¤å€¼å’ŒèŒƒå›´
            if hasattr(self, 'future_days_min'):
                # æ ¹æ®æ•°æ®ä¸­çš„æ³Œä¹³å¤©æ•°èŒƒå›´è®¾ç½®åˆç†çš„æœªæ¥æ³Œä¹³å¤©æ•°èŒƒå›´
                if 'lactation_days' in data_ranges:
                    lactation_range = data_ranges['lactation_days']
                    # æœªæ¥æ³Œä¹³å¤©æ•°é€šå¸¸åœ¨50-350å¤©ä¹‹é—´
                    self.future_days_min.setValue(50)
                    self.future_days_max.setValue(min(350, int(lactation_range['max'] * 1.2)))
                else:
                    self.future_days_min.setValue(50)
                    self.future_days_max.setValue(350)
            
            # åœ¨çŠ¶æ€æ æ˜¾ç¤ºæ•°æ®èŒƒå›´æ‘˜è¦
            range_summary = f"æ•°æ®è·¨è¶Š{max_months}ä¸ªæœˆ"
            if 'protein_pct' in data_ranges:
                protein_info = data_ranges['protein_pct']
                range_summary += f"ï¼Œè›‹ç™½ç‡{protein_info['min']:.1f}-{protein_info['max']:.1f}%"
            if 'somatic_cell_count' in data_ranges:
                somatic_info = data_ranges['somatic_cell_count']
                range_summary += f"ï¼Œä½“ç»†èƒæ•°{somatic_info['min']:.1f}-{somatic_info['max']:.1f}ä¸‡/ml"
            self.statusBar().showMessage(f"ç­›é€‰æ¡ä»¶å·²è®¾ç½®ä¸ºå®é™…æ•°æ®èŒƒå›´ - {range_summary}")
            
        except Exception as e:
            print(f"æ›´æ–°ç­›é€‰èŒƒå›´æ—¶å‡ºé”™: {e}")
            import traceback
            traceback.print_exc()
            # å¦‚æœå‡ºé”™ï¼Œä½¿ç”¨æ—§çš„é€»è¾‘ä½œä¸ºå¤‡ä»½
            self._update_filter_ranges_fallback(df)
    
    def _update_filter_ranges_fallback(self, df):
        """å¤‡ç”¨çš„èŒƒå›´æ›´æ–°é€»è¾‘"""
        try:
            # åŸæœ‰çš„ç®€å•é€»è¾‘
            if 'parity' in df.columns:
                parity_data = df['parity'].dropna()
                if not parity_data.empty:
                    min_parity = int(parity_data.min())
                    max_parity = int(parity_data.max())
                    
                    if hasattr(self, 'parity_min') and hasattr(self, 'parity_max'):
                        self.parity_min.setRange(min_parity, 99)
                        self.parity_max.setRange(min_parity, 99)
                        self.parity_min.setValue(min_parity)
                        self.parity_max.setValue(max_parity)
            
            # æ›´æ–°è›‹ç™½ç‡ç­›é€‰æ§ä»¶ï¼ˆå¤‡ç”¨é€»è¾‘ï¼‰
            if 'protein_pct' in df.columns and hasattr(self, 'protein_min'):
                protein_data = df['protein_pct'].dropna()
                if not protein_data.empty:
                    min_protein = float(protein_data.min())
                    max_protein = float(protein_data.max())
                    
                    self.protein_min.setRange(0.0, 999999.99)
                    self.protein_max.setRange(0.0, 999999.99)
                    self.protein_min.setValue(min_protein)
                    self.protein_max.setValue(max_protein)
                    print(f"  è›‹ç™½ç‡æ§ä»¶æ›´æ–°(å¤‡ç”¨): {min_protein}-{max_protein}% (å®é™…æ•°æ®èŒƒå›´)")
            
            # æ›´æ–°ä½“ç»†èƒæ•°ç­›é€‰æ§ä»¶ï¼ˆå¤‡ç”¨é€»è¾‘ï¼‰
            if 'somatic_cell_count' in df.columns and hasattr(self, 'somatic_min'):
                somatic_data = df['somatic_cell_count'].dropna()
                if not somatic_data.empty:
                    min_somatic = float(somatic_data.min())
                    max_somatic = float(somatic_data.max())
                    
                    self.somatic_min.setRange(0.0, 999999.99)
                    self.somatic_max.setRange(0.0, 999999.99)
                    self.somatic_min.setValue(min_somatic)
                    self.somatic_max.setValue(max_somatic)
                    print(f"  ä½“ç»†èƒæ•°æ§ä»¶æ›´æ–°(å¤‡ç”¨): {min_somatic}-{max_somatic}ä¸‡/ml (å®é™…æ•°æ®èŒƒå›´)")
            
            # æ›´æ–°å…¶ä»–ç­›é€‰é¡¹æ§ä»¶ï¼ˆå¤‡ç”¨é€»è¾‘ï¼‰
            if hasattr(self, 'added_other_filters'):
                # æ”¯æŒçš„æ•°å€¼å­—æ®µåˆ—è¡¨
                numeric_fields = [
                    'fat_pct', 'lactose_pct', 'milk_yield', 'lactation_days', 'solids_pct',
                    'fat_protein_ratio', 'urea_nitrogen', 'total_fat_pct', 'total_protein_pct',
                    'mature_equivalent', 'somatic_cell_score', 'freezing_point', 'total_bacterial_count'
                ]
                
                for filter_key, widget in self.added_other_filters.items():
                    if filter_key in numeric_fields and filter_key in df.columns:
                        field_data = df[filter_key].dropna()
                        if not field_data.empty:
                            min_val = float(field_data.min())
                            max_val = float(field_data.max())
                            
                            widget.range_min.setRange(-999999.99, 999999.99)
                            widget.range_max.setRange(-999999.99, 999999.99)
                            widget.range_min.setValue(min_val)
                            widget.range_max.setValue(max_val)
                            print(f"  {filter_key}æ§ä»¶æ›´æ–°(å¤‡ç”¨): {min_val}-{max_val} (å®é™…æ•°æ®èŒƒå›´)")
            
            # è®¾ç½®åŸºæœ¬çš„æœˆæ•°èŒƒå›´
            if hasattr(self, 'protein_months'):
                self.protein_months.setRange(1, 12)
                self.protein_months.setValue(3)
                
        except Exception as e:
            print(f"å¤‡ç”¨èŒƒå›´æ›´æ–°ä¹Ÿå‡ºé”™: {e}")
            # æœ€åçš„ä¿é™©ï¼Œè®¾ç½®æœ€åŸºæœ¬çš„é»˜è®¤å€¼
            pass
    
    def start_filtering(self):
        """å¼€å§‹ç­›é€‰"""
        # é‡ç½®ç­›é€‰å®Œæˆæ ‡å¿—
        self._filtering_completed = False
        
        if not self.data_list:
            self.show_warning("è­¦å‘Š", "è¯·å…ˆå¤„ç†æ–‡ä»¶")
            return
        
        # æ„å»ºç­›é€‰æ¡ä»¶
        filters = self.build_filters()
        selected_files = [item['filename'] for item in self.data_list]
        
        # æ£€æŸ¥æ˜¯å¦æœ‰å¯ç”¨çš„ç‰¹æ®Šç­›é€‰é¡¹
        special_filters_enabled = False
        special_filter_names = []
        
        for filter_name, filter_config in filters.items():
            if (filter_name not in ['parity', 'date_range', 'future_lactation_days'] and 
                filter_config.get('enabled', False)):
                special_filters_enabled = True
                # è·å–ä¸­æ–‡åç§°
                if filter_name == 'protein_pct':
                    special_filter_names.append('è›‹ç™½ç‡')
                elif filter_name == 'somatic_cell_count':
                    special_filter_names.append('ä½“ç»†èƒæ•°')
                else:
                    # ä»å…¶ä»–ç­›é€‰é¡¹ä¸­è·å–ä¸­æ–‡åç§°
                    if hasattr(self, 'added_other_filters') and filter_name in self.added_other_filters:
                        widget = self.added_other_filters[filter_name]
                        special_filter_names.append(widget.chinese_name)
                    else:
                        special_filter_names.append(filter_name)
        
        if not special_filters_enabled:
            reply = QMessageBox.question(
                self, 
                "ç¡®è®¤ç­›é€‰", 
                "æ‚¨æ²¡æœ‰å¯ç”¨ä»»ä½•ç‰¹æ®Šç­›é€‰é¡¹ï¼ˆè›‹ç™½ç‡ã€ä½“ç»†èƒæ•°ç­‰ï¼‰ï¼Œåªä¼šåº”ç”¨åŸºç¡€ç­›é€‰æ¡ä»¶ã€‚\nç¡®å®šè¦ç»§ç»­å—ï¼Ÿ",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
        else:
            filter_list = "ã€".join(special_filter_names)
            reply = QMessageBox.question(
                self, 
                "ç¡®è®¤ç­›é€‰", 
                f"å°†åº”ç”¨ä»¥ä¸‹ç­›é€‰é¡¹ï¼š{filter_list}\nç¡®å®šè¦å¼€å§‹ç­›é€‰å—ï¼Ÿ",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
        
        # æ˜¾ç¤º/éšè—æŒ‰é’®ï¼ˆå¦‚æœå­˜åœ¨ï¼‰
        if hasattr(self, 'filter_btn'):
            self.filter_btn.setEnabled(False)
        
        # åˆ›å»ºå¢å¼ºè¿›åº¦æ¡å¯¹è¯æ¡†ï¼ˆä½¿ç”¨æ–°çš„æµç•…è¿›åº¦æ¡ï¼‰
        self.filter_progress_dialog = SmoothProgressDialog(
            "æ­£åœ¨ç­›é€‰æ•°æ®...",
            "å–æ¶ˆ",
            0, 100,
            self
        )
        self.filter_progress_dialog.setWindowTitle("æ•°æ®ç­›é€‰")
        self.filter_progress_dialog.canceled.connect(self.cancel_filtering)
        self.filter_progress_dialog.show()
        
        # å¯åŠ¨ç­›é€‰çº¿ç¨‹ï¼ˆä¼ é€’processorå®ä¾‹ä»¥å…±äº«åœ¨ç¾¤ç‰›æ•°æ®ï¼‰
        self.filter_thread = FilterThread(self.data_list, filters, selected_files, self.processor, self.urea_tracker)
        self.filter_thread.progress_updated.connect(self.update_filter_progress_dialog)
        self.filter_thread.filtering_completed.connect(self.filtering_completed)
        self.filter_thread.log_updated.connect(self.update_process_log)
        self.filter_thread.start()
        
        # åˆ‡æ¢åˆ°å¤„ç†è¿‡ç¨‹æ ‡ç­¾é¡µ
        self.tab_widget.setCurrentWidget(self.process_log_widget)
    
    def build_filters(self):
        """æ„å»ºç­›é€‰æ¡ä»¶"""
        filters = {}
        
        # ç‰›åœºç¼–å·ç­›é€‰å·²ç§»é™¤ - å•ç‰›åœºä¸Šä¼ æ¨¡å¼
        
        # èƒæ¬¡ï¼ˆå¦‚æœæ§ä»¶å­˜åœ¨ï¼‰
        if hasattr(self, 'parity_min') and hasattr(self, 'parity_max'):
            filters['parity'] = {
                'field': 'parity',
                'enabled': True,
                'min': self.parity_min.value(),
                'max': self.parity_max.value()
            }
        else:
            # ä½¿ç”¨é»˜è®¤å€¼
            filters['parity'] = {
                'field': 'parity',
                'enabled': True,
                'min': 1,
                'max': 99
            }
        
        # æ—¥æœŸèŒƒå›´ï¼ˆå¦‚æœæ§ä»¶å­˜åœ¨ï¼‰
        if hasattr(self, 'date_start') and hasattr(self, 'date_end'):
            filters['date_range'] = {
                'field': 'sample_date',
                'enabled': True,
                'start_date': self.date_start.date().toString("yyyy-MM-dd"),
                'end_date': self.date_end.date().toString("yyyy-MM-dd")
            }
        else:
            # ä½¿ç”¨é»˜è®¤å€¼ï¼ˆä¸é™åˆ¶æ—¥æœŸï¼‰
            filters['date_range'] = {
                'field': 'sample_date',
                'enabled': False,
                'start_date': "1900-01-01",
                'end_date': "2099-12-31"
            }
        
        # è›‹ç™½ç‡ç­›é€‰ï¼ˆæ–°çš„ç‹¬ç«‹ç­›é€‰é¡¹ï¼‰
        if hasattr(self, 'protein_enabled') and self.protein_enabled.isChecked():
            filters['protein_pct'] = {
                'field': 'protein_pct',
                'enabled': True,
                'min': self.protein_min.value(),
                'max': self.protein_max.value(),
                'min_match_months': self.protein_months.value(),
                'empty_handling': self.protein_empty.currentText()
            }
        
        # ä½“ç»†èƒæ•°ç­›é€‰ï¼ˆæ–°çš„ç‹¬ç«‹ç­›é€‰é¡¹ï¼‰
        if hasattr(self, 'somatic_enabled') and self.somatic_enabled.isChecked():
            filters['somatic_cell_count'] = {
                'field': 'somatic_cell_count',
                'enabled': True,
                'min': self.somatic_min.value(),
                'max': self.somatic_max.value(),
                'min_match_months': self.somatic_months.value(),
                'empty_handling': self.somatic_empty.currentText()
            }
        
        # å…¶ä»–ç­›é€‰é¡¹
        for filter_key, widget in self.added_other_filters.items():
            if widget.enabled_checkbox.isChecked():
                filters[filter_key] = {
                    'field': filter_key,
                    'enabled': True,
                    'min': widget.range_min.value(),
                    'max': widget.range_max.value(),
                    'min_match_months': widget.months_spinbox.value(),
                    'empty_handling': widget.empty_combo.currentText()
                }
        
        # æœªæ¥æ³Œä¹³å¤©æ•° - æ ¹æ®å¤é€‰æ¡†çŠ¶æ€å†³å®šæ˜¯å¦å¯ç”¨ï¼ˆå¦‚æœæ§ä»¶å­˜åœ¨ï¼‰
        if hasattr(self, 'future_days_enabled') and hasattr(self, 'plan_date'):
            filters['future_lactation_days'] = {
                'field': 'future_lactation_days',
                'enabled': self.future_days_enabled.isChecked(),
                'min': self.future_days_min.value(),
                'max': self.future_days_max.value(),
                'plan_date': self.plan_date.date().toString("yyyy-MM-dd")
            }
        else:
            # é»˜è®¤ç¦ç”¨æœªæ¥æ³Œä¹³å¤©æ•°ç­›é€‰
            filters['future_lactation_days'] = {
                'field': 'future_lactation_days',
                'enabled': False,
                'min': 50,
                'max': 350,
                'plan_date': QDate.currentDate().addDays(30).toString("yyyy-MM-dd")
            }
        
        # å°¿ç´ æ°®è¿½è¸ªåˆ†æ - ä¸åœ¨åŸºç¡€ç­›é€‰ä¸­æ‰§è¡Œ
        # å°¿ç´ æ°®è¿½è¸ªæœ‰ç‹¬ç«‹çš„"å¼€å§‹åˆ†æ"æŒ‰é’®ï¼Œä¸åº”åœ¨åŸºç¡€ç­›é€‰ä¸­è‡ªåŠ¨æ‰§è¡Œ
        filters['urea_tracking'] = {
            'enabled': False  # åŸºç¡€ç­›é€‰ä¸æ‰§è¡Œå°¿ç´ æ°®è¿½è¸ª
        }
        
        return filters
    
    def update_filter_progress_dialog(self, status, progress):
        """æ›´æ–°ç­›é€‰è¿›åº¦å¯¹è¯æ¡†"""
        if hasattr(self, 'filter_progress_dialog'):
            self.filter_progress_dialog.setValue(progress)
            self.filter_progress_dialog.setLabelText(status)
        self.statusBar().showMessage(status)
    
    def cancel_filtering(self):
        """å–æ¶ˆç­›é€‰"""
        # å¦‚æœç­›é€‰å·²å®Œæˆï¼Œç›´æ¥è¿”å›
        if hasattr(self, '_filtering_completed') and self._filtering_completed:
            return
            
        # æ£€æŸ¥æ˜¯å¦æœ‰æ­£åœ¨è¿è¡Œçš„ç­›é€‰çº¿ç¨‹
        if hasattr(self, 'filter_thread') and self.filter_thread.isRunning():
            # å¼¹å‡ºç¡®è®¤å¯¹è¯æ¡†
            reply = QMessageBox.question(
                self,
                "ç¡®è®¤å–æ¶ˆ",
                "ç¡®å®šè¦å–æ¶ˆå½“å‰ç­›é€‰å—ï¼Ÿ",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                self.filter_thread.request_cancel()
                self.filter_thread.wait()
                if hasattr(self, 'filter_btn'):
                    self.filter_btn.setEnabled(True)
                self.statusBar().showMessage("æ•°æ®ç­›é€‰å·²å–æ¶ˆ")
    
    def update_filter_progress(self, status, progress):
        """æ›´æ–°ç­›é€‰è¿›åº¦ï¼ˆå…¼å®¹æ—§ä»£ç ï¼‰"""
        if hasattr(self, 'filter_label'):
            self.filter_label.setText(status)
        if hasattr(self, 'filter_progress'):
            self.filter_progress.setValue(progress)
        self.statusBar().showMessage(status)
    
    def filtering_completed(self, success, message, results_df, stats=None):
        """ç­›é€‰å®Œæˆ"""
        # è®¾ç½®æ ‡å¿—ï¼Œè¡¨ç¤ºç­›é€‰å·²å®Œæˆ
        self._filtering_completed = True
        
        # å…³é—­è¿›åº¦æ¡å¯¹è¯æ¡†
        if hasattr(self, 'filter_progress_dialog'):
            # å…ˆæ–­å¼€ä¿¡å·è¿æ¥
            try:
                self.filter_progress_dialog.canceled.disconnect()
            except:
                pass
            self.filter_progress_dialog.close()
        
        self._reset_filter_ui_state()
        
        if success:
            # è°ƒè¯•ï¼šéªŒè¯æ¥æ”¶åˆ°çš„æ•°æ®
            print(f"DEBUG: ===== æ¥æ”¶åˆ°çš„æ•°æ®éªŒè¯ =====")
            print(f"DEBUG: æ¥æ”¶åˆ°çš„results_dfè¡Œæ•°: {len(results_df)}")
            if not results_df.empty and 'æœªæ¥æ³Œä¹³å¤©æ•°(å¤©)' in results_df.columns:
                received_values = results_df['æœªæ¥æ³Œä¹³å¤©æ•°(å¤©)'].dropna()
                if len(received_values) > 0:
                    print(f"DEBUG: æ¥æ”¶åˆ°çš„æœªæ¥æ³Œä¹³å¤©æ•°èŒƒå›´: {received_values.min()}-{received_values.max()}")
                    print(f"DEBUG: è¶…å‡ºèŒƒå›´çš„è®°å½•æ•°: {((received_values < 1) | (received_values > 300)).sum()}")
                    # æ˜¾ç¤ºå‡ æ¡å…·ä½“è®°å½•
                    print(f"DEBUG: å‰5æ¡è®°å½•çš„æœªæ¥æ³Œä¹³å¤©æ•°:")
                    for i in range(min(5, len(results_df))):
                        row = results_df.iloc[i]
                        future_days = row.get('æœªæ¥æ³Œä¹³å¤©æ•°(å¤©)', 'N/A')
                        mgmt_id = row.get('management_id', 'N/A')
                        print(f"  ç®¡ç†å·{mgmt_id}: {future_days}å¤©")
            print(f"DEBUG: =========================")
            
            print(f"ç­›é€‰æˆåŠŸï¼Œç»“æœè¡Œæ•°: {len(results_df)}")  # è°ƒè¯•ä¿¡æ¯
            self.current_results = results_df
            
            # è·å–å¯ç”¨çš„æ€§çŠ¶åˆ—è¡¨ç”¨äºåŠ¨æ€åˆ›å»ºç»Ÿè®¡é€‰é¡¹å¡
            enabled_traits = self._get_enabled_traits()
            if enabled_traits:
                # é‡æ–°åˆ›å»ºç»Ÿè®¡é€‰é¡¹å¡
                self.create_statistics_tabs(enabled_traits)
            
            # å¼ºåˆ¶åˆ·æ–°ç•Œé¢æ˜¾ç¤º
            self.refresh_results_display(results_df)
            
            self.export_btn.setEnabled(True)

            # æ˜¾ç¤ºç»Ÿè®¡ä¿¡æ¯
            self.show_statistics(results_df)
            
            # æ›´æ–°ç­›é€‰åˆ†æ
            if stats:
                print(f"ç»Ÿè®¡ä¿¡æ¯: {stats}")  # è°ƒè¯•ä¿¡æ¯
                self.update_analysis_panel(stats)
                
                # å¤„ç†å°¿ç´ æ°®è¿½è¸ªç»“æœ
                if 'urea_tracking' in stats:
                    self.urea_tracking_results = stats['urea_tracking']
                    # åœ¨ç»“æœæ ‡ç­¾é¡µä¸­æ·»åŠ å°¿ç´ æ°®è¿½è¸ªæ ‡ç­¾
                    self.add_urea_tracking_tab()
            else:
                print("æ²¡æœ‰æ”¶åˆ°ç»Ÿè®¡ä¿¡æ¯")  # è°ƒè¯•ä¿¡æ¯
            
        else:
            print(f"ç­›é€‰å¤±è´¥: {message}")  # è°ƒè¯•ä¿¡æ¯
            QMessageBox.critical(self, "ç­›é€‰å¤±è´¥", message)
        
        self.statusBar().showMessage(message)
    
    def refresh_results_display(self, df):
        """å¼ºåˆ¶åˆ·æ–°ç»“æœæ˜¾ç¤º"""
        try:
            # æ¸…ç©ºå½“å‰è¡¨æ ¼
            self.result_table.clear()
            self.result_table.setRowCount(0)
            self.result_table.setColumnCount(0)
            
            # å¼ºåˆ¶å¤„ç†å¾…å¤„ç†çš„äº‹ä»¶
            QApplication.processEvents()
            
            # é‡æ–°æ˜¾ç¤ºç»“æœ
            self.display_results(df)
            
            # å†æ¬¡å¼ºåˆ¶å¤„ç†äº‹ä»¶ï¼Œç¡®ä¿ç•Œé¢æ›´æ–°
            QApplication.processEvents()
            
            # å¼ºåˆ¶é‡ç»˜è¡¨æ ¼
            self.result_table.viewport().update()
            self.result_table.update()
            
            print(f"DEBUG: ç•Œé¢åˆ·æ–°å®Œæˆï¼Œè¡¨æ ¼å½“å‰è¡Œæ•°: {self.result_table.rowCount()}")
            
        except Exception as e:
            print(f"DEBUG: åˆ·æ–°ç•Œé¢æ—¶å‡ºé”™: {e}")
            # å¦‚æœå¼ºåˆ¶åˆ·æ–°å¤±è´¥ï¼Œä»ç„¶å°è¯•æ™®é€šæ˜¾ç¤º
            self.display_results(df)
    
    def update_analysis_panel(self, stats):
        """æ›´æ–°ç­›é€‰åˆ†æé¢æ¿"""
        # æ›´æ–°å„ä¸ªç»Ÿè®¡å¡ç‰‡çš„æ•°å€¼
        getattr(self.total_data_card, 'value_label').setText(str(stats.get('total_cows', 0)))
        getattr(self.range_data_card, 'value_label').setText(str(stats.get('range_cows', 0)))
        getattr(self.result_data_card, 'value_label').setText(str(stats.get('result_cows', 0)))
        getattr(self.rate_data_card, 'value_label').setText(f"{stats.get('filter_rate', 0):.1f}%")
        
        # åˆ‡æ¢åˆ°ç­›é€‰åˆ†ææ ‡ç­¾é¡µ
        self.tab_widget.setCurrentIndex(2)  # ç­›é€‰åˆ†ææ˜¯ç¬¬3ä¸ªæ ‡ç­¾é¡µï¼ˆç´¢å¼•ä¸º2ï¼‰
    
    def display_results(self, df):
        """æ˜¾ç¤ºç­›é€‰ç»“æœ"""
        if df.empty:
            self.result_table.setRowCount(0)
            self.result_table.setColumnCount(0)
            return
        
        # é¦–å…ˆè¿‡æ»¤æ‰æ‰€æœ‰æ•°æ®éƒ½ä¸ºç©ºçš„è¡Œ
        # æ£€æŸ¥æ¯è¡Œæ˜¯å¦è‡³å°‘æœ‰ä¸€ä¸ªéç©ºçš„å…³é”®å­—æ®µ
        key_fields = ['management_id', 'parity']
        valid_rows = []
        
        for i, row in df.iterrows():
            # æ£€æŸ¥å…³é”®å­—æ®µæ˜¯å¦è‡³å°‘æœ‰ä¸€ä¸ªä¸ä¸ºç©º
            has_key_data = any(pd.notna(row.get(field)) and str(row.get(field)).strip() != '' for field in key_fields)
            if has_key_data:
                valid_rows.append(i)
        
        # åªæ˜¾ç¤ºæœ‰æ•ˆè¡Œ
        if valid_rows:
            filtered_df = df.loc[valid_rows].copy()
        else:
            # å¦‚æœæ²¡æœ‰æœ‰æ•ˆè¡Œï¼Œæ˜¾ç¤ºç©ºè¡¨æ ¼
            self.result_table.setRowCount(0)
            self.result_table.setColumnCount(0)
            return
        
        print(f"DEBUG: åŸå§‹æ•°æ®è¡Œæ•°: {len(df)}, æœ‰æ•ˆæ•°æ®è¡Œæ•°: {len(filtered_df)}")
        
        # åˆ—åä¸­è‹±æ–‡æ˜ å°„
        column_mapping = {
            'management_id': 'ç®¡ç†å·',
            'parity': 'æœ€åä¸€æ¬¡å–æ ·æ—¶çš„èƒæ¬¡',
            'å¹³å‡è›‹ç™½ç‡(%)': 'å¹³å‡è›‹ç™½ç‡(%)',
            'æœ€åä¸€ä¸ªæœˆæ³Œä¹³å¤©æ•°(å¤©)': 'æœ€åä¸€ä¸ªæœˆæ³Œä¹³å¤©æ•°(å¤©)',
            'æœªæ¥æ³Œä¹³å¤©æ•°(å¤©)': 'æœªæ¥æ³Œä¹³å¤©æ•°(å¤©)'
        }
        
        # è·å–ä¸­æ–‡åˆ—å
        chinese_columns = []
        for col in filtered_df.columns:
            if col in column_mapping:
                chinese_columns.append(column_mapping[col])
            else:
                chinese_columns.append(col)  # æœˆä»½åˆ—åå·²ç»æ˜¯ä¸­æ–‡æ ¼å¼
        
        # è·å–æ€»å¹³å‡å€¼å’Œæœˆåº¦å¹³å‡å€¼
        overall_avg = getattr(filtered_df, 'attrs', {}).get('overall_protein_avg', None)
        monthly_averages = getattr(filtered_df, 'attrs', {}).get('monthly_averages', {})
        parity_avg = getattr(filtered_df, 'attrs', {}).get('parity_avg', None)
        
        # è®¾ç½®è¡¨æ ¼è¡Œæ•°ï¼ˆæ€»å¹³å‡å€¼è¡Œ + æœˆåº¦å¹³å‡å€¼è¡Œ + æœ‰æ•ˆæ•°æ®è¡Œï¼‰
        extra_rows = 0
        if overall_avg is not None:
            extra_rows += 1
        if monthly_averages:
            extra_rows += 1
            
        self.result_table.setRowCount(len(filtered_df) + extra_rows)
        self.result_table.setColumnCount(len(filtered_df.columns))
        self.result_table.setHorizontalHeaderLabels(chinese_columns)
        
        # æ·»åŠ æ±‡æ€»è¡Œ
        row_offset = 0
        
        # ç¬¬ä¸€è¡Œï¼šæ€»å¹³å‡å€¼
        if overall_avg is not None:
            # ä¸ä½¿ç”¨åˆå¹¶å•å…ƒæ ¼ï¼Œåœ¨æ¯åˆ—è®¾ç½®é€‚å½“çš„å†…å®¹
            for j in range(len(filtered_df.columns)):
                if j == 0:
                    avg_item = QTableWidgetItem(f"æ‰€æœ‰æœˆä»½è›‹ç™½ç‡æ€»å¹³å‡å€¼: {overall_avg}%")
                    avg_item.setFont(QFont("å¾®è½¯é›…é»‘", 10, QFont.Weight.Bold))
                else:
                    avg_item = QTableWidgetItem("")
                    
                avg_item.setBackground(QColor(255, 235, 59))  # é»„è‰²èƒŒæ™¯çªå‡ºæ˜¾ç¤º
                self.result_table.setItem(row_offset, j, avg_item)
            
            row_offset += 1
        
        # ç¬¬äºŒè¡Œï¼šå„æœˆå¹³å‡å€¼
        if monthly_averages:
            for j, column_name in enumerate(chinese_columns):
                original_col = filtered_df.columns[j]
                
                if original_col in monthly_averages and monthly_averages[original_col] is not None:
                    # æ˜¾ç¤ºè¯¥æœˆçš„å¹³å‡å€¼
                    avg_value = monthly_averages[original_col]
                    if 'è›‹ç™½ç‡' in column_name:
                        # è›‹ç™½ç‡æ˜¾ç¤ºç™¾åˆ†å·
                        month_avg_item = QTableWidgetItem(f"{avg_value}%")
                        month_avg_item.setBackground(QColor(255, 248, 220))  # æµ…é»„è‰²èƒŒæ™¯
                    elif 'æ³Œä¹³å¤©æ•°' in column_name:
                        # æ³Œä¹³å¤©æ•°æ˜¾ç¤ºå¤©æ•°
                        month_avg_item = QTableWidgetItem(f"{avg_value}å¤©")
                        month_avg_item.setBackground(QColor(240, 248, 255))  # æµ…è“è‰²èƒŒæ™¯
                    elif 'äº§å¥¶é‡' in column_name:
                        # äº§å¥¶é‡æ˜¾ç¤ºå•ä½
                        month_avg_item = QTableWidgetItem(f"{avg_value}Kg")
                        month_avg_item.setBackground(QColor(240, 255, 240))  # æµ…ç»¿è‰²èƒŒæ™¯
                    else:
                        month_avg_item = QTableWidgetItem(str(avg_value))
                        month_avg_item.setBackground(QColor(255, 248, 220))
                    
                    month_avg_item.setFont(QFont("å¾®è½¯é›…é»‘", 9, QFont.Weight.Bold))
                    month_avg_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                elif 'è›‹ç™½ç‡' in column_name:
                    # è›‹ç™½ç‡åˆ—ä½†æ²¡æœ‰æ•°æ®
                    month_avg_item = QTableWidgetItem("--")
                    month_avg_item.setBackground(QColor(255, 248, 220))
                    month_avg_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                elif 'æ³Œä¹³å¤©æ•°' in column_name:
                    # æ³Œä¹³å¤©æ•°åˆ—ä½†æ²¡æœ‰æ•°æ®
                    month_avg_item = QTableWidgetItem("--")
                    month_avg_item.setBackground(QColor(240, 248, 255))
                    month_avg_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                elif column_name == 'æœ€åä¸€æ¬¡å–æ ·æ—¶çš„èƒæ¬¡' and parity_avg is not None:
                    # èƒæ¬¡åˆ—æ˜¾ç¤ºå¹³å‡èƒæ¬¡
                    month_avg_item = QTableWidgetItem(f"{parity_avg}èƒ")
                    month_avg_item.setBackground(QColor(240, 255, 240))  # æµ…ç»¿è‰²èƒŒæ™¯
                    month_avg_item.setFont(QFont("å¾®è½¯é›…é»‘", 9, QFont.Weight.Bold))
                    month_avg_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                elif j == 0:
                    # ç¬¬ä¸€åˆ—æ˜¾ç¤ºæ ‡ç­¾
                    month_avg_item = QTableWidgetItem("å½“æœˆå¹³å‡å€¼")
                    month_avg_item.setBackground(QColor(240, 240, 240))
                    month_avg_item.setFont(QFont("å¾®è½¯é›…é»‘", 9, QFont.Weight.Bold))
                else:
                    # å…¶ä»–åˆ—ä¸ºç©º
                    month_avg_item = QTableWidgetItem("")
                    month_avg_item.setBackground(QColor(248, 248, 248))
                
                self.result_table.setItem(row_offset, j, month_avg_item)
            
            row_offset += 1
        
        # å¡«å……æœ‰æ•ˆæ•°æ®ï¼ˆä¸å†ä½¿ç”¨åŸå§‹ç´¢å¼•ï¼Œè€Œæ˜¯è¿ç»­å¡«å……ï¼‰
        display_row = row_offset
        for _, row in filtered_df.iterrows():
            for j, value in enumerate(row):
                if pd.notna(value):
                    item = QTableWidgetItem(str(value))
                else:
                    item = QTableWidgetItem("")
                
                # åˆ¤æ–­æ˜¯å¦ä¸ºè›‹ç™½ç‡ç›¸å…³åˆ—ï¼Œè®¾ç½®é»„è‰²èƒŒæ™¯
                column_name = chinese_columns[j]
                is_protein_column = (
                    'è›‹ç™½ç‡' in column_name or 
                    column_name == 'å¹³å‡è›‹ç™½ç‡(%)'
                )
                is_lactation_column = (
                    'æ³Œä¹³å¤©æ•°' in column_name
                )
                is_milk_column = (
                    'äº§å¥¶é‡' in column_name
                )
                
                if is_protein_column:
                    # è›‹ç™½ç‡ç›¸å…³åˆ—æ ‡é»„
                    item.setBackground(QColor(255, 248, 220))  # æµ…é»„è‰²èƒŒæ™¯
                    if column_name == 'å¹³å‡è›‹ç™½ç‡(%)' and pd.notna(value):
                        item.setFont(QFont("å¾®è½¯é›…é»‘", 9, QFont.Weight.Bold))
                elif is_lactation_column:
                    # æ³Œä¹³å¤©æ•°ç›¸å…³åˆ—æ ‡è“
                    item.setBackground(QColor(240, 248, 255))  # æµ…è“è‰²èƒŒæ™¯
                elif is_milk_column:
                    # äº§å¥¶é‡ç›¸å…³åˆ—æ ‡ç»¿
                    item.setBackground(QColor(240, 255, 240))  # æµ…ç»¿è‰²èƒŒæ™¯
                
                self.result_table.setItem(display_row, j, item)
            
            display_row += 1
        
        # è®¾ç½®è¡¨å¤´æ ·å¼ - ä¸ºè›‹ç™½ç‡åˆ—è®¾ç½®é»„è‰²èƒŒæ™¯
        protein_columns = []
        for j, column_name in enumerate(chinese_columns):
            is_protein_column = (
                'è›‹ç™½ç‡' in column_name or 
                column_name == 'å¹³å‡è›‹ç™½ç‡(%)'
            )
            if is_protein_column:
                protein_columns.append(j)
        
        # æ”¶é›†æ³Œä¹³å¤©æ•°åˆ—ç´¢å¼•
        lactation_columns = []
        for j, column_name in enumerate(chinese_columns):
            is_lactation_column = 'æ³Œä¹³å¤©æ•°' in column_name
            if is_lactation_column:
                lactation_columns.append(j)
        
        # æ”¶é›†äº§å¥¶é‡åˆ—ç´¢å¼•
        milk_columns = []
        for j, column_name in enumerate(chinese_columns):
            is_milk_column = 'äº§å¥¶é‡' in column_name
            if is_milk_column:
                milk_columns.append(j)
        
        # ä¸ºç‰¹æ®Šåˆ—è®¾ç½®æ ·å¼
        if protein_columns or lactation_columns or milk_columns:
            header_style = """
                QHeaderView::section {
                    background-color: #f8f9fa;
                    border: 1px solid #dee2e6;
                    padding: 4px;
                    font-weight: bold;
                }
            """
            
            # ä¸ºè›‹ç™½ç‡åˆ—æ·»åŠ é»„è‰²èƒŒæ™¯çš„æ ·å¼
            for col_index in protein_columns:
                header_style += f"""
                    QHeaderView::section:nth-child({col_index + 1}) {{
                        background-color: #fff3cd;
                        color: #856404;
                        font-weight: bold;
                    }}
                """
            
            # ä¸ºæ³Œä¹³å¤©æ•°åˆ—æ·»åŠ è“è‰²èƒŒæ™¯çš„æ ·å¼
            for col_index in lactation_columns:
                header_style += f"""
                    QHeaderView::section:nth-child({col_index + 1}) {{
                        background-color: #e3f2fd;
                        color: #0d47a1;
                        font-weight: bold;
                    }}
                """
            
            # ä¸ºäº§å¥¶é‡åˆ—æ·»åŠ ç»¿è‰²èƒŒæ™¯çš„æ ·å¼
            for col_index in milk_columns:
                header_style += f"""
                    QHeaderView::section:nth-child({col_index + 1}) {{
                        background-color: #d3f5d3;
                        color: #155724;
                        font-weight: bold;
                    }}
                """
            
            self.result_table.horizontalHeader().setStyleSheet(header_style)
        
        # è°ƒæ•´åˆ—å®½
        self.result_table.horizontalHeader().setStretchLastSection(True)
        self.result_table.resizeColumnsToContents()
        
        # åˆ‡æ¢åˆ°ç»“æœæ ‡ç­¾é¡µ
        self.tab_widget.setCurrentIndex(1)
        
        print(f"DEBUG: æœ€ç»ˆæ˜¾ç¤ºè¡Œæ•°: {self.result_table.rowCount()}, åˆ—æ•°: {self.result_table.columnCount()}")
    
    def show_statistics(self, df):
        """æ˜¾ç¤ºç»Ÿè®¡ä¿¡æ¯åˆ°ä¸åŒçš„é€‰é¡¹å¡"""
        if df.empty:
            empty_message = "æš‚æ— ç­›é€‰ç»“æœ\nè¯·æ£€æŸ¥ç­›é€‰æ¡ä»¶æ˜¯å¦è¿‡äºä¸¥æ ¼ã€‚"
            self.overall_stats_widget.setText(f"ğŸ“Š æ€»ä½“ç»Ÿè®¡ä¿¡æ¯\n\n{empty_message}")
            
            # æ›´æ–°èƒæ¬¡åˆ†æ
            if hasattr(self, 'parity_stats_widget'):
                self.parity_stats_widget.setText(f"ğŸ„ èƒæ¬¡åˆ†æ\n\n{empty_message}")
            
            # æ›´æ–°åŠ¨æ€æ€§çŠ¶é€‰é¡¹å¡
            if hasattr(self, 'trait_stats_widgets'):
                for trait, widget in self.trait_stats_widgets.items():
                    widget.setText(f"æ€§çŠ¶åˆ†æ\n\n{empty_message}")
            
            # å…¼å®¹æ—§ç‰ˆæœ¬çš„å›ºå®šé€‰é¡¹å¡
            if hasattr(self, 'protein_stats_widget'):
                self.protein_stats_widget.setText(f"ğŸ¥› è›‹ç™½ç‡åˆ†æ\n\n{empty_message}")
            if hasattr(self, 'somatic_stats_widget'):
                self.somatic_stats_widget.setText(f"ğŸ”¬ ä½“ç»†èƒæ•°åˆ†æ\n\n{empty_message}")
            if hasattr(self, 'other_traits_stats_widget'):
                self.other_traits_stats_widget.setText(f"ğŸ“ˆ å…¶ä»–æ€§çŠ¶åˆ†æ\n\n{empty_message}")
            return
        
        # 1. æ€»ä½“ç»Ÿè®¡
        self.update_overall_statistics(df)
        
        # 2. èƒæ¬¡åˆ†æ
        self.update_parity_statistics(df)
        
        # 3. åŠ¨æ€æ€§çŠ¶åˆ†æ
        if hasattr(self, 'trait_stats_widgets'):
            for trait, widget in self.trait_stats_widgets.items():
                self.update_trait_statistics(df, trait, widget)
        else:
            # å…¼å®¹æ—§ç‰ˆæœ¬ - è›‹ç™½ç‡åˆ†æ
            if hasattr(self, 'protein_stats_widget'):
                self.update_protein_statistics(df)
        
        # 3. ä½“ç»†èƒæ•°åˆ†æ
        self.update_somatic_statistics(df)
        
        # 4. èƒæ¬¡åˆ†æ
        self.update_parity_statistics(df)
        
        # 5. å…¶ä»–æ€§çŠ¶åˆ†æ
        self.update_other_traits_statistics(df)
    
    def update_overall_statistics(self, df):
        """æ›´æ–°æ€»ä½“ç»Ÿè®¡ä¿¡æ¯"""
        stats = "ğŸ“Š æ€»ä½“ç»Ÿè®¡ä¿¡æ¯\n\n"
        
        # æ˜¾ç¤ºæ€»å¹³å‡è›‹ç™½ç‡ï¼ˆæ¥è‡ªDataFrameå±æ€§ï¼‰
        overall_avg = getattr(df, 'attrs', {}).get('overall_protein_avg', None)
        if overall_avg is not None:
            stats += f"ğŸ¯ æ‰€æœ‰æœˆä»½è›‹ç™½ç‡æ€»å¹³å‡å€¼: {overall_avg}%\n"
        
        # æ˜¾ç¤ºå¹³å‡èƒæ¬¡ï¼ˆæ¥è‡ªDataFrameå±æ€§ï¼‰
        parity_avg = getattr(df, 'attrs', {}).get('parity_avg', None)
        if parity_avg is not None:
            stats += f"ğŸ„ ç¾¤ä½“å¹³å‡èƒæ¬¡: {parity_avg}èƒ\n\n"
        else:
            stats += "\n"
        
        stats += f"ğŸ“Š æ€»è®°å½•æ•°: {len(df)}\n"
        
        # ç»Ÿè®¡å”¯ä¸€ç‰›åªæ•°ï¼ˆåŸºäºmanagement_idï¼‰
        if 'management_id' in df.columns:
            unique_cows = df['management_id'].dropna().nunique()
            stats += f"ğŸ„ ç¬¦åˆæ¡ä»¶ç‰›åªæ•°: {unique_cows}å¤´\n"
        
        # æœ€åä¸€ä¸ªæœˆæ³Œä¹³å¤©æ•°ç»Ÿè®¡
        if 'æœ€åä¸€ä¸ªæœˆæ³Œä¹³å¤©æ•°' in df.columns:
            lactation_data = df['æœ€åä¸€ä¸ªæœˆæ³Œä¹³å¤©æ•°'].dropna()
            if not lactation_data.empty:
                stats += f"\nâ° æœ€åä¸€ä¸ªæœˆæ³Œä¹³å¤©æ•°:\n"
                stats += f"  å¹³å‡: {lactation_data.mean():.1f}å¤©\n"
                stats += f"  èŒƒå›´: {lactation_data.min():.0f}-{lactation_data.max():.0f}å¤©\n"
        
        self.overall_stats_widget.setText(stats)
    
    def update_protein_statistics(self, df):
        """æ›´æ–°è›‹ç™½ç‡åˆ†æä¿¡æ¯"""
        stats = "ğŸ¥› è›‹ç™½ç‡åˆ†æ\n\n"
        
        # ä¸ªä½“å¹³å‡è›‹ç™½ç‡ç»Ÿè®¡
        if 'å¹³å‡è›‹ç™½ç‡(%)' in df.columns:
            individual_avg = df['å¹³å‡è›‹ç™½ç‡(%)'].dropna()
            if not individual_avg.empty:
                stats += f"ğŸ“ˆ ä¸ªä½“å¹³å‡è›‹ç™½ç‡åˆ†å¸ƒ:\n"
                stats += f"  æœ€ä½ä¸ªä½“å¹³å‡å€¼: {individual_avg.min():.2f}%\n"
                stats += f"  æœ€é«˜ä¸ªä½“å¹³å‡å€¼: {individual_avg.max():.2f}%\n"
                stats += f"  ç¾¤ä½“å¹³å‡å€¼: {individual_avg.mean():.2f}%\n"
                stats += f"  æ ‡å‡†å·®: {individual_avg.std():.2f}%\n\n"
        
        # è›‹ç™½ç‡è¯¦ç»†ç»Ÿè®¡ï¼ˆä»…å½“æœ‰è›‹ç™½ç‡æ•°æ®æ—¶æ˜¾ç¤ºï¼‰
        protein_columns = [col for col in df.columns if col.endswith('æœˆè›‹ç™½ç‡(%)')]
        if protein_columns:
            stats += f"ğŸ“… è›‹ç™½ç‡æœˆåº¦æ˜ç»†:\n"
            # æŒ‰æ—¶é—´é¡ºåºæ’åˆ—æœˆä»½åˆ—
            sorted_protein_cols = sorted(protein_columns, key=lambda x: x[:7])  # æŒ‰YYYYå¹´MMæœˆæ’åº
            
            for col in sorted_protein_cols:
                protein_data = df[col].dropna()
                if not protein_data.empty:
                    month_name = col.replace('è›‹ç™½ç‡(%)', '')
                    stats += f"  {month_name}:\n"
                    stats += f"    å¹³å‡: {protein_data.mean():.2f}%\n"
                    stats += f"    èŒƒå›´: {protein_data.min():.2f}%-{protein_data.max():.2f}%\n"
                    stats += f"    æ ‡å‡†å·®: {protein_data.std():.2f}%\n"
                    stats += f"    æ ·æœ¬æ•°: {len(protein_data)}å¤´\n\n"
        else:
            stats += f"ğŸ“… å½“å‰ç­›é€‰ç»“æœä¸­æ— è›‹ç™½ç‡æ•°æ®\n\n"
        
        # æ€»ä½“å¹³å‡è›‹ç™½ç‡ï¼ˆæ¥è‡ªDataFrameå±æ€§ï¼Œä»…å½“æœ‰è›‹ç™½ç‡æ•°æ®æ—¶æ˜¾ç¤ºï¼‰
        overall_avg = getattr(df, 'attrs', {}).get('overall_protein_avg', None)
        if overall_avg is not None:
            stats += f"ğŸ’¡ è®¡ç®—è¯´æ˜:\n"
            stats += f"  â€¢ æ‰€æœ‰æœˆä»½è›‹ç™½ç‡æ€»å¹³å‡å€¼: {overall_avg}%\n"
            stats += f"  â€¢ é‡‡ç”¨äº§å¥¶é‡åŠ æƒå¹³å‡è®¡ç®—\n"
            stats += f"  â€¢ ä¸ªä½“å¹³å‡å€¼é‡‡ç”¨ç®—æœ¯å¹³å‡\n"
        elif protein_columns:  # æœ‰è›‹ç™½ç‡åˆ—ä½†æ²¡æœ‰æ€»å¹³å‡å€¼
            stats += f"ğŸ’¡ è®¡ç®—è¯´æ˜:\n"
            stats += f"  â€¢ æ— æ³•è®¡ç®—æ€»å¹³å‡å€¼ï¼ˆç¼ºå°‘äº§å¥¶é‡æ•°æ®ï¼‰\n"
            stats += f"  â€¢ ä¸ªä½“å¹³å‡å€¼é‡‡ç”¨ç®—æœ¯å¹³å‡\n"
        
        self.protein_stats_widget.setText(stats)
    
    def update_somatic_statistics(self, df):
        """æ›´æ–°ä½“ç»†èƒæ•°åˆ†æä¿¡æ¯"""
        stats = "ğŸ”¬ ä½“ç»†èƒæ•°åˆ†æ\n\n"
        
        # ä½“ç»†èƒæ•°è¯¦ç»†ç»Ÿè®¡
        somatic_columns = [col for col in df.columns if 'ä½“ç»†èƒæ•°' in col and col.endswith('ä¸‡/ml)')]
        if somatic_columns:
            stats += f"ğŸ“… ä½“ç»†èƒæ•°æœˆåº¦æ˜ç»†:\n"
            # æŒ‰æ—¶é—´é¡ºåºæ’åˆ—æœˆä»½åˆ—
            sorted_somatic_cols = sorted(somatic_columns, key=lambda x: x[:7])  # æŒ‰YYYYå¹´MMæœˆæ’åº
            
            for col in sorted_somatic_cols:
                somatic_data = df[col].dropna()
                if not somatic_data.empty:
                    month_name = col.replace('ä½“ç»†èƒæ•°(ä¸‡/ml)', '')
                    stats += f"  {month_name}:\n"
                    stats += f"    å¹³å‡: {somatic_data.mean():.1f}ä¸‡/ml\n"
                    stats += f"    èŒƒå›´: {somatic_data.min():.1f}-{somatic_data.max():.1f}ä¸‡/ml\n"
                    stats += f"    æ ‡å‡†å·®: {somatic_data.std():.1f}ä¸‡/ml\n"
                    stats += f"    æ ·æœ¬æ•°: {len(somatic_data)}å¤´\n\n"
        
        if not somatic_columns:
            stats += "æœ¬æ¬¡ç­›é€‰ç»“æœä¸­æ— ä½“ç»†èƒæ•°æ•°æ®ã€‚\n\n"
        
        stats += f"ğŸ’¡ è¯´æ˜:\n"
        stats += f"  â€¢ ä½“ç»†èƒæ•°ä»¥ä¸‡/mlä¸ºå•ä½\n"
        stats += f"  â€¢ æ•°å€¼è¶Šä½è¡¨ç¤ºä¹³æˆ¿å¥åº·çŠ¶å†µè¶Šå¥½\n"
        stats += f"  â€¢ æ­£å¸¸èŒƒå›´é€šå¸¸åœ¨10-40ä¸‡/mlä¹‹é—´\n"
        
        self.somatic_stats_widget.setText(stats)
    
    def update_parity_statistics(self, df):
        """æ›´æ–°èƒæ¬¡åˆ†æä¿¡æ¯"""
        stats = "ğŸ„ èƒæ¬¡åˆ†æ\n\n"
        
        # æŒ‰èƒæ¬¡ç»Ÿè®¡  
        if 'parity' in df.columns:
            parity_counts = df['parity'].value_counts().sort_index()
            stats += f"ğŸ“Š å„èƒæ¬¡åˆ†å¸ƒ:\n"
            total_cows = parity_counts.sum()
            for parity, count in parity_counts.items():
                percentage = (count / total_cows) * 100
                stats += f"  {parity}èƒ: {count}å¤´ ({percentage:.1f}%)\n"
            
            # èƒæ¬¡ç»Ÿè®¡æŒ‡æ ‡
            parity_data = df['parity'].dropna()
            if not parity_data.empty:
                stats += f"\nğŸ“ˆ èƒæ¬¡ç»Ÿè®¡æŒ‡æ ‡:\n"
                stats += f"  å¹³å‡èƒæ¬¡: {parity_data.mean():.1f}èƒ\n"
                stats += f"  èƒæ¬¡èŒƒå›´: {parity_data.min():.0f}-{parity_data.max():.0f}èƒ\n"
                stats += f"  ä¸­ä½æ•°èƒæ¬¡: {parity_data.median():.0f}èƒ\n"
                stats += f"  æ€»ç‰›å¤´æ•°: {len(parity_data)}å¤´\n"
                
                # èƒæ¬¡åˆ†ç»„åˆ†æ
                stats += f"\nğŸ” èƒæ¬¡åˆ†ç»„åˆ†æ:\n"
                first_parity = parity_counts.get(1, 0)
                second_parity = parity_counts.get(2, 0)
                third_plus = parity_counts[parity_counts.index >= 3].sum() if len(parity_counts[parity_counts.index >= 3]) > 0 else 0
                
                stats += f"  å¤´èƒç‰›: {first_parity}å¤´ ({(first_parity/total_cows)*100:.1f}%)\n"
                stats += f"  äºŒèƒç‰›: {second_parity}å¤´ ({(second_parity/total_cows)*100:.1f}%)\n"
                stats += f"  ä¸‰èƒåŠä»¥ä¸Š: {third_plus}å¤´ ({(third_plus/total_cows)*100:.1f}%)\n"
        
        self.parity_stats_widget.setText(stats)
    
    def update_other_traits_statistics(self, df):
        """æ›´æ–°å…¶ä»–æ€§çŠ¶åˆ†æä¿¡æ¯"""
        stats = "ğŸ“ˆ å…¶ä»–æ€§çŠ¶åˆ†æ\n\n"
        
        # äº§å¥¶é‡è¯¦ç»†ç»Ÿè®¡
        milk_columns = [col for col in df.columns if col.endswith('æœˆäº§å¥¶é‡(Kg)')]
        if milk_columns:
            stats += f"ğŸ¥› äº§å¥¶é‡æœˆåº¦æ˜ç»†:\n"
            # æŒ‰æ—¶é—´é¡ºåºæ’åˆ—æœˆä»½åˆ—
            sorted_milk_cols = sorted(milk_columns, key=lambda x: x[:7])  # æŒ‰YYYYå¹´MMæœˆæ’åº
            
            for col in sorted_milk_cols:
                milk_data = df[col].dropna()
                if not milk_data.empty:
                    month_name = col.replace('äº§å¥¶é‡(Kg)', '')
                    stats += f"  {month_name}:\n"
                    stats += f"    å¹³å‡: {milk_data.mean():.1f}Kg\n"
                    stats += f"    èŒƒå›´: {milk_data.min():.1f}-{milk_data.max():.1f}Kg\n"
                    stats += f"    æ ‡å‡†å·®: {milk_data.std():.1f}Kg\n"
                    stats += f"    æ ·æœ¬æ•°: {len(milk_data)}å¤´\n\n"
        
        # å…¶ä»–æ€§çŠ¶ç»Ÿè®¡ï¼ˆä¹³è„‚ç‡ã€æ³Œä¹³å¤©æ•°ç­‰ï¼‰
        other_trait_patterns = [
            ('ä¹³è„‚ç‡', 'æœˆä¹³è„‚ç‡(%)', '%'),
            ('æ³Œä¹³å¤©æ•°', 'æœˆæ³Œä¹³å¤©æ•°(å¤©)', 'å¤©'),
            ('ä¹³ç³–ç‡', 'æœˆä¹³ç³–ç‡(%)', '%'),
            ('å›ºå½¢ç‰©', 'æœˆå›ºå½¢ç‰©(%)', '%'),
            ('è„‚è›‹æ¯”', 'æœˆè„‚è›‹æ¯”', ''),
            ('å°¿ç´ æ°®', 'æœˆå°¿ç´ æ°®(mg/dl)', 'mg/dl')
        ]
        
        for trait_name, pattern, unit in other_trait_patterns:
            trait_columns = [col for col in df.columns if pattern in col]
            if trait_columns:
                stats += f"ğŸ“Š {trait_name}åˆ†æ:\n"
                # æŒ‰æ—¶é—´é¡ºåºæ’åˆ—
                sorted_cols = sorted(trait_columns, key=lambda x: x[:7])
                
                for col in sorted_cols:
                    trait_data = df[col].dropna()
                    if not trait_data.empty:
                        month_name = col.replace(pattern, '')
                        stats += f"  {month_name}: å¹³å‡ {trait_data.mean():.2f}{unit} "
                        stats += f"(èŒƒå›´: {trait_data.min():.2f}-{trait_data.max():.2f}{unit})\n"
                stats += "\n"
        
        if not milk_columns and not any(df.columns.str.contains(pattern).any() for _, pattern, _ in other_trait_patterns):
            stats += "æœ¬æ¬¡ç­›é€‰ç»“æœä¸­æš‚æ— å…¶ä»–æ€§çŠ¶æ•°æ®ã€‚\n"
        
        self.other_traits_stats_widget.setText(stats)
    
    def export_results(self):
        """å¯¼å‡ºç»“æœ"""
        if self.current_results.empty:
            self.show_warning("è­¦å‘Š", "æ²¡æœ‰å¯å¯¼å‡ºçš„ç»“æœ")
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "ä¿å­˜ç­›é€‰ç»“æœ",
            f"DHIç­›é€‰ç»“æœ_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excelæ–‡ä»¶ (*.xlsx)"
        )
        
        if filename:
            try:
                self._export_formatted_excel(filename)
                
                # ä½¿ç”¨ç¾è§‚çš„è‡ªå®šä¹‰å¯¹è¯æ¡†
                self.show_export_success_dialog("DHIç­›é€‰ç»“æœå·²ä¿å­˜åˆ°ï¼š", filename)
                
                self.statusBar().showMessage(f"å·²å¯¼å‡ºåˆ°: {filename}")
                
            except Exception as e:
                QMessageBox.critical(self, "å¯¼å‡ºå¤±è´¥", f"å¯¼å‡ºå¤±è´¥:\n{str(e)}")

    def open_file(self, file_path):
        """æ‰“å¼€æ–‡ä»¶"""
        try:
            import subprocess
            import os
            
            if os.path.exists(file_path):
                if os.name == 'nt':  # Windows
                    os.startfile(file_path)
                elif os.name == 'posix':  # macOS and Linux
                    if sys.platform == 'darwin':  # macOS
                        subprocess.run(['open', file_path])
                    else:  # Linux
                        subprocess.run(['xdg-open', file_path])
                
                print(f"æ­£åœ¨æ‰“å¼€æ–‡ä»¶: {file_path}")
            else:
                QMessageBox.warning(self, "é”™è¯¯", f"æ–‡ä»¶ä¸å­˜åœ¨:\n{file_path}")
        except Exception as e:
            QMessageBox.warning(self, "æ‰“å¼€å¤±è´¥", f"æ— æ³•æ‰“å¼€æ–‡ä»¶:\n{str(e)}")

    def open_file_folder(self, file_path):
        """æ‰“å¼€æ–‡ä»¶æ‰€åœ¨æ–‡ä»¶å¤¹"""
        try:
            import subprocess
            import os
            
            if os.path.exists(file_path):
                folder_path = os.path.dirname(file_path)
                
                if os.name == 'nt':  # Windows
                    # åœ¨Windowsä¸­æ‰“å¼€æ–‡ä»¶å¤¹å¹¶é€‰ä¸­æ–‡ä»¶
                    subprocess.run(['explorer', '/select,', file_path])
                elif os.name == 'posix':  # macOS and Linux
                    if sys.platform == 'darwin':  # macOS
                        # åœ¨macOSä¸­æ‰“å¼€Finderå¹¶é€‰ä¸­æ–‡ä»¶
                        subprocess.run(['open', '-R', file_path])
                    else:  # Linux
                        # åœ¨Linuxä¸­æ‰“å¼€æ–‡ä»¶ç®¡ç†å™¨
                        subprocess.run(['xdg-open', folder_path])
                
                print(f"æ­£åœ¨æ‰“å¼€æ–‡ä»¶å¤¹: {folder_path}")
            else:
                QMessageBox.warning(self, "é”™è¯¯", f"æ–‡ä»¶ä¸å­˜åœ¨:\n{file_path}")
        except Exception as e:
            QMessageBox.warning(self, "æ‰“å¼€å¤±è´¥", f"æ— æ³•æ‰“å¼€æ–‡ä»¶å¤¹:\n{str(e)}")
    
    def _export_formatted_excel(self, filename):
        """å¯¼å‡ºæ ¼å¼åŒ–çš„Excelæ–‡ä»¶"""
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # åˆ›å»ºå·¥ä½œç°¿
        wb = Workbook()
        ws = wb.active
        ws.title = "ç­›é€‰ç»“æœ"
        
        df = self.current_results
        
        # åˆ—åä¸­è‹±æ–‡æ˜ å°„ï¼ˆä¸ç•Œé¢æ˜¾ç¤ºä¸€è‡´ï¼‰
        column_mapping = {
            'management_id': 'ç®¡ç†å·',
            'parity': 'æœ€åä¸€æ¬¡å–æ ·æ—¶çš„èƒæ¬¡',
            'å¹³å‡è›‹ç™½ç‡(%)': 'å¹³å‡è›‹ç™½ç‡(%)',
            'æœ€åä¸€ä¸ªæœˆæ³Œä¹³å¤©æ•°(å¤©)': 'æœ€åä¸€ä¸ªæœˆæ³Œä¹³å¤©æ•°(å¤©)',
            'æœªæ¥æ³Œä¹³å¤©æ•°(å¤©)': 'æœªæ¥æ³Œä¹³å¤©æ•°(å¤©)'
        }
        
        # è·å–ä¸­æ–‡åˆ—å
        chinese_columns = []
        for col in df.columns:
            if col in column_mapping:
                chinese_columns.append(column_mapping[col])
            else:
                chinese_columns.append(col)  # æœˆä»½åˆ—åå·²ç»æ˜¯ä¸­æ–‡æ ¼å¼
        
        # è·å–æ€»å¹³å‡å€¼å’Œæœˆåº¦å¹³å‡å€¼
        overall_avg = getattr(df, 'attrs', {}).get('overall_protein_avg', None)
        monthly_averages = getattr(df, 'attrs', {}).get('monthly_averages', {})
        parity_avg = getattr(df, 'attrs', {}).get('parity_avg', None)
        
        # æ ·å¼å®šä¹‰
        yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # æµ…é»„è‰²
        dark_yellow_fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")  # æ·±é»„è‰²
        gray_fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")  # ç°è‰²
        light_gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")  # æµ…ç°è‰²
        
        header_font = Font(bold=True, size=11)
        bold_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # å½“å‰è¡Œä½ç½®
        current_row = 1
        
        # 1. å†™å…¥æ€»å¹³å‡å€¼è¡Œ
        if overall_avg is not None:
            # ä¸ä½¿ç”¨åˆå¹¶å•å…ƒæ ¼ï¼Œç›´æ¥åœ¨ç¬¬ä¸€åˆ—æ˜¾ç¤ºæ€»å¹³å‡å€¼
            for j in range(1, len(chinese_columns) + 1):
                cell = ws.cell(row=current_row, column=j)
                if j == 1:
                    cell.value = f"æ‰€æœ‰æœˆä»½è›‹ç™½ç‡æ€»å¹³å‡å€¼: {overall_avg}%"
                    cell.font = header_font
                else:
                    cell.value = ""
                    cell.font = normal_font
                cell.fill = dark_yellow_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            current_row += 1
        
        # 2. å†™å…¥å„æœˆå¹³å‡å€¼è¡Œ
        if monthly_averages:
            # å®šä¹‰æµ…è“è‰²å¡«å……ï¼ˆç”¨äºæ³Œä¹³å¤©æ•°ï¼‰
            light_blue_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            # å®šä¹‰æµ…ç»¿è‰²å¡«å……ï¼ˆç”¨äºèƒæ¬¡ï¼‰
            light_green_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            
            for j, column_name in enumerate(chinese_columns, 1):
                original_col = df.columns[j-1]
                cell = ws.cell(row=current_row, column=j)
                
                if original_col in monthly_averages and monthly_averages[original_col] is not None:
                    # æ˜¾ç¤ºè¯¥æœˆçš„å¹³å‡å€¼
                    avg_value = monthly_averages[original_col]
                    if 'è›‹ç™½ç‡' in column_name:
                        # è›‹ç™½ç‡æ˜¾ç¤ºç™¾åˆ†å·
                        cell.value = f"{avg_value}%"
                        cell.fill = yellow_fill
                    elif 'æ³Œä¹³å¤©æ•°' in column_name:
                        # æ³Œä¹³å¤©æ•°æ˜¾ç¤ºå¤©æ•°
                        cell.value = f"{avg_value}å¤©"
                        cell.fill = light_blue_fill
                    elif 'äº§å¥¶é‡' in column_name:
                        # äº§å¥¶é‡æ˜¾ç¤ºå•ä½
                        cell.value = f"{avg_value}Kg"
                        cell.fill = light_green_fill
                    else:
                        cell.value = str(avg_value)
                        cell.fill = yellow_fill
                    
                    cell.font = bold_font
                    cell.alignment = center_alignment
                elif 'è›‹ç™½ç‡' in column_name:
                    # è›‹ç™½ç‡åˆ—ä½†æ²¡æœ‰æ•°æ®
                    cell.value = "--"
                    cell.fill = yellow_fill
                    cell.alignment = center_alignment
                elif 'æ³Œä¹³å¤©æ•°' in column_name:
                    # æ³Œä¹³å¤©æ•°åˆ—ä½†æ²¡æœ‰æ•°æ®
                    cell.value = "--"
                    cell.fill = light_blue_fill
                    cell.alignment = center_alignment
                elif column_name == 'æœ€åä¸€æ¬¡å–æ ·æ—¶çš„èƒæ¬¡' and parity_avg is not None:
                    # èƒæ¬¡åˆ—æ˜¾ç¤ºå¹³å‡èƒæ¬¡
                    cell.value = f"{parity_avg}èƒ"
                    cell.fill = light_green_fill
                    cell.font = bold_font
                    cell.alignment = center_alignment
                elif j == 1:
                    # ç¬¬ä¸€åˆ—æ˜¾ç¤ºæ ‡ç­¾
                    cell.value = "å½“æœˆå¹³å‡å€¼"
                    cell.fill = light_gray_fill
                    cell.font = bold_font
                else:
                    # å…¶ä»–åˆ—ä¸ºç©º
                    cell.value = ""
                    cell.fill = gray_fill
                
                cell.border = thin_border
            
            current_row += 1
        
        # 3. å†™å…¥è¡¨å¤´
        for j, column_name in enumerate(chinese_columns, 1):
            cell = ws.cell(row=current_row, column=j, value=column_name)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # è›‹ç™½ç‡ç›¸å…³åˆ—æ ‡é»„
            is_protein_column = ('è›‹ç™½ç‡' in column_name or column_name == 'å¹³å‡è›‹ç™½ç‡(%)')
            is_milk_column = 'äº§å¥¶é‡' in column_name
            if is_protein_column:
                cell.fill = yellow_fill
            elif is_milk_column:
                cell.fill = light_green_fill
        
        current_row += 1
        
        # 4. å†™å…¥æ•°æ®
        for _, row in df.iterrows():
            for j, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=j)
                
                # ç‰¹æ®Šå¤„ç†æ³Œä¹³å¤©æ•°åˆ—ï¼Œç¡®ä¿æ˜¾ç¤ºä¸ºæ•´æ•°
                column_name = chinese_columns[j-1]
                if 'æ³Œä¹³å¤©æ•°' in column_name and pd.notna(value):
                    try:
                        cell.value = int(float(value))
                    except (ValueError, TypeError):
                        cell.value = value if pd.notna(value) else ""
                else:
                    cell.value = value if pd.notna(value) else ""
                
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = center_alignment
                
                # æ ¹æ®åˆ—ç±»å‹è®¾ç½®èƒŒæ™¯è‰²
                is_protein_column = ('è›‹ç™½ç‡' in column_name or column_name == 'å¹³å‡è›‹ç™½ç‡(%)')
                is_lactation_column = 'æ³Œä¹³å¤©æ•°' in column_name
                is_milk_column = 'äº§å¥¶é‡' in column_name
                
                if is_protein_column:
                    cell.fill = yellow_fill
                    if column_name == 'å¹³å‡è›‹ç™½ç‡(%)' and pd.notna(value):
                        cell.font = bold_font
                elif is_lactation_column:
                    cell.fill = light_blue_fill
                elif is_milk_column:
                    cell.fill = light_green_fill
            
            current_row += 1
        
        # 5. è°ƒæ•´åˆ—å®½
        for col_num in range(1, len(chinese_columns) + 1):
            max_length = 0
            column_letter = ws.cell(row=1, column=col_num).column_letter
            
            # éå†è¯¥åˆ—çš„æ‰€æœ‰å•å…ƒæ ¼ï¼ˆè·³è¿‡åˆå¹¶å•å…ƒæ ¼ï¼‰
            for row_num in range(1, current_row):
                try:
                    cell = ws.cell(row=row_num, column=col_num)
                    # è·³è¿‡åˆå¹¶å•å…ƒæ ¼
                    if hasattr(cell, 'value') and cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # è®¾ç½®åˆ—å®½ï¼Œè€ƒè™‘ä¸­æ–‡å­—ç¬¦
            adjusted_width = min(max(max_length * 1.2, 10), 25)  # ä¸­æ–‡å­—ç¬¦éœ€è¦æ›´å¤šç©ºé—´
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # ä¿å­˜æ–‡ä»¶
        wb.save(filename)

    def create_special_filter_group(self, title: str, filter_type: str):
        """åˆ›å»ºç‰¹æ®Šç­›é€‰ç»„ï¼ˆè›‹ç™½ç‡ã€ä½“ç»†èƒæ•°ç­‰ï¼‰
        
        Args:
            title: ç»„æ ‡é¢˜
            filter_type: ç­›é€‰ç±»å‹ ('protein' æˆ– 'somatic')
        """
        group = self.create_card_widget(title)
        layout = QVBoxLayout(group.content_widget)
        
        # è‡ªé€‚åº”è¾¹è·
        screen = QApplication.primaryScreen()
        dpi_ratio = screen.devicePixelRatio()
        scale_factor = self.display_scale / 100.0
        margin = max(int(15 * dpi_ratio * scale_factor), 8)
        
        layout.setContentsMargins(margin, margin, margin, margin)
        
        # è·å–æ ·å¼
        form_styles = self.get_responsive_form_styles()
        checkbox_font_size = max(int(13 * dpi_ratio * 0.7), 12)
        checkbox_spacing = max(int(8 * dpi_ratio * 0.6), 6)
        checkbox_size = max(int(16 * dpi_ratio * 0.6), 14)
        checkbox_border_radius = max(int(3 * dpi_ratio * 0.6), 2)
        dash_font_size = max(int(14 * dpi_ratio), 12)
        dash_margin = max(int(8 * dpi_ratio), 6)
        
        # å¯ç”¨å¼€å…³
        filter_enabled = QCheckBox(f"å¯ç”¨{title}")
        filter_enabled.setChecked(False)
        filter_enabled.setToolTip(f"å‹¾é€‰åå¯ç”¨{title}ç­›é€‰")
        filter_enabled.setStyleSheet(f"""
            QCheckBox {{
                font-size: {checkbox_font_size}px;
                color: #495057;
                spacing: {checkbox_spacing}px;
                font-weight: bold;
            }}
            QCheckBox::indicator {{
                width: {checkbox_size}px;
                height: {checkbox_size}px;
                border: 2px solid #ced4da;
                border-radius: {checkbox_border_radius}px;
                background-color: white;
            }}
            QCheckBox::indicator:hover {{
                border-color: #80bdff;
                background-color: #f8f9fa;
            }}
            QCheckBox::indicator:checked {{
                background-color: #28a745 !important;
                border-color: #28a745 !important;
                image: url(data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iMTIiIGhlaWdodD0iOSIgdmlld0JveD0iMCAwIDEyIDkiIGZpbGw9Im5vbmUiIHhtbG5zPSJodHRwOi8vd3d3LnczLm9yZy8yMDAwL3N2ZyI+CjxwYXRoIGQ9Ik0xIDQuNUw0LjUgOEwxMSAxIiBzdHJva2U9IndoaXRlIiBzdHJva2Utd2lkdGg9IjIiIHN0cm9rZS1saW5lY2FwPSJyb3VuZCIgc3Ryb2tlLWxpbmVqb2luPSJyb3VuZCIvPgo8L3N2Zz4K);
            }}
            QCheckBox::indicator:checked:hover {{
                background-color: #218838 !important;
                border-color: #218838 !important;
            }}
            QCheckBox::indicator:checked:pressed {{
                background-color: #1e7e34 !important;
                border-color: #1e7e34 !important;
            }}
        """)
        layout.addWidget(filter_enabled)
        
        # ç­›é€‰èŒƒå›´
        range_layout = QHBoxLayout()
        range_min = QDoubleSpinBox()
        range_max = QDoubleSpinBox()
        
        if filter_type == "protein":
            range_min.setRange(0.0, 999999.99)
            range_min.setSingleStep(0.01)
            range_min.setDecimals(2)
            range_max.setRange(0.0, 999999.99)
            range_max.setSingleStep(0.01)
            range_max.setDecimals(2)
            
            # ä½¿ç”¨å®é™…æ•°æ®èŒƒå›´
            if hasattr(self, 'current_data_ranges') and 'protein_pct' in self.current_data_ranges:
                field_range = self.current_data_ranges['protein_pct']
                actual_min = field_range['min']
                actual_max = field_range['max']
                range_min.setValue(actual_min)
                range_max.setValue(actual_max)
                print(f"è›‹ç™½ç‡ç­›é€‰å®é™…æ•°æ®èŒƒå›´: {actual_min}-{actual_max}%")
            else:
                range_min.setValue(3.00)
                range_max.setValue(4.50)
            
            range_label = QLabel("è›‹ç™½ç‡èŒƒå›´(%):")
        elif filter_type == "somatic":
            range_min.setRange(0.0, 999999.99)
            range_min.setSingleStep(0.1)
            range_min.setDecimals(1)
            range_max.setRange(0.0, 999999.99)
            range_max.setSingleStep(0.1)
            range_max.setDecimals(1)
            
            # ä½¿ç”¨å®é™…æ•°æ®èŒƒå›´
            if hasattr(self, 'current_data_ranges') and 'somatic_cell_count' in self.current_data_ranges:
                field_range = self.current_data_ranges['somatic_cell_count']
                actual_min = field_range['min']
                actual_max = field_range['max']
                range_min.setValue(actual_min)
                range_max.setValue(actual_max)
                print(f"ä½“ç»†èƒæ•°ç­›é€‰å®é™…æ•°æ®èŒƒå›´: {actual_min}-{actual_max}ä¸‡/ml")
            else:
                range_min.setValue(0.0)
                range_max.setValue(50.0)
            
            range_label = QLabel("ä½“ç»†èƒæ•°èŒƒå›´(ä¸‡/ml):")
        
        range_min.setStyleSheet(form_styles)
        range_max.setStyleSheet(form_styles)
        range_label.setStyleSheet("color: #495057; font-weight: bold;")
        
        range_layout.addWidget(range_label)
        range_layout.addWidget(range_min)
        dash_label = QLabel("â€”")
        dash_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        dash_label.setStyleSheet(f"color: #6c757d; margin: 0 {dash_margin}px; font-size: {dash_font_size}px;")
        range_layout.addWidget(dash_label)
        range_layout.addWidget(range_max)
        range_layout.addStretch()
        
        range_widget = QWidget()
        range_widget.setLayout(range_layout)
        layout.addWidget(range_widget)
        
        # æœ€å°‘ç¬¦åˆæœˆæ•°
        months_layout = QHBoxLayout()
        months_spinbox = QSpinBox()
        
        # ä½¿ç”¨æ™ºèƒ½æœˆæ•°ä¸Šé™
        max_months = 12  # é»˜è®¤å€¼
        if hasattr(self, 'current_data_ranges'):
            max_months = self.current_data_ranges.get('months', {}).get('max', 12)
        
        months_spinbox.setRange(0, max_months)
        default_months = min(3, max_months // 2) if max_months > 0 else 1
        months_spinbox.setValue(default_months)
        months_spinbox.setStyleSheet(form_styles)
        months_label = QLabel(f"{title}æœ€å°‘ç¬¦åˆæœˆæ•° (0-{max_months}):")
        months_label.setStyleSheet("color: #495057; font-weight: bold;")
        
        months_layout.addWidget(months_label)
        months_layout.addWidget(months_spinbox)
        months_layout.addStretch()
        
        months_widget = QWidget()
        months_widget.setLayout(months_layout)
        layout.addWidget(months_widget)
        
        # ç©ºå€¼å¤„ç†ç­–ç•¥é€‰æ‹©
        empty_layout = QHBoxLayout()
        empty_label = QLabel(f"{title}ç©ºå€¼å¤„ç†ç­–ç•¥:")
        empty_label.setStyleSheet("color: #495057; font-weight: bold;")
        
        empty_combo = QComboBox()
        empty_combo.addItems(["è§†ä¸ºä¸ç¬¦åˆ", "è§†ä¸ºç¬¦åˆ", "å†å²æ•°æ®å¡«å……"])
        empty_combo.setCurrentText("è§†ä¸ºä¸ç¬¦åˆ")  # é»˜è®¤é€‰æ‹©
        empty_combo.setStyleSheet(form_styles)
        empty_combo.setToolTip(f"é€‰æ‹©{title}æ•°æ®ä¸ºç©ºæ—¶çš„å¤„ç†æ–¹å¼")
        
        empty_layout.addWidget(empty_label)
        empty_layout.addWidget(empty_combo)
        empty_layout.addStretch()
        
        empty_widget = QWidget()
        empty_widget.setLayout(empty_layout)
        layout.addWidget(empty_widget)
        
        # æ§åˆ¶ç»„ä»¶å¯ç”¨çŠ¶æ€
        def toggle_filter_controls():
            enabled = filter_enabled.isChecked()
            range_widget.setEnabled(enabled)
            months_widget.setEnabled(enabled)
            empty_widget.setEnabled(enabled)
        
        filter_enabled.toggled.connect(toggle_filter_controls)
        toggle_filter_controls()  # åˆå§‹åŒ–çŠ¶æ€
        
        # å­˜å‚¨æ§ä»¶å¼•ç”¨ä»¥ä¾¿åç»­è®¿é—®
        if filter_type == "protein":
            self.protein_enabled = filter_enabled
            self.protein_min = range_min
            self.protein_max = range_max
            self.protein_months = months_spinbox
            self.protein_empty = empty_combo
        elif filter_type == "somatic":
            self.somatic_enabled = filter_enabled
            self.somatic_min = range_min
            self.somatic_max = range_max
            self.somatic_months = months_spinbox
            self.somatic_empty = empty_combo
        
        return group
    
    def select_active_cattle_file(self):
        """é€‰æ‹©åœ¨ç¾¤ç‰›æ–‡ä»¶"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "é€‰æ‹©åœ¨ç¾¤ç‰›æ–‡ä»¶",
            "",
            "Excelæ–‡ä»¶ (*.xlsx *.xls)"
        )
        
        if file_path:
            filename = os.path.basename(file_path)
            success, message, cattle_list = self.processor.process_active_cattle_file(file_path, filename)
            
            if success:
                cattle_count = len(cattle_list) if cattle_list else 0
                self.active_cattle_label.setText(f"å·²åŠ è½½: {filename} ({cattle_count}å¤´ç‰›)")
                self.active_cattle_label.setStyleSheet("color: #28a745; font-size: 12px; font-weight: bold;")
                self.clear_active_cattle_btn.setVisible(True)
                QMessageBox.information(self, "æˆåŠŸ", message)
            else:
                QMessageBox.critical(self, "é”™è¯¯", message)
    
    def clear_active_cattle(self):
        """æ¸…é™¤åœ¨ç¾¤ç‰›æ•°æ®"""
        reply = QMessageBox.question(
            self, 
            "ç¡®è®¤æ¸…é™¤", 
            "ç¡®å®šè¦æ¸…é™¤åœ¨ç¾¤ç‰›æ•°æ®å—ï¼Ÿ",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.processor.clear_active_cattle_data()
            self.active_cattle_label.setText("æœªä¸Šä¼ åœ¨ç¾¤ç‰›æ–‡ä»¶")
            self.active_cattle_label.setStyleSheet("color: #6c757d; font-size: 12px;")
            self.clear_active_cattle_btn.setVisible(False)
            QMessageBox.information(self, "æˆåŠŸ", "å·²æ¸…é™¤åœ¨ç¾¤ç‰›æ•°æ®")
    
    def on_filter_checkbox_toggled(self, filter_key, checked):
        """å½“ç­›é€‰é¡¹å¤é€‰æ¡†çŠ¶æ€æ”¹å˜æ—¶è§¦å‘"""
        if checked:
            # æ·»åŠ ç­›é€‰é¡¹
            if filter_key not in self.added_other_filters:
                # ä½¿ç”¨æ–°çš„é…ç½®åŠ è½½æ–¹æ³•
                filter_config = self.load_filter_config(filter_key)
                
                filter_widget = self.create_other_filter_widget(filter_key, filter_config)
                self.other_filters_layout.insertWidget(self.other_filters_layout.count() - 1, filter_widget)
                self.added_other_filters[filter_key] = filter_widget
                # åŠ¨æ€è°ƒæ•´å®¹å™¨é«˜åº¦
                self.adjust_filters_container_height()
        else:
            # ç§»é™¤ç­›é€‰é¡¹
            if filter_key in self.added_other_filters:
                widget = self.added_other_filters[filter_key]
                self.other_filters_layout.removeWidget(widget)
                widget.deleteLater()
                del self.added_other_filters[filter_key]
                # åŠ¨æ€è°ƒæ•´å®¹å™¨é«˜åº¦
                self.adjust_filters_container_height()
    
    def quick_add_common_filters(self):
        """ä¸€é”®æ·»åŠ å¸¸ç”¨ç­›é€‰é¡¹"""
        common_filters = ['fat_pct', 'milk_yield', 'lactose_pct', 'solid_pct']  # å¸¸ç”¨ç­›é€‰é¡¹
        
        for filter_key in common_filters:
            if filter_key in self.filter_checkboxes:
                checkbox = self.filter_checkboxes[filter_key]
                if not checkbox.isChecked():
                    checkbox.setChecked(True)
    
    def select_all_filters(self):
        """å…¨é€‰æ‰€æœ‰ç­›é€‰é¡¹"""
        for checkbox in self.filter_checkboxes.values():
            checkbox.setChecked(True)
    
    def clear_all_filters(self):
        """æ¸…ç©ºæ‰€æœ‰ç­›é€‰é¡¹é€‰æ‹©"""
        for checkbox in self.filter_checkboxes.values():
            checkbox.setChecked(False)
    
    def apply_selected_filters(self):
        """åº”ç”¨é€‰ä¸­çš„ç­›é€‰é¡¹"""
        selected_count = sum(1 for checkbox in self.filter_checkboxes.values() if checkbox.isChecked())
        
        if selected_count == 0:
            QMessageBox.information(self, "æç¤º", "è¯·å…ˆé€‰æ‹©è¦æ·»åŠ çš„ç­›é€‰é¡¹ç›®")
            return
        
        QMessageBox.information(
            self, 
            "åº”ç”¨æˆåŠŸ", 
            f"å·²åº”ç”¨ {selected_count} ä¸ªç­›é€‰é¡¹ç›®\n\nè¯·åœ¨ä¸‹æ–¹é…ç½®ç›¸åº”çš„ç­›é€‰æ¡ä»¶ã€‚"
        )
    
    def adjust_filters_container_height(self):
        """æ ¹æ®æ·»åŠ çš„ç­›é€‰é¡¹æ•°é‡åŠ¨æ€è°ƒæ•´å®¹å™¨é«˜åº¦"""
        if not hasattr(self, 'filters_container') or not hasattr(self, 'added_other_filters'):
            return
        
        # è®¡ç®—æ‰€éœ€é«˜åº¦ï¼šæ¯ä¸ªç­›é€‰é¡¹çº¦120pxé«˜åº¦ï¼ŒåŠ ä¸Šä¸€äº›è¾¹è·
        filter_count = len(self.added_other_filters)
        if filter_count == 0:
            # æ²¡æœ‰ç­›é€‰é¡¹æ—¶ï¼Œä¿æŒæœ€å°é«˜åº¦
            min_height = 50
            self.filters_container.setMinimumHeight(min_height)
            self.filters_container.setMaximumHeight(min_height)
        else:
            # æ ¹æ®ç­›é€‰é¡¹æ•°é‡è®¡ç®—é«˜åº¦
            item_height = 120  # æ¯ä¸ªç­›é€‰é¡¹çš„çº¦å®šé«˜åº¦
            padding = 20       # ä¸Šä¸‹è¾¹è·
            total_height = filter_count * item_height + padding
            
            # åªè®¾ç½®æœ€å°é«˜åº¦ï¼Œæœ€å¤§é«˜åº¦ä¸é™åˆ¶ï¼Œè®©å®¹å™¨è‡ªåŠ¨æ‰©å±•
            self.filters_container.setMinimumHeight(total_height)
            self.filters_container.setMaximumHeight(16777215)  # è®¾ç½®ä¸ºæœ€å¤§å€¼ï¼Œä¸é™åˆ¶é«˜åº¦
        
        # å¼ºåˆ¶é‡æ–°å¸ƒå±€
        self.filters_container.updateGeometry()
        if hasattr(self, 'filters_container') and self.filters_container.parent():
            self.filters_container.parent().updateGeometry()
    
    def add_other_filter(self, text):
        """æ·»åŠ å…¶ä»–ç­›é€‰é¡¹ï¼ˆä¿ç•™å…¼å®¹æ€§ï¼‰"""
        # è¿™ä¸ªæ–¹æ³•ç°åœ¨ä¸»è¦ç”¨äºä¿æŒå…¼å®¹æ€§
        pass
    
    def create_other_filter_widget(self, filter_key: str, filter_config: Dict):
        """åˆ›å»ºå…¶ä»–ç­›é€‰é¡¹çš„ç•Œé¢ç»„ä»¶"""
        chinese_name = filter_config.get("chinese_name", filter_key)
        
        # è·å–form_styles
        form_styles = self.get_responsive_form_styles()
        
        # ä¸»å®¹å™¨
        widget = QWidget()
        widget.setMinimumWidth(550)  # å¢åŠ æœ€å°å®½åº¦ç¡®ä¿å†…å®¹æ˜¾ç¤ºå®Œæ•´
        widget.setMinimumHeight(140)  # è®¾ç½®æœ€å°é«˜åº¦ç¡®ä¿è¶³å¤Ÿç©ºé—´
        widget.setStyleSheet("""
            QWidget {
                border: 1px solid #dee2e6;
                border-radius: 6px;
                padding: 10px;
                margin: 4px;
                background-color: #f8f9fa;
            }
        """)
        layout = QVBoxLayout(widget)
        
        # æ ‡é¢˜è¡Œï¼ˆåŒ…å«åˆ é™¤æŒ‰é’®ï¼‰
        title_layout = QHBoxLayout()
        
        # å¯ç”¨å¤é€‰æ¡†
        enabled_checkbox = QCheckBox(f"å¯ç”¨{chinese_name}ç­›é€‰")
        enabled_checkbox.setChecked(False)
        
        # è·å–å±å¹•DPIä¿¡æ¯ç”¨äºè‡ªé€‚åº”æ ·å¼
        screen_info = self.get_safe_screen_info()
        dpi_ratio = screen_info['dpi_ratio']
        checkbox_font_size = max(int(13 * dpi_ratio * 0.7), 12)
        checkbox_spacing = max(int(8 * dpi_ratio * 0.6), 6)
        checkbox_size = max(int(16 * dpi_ratio * 0.6), 14)
        checkbox_border_radius = max(int(3 * dpi_ratio * 0.6), 2)
        
        # è®¾ç½®å®Œæ•´çš„å¤é€‰æ¡†æ ·å¼ï¼ŒåŒ…å«hoverçŠ¶æ€
        enabled_checkbox.setStyleSheet(f"""
            QCheckBox {{
                font-size: {checkbox_font_size}px;
                color: #495057;
                spacing: {checkbox_spacing}px;
                font-weight: bold;
            }}
            QCheckBox::indicator {{
                width: {checkbox_size}px;
                height: {checkbox_size}px;
                border: 2px solid #ced4da;
                border-radius: {checkbox_border_radius}px;
                background-color: white;
            }}
            QCheckBox::indicator:hover {{
                border-color: #80bdff;
                background-color: #f8f9fa;
            }}
            QCheckBox::indicator:checked {{
                background-color: #28a745 !important;
                border-color: #28a745 !important;
                image: url(data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iMTIiIGhlaWdodD0iOSIgdmlld0JveD0iMCAwIDEyIDkiIGZpbGw9Im5vbmUiIHhtbG5zPSJodHRwOi8vd3d3LnczLm9yZy8yMDAwL3N2ZyI+CjxwYXRoIGQ9Ik0xIDQuNUw0LjUgOEwxMSAxIiBzdHJva2U9IndoaXRlIiBzdHJva2Utd2lkdGg9IjIiIHN0cm9rZS1saW5lY2FwPSJyb3VuZCIgc3Ryb2tlLWxpbmVqb2luPSJyb3VuZCIvPgo8L3N2Zz4K);
            }}
            QCheckBox::indicator:checked:hover {{
                background-color: #218838 !important;
                border-color: #218838 !important;
            }}
            QCheckBox::indicator:checked:pressed {{
                background-color: #1e7e34 !important;
                border-color: #1e7e34 !important;
            }}
        """)
        
        title_layout.addWidget(enabled_checkbox)
        
        title_layout.addStretch()
        
        # åˆ é™¤æŒ‰é’®
        remove_btn = QPushButton("âŒ")
        remove_btn.setFixedSize(24, 24)
        remove_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                border-radius: 12px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
        """)
        remove_btn.clicked.connect(lambda: self.remove_other_filter(filter_key))
        title_layout.addWidget(remove_btn)
        
        layout.addLayout(title_layout)
        
        # ç­›é€‰èŒƒå›´
        range_layout = QHBoxLayout()
        
        range_min = QDoubleSpinBox()
        range_min.setRange(-999999.99, 999999.99)
        range_min.setDecimals(2)
        
        range_max = QDoubleSpinBox()
        range_max.setRange(-999999.99, 999999.99)
        range_max.setDecimals(2)
        
        # ä½¿ç”¨å®é™…æ•°æ®èŒƒå›´ - å¦‚æœæœ‰æ•°æ®èŒƒå›´ï¼Œä½¿ç”¨å®é™…çš„æœ€å°å€¼å’Œæœ€å¤§å€¼
        if hasattr(self, 'current_data_ranges') and filter_key in self.current_data_ranges:
            field_range = self.current_data_ranges[filter_key]
            actual_min = field_range['min']
            actual_max = field_range['max']
            range_min.setValue(actual_min)
            range_max.setValue(actual_max)
            print(f"ä¸º{chinese_name}è®¾ç½®å®é™…æ•°æ®èŒƒå›´: {actual_min}-{actual_max}")
        else:
            # ä½¿ç”¨é…ç½®ä¸­çš„é»˜è®¤å€¼
            range_min.setValue(filter_config.get("min", 0.0))
            range_max.setValue(filter_config.get("max", 100.0))
        
        range_layout.addWidget(QLabel(f"{chinese_name}èŒƒå›´:"))
        range_layout.addWidget(range_min)
        range_layout.addWidget(QLabel("â€”"))
        range_layout.addWidget(range_max)
        range_layout.addStretch()
        
        layout.addLayout(range_layout)
        
        # æœ€å°‘ç¬¦åˆæœˆæ•°å’Œç©ºå€¼å¤„ç†
        options_layout = QHBoxLayout()
        
        months_spinbox = QSpinBox()
        # ä½¿ç”¨æ™ºèƒ½æœˆæ•°ä¸Šé™
        max_months = 12  # é»˜è®¤å€¼
        if hasattr(self, 'current_data_ranges'):
            max_months = self.current_data_ranges.get('months', {}).get('max', 12)
        
        months_spinbox.setRange(0, max_months)
        default_months = min(3, max_months // 2) if max_months > 0 else 1
        months_spinbox.setValue(filter_config.get("min_match_months", default_months))
        
        # ç©ºå€¼å¤„ç†ç­–ç•¥ä¸‹æ‹‰é€‰æ‹©
        empty_combo = QComboBox()
        empty_combo.addItems(["è§†ä¸ºä¸ç¬¦åˆ", "è§†ä¸ºç¬¦åˆ", "å†å²æ•°æ®å¡«å……"])
        # æ ¹æ®é…ç½®è®¾ç½®é»˜è®¤å€¼
        if filter_config.get("treat_empty_as_match", False):
            empty_combo.setCurrentText("è§†ä¸ºç¬¦åˆ")
        else:
            empty_combo.setCurrentText("è§†ä¸ºä¸ç¬¦åˆ")
        empty_combo.setStyleSheet(form_styles)
        empty_combo.setToolTip("é€‰æ‹©æ•°æ®ä¸ºç©ºæ—¶çš„å¤„ç†æ–¹å¼")
        
        options_layout.addWidget(QLabel("æœ€å°‘ç¬¦åˆæœˆæ•°:"))
        options_layout.addWidget(months_spinbox)
        options_layout.addWidget(QLabel("ç©ºå€¼å¤„ç†:"))
        options_layout.addWidget(empty_combo)
        options_layout.addStretch()
        
        layout.addLayout(options_layout)
        
        # æ§åˆ¶å¯ç”¨çŠ¶æ€
        def toggle_controls():
            enabled = enabled_checkbox.isChecked()
            for i in range(range_layout.count()):
                item = range_layout.itemAt(i)
                if item.widget():
                    item.widget().setEnabled(enabled)
            for i in range(options_layout.count()):
                item = options_layout.itemAt(i)
                if item.widget():
                    item.widget().setEnabled(enabled)
        
        enabled_checkbox.toggled.connect(toggle_controls)
        toggle_controls()
        
        # å­˜å‚¨æ§ä»¶å¼•ç”¨
        widget.enabled_checkbox = enabled_checkbox
        widget.range_min = range_min
        widget.range_max = range_max
        widget.months_spinbox = months_spinbox
        widget.empty_combo = empty_combo
        widget.filter_key = filter_key
        widget.chinese_name = chinese_name
        
        return widget
    
    def remove_other_filter(self, filter_key: str):
        """ç§»é™¤å…¶ä»–ç­›é€‰é¡¹"""
        if filter_key in self.added_other_filters:
            widget = self.added_other_filters[filter_key]
            self.other_filters_layout.removeWidget(widget)
            widget.deleteLater()
            del self.added_other_filters[filter_key]
            
            # åŒæ­¥æ›´æ–°å¤é€‰æ¡†çŠ¶æ€
            if filter_key in self.filter_checkboxes:
                checkbox = self.filter_checkboxes[filter_key]
                checkbox.blockSignals(True)  # é˜»æ­¢ä¿¡å·é¿å…é€’å½’è°ƒç”¨
                checkbox.setChecked(False)
                checkbox.blockSignals(False)
            
            # åŠ¨æ€è°ƒæ•´å®¹å™¨é«˜åº¦
            self.adjust_filters_container_height()
    
    
    def _check_filter_thread_stopped(self):
        """æ£€æŸ¥ç­›é€‰çº¿ç¨‹æ˜¯å¦å·²åœæ­¢"""
        if hasattr(self, 'filter_thread') and self.filter_thread.isRunning():
            # çº¿ç¨‹è¿˜åœ¨è¿è¡Œï¼Œç»§ç»­ç­‰å¾…
            QTimer.singleShot(100, self._check_filter_thread_stopped)
        else:
            # çº¿ç¨‹å·²åœæ­¢ï¼Œé‡ç½®UIçŠ¶æ€
            self._reset_filter_ui_state()
            self.statusBar().showMessage("ç­›é€‰å·²è¢«ç”¨æˆ·å–æ¶ˆ")
    
    def _reset_filter_ui_state(self):
        """é‡ç½®ç­›é€‰UIçŠ¶æ€"""
        if hasattr(self, 'filter_progress'):
            self.filter_progress.setVisible(False)
        if hasattr(self, 'filter_btn'):
            self.filter_btn.setEnabled(True)
            self.filter_btn.setVisible(True)
        if hasattr(self, 'cancel_filter_btn'):
            self.cancel_filter_btn.setEnabled(False)
            self.cancel_filter_btn.setVisible(False)
    
    def _get_enabled_traits(self):
        """è·å–å½“å‰å¯ç”¨çš„æ€§çŠ¶åˆ—è¡¨"""
        enabled_traits = []
        
        # æ£€æŸ¥è›‹ç™½ç‡ç­›é€‰
        if hasattr(self, 'protein_enabled') and self.protein_enabled.isChecked():
            enabled_traits.append('protein_pct')
        
        # æ£€æŸ¥ä½“ç»†èƒæ•°ç­›é€‰
        if hasattr(self, 'somatic_enabled') and self.somatic_enabled.isChecked():
            enabled_traits.append('somatic_cell_count')
        
        # æ£€æŸ¥å…¶ä»–ç­›é€‰é¡¹
        for filter_key, widget in self.added_other_filters.items():
            if widget.enabled_checkbox.isChecked():
                enabled_traits.append(filter_key)
        
        return enabled_traits
    
    def update_trait_statistics(self, df, trait, widget):
        """æ›´æ–°ç‰¹å®šæ€§çŠ¶çš„ç»Ÿè®¡ä¿¡æ¯"""
        if df.empty:
            widget.setText(f"æ€§çŠ¶åˆ†æ\n\næš‚æ— ç­›é€‰ç»“æœ")
            return
        
        # æ€§çŠ¶ä¸­æ–‡åç§°æ˜ å°„
        trait_names = {
            'protein_pct': 'ğŸ¥› è›‹ç™½ç‡',
            'somatic_cell_count': 'ğŸ”¬ ä½“ç»†èƒæ•°',
            'fat_pct': 'ğŸ§ˆ ä¹³è„‚ç‡',
            'lactose_pct': 'ğŸ¬ ä¹³ç³–ç‡',
            'milk_yield': 'ğŸ¥› äº§å¥¶é‡',
            'lactation_days': 'ğŸ“… æ³Œä¹³å¤©æ•°',
            'solids_pct': 'ğŸ§ª å›ºå½¢ç‰©',
            'fat_protein_ratio': 'âš–ï¸ è„‚è›‹æ¯”',
            'urea_nitrogen': 'ğŸ§¬ å°¿ç´ æ°®',
            'total_fat_pct': 'ğŸ§ˆ æ€»ä¹³è„‚',
            'total_protein_pct': 'ğŸ¥› æ€»è›‹ç™½',
            'mature_equivalent': 'ğŸ„ æˆå¹´å½“é‡'
        }
        
        trait_name = trait_names.get(trait, trait)
        stats_text = f"{trait_name}åˆ†æ\n\n"
        
        # æŸ¥æ‰¾å¯¹åº”çš„æœˆåº¦æ˜ç»†åˆ—
        monthly_columns = []
        for col in df.columns:
            if trait == 'protein_pct' and 'è›‹ç™½ç‡(%)' in col and 'å¹´' in col and 'æœˆ' in col:
                monthly_columns.append(col)
            elif trait == 'somatic_cell_count' and 'ä½“ç»†èƒæ•°(ä¸‡/ml)' in col and 'å¹´' in col and 'æœˆ' in col:
                monthly_columns.append(col)
            elif trait == 'fat_pct' and 'ä¹³è„‚ç‡(%)' in col and 'å¹´' in col and 'æœˆ' in col:
                monthly_columns.append(col)
            elif trait == 'lactose_pct' and 'ä¹³ç³–ç‡(%)' in col and 'å¹´' in col and 'æœˆ' in col:
                monthly_columns.append(col)
            elif trait == 'milk_yield' and 'äº§å¥¶é‡(Kg)' in col and 'å¹´' in col and 'æœˆ' in col:
                monthly_columns.append(col)
            elif trait == 'lactation_days' and 'æ³Œä¹³å¤©æ•°(å¤©)' in col and 'å¹´' in col and 'æœˆ' in col:
                monthly_columns.append(col)
            elif trait == 'solids_pct' and 'å›ºå½¢ç‰©(%)' in col and 'å¹´' in col and 'æœˆ' in col:
                monthly_columns.append(col)
            elif trait == 'fat_protein_ratio' and 'è„‚è›‹æ¯”' in col and 'å¹´' in col and 'æœˆ' in col:
                monthly_columns.append(col)
            elif trait == 'urea_nitrogen' and 'å°¿ç´ æ°®(mg/dl)' in col and 'å¹´' in col and 'æœˆ' in col:
                monthly_columns.append(col)
            elif trait == 'total_fat_pct' and 'æ€»ä¹³è„‚(%)' in col and 'å¹´' in col and 'æœˆ' in col:
                monthly_columns.append(col)
            elif trait == 'total_protein_pct' and 'æ€»è›‹ç™½(%)' in col and 'å¹´' in col and 'æœˆ' in col:
                monthly_columns.append(col)
            elif trait == 'mature_equivalent' and 'æˆå¹´å½“é‡(Kg)' in col and 'å¹´' in col and 'æœˆ' in col:
                monthly_columns.append(col)
        
        if not monthly_columns:
            stats_text += f"ğŸ“Š ç­›é€‰ç»“æœ: {len(df)} æ¡è®°å½•\n"
            stats_text += f"âŒ æœªæ‰¾åˆ° {trait_name} çš„æœˆåº¦æ•°æ®åˆ—\n"
            stats_text += "è¯·ç¡®è®¤è¯¥æ€§çŠ¶å·²åŒ…å«åœ¨ç­›é€‰ç»“æœä¸­ã€‚"
        else:
            # æŒ‰æ—¶é—´æ’åºæœˆåº¦åˆ—
            monthly_columns.sort()
            
            # åŸºç¡€ç»Ÿè®¡
            stats_text += f"ğŸ“Š ç­›é€‰ç»“æœ: {len(df)} æ¡è®°å½•\n"
            stats_text += f"ğŸ“… æœˆåº¦æ•°æ®: {len(monthly_columns)} ä¸ªæœˆ\n\n"
            
            # å„æœˆç»Ÿè®¡
            stats_text += f"ğŸ“ˆ å„æœˆ {trait_name} ç»Ÿè®¡:\n"
            for col in monthly_columns:
                month_data = df[col].dropna()
                if len(month_data) > 0:
                    avg_val = month_data.mean()
                    min_val = month_data.min()
                    max_val = month_data.max()
                    count = len(month_data)
                    
                    # æ ¹æ®æ€§çŠ¶ç±»å‹æ ¼å¼åŒ–æ•°å€¼
                    if trait in ['protein_pct', 'fat_pct', 'lactose_pct', 'solids_pct', 'total_fat_pct', 'total_protein_pct']:
                        stats_text += f"  {col}: å¹³å‡ {avg_val:.2f}%ï¼ŒèŒƒå›´ {min_val:.2f}%-{max_val:.2f}%ï¼Œ{count}å¤´ç‰›\n"
                    elif trait in ['milk_yield', 'mature_equivalent']:
                        stats_text += f"  {col}: å¹³å‡ {avg_val:.1f}Kgï¼ŒèŒƒå›´ {min_val:.1f}-{max_val:.1f}Kgï¼Œ{count}å¤´ç‰›\n"
                    elif trait in ['lactation_days']:
                        stats_text += f"  {col}: å¹³å‡ {avg_val:.1f}å¤©ï¼ŒèŒƒå›´ {min_val:.0f}-{max_val:.0f}å¤©ï¼Œ{count}å¤´ç‰›\n"
                    elif trait in ['somatic_cell_count']:
                        stats_text += f"  {col}: å¹³å‡ {avg_val:.1f}ä¸‡/mlï¼ŒèŒƒå›´ {min_val:.1f}-{max_val:.1f}ä¸‡/mlï¼Œ{count}å¤´ç‰›\n"
                    elif trait in ['urea_nitrogen']:
                        stats_text += f"  {col}: å¹³å‡ {avg_val:.1f}mg/dlï¼ŒèŒƒå›´ {min_val:.1f}-{max_val:.1f}mg/dlï¼Œ{count}å¤´ç‰›\n"
                    elif trait in ['fat_protein_ratio']:
                        stats_text += f"  {col}: å¹³å‡ {avg_val:.2f}ï¼ŒèŒƒå›´ {min_val:.2f}-{max_val:.2f}ï¼Œ{count}å¤´ç‰›\n"
                    else:
                        stats_text += f"  {col}: å¹³å‡ {avg_val:.2f}ï¼ŒèŒƒå›´ {min_val:.2f}-{max_val:.2f}ï¼Œ{count}å¤´ç‰›\n"
                else:
                    stats_text += f"  {col}: æ— æœ‰æ•ˆæ•°æ®\n"
            
            # æ•´ä½“ç»Ÿè®¡
            all_monthly_data = []
            for col in monthly_columns:
                month_data = df[col].dropna()
                all_monthly_data.extend(month_data.tolist())
            
            if all_monthly_data:
                overall_avg = sum(all_monthly_data) / len(all_monthly_data)
                overall_min = min(all_monthly_data)
                overall_max = max(all_monthly_data)
                
                stats_text += f"\nğŸ¯ æ•´ä½“ {trait_name} ç»Ÿè®¡:\n"
                if trait in ['protein_pct', 'fat_pct', 'lactose_pct', 'solids_pct', 'total_fat_pct', 'total_protein_pct']:
                    stats_text += f"  æ€»ä½“å¹³å‡: {overall_avg:.2f}%\n"
                    stats_text += f"  æ€»ä½“èŒƒå›´: {overall_min:.2f}% - {overall_max:.2f}%\n"
                elif trait in ['milk_yield', 'mature_equivalent']:
                    stats_text += f"  æ€»ä½“å¹³å‡: {overall_avg:.1f}Kg\n"
                    stats_text += f"  æ€»ä½“èŒƒå›´: {overall_min:.1f}Kg - {overall_max:.1f}Kg\n"
                elif trait in ['lactation_days']:
                    stats_text += f"  æ€»ä½“å¹³å‡: {overall_avg:.1f}å¤©\n"
                    stats_text += f"  æ€»ä½“èŒƒå›´: {overall_min:.0f}å¤© - {overall_max:.0f}å¤©\n"
                elif trait in ['somatic_cell_count']:
                    stats_text += f"  æ€»ä½“å¹³å‡: {overall_avg:.1f}ä¸‡/ml\n"
                    stats_text += f"  æ€»ä½“èŒƒå›´: {overall_min:.1f}ä¸‡/ml - {overall_max:.1f}ä¸‡/ml\n"
                elif trait in ['urea_nitrogen']:
                    stats_text += f"  æ€»ä½“å¹³å‡: {overall_avg:.1f}mg/dl\n"
                    stats_text += f"  æ€»ä½“èŒƒå›´: {overall_min:.1f}mg/dl - {overall_max:.1f}mg/dl\n"
                elif trait in ['fat_protein_ratio']:
                    stats_text += f"  æ€»ä½“å¹³å‡: {overall_avg:.2f}\n"
                    stats_text += f"  æ€»ä½“èŒƒå›´: {overall_min:.2f} - {overall_max:.2f}\n"
                else:
                    stats_text += f"  æ€»ä½“å¹³å‡: {overall_avg:.2f}\n"
                    stats_text += f"  æ€»ä½“èŒƒå›´: {overall_min:.2f} - {overall_max:.2f}\n"
                
                stats_text += f"  æœ‰æ•ˆæ•°æ®ç‚¹: {len(all_monthly_data)} ä¸ª\n"
        
        widget.setText(stats_text)
    
    # ==================== æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥ç›¸å…³æ–¹æ³• ====================
    
    def on_mastitis_system_selected(self, system_type: str, checked: bool):
        """ç³»ç»Ÿç±»å‹é€‰æ‹©äº‹ä»¶å¤„ç†"""
        if not checked:
            return
        
        # ç¡®ä¿åªèƒ½é€‰æ‹©ä¸€ä¸ªç³»ç»Ÿï¼ˆå•é€‰æ¨¡å¼ï¼‰
        if system_type == 'yiqiniu':
            self.huimuyun_radio.setChecked(False)
            self.custom_radio.setChecked(False)
        elif system_type == 'huimuyun':
            self.yiqiniu_radio.setChecked(False)
            self.custom_radio.setChecked(False)
        elif system_type == 'custom':
            self.yiqiniu_radio.setChecked(False)
            self.huimuyun_radio.setChecked(False)
        
        self.current_mastitis_system = system_type
        self.create_mastitis_file_upload_widgets(system_type)
        self.mastitis_upload_group.setVisible(True)
        self.update_mastitis_screen_button_state()
    
    def create_mastitis_file_upload_widgets(self, system_type: str):
        """æ ¹æ®ç³»ç»Ÿç±»å‹åˆ›å»ºæ–‡ä»¶ä¸Šä¼ æ§ä»¶"""
        # æ¸…é™¤ç°æœ‰æ§ä»¶
        for widget in self.mastitis_file_uploads.values():
            widget.setParent(None)
        self.mastitis_file_uploads.clear()
        
        # æ ¹æ®ç³»ç»Ÿç±»å‹åˆ›å»ºå¯¹åº”çš„ä¸Šä¼ æ§ä»¶å’Œå­—æ®µæ˜ å°„
        if system_type == 'yiqiniu':
            files_config = {
                'cattle_info': {
                    'name': 'ç‰›ç¾¤åŸºç¡€ä¿¡æ¯è¡¨',
                    'fields': {
                        'è€³å·': 'è€³å·ï¼ˆå»æ‰æœ€å‰é¢çš„"0"ï¼‰',
                        'èƒæ¬¡': 'èƒæ¬¡ï¼ˆå»æ‰æœ€å‰é¢çš„"0"ï¼‰', 
                        'æ³Œä¹³å¤©æ•°': 'æ³Œä¹³å¤©æ•°',
                        'ç¹è‚²çŠ¶æ€': 'ç¹è‚²çŠ¶æ€',
                        'åœ¨èƒå¤©æ•°': 'åœ¨èƒå¤©æ•°',
                        'æœ€è¿‘äº§çŠŠæ—¥æœŸ': 'æœ€è¿‘äº§çŠŠæ—¥æœŸ'
                    }
                },
                'milk_yield': {
                    'name': 'å¥¶ç‰›äº§å¥¶æ—¥æ±‡æ€»è¡¨',
                    'fields': {
                        'è€³å·': 'è€³å·ï¼ˆå»æ‰æœ€å‰é¢çš„"0"ï¼‰',
                        'æŒ¤å¥¶æ—¥æœŸ': 'æŒ¤å¥¶æ—¥æœŸ',
                        'æ—¥äº§é‡(kg)': 'æ—¥äº§é‡ï¼ˆkgï¼‰'
                    }
                },
                'disease': {
                    'name': 'å‘ç—…æŸ¥è¯¢å¯¼å‡ºè¡¨',
                    'fields': {
                        'è€³å·': 'è€³å·ï¼ˆå»æ‰æœ€å‰é¢çš„"0"ï¼‰',
                        'å‘ç—…æ—¥æœŸ': 'å‘ç—…æ—¥æœŸ',
                        'ç–¾ç—…ç§ç±»': 'ç–¾ç—…ç§ç±»'
                    }
                }
            }
        elif system_type == 'huimuyun':
            files_config = {
                'cattle_info': {
                    'name': 'ç‰›ç¾¤åŸºç¡€ä¿¡æ¯è¡¨',
                    'fields': {
                        'è€³å·': 'è€³å·ï¼ˆå»æ‰æœ€å‰é¢çš„"0"ï¼‰',
                        'èƒæ¬¡': 'èƒæ¬¡ï¼ˆå»æ‰æœ€å‰é¢çš„"0"ï¼‰',
                        'æ³Œä¹³å¤©æ•°': 'æ³Œä¹³å¤©æ•°',
                        'ç¹è‚²çŠ¶æ€': 'ç¹è‚²çŠ¶æ€',
                        'åœ¨èƒå¤©æ•°': 'æ€€å­•å¤©æ•°',
                        'æœ€è¿‘äº§çŠŠæ—¥æœŸ': 'äº§çŠŠæ—¥æœŸ',
                        'æœ€è¿‘ä¸ƒå¤©å¥¶å…å¹³å‡å¥¶é‡': 'æœ€è¿‘ä¸ƒå¤©å¥¶å…å¹³å‡å¥¶é‡'
                    }
                },
                'disease': {
                    'name': 'å‘ç—…äº‹ä»¶ç®¡ç†è¡¨',
                    'fields': {
                        'è€³å·': 'è€³å·ï¼ˆå»æ‰æœ€å‰é¢çš„"0"ï¼‰',
                        'äº‹ä»¶æ—¥æœŸ': 'äº‹ä»¶æ—¥æœŸ',
                        'äº‹ä»¶ç±»å‹': 'äº‹ä»¶ç±»å‹'
                    }
                }
            }
        elif system_type == 'custom':
            files_config = {
                'cattle_info': {
                    'name': 'ç‰›ç¾¤åŸºç¡€ä¿¡æ¯è¡¨',
                    'fields': {
                        'è€³å·': 'è€³å·',
                        'èƒæ¬¡': 'èƒæ¬¡',
                        'æ³Œä¹³å¤©æ•°': 'æ³Œä¹³å¤©æ•°',
                        'ç¹è‚²çŠ¶æ€': 'ç¹è‚²çŠ¶æ€',
                        'åœ¨èƒå¤©æ•°': 'åœ¨èƒå¤©æ•°',
                        'æœ€è¿‘äº§çŠŠæ—¥æœŸ': 'æœ€è¿‘äº§çŠŠæ—¥æœŸ',
                        'æœ€è¿‘ä¸ƒå¤©å¥¶å…å¹³å‡å¥¶é‡': 'æœ€è¿‘ä¸ƒå¤©å¥¶å…å¹³å‡å¥¶é‡'
                    },
                    'custom': True
                },
                'disease': {
                    'name': 'å‘ç—…æŸ¥è¯¢å¯¼å‡ºè¡¨',
                    'fields': {
                        'è€³å·': 'è€³å·',
                        'å‘ç—…æ—¥æœŸ': 'äº‹ä»¶æ—¥æœŸ',
                        'ç–¾ç—…ç§ç±»': 'äº‹ä»¶ç±»å‹'
                    },
                    'custom': True
                }
            }
        else:
            return
        
        # åˆ›å»ºæ–‡ä»¶ä¸Šä¼ æ§ä»¶
        for file_key, file_config in files_config.items():
            file_widget = self.create_mastitis_file_upload_widget(file_key, file_config, system_type)
            self.mastitis_file_uploads[file_key] = file_widget
            self.mastitis_upload_layout.addWidget(file_widget)
    
    def create_mastitis_file_upload_widget(self, file_key: str, file_config: dict, system_type: str):
        """åˆ›å»ºå•ä¸ªæ–‡ä»¶ä¸Šä¼ æ§ä»¶"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(15, 15, 15, 15)
        
        # æ–‡ä»¶åæ ‡ç­¾
        name_label = QLabel(f"ğŸ“„ {file_config['name']}")
        name_label.setStyleSheet("font-weight: bold; font-size: 14px; color: black; background-color: white;")
        layout.addWidget(name_label)
        
        # å­—æ®µæ˜ å°„æ˜¾ç¤º
        if file_config.get('custom', False):
            # è‡ªå®šä¹‰ç³»ç»Ÿï¼Œæ˜¾ç¤ºå¯ç¼–è¾‘çš„å­—æ®µæ˜ å°„
            mapping_label = QLabel("å­—æ®µæ˜ å°„ï¼ˆæ‰€éœ€æ•°æ® â†’ è¡¨å¤´åˆ—åï¼‰ï¼š")
            mapping_label.setStyleSheet("font-weight: bold; color: black; margin-top: 8px; background-color: white;")
            layout.addWidget(mapping_label)
            
            # åˆ›å»ºå­—æ®µæ˜ å°„ç¼–è¾‘åŒºåŸŸ
            mapping_widget = QWidget()
            mapping_layout = QVBoxLayout(mapping_widget)
            mapping_layout.setContentsMargins(0, 0, 0, 0)
            
            # å­˜å‚¨æ˜ å°„è¾“å…¥æ§ä»¶
            mapping_inputs = {}
            
            for required_field, default_value in file_config['fields'].items():
                field_layout = QHBoxLayout()
                
                # æ‰€éœ€æ•°æ®æ ‡ç­¾
                field_label = QLabel(f"{required_field}:")
                field_label.setFixedWidth(120)
                field_label.setStyleSheet("font-weight: bold; color: black; background-color: white;")
                field_layout.addWidget(field_label)
                
                # ç®­å¤´
                arrow_label = QLabel("â†’")
                arrow_label.setStyleSheet("color: black; font-size: 14px; background-color: white; font-weight: bold;")
                field_layout.addWidget(arrow_label)
                
                # è¾“å…¥æ¡†
                input_edit = QLineEdit()
                input_edit.setPlaceholderText(f"è¯·è¾“å…¥è¡¨å¤´åˆ—åï¼ˆå¦‚ï¼š{default_value}ï¼‰")
                input_edit.setStyleSheet("""
                    QLineEdit {
                        padding: 4px 8px;
                        border: 1px solid #bdc3c7;
                        border-radius: 4px;
                        background-color: white;
                        color: black;
                        font-weight: bold;
                    }
                """)
                field_layout.addWidget(input_edit)
                
                mapping_inputs[required_field] = input_edit
                mapping_layout.addLayout(field_layout)
            
            mapping_widget.setLayout(mapping_layout)
            layout.addWidget(mapping_widget)
            
            # ä¿å­˜æ˜ å°„è¾“å…¥æ§ä»¶çš„å¼•ç”¨
            widget.mapping_inputs = mapping_inputs
            
        else:
            # å›ºå®šç³»ç»Ÿï¼Œåªæ˜¾ç¤ºå­—æ®µæ˜ å°„å…³ç³»
            mapping_label = QLabel("å­—æ®µæ˜ å°„ï¼ˆæ‰€éœ€æ•°æ® â†’ è¡¨å¤´åˆ—åï¼‰ï¼š")
            mapping_label.setStyleSheet("font-weight: bold; color: black; margin-top: 8px; background-color: white;")
            layout.addWidget(mapping_label)
            
            # åˆ›å»ºå­—æ®µæ˜ å°„æ˜¾ç¤ºåŒºåŸŸ
            mapping_text = []
            for required_field, table_header in file_config['fields'].items():
                mapping_text.append(f"â€¢ {required_field} â†’ {table_header}")
            
            mapping_display = QLabel("\n".join(mapping_text))
            mapping_display.setStyleSheet("""
                QLabel {
                    color: black;
                    font-size: 12px;
                    padding: 8px;
                    background-color: #f8f9fa;
                    border: 1px solid #e9ecef;
                    border-radius: 4px;
                    margin-bottom: 8px;
                    font-weight: bold;
                }
            """)
            layout.addWidget(mapping_display)
        
        # åˆ†éš”çº¿
        separator = QWidget()
        separator.setStyleSheet("border-top: 1px solid #dee2e6; margin: 8px 0;")
        separator.setFixedHeight(1)
        layout.addWidget(separator)
        
        # æ–‡ä»¶ä¸Šä¼ è¡Œ
        upload_layout = QHBoxLayout()
        
        # æ–‡ä»¶è·¯å¾„æ˜¾ç¤º
        file_path_edit = QLineEdit()
        file_path_edit.setPlaceholderText("è¯·é€‰æ‹©æ–‡ä»¶...")
        file_path_edit.setReadOnly(True)
        file_path_edit.setStyleSheet("""
            QLineEdit {
                padding: 8px;
                border: 2px solid #bdc3c7;
                border-radius: 4px;
                background-color: white;
                color: black;
                font-weight: bold;
            }
        """)
        upload_layout.addWidget(file_path_edit)
        
        # é€‰æ‹©æ–‡ä»¶æŒ‰é’®
        select_btn = QPushButton("é€‰æ‹©æ–‡ä»¶")
        select_btn.setStyleSheet("""
            QPushButton {
                padding: 8px 16px;
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        select_btn.clicked.connect(lambda: self.select_mastitis_file(file_key, file_config['name']))
        upload_layout.addWidget(select_btn)
        
        layout.addLayout(upload_layout)
        
        # çŠ¶æ€æ ‡ç­¾
        status_label = QLabel("æœªé€‰æ‹©")
        status_label.setStyleSheet("color: black; font-size: 12px; margin-top: 4px; background-color: white; font-weight: bold;")
        layout.addWidget(status_label)
        
        # ä¿å­˜å¼•ç”¨
        widget.file_path_edit = file_path_edit
        widget.select_btn = select_btn
        widget.status_label = status_label
        widget.file_config = file_config
        widget.file_path = None
        
        # è®¾ç½®æ ·å¼
        widget.setStyleSheet("""
            QWidget {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 8px;
                margin: 4px;
            }
        """)
        
        return widget
    
    def select_mastitis_file(self, file_key: str, file_name: str):
        """é€‰æ‹©æ–‡ä»¶"""
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getOpenFileName(
            self, f"é€‰æ‹©{file_name}", "", 
            "Excelæ–‡ä»¶ (*.xlsx *.xls);;æ‰€æœ‰æ–‡ä»¶ (*)"
        )
        
        if file_path:
            # åˆ›å»ºè¿›åº¦å¯¹è¯æ¡†
            file_process_dialog = SmoothProgressDialog(
                f"å¤„ç†æ–‡ä»¶: {os.path.basename(file_path)}",
                "å–æ¶ˆ",
                0, 100,
                self
            )
            file_process_dialog.show()
            file_process_dialog.setLabelText("æ­£åœ¨å¤„ç†æ–‡ä»¶...")
            
            widget = self.mastitis_file_uploads[file_key]
            widget.file_path = file_path
            widget.file_path_edit.setText(os.path.basename(file_path))
            widget.status_label.setText(f"å·²é€‰æ‹©: {os.path.basename(file_path)}")
            widget.status_label.setStyleSheet("color: #28a745; font-size: 12px;")
            
            # æ›´æ–°è¿›åº¦ - æ–‡ä»¶é€‰æ‹©å®Œæˆ
            file_process_dialog.setValue(20)
            file_process_dialog.setLabelText("æ­£åœ¨è¯»å–æ–‡ä»¶ä¿¡æ¯...")
            QApplication.processEvents()
            
            # æ˜¾ç¤ºæ–‡ä»¶ä¿¡æ¯åˆ°å³ä¾§é¢æ¿
            self.display_mastitis_file_info(file_key, file_name, file_path)
            
            # æ›´æ–°è¿›åº¦ - æ–‡ä»¶ä¿¡æ¯è¯»å–å®Œæˆ
            file_process_dialog.setValue(50)
            
            # å¦‚æœæ˜¯ç‰›ç¾¤åŸºç¡€ä¿¡æ¯è¡¨ï¼Œç«‹å³å¤„ç†å¹¶ä¿å­˜æ•°æ®
            if file_key == 'cattle_info':
                file_process_dialog.setLabelText("æ­£åœ¨å¤„ç†ç‰›ç¾¤åŸºç¡€ä¿¡æ¯...")
                file_process_dialog.setValue(60)
                QApplication.processEvents()
                
                # ç«‹å³å¤„ç†ç‰›ç¾¤åŸºç¡€ä¿¡æ¯è¡¨
                success = self.process_and_save_cattle_basic_info(file_path)
                
                if success:
                    file_process_dialog.setLabelText("æ­£åœ¨æå–ç¹è‚²çŠ¶æ€...")
                    file_process_dialog.setValue(80)
                    QApplication.processEvents()
                    self.extract_and_update_breeding_status(file_path)
                    file_process_dialog.setValue(100)
                    file_process_dialog.setLabelText("ç‰›ç¾¤åŸºç¡€ä¿¡æ¯å¤„ç†å®Œæˆ")
                else:
                    file_process_dialog.setValue(100)
                    file_process_dialog.setLabelText("ç‰›ç¾¤åŸºç¡€ä¿¡æ¯å¤„ç†å¤±è´¥")
            else:
                file_process_dialog.setValue(100)
                file_process_dialog.setLabelText("æ–‡ä»¶å¤„ç†å®Œæˆ")
            
            # å»¶è¿Ÿå…³é—­è¿›åº¦å¯¹è¯æ¡†
            QTimer.singleShot(2000, lambda: file_process_dialog.close())
            
            self.update_mastitis_screen_button_state()
    
    

    def display_mastitis_file_info(self, file_key: str, file_name: str, file_path: str):
        """æ˜¾ç¤ºæ…¢æ€§ä¹³æˆ¿ç‚æ–‡ä»¶ä¿¡æ¯åˆ°å³ä¾§é¢æ¿"""
        try:
            import os
            import pandas as pd
            from datetime import datetime
            
            # åˆ›å»ºè¿›åº¦å¯¹è¯æ¡†
            progress_dialog = SmoothProgressDialog(
                f"æ­£åœ¨è¯»å–æ–‡ä»¶: {os.path.basename(file_path)}",
                "å–æ¶ˆ",
                0, 100,
                self
            )
            progress_dialog.show()
            
            # è·å–æ–‡ä»¶åŸºæœ¬ä¿¡æ¯
            progress_dialog.setLabelText("è·å–æ–‡ä»¶ä¿¡æ¯...")
            progress_dialog.setValue(20)
            QApplication.processEvents()
            
            file_size = os.path.getsize(file_path)
            file_size_mb = file_size / (1024 * 1024)
            modified_time = datetime.fromtimestamp(os.path.getmtime(file_path))
            
            # è¯»å–Excelæ–‡ä»¶è·å–æ•°æ®ä¿¡æ¯
            progress_dialog.setLabelText("è¯»å–æ–‡ä»¶å†…å®¹...")
            progress_dialog.setValue(50)
            QApplication.processEvents()
            
            try:
                if file_key == 'milk_yield' and self.current_mastitis_system == 'yiqiniu':
                    # å¥¶ç‰›äº§å¥¶æ—¥æ±‡æ€»è¡¨å¯èƒ½æœ‰å¤šä¸ªsheet
                    with pd.ExcelFile(file_path) as xls:
                        sheet_names = xls.sheet_names
                        total_rows = 0
                        sheet_info = []
                        for i, sheet_name in enumerate(sheet_names):
                            progress_value = 50 + int(40 * (i + 1) / len(sheet_names))
                            progress_dialog.setValue(progress_value)
                            progress_dialog.setLabelText(f"è¯»å–å·¥ä½œè¡¨: {sheet_name}")
                            QApplication.processEvents()
                            
                            df = pd.read_excel(file_path, sheet_name=sheet_name)
                            total_rows += len(df)
                            sheet_info.append(f"  - {sheet_name}: {len(df)}è¡Œ")
                        
                        data_info = f"æ•°æ®ä¿¡æ¯: {len(sheet_names)}ä¸ªå·¥ä½œè¡¨ï¼Œå…±{total_rows}è¡Œæ•°æ®\n"
                        data_info += "\n".join(sheet_info)
                else:
                    # å•ä¸ªsheet
                    df = pd.read_excel(file_path)
                    data_info = f"æ•°æ®ä¿¡æ¯: {len(df)}è¡Œ Ã— {len(df.columns)}åˆ—"
                    if len(df) > 0:
                        # æ˜¾ç¤ºå‰å‡ ä¸ªåˆ—å
                        columns_preview = ", ".join(df.columns[:5].tolist())
                        if len(df.columns) > 5:
                            columns_preview += "..."
                        data_info += f"\nåˆ—åé¢„è§ˆ: {columns_preview}"
            except Exception as e:
                data_info = f"æ•°æ®ä¿¡æ¯: è¯»å–å¤±è´¥ - {str(e)}"
            
            progress_dialog.setValue(100)
            progress_dialog.close()
            
            # æ„å»ºä¿¡æ¯æ–‡æœ¬
            info_text = f"""
ğŸ†• æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥æ–‡ä»¶ä¸Šä¼ 
â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”

ğŸ“‹ æ–‡ä»¶ç±»å‹: {file_name}
ğŸ“ æ–‡ä»¶å: {os.path.basename(file_path)}
ğŸ“Š æ–‡ä»¶å¤§å°: {file_size_mb:.2f} MB
ğŸ“… ä¿®æ”¹æ—¶é—´: {modified_time.strftime('%Y-%m-%d %H:%M:%S')}
ğŸ—‚ï¸ å®Œæ•´è·¯å¾„: {file_path}

{data_info}

âš™ï¸ ç³»ç»Ÿç±»å‹: {self.current_mastitis_system}
ğŸ”„ çŠ¶æ€: å·²ä¸Šä¼ ï¼Œç­‰å¾…å¤„ç†

"""
            
            # æ˜¾ç¤ºåˆ°å³ä¾§æ–‡ä»¶ä¿¡æ¯é¢æ¿
            self.file_info_widget.append(info_text)
            
            # è‡ªåŠ¨åˆ‡æ¢åˆ°æ–‡ä»¶ä¿¡æ¯æ ‡ç­¾é¡µ
            self.tab_widget.setCurrentWidget(self.file_info_widget)
            
        except Exception as e:
            error_text = f"""
âŒ æ–‡ä»¶ä¿¡æ¯è·å–å¤±è´¥
â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”

ğŸ“ æ–‡ä»¶å: {os.path.basename(file_path)}
âŒ é”™è¯¯: {str(e)}

"""
            self.file_info_widget.append(error_text)
    
    def process_and_save_cattle_basic_info(self, file_path: str) -> bool:
        """ç«‹å³å¤„ç†å¹¶ä¿å­˜ç‰›ç¾¤åŸºç¡€ä¿¡æ¯åˆ°ä¸»çª—å£"""
        try:
            print(f"\nğŸ”„ [ç«‹å³å¤„ç†] å¼€å§‹å¤„ç†ç‰›ç¾¤åŸºç¡€ä¿¡æ¯è¡¨...")
            print(f"   æ–‡ä»¶è·¯å¾„: {file_path}")
            print(f"   å½“å‰ç³»ç»Ÿ: {self.current_mastitis_system}")
            print(f"   æ–‡ä»¶æ˜¯å¦å­˜åœ¨: {os.path.exists(file_path)}")
            
            # æ£€æŸ¥å½“å‰ä¸»çª—å£çŠ¶æ€
            print(f"ğŸ” [å¤„ç†å‰] ä¸»çª—å£çŠ¶æ€æ£€æŸ¥...")
            print(f"   hasattr(self, 'cattle_basic_info'): {hasattr(self, 'cattle_basic_info')}")
            print(f"   hasattr(self, 'current_system'): {hasattr(self, 'current_system')}")
            print(f"   hasattr(self, 'data_processor'): {hasattr(self, 'data_processor')}")
            
            # ç›´æ¥å¤„ç†ç‰›ç¾¤åŸºç¡€ä¿¡æ¯æ–‡ä»¶ï¼ˆä¸ä¾èµ–å…¶ä»–æ–‡ä»¶ï¼‰
            print(f"   ğŸ”„ ç›´æ¥å¤„ç†ç‰›ç¾¤åŸºç¡€ä¿¡æ¯æ–‡ä»¶...")
            
            if self.current_mastitis_system == 'yiqiniu':
                # ä¼Šèµ·ç‰›ç³»ç»Ÿï¼šç›´æ¥è°ƒç”¨ç‰›ç¾¤åŸºç¡€ä¿¡æ¯å¤„ç†æ–¹æ³•
                success, message, cattle_df = self.data_processor._process_yiqiniu_cattle_info(file_path)
                processed_data = {'cattle_info': cattle_df} if success else {}
                
            elif self.current_mastitis_system == 'huimuyun':
                # æ…§ç‰§äº‘ç³»ç»Ÿï¼šç›´æ¥è°ƒç”¨ç‰›ç¾¤åŸºç¡€ä¿¡æ¯å¤„ç†æ–¹æ³•
                success, message, cattle_df = self.data_processor._process_huimuyun_cattle_info(file_path)
                processed_data = {'cattle_info': cattle_df} if success else {}
                
            elif self.current_mastitis_system == 'custom':
                # è‡ªå®šä¹‰ç³»ç»Ÿéœ€è¦å­—æ®µæ˜ å°„
                widget = self.mastitis_file_uploads.get('cattle_info')
                if widget and hasattr(widget, 'mapping_inputs'):
                    field_mappings = {}
                    for field, input_widget in widget.mapping_inputs.items():
                        column_name = input_widget.text().strip()
                        if column_name:
                            field_mappings[field] = column_name
                    
                    success, message, cattle_df = self.data_processor._process_custom_cattle_info(
                        file_path, field_mappings
                    )
                    processed_data = {'cattle_info': cattle_df} if success else {}
                else:
                    print(f"   âŒ è‡ªå®šä¹‰ç³»ç»Ÿç¼ºå°‘å­—æ®µæ˜ å°„é…ç½®")
                    return False
            else:
                print(f"   âŒ æœªçŸ¥ç³»ç»Ÿç±»å‹: {self.current_mastitis_system}")
                return False
            
            if success and 'cattle_info' in processed_data:
                # ä¿å­˜ç‰›ç¾¤åŸºç¡€ä¿¡æ¯åˆ°ä¸»çª—å£
                self.cattle_basic_info = processed_data['cattle_info']
                self.current_system = self.current_mastitis_system
                
                print(f"âœ… [ç«‹å³å¤„ç†] ç‰›ç¾¤åŸºç¡€ä¿¡æ¯å·²ä¿å­˜åˆ°ä¸»çª—å£: {len(self.cattle_basic_info)}å¤´ç‰›")
                print(f"âœ… [ç«‹å³å¤„ç†] ç³»ç»Ÿç±»å‹å·²ä¿å­˜: {self.current_system}")
                
                # æ›´æ–°éšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹çš„æ•°æ®çŠ¶æ€æ˜¾ç¤º
                if hasattr(self, 'update_monitoring_data_status'):
                    self.update_monitoring_data_status()
                
                # åœ¨å¤„ç†è¿‡ç¨‹é¢æ¿ä¸­æ˜¾ç¤ºæˆåŠŸä¿¡æ¯
                self.process_log_widget.append(f"""
ğŸ‰ ç‰›ç¾¤åŸºç¡€ä¿¡æ¯ç«‹å³å¤„ç†æˆåŠŸ
â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”

ğŸ“ æ–‡ä»¶: {os.path.basename(file_path)}
âš™ï¸ ç³»ç»Ÿ: {self.current_mastitis_system}
ğŸ„ ç‰›åªæ•°é‡: {len(self.cattle_basic_info)}å¤´
âœ… çŠ¶æ€: å·²è‡ªåŠ¨ä¿å­˜ï¼Œå¯ç”¨äºéšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹

ğŸ’¡ ç°åœ¨æ‚¨å¯ä»¥åˆ°"éšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹"åŠŸèƒ½ä¸­ä¸Šä¼ DHIæ•°æ®è¿›è¡Œåˆ†æäº†ï¼
""")
                
                return True
            else:
                print(f"   âŒ ç‰›ç¾¤åŸºç¡€ä¿¡æ¯å¤„ç†å¤±è´¥: {message}")
                return False
                
        except Exception as e:
            print(f"   âŒ å¤„ç†ç‰›ç¾¤åŸºç¡€ä¿¡æ¯æ—¶å‡ºé”™: {str(e)}")
            return False

    def extract_and_update_breeding_status(self, file_path: str):
        """æå–ç‰›ç¾¤åŸºç¡€ä¿¡æ¯è¡¨ä¸­çš„ç¹æ®–çŠ¶æ€å¹¶æ›´æ–°é€‰é¡¹"""
        try:
            # åœ¨å¤„ç†è¿‡ç¨‹ä¸­æ˜¾ç¤ºæ“ä½œä¿¡æ¯
            self.process_log_widget.append(f"""
ğŸ”„ è‡ªåŠ¨æå–ç¹è‚²çŠ¶æ€
â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”

ğŸ“ æ–‡ä»¶: {os.path.basename(file_path)}
âš™ï¸ ç³»ç»Ÿ: {self.current_mastitis_system}
ğŸ” å¼€å§‹æå–...
""")
            
            # åˆ‡æ¢åˆ°å¤„ç†è¿‡ç¨‹æ ‡ç­¾é¡µ
            self.tab_widget.setCurrentWidget(self.process_log_widget)
            
            # è¯»å–Excelæ–‡ä»¶
            self.process_log_widget.append("ğŸ“– æ­£åœ¨è¯»å–Excelæ–‡ä»¶...")
            df = pd.read_excel(file_path)
            self.process_log_widget.append(f"âœ… æˆåŠŸè¯»å–æ–‡ä»¶ï¼Œå…± {len(df)} è¡Œæ•°æ®")
            
            # æ ¹æ®å½“å‰ç³»ç»Ÿç±»å‹ç¡®å®šç¹æ®–çŠ¶æ€åˆ—å
            breeding_status_col = None
            if self.current_mastitis_system == 'yiqiniu':
                breeding_status_col = 'ç¹è‚²çŠ¶æ€'
            elif self.current_mastitis_system == 'huimuyun':
                breeding_status_col = 'ç¹è‚²çŠ¶æ€'
            elif self.current_mastitis_system == 'custom':
                # è‡ªå®šä¹‰ç³»ç»Ÿï¼Œä»å­—æ®µæ˜ å°„ä¸­è·å–
                widget = self.mastitis_file_uploads.get('cattle_info')
                if widget and hasattr(widget, 'mapping_inputs'):
                    input_widget = widget.mapping_inputs.get('ç¹è‚²çŠ¶æ€')
                    if input_widget:
                        breeding_status_col = input_widget.text().strip()
            
            self.process_log_widget.append(f"ğŸ” æŸ¥æ‰¾ç¹è‚²çŠ¶æ€åˆ—: {breeding_status_col}")
            
            if not breeding_status_col:
                error_msg = "âŒ æœªæ‰¾åˆ°ç¹æ®–çŠ¶æ€åˆ—åæ˜ å°„"
                self.process_log_widget.append(error_msg)
                return
                
            if breeding_status_col not in df.columns:
                error_msg = f"âŒ æ–‡ä»¶ä¸­æœªæ‰¾åˆ°åˆ— '{breeding_status_col}'"
                self.process_log_widget.append(f"ğŸ“‹ å¯ç”¨åˆ—å: {list(df.columns)}")
                self.process_log_widget.append(error_msg)
                return
            
            # æå–æ‰€æœ‰ä¸åŒçš„ç¹æ®–çŠ¶æ€å€¼
            self.process_log_widget.append("ğŸ” æ­£åœ¨åˆ†æç¹è‚²çŠ¶æ€æ•°æ®...")
            unique_statuses = df[breeding_status_col].dropna().unique()
            unique_statuses = [str(status).strip() for status in unique_statuses if str(status).strip()]
            unique_statuses = sorted(set(unique_statuses))  # å»é‡å¹¶æ’åº
            
            if unique_statuses:
                success_msg = f"âœ… æˆåŠŸæå–ç¹æ®–çŠ¶æ€é€‰é¡¹: {', '.join(unique_statuses)}"
                self.process_log_widget.append(success_msg)
                
                # æ›´æ–°æ‰€æœ‰å¤„ç½®åŠæ³•çš„ç¹æ®–çŠ¶æ€é€‰é¡¹
                self.process_log_widget.append("ğŸ”„ æ­£åœ¨æ›´æ–°å¤„ç½®åŠæ³•é…ç½®...")
                try:
                    self.update_breeding_status_options(unique_statuses)
                    self.process_log_widget.append("âœ… å¤„ç½®åŠæ³•é…ç½®æ›´æ–°å®Œæˆ")
                except Exception as update_error:
                    error_msg = f"âŒ æ›´æ–°å¤„ç½®åŠæ³•é…ç½®æ—¶å‡ºé”™: {str(update_error)}"
                    self.process_log_widget.append(error_msg)
                    print(f"æ›´æ–°å¤„ç½®åŠæ³•é…ç½®æ—¶å‡ºé”™: {update_error}")
            else:
                warning_msg = "âš ï¸ æœªæ‰¾åˆ°æœ‰æ•ˆçš„ç¹æ®–çŠ¶æ€æ•°æ®"
                self.process_log_widget.append(warning_msg)
                
        except Exception as e:
            error_msg = f"âŒ æå–ç¹æ®–çŠ¶æ€æ—¶å‡ºé”™: {str(e)}"
            self.process_log_widget.append(error_msg)
            print(f"æå–ç¹æ®–çŠ¶æ€æ—¶å‡ºé”™: {e}")
            import traceback
            traceback.print_exc()
    
    def update_breeding_status_options(self, statuses: List[str]):
        """æ›´æ–°æ‰€æœ‰å¤„ç½®åŠæ³•çš„ç¹æ®–çŠ¶æ€é€‰é¡¹"""
        try:
            print(f"å¼€å§‹æ›´æ–°ç¹æ®–çŠ¶æ€é€‰é¡¹: {statuses}")
            
            for method_key, widget in self.treatment_configs.items():
                print(f"å¤„ç†å¤„ç½®åŠæ³•: {method_key}")
                
                if hasattr(widget, 'breeding_checkboxes'):
                    # æ¸…é™¤ç°æœ‰çš„å¤é€‰æ¡†
                    print(f"æ¸…é™¤ç°æœ‰å¤é€‰æ¡†: {len(widget.breeding_checkboxes)}ä¸ª")
                    for cb in widget.breeding_checkboxes.values():
                        if cb is not None:
                            cb.setParent(None)
                    widget.breeding_checkboxes.clear()
                    
                    # æ‰¾åˆ°ç¹æ®–çŠ¶æ€å¸ƒå±€å®¹å™¨ - ä½¿ç”¨æ›´ç®€å•çš„æ–¹æ³•
                    breeding_status_widget = None
                    config_layout = widget.config_widget.layout()
                    
                    if config_layout is not None:
                        # ç›´æ¥æŸ¥æ‰¾æ ‡ç­¾ä¸º"ç¹æ®–çŠ¶æ€:"çš„è¡Œ
                        for i in range(config_layout.rowCount()):
                            label_item = config_layout.itemAt(i, QFormLayout.ItemRole.LabelRole)
                            field_item = config_layout.itemAt(i, QFormLayout.ItemRole.FieldRole)
                            
                            if label_item and label_item.widget():
                                label_text = label_item.widget().text()
                                if "ç¹æ®–çŠ¶æ€" in label_text and field_item and field_item.widget():
                                    breeding_status_widget = field_item.widget()
                                    print(f"æ‰¾åˆ°ç¹æ®–çŠ¶æ€æ§ä»¶: {breeding_status_widget}")
                                    break
                    
                    if breeding_status_widget:
                        # è·å–æˆ–åˆ›å»ºå¸ƒå±€
                        layout = breeding_status_widget.layout()
                        if layout is None:
                            # å¦‚æœæ²¡æœ‰å¸ƒå±€ï¼Œåˆ›å»ºæ–°çš„
                            layout = QGridLayout(breeding_status_widget)
                            layout.setContentsMargins(0, 0, 0, 0)
                            print("åˆ›å»ºæ–°çš„ç½‘æ ¼å¸ƒå±€")
                        else:
                            # å¦‚æœæœ‰å¸ƒå±€ï¼Œæ¸…ç©ºå†…å®¹
                            print(f"æ¸…ç©ºç°æœ‰å¸ƒå±€ï¼Œæœ‰ {layout.count()} ä¸ªé¡¹ç›®")
                            while layout.count():
                                item = layout.takeAt(0)
                                if item and item.widget():
                                    item.widget().setParent(None)
                        
                        # é‡æ–°åˆ›å»ºå¤é€‰æ¡†
                        widget.breeding_checkboxes = {}
                        print(f"åˆ›å»º {len(statuses)} ä¸ªæ–°çš„å¤é€‰æ¡†")
                        for i, status in enumerate(statuses):
                            cb = QCheckBox(status)
                            cb.setChecked(True)  # é»˜è®¤å…¨é€‰
                            widget.breeding_checkboxes[status] = cb
                            
                            # è®¡ç®—è¡Œåˆ—ä½ç½®ï¼ˆæ¯è¡Œ3ä¸ªï¼‰
                            row = i // 3
                            col = i % 3
                            layout.addWidget(cb, row, col)
                        
                        print(f"æˆåŠŸæ›´æ–° {method_key} çš„ç¹æ®–çŠ¶æ€é€‰é¡¹")
                    else:
                        print(f"æœªæ‰¾åˆ° {method_key} çš„ç¹æ®–çŠ¶æ€æ§ä»¶")
                else:
                    print(f"{method_key} æ²¡æœ‰ breeding_checkboxes å±æ€§")
            
            print("ç¹æ®–çŠ¶æ€é€‰é¡¹æ›´æ–°å®Œæˆ")
            
        except Exception as e:
            print(f"update_breeding_status_options å‡ºé”™: {e}")
            import traceback
            traceback.print_exc()
            raise

    def update_chronic_months_options(self, available_months: List[str]):
        """æ›´æ–°æ…¢æ€§æ„ŸæŸ“ç‰›è¯†åˆ«çš„æœˆä»½é€‰æ‹©é€‰é¡¹"""
        print(f"å¼€å§‹æ›´æ–°æ…¢æ€§æ„ŸæŸ“ç‰›æœˆä»½é€‰é¡¹: {available_months}")
        
        # æ£€æŸ¥chronic_months_widgetæ˜¯å¦å­˜åœ¨
        if not hasattr(self, 'chronic_months_widget'):
            print("chronic_months_widgetä¸å­˜åœ¨ï¼Œè·³è¿‡æœˆä»½é€‰é¡¹æ›´æ–°")
            return
        
        # æ¸…ç©ºç°æœ‰å¸ƒå±€
        layout = self.chronic_months_widget.layout()
        for i in reversed(range(layout.count())):
            item = layout.takeAt(i)
            if item and item.widget():
                item.widget().deleteLater()
        
        # æ¸…ç©ºå¤é€‰æ¡†å­—å…¸
        self.chronic_month_checkboxes = {}
        
        if not available_months:
            # æ²¡æœ‰æ•°æ®æ—¶æ˜¾ç¤ºæç¤ºä¿¡æ¯
            no_data_label = QLabel("è¯·å…ˆä¸Šä¼ DHIæ•°æ®ä»¥é€‰æ‹©æœˆä»½")
            no_data_label.setStyleSheet("color: #6c757d; font-style: italic;")
            layout.addWidget(no_data_label, 0, 0, 1, 3)
            print("æ²¡æœ‰å¯ç”¨æœˆä»½ï¼Œæ˜¾ç¤ºæç¤ºä¿¡æ¯")
            return
        
        # åˆ›å»ºæœˆä»½å¤é€‰æ¡†ï¼ŒæŒ‰å¹´-æœˆæ’åº
        sorted_months = sorted(available_months)
        last_month = sorted_months[-1] if sorted_months else None  # è·å–æœ€åä¸€ä¸ªæœˆä»½ï¼ˆæœ€æ–°æ•°æ®ï¼‰
        print(f"æŒ‰é¡ºåºåˆ›å»º {len(sorted_months)} ä¸ªæœˆä»½å¤é€‰æ¡†ï¼Œé»˜è®¤é€‰æ‹©æœ€æ–°æœˆä»½: {last_month}")

        for i, month in enumerate(sorted_months):
            cb = QCheckBox(month)
            cb.setChecked(month == last_month)  # é»˜è®¤åªé€‰æ‹©æœ€æ–°æœˆä»½
            self.chronic_month_checkboxes[month] = cb

            # è®¡ç®—è¡Œåˆ—ä½ç½®ï¼ˆæ¯è¡Œ4ä¸ªï¼‰
            row = i // 4
            col = i % 4
            layout.addWidget(cb, row, col)

        print(f"æˆåŠŸåˆ›å»º {len(sorted_months)} ä¸ªæœˆä»½å¤é€‰æ¡†ï¼Œé»˜è®¤é€‰æ‹©: {last_month}")
        
        # æ·»åŠ å…¨é€‰/å…¨ä¸é€‰æŒ‰é’®
        button_row = (len(sorted_months) - 1) // 4 + 1
        
        select_all_btn = QPushButton("å…¨é€‰")
        select_all_btn.setMaximumWidth(60)
        select_all_btn.clicked.connect(lambda: self._set_all_chronic_months(True))
        layout.addWidget(select_all_btn, button_row, 0)
        
        select_none_btn = QPushButton("å…¨ä¸é€‰")
        select_none_btn.setMaximumWidth(60)
        select_none_btn.clicked.connect(lambda: self._set_all_chronic_months(False))
        layout.addWidget(select_none_btn, button_row, 1)
        
        print("æœˆä»½é€‰é¡¹æ›´æ–°å®Œæˆ")

    def _set_all_chronic_months(self, checked: bool):
        """è®¾ç½®æ‰€æœ‰æ…¢æ€§æ„ŸæŸ“ç‰›æœˆä»½å¤é€‰æ¡†çš„çŠ¶æ€"""
        for month, cb in self.chronic_month_checkboxes.items():
            cb.setChecked(checked)
    
    def create_treatment_config_widget(self, method_key: str, method_name: str, icon: str, form_styles: str):
        """åˆ›å»ºå¤„ç½®åŠæ³•é…ç½®æ§ä»¶"""
        widget = QWidget()
        widget.setStyleSheet("""
            QWidget {
                border: 1px solid #e0e0e0;
                border-radius: 8px;
                padding: 10px;
                margin: 5px 0;
                background-color: #fafafa;
            }
        """)
        
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(15, 15, 15, 15)

        # ç¹æ®–çŠ¶æ€æç¤ºä¿¡æ¯
        breeding_hints = {
            'cull': 'é€‚ç”¨äºï¼šæœªå­•ç‰›',
            'isolate': 'é€‚ç”¨äºï¼šæœªå­•ç‰›',
            'blind_quarter': 'é€‚ç”¨äºï¼šæ€€å­•ç‰›',
            'early_dry': 'é€‚ç”¨äºï¼šæ€€å­•ç‰›',
            'treatment': 'é€‚ç”¨äºï¼šæ€€å­•ç‰›'
        }

        # æ ‡é¢˜è¡Œ
        title_layout = QHBoxLayout()

        # å¯ç”¨å¤é€‰æ¡†
        enabled_cb = QCheckBox(f"{icon} {method_name}")
        enabled_cb.setStyleSheet("font-weight: bold; color: black; background-color: white;")
        enabled_cb.setChecked(True)  # é»˜è®¤å¯ç”¨
        title_layout.addWidget(enabled_cb)

        # æ·»åŠ ç¹æ®–çŠ¶æ€æç¤º
        hint_label = QLabel(breeding_hints.get(method_key, ''))
        hint_label.setStyleSheet("color: #6c757d; font-size: 11px; font-style: italic; background-color: transparent;")
        title_layout.addWidget(hint_label)

        title_layout.addStretch()

        layout.addLayout(title_layout)
        
        # é…ç½®åŒºåŸŸï¼ˆé»˜è®¤æ˜¾ç¤ºï¼‰
        config_widget = QWidget()
        config_widget.setStyleSheet("""
            QWidget {
                text-align: left;
                alignment: left;
            }
        """)
        config_layout = QFormLayout(config_widget)
        config_layout.setContentsMargins(20, 10, 10, 10)
        config_layout.setLabelAlignment(Qt.AlignmentFlag.AlignLeft)  # è®¾ç½®æ ‡ç­¾å·¦å¯¹é½
        config_layout.setFormAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)  # è®¾ç½®è¡¨å•å·¦å¯¹é½
        config_layout.setFieldGrowthPolicy(QFormLayout.FieldGrowthPolicy.ExpandingFieldsGrow)  # å­—æ®µæ‰©å±•ç­–ç•¥
        
        # æ ¹æ®å¤„ç½®åŠæ³•ç±»å‹åˆ›å»ºå¯¹åº”çš„é…ç½®é¡¹
        if method_key == 'cull':  # æ·˜æ±°
            # äº§å¥¶é‡æ¡ä»¶
            yield_layout = QHBoxLayout()
            yield_combo = QComboBox()
            yield_combo.addItems(["<", "<=", "=", ">=", ">"])
            yield_combo.setCurrentText("<=")
            yield_combo.setStyleSheet(form_styles)
            yield_combo.setFixedWidth(70)
            
            yield_spin = QDoubleSpinBox()
            yield_spin.setRange(0, 100)
            yield_spin.setValue(15)
            yield_spin.setSuffix("kg")
            yield_spin.setStyleSheet(form_styles)
            
            yield_layout.addWidget(yield_combo)
            yield_layout.addWidget(yield_spin)
            yield_layout.addStretch()
            
            yield_widget = QWidget()
            yield_widget.setLayout(yield_layout)
            yield_label = QLabel("äº§å¥¶é‡:")
            yield_label.setStyleSheet("color: black; background-color: white; font-weight: bold;")
            config_layout.addRow(yield_label, yield_widget)
            widget.yield_combo = yield_combo
            widget.yield_spin = yield_spin
            
        elif method_key == 'isolate':  # ç¦é…éš”ç¦»
            # äº§å¥¶é‡æ¡ä»¶
            yield_layout = QHBoxLayout()
            yield_combo = QComboBox()
            yield_combo.addItems(["<", "<=", "=", ">=", ">"])
            yield_combo.setCurrentText(">=")
            yield_combo.setStyleSheet(form_styles)
            yield_combo.setFixedWidth(70)
            
            yield_spin = QDoubleSpinBox()
            yield_spin.setRange(0, 100)
            yield_spin.setValue(15)
            yield_spin.setSuffix("kg")
            yield_spin.setStyleSheet(form_styles)
            
            yield_layout.addWidget(yield_combo)
            yield_layout.addWidget(yield_spin)
            yield_layout.addStretch()
            
            yield_widget = QWidget()
            yield_widget.setLayout(yield_layout)
            yield_label2 = QLabel("äº§å¥¶é‡:")
            yield_label2.setStyleSheet("color: black; background-color: white; font-weight: bold;")
            config_layout.addRow(yield_label2, yield_widget)
            widget.yield_combo = yield_combo
            widget.yield_spin = yield_spin
            
        elif method_key == 'blind_quarter':  # çä¹³åŒº
            # äº§å¥¶é‡æ¡ä»¶ï¼ˆéœ€æ±‚ï¼š>=15kgï¼‰
            yield_layout = QHBoxLayout()
            yield_combo = QComboBox()
            yield_combo.addItems(["<", "<=", "=", ">=", ">"])
            yield_combo.setCurrentText(">=")
            yield_combo.setStyleSheet(form_styles)
            yield_combo.setFixedWidth(70)

            yield_spin = QDoubleSpinBox()
            yield_spin.setRange(0, 100)
            yield_spin.setValue(15)
            yield_spin.setSuffix("kg")
            yield_spin.setStyleSheet(form_styles)

            yield_layout.addWidget(yield_combo)
            yield_layout.addWidget(yield_spin)
            yield_layout.addStretch()

            yield_widget = QWidget()
            yield_widget.setLayout(yield_layout)
            yield_label = QLabel("äº§å¥¶é‡:")
            yield_label.setStyleSheet("color: black; background-color: white; font-weight: bold;")
            config_layout.addRow(yield_label, yield_widget)
            widget.yield_combo = yield_combo
            widget.yield_spin = yield_spin

            # åœ¨èƒå¤©æ•°æ¡ä»¶ï¼ˆéœ€æ±‚ï¼š<180å¤©ï¼‰
            gestation_layout = QHBoxLayout()
            gestation_combo = QComboBox()
            gestation_combo.addItems(["<", "<=", "=", ">=", ">"])
            gestation_combo.setCurrentText("<")
            gestation_combo.setStyleSheet(form_styles)
            gestation_combo.setFixedWidth(70)

            gestation_spin = QSpinBox()
            gestation_spin.setRange(0, 300)
            gestation_spin.setValue(180)
            gestation_spin.setSuffix("å¤©")
            gestation_spin.setStyleSheet(form_styles)

            gestation_layout.addWidget(gestation_combo)
            gestation_layout.addWidget(gestation_spin)
            gestation_layout.addStretch()

            gestation_widget = QWidget()
            gestation_widget.setLayout(gestation_layout)
            config_layout.addRow("åœ¨èƒå¤©æ•°:", gestation_widget)
            widget.gestation_combo = gestation_combo
            widget.gestation_spin = gestation_spin
            
        elif method_key == 'early_dry':  # æå‰å¹²å¥¶
            # äº§å¥¶é‡æ¡ä»¶ï¼ˆéœ€æ±‚ï¼š>=0kgï¼‰
            yield_layout = QHBoxLayout()
            yield_combo = QComboBox()
            yield_combo.addItems(["<", "<=", "=", ">=", ">"])
            yield_combo.setCurrentText(">=")
            yield_combo.setStyleSheet(form_styles)
            yield_combo.setFixedWidth(70)

            yield_spin = QDoubleSpinBox()
            yield_spin.setRange(0, 100)
            yield_spin.setValue(0)
            yield_spin.setSuffix("kg")
            yield_spin.setStyleSheet(form_styles)

            yield_layout.addWidget(yield_combo)
            yield_layout.addWidget(yield_spin)
            yield_layout.addStretch()

            yield_widget = QWidget()
            yield_widget.setLayout(yield_layout)
            yield_label = QLabel("äº§å¥¶é‡:")
            yield_label.setStyleSheet("color: black; background-color: white; font-weight: bold;")
            config_layout.addRow(yield_label, yield_widget)
            widget.yield_combo = yield_combo
            widget.yield_spin = yield_spin

            # åœ¨èƒå¤©æ•°æ¡ä»¶ï¼ˆéœ€æ±‚ï¼š>=180å¤©ï¼‰
            gestation_layout = QHBoxLayout()
            gestation_combo = QComboBox()
            gestation_combo.addItems(["<", "<=", "=", ">=", ">"])
            gestation_combo.setCurrentText(">=")
            gestation_combo.setStyleSheet(form_styles)
            gestation_combo.setFixedWidth(70)

            gestation_spin = QSpinBox()
            gestation_spin.setRange(0, 300)
            gestation_spin.setValue(180)
            gestation_spin.setSuffix("å¤©")
            gestation_spin.setStyleSheet(form_styles)

            gestation_layout.addWidget(gestation_combo)
            gestation_layout.addWidget(gestation_spin)
            gestation_layout.addStretch()

            gestation_widget = QWidget()
            gestation_widget.setLayout(gestation_layout)
            config_layout.addRow("åœ¨èƒå¤©æ•°:", gestation_widget)
            widget.gestation_combo = gestation_combo
            widget.gestation_spin = gestation_spin

        elif method_key == 'treatment':  # æ²»ç–—
            # äº§å¥¶é‡æ¡ä»¶ï¼ˆéœ€æ±‚ï¼š>=0kgï¼‰
            yield_layout = QHBoxLayout()
            yield_combo = QComboBox()
            yield_combo.addItems(["<", "<=", "=", ">=", ">"])
            yield_combo.setCurrentText(">=")
            yield_combo.setStyleSheet(form_styles)
            yield_combo.setFixedWidth(70)

            yield_spin = QDoubleSpinBox()
            yield_spin.setRange(0, 100)
            yield_spin.setValue(0)
            yield_spin.setSuffix("kg")
            yield_spin.setStyleSheet(form_styles)

            yield_layout.addWidget(yield_combo)
            yield_layout.addWidget(yield_spin)
            yield_layout.addStretch()

            yield_widget = QWidget()
            yield_widget.setLayout(yield_layout)
            yield_label = QLabel("äº§å¥¶é‡:")
            yield_label.setStyleSheet("color: black; background-color: white; font-weight: bold;")
            config_layout.addRow(yield_label, yield_widget)
            widget.yield_combo = yield_combo
            widget.yield_spin = yield_spin

            # åœ¨èƒå¤©æ•°æ¡ä»¶ï¼ˆéœ€æ±‚ï¼š>=0å¤©ï¼‰
            gestation_layout = QHBoxLayout()
            gestation_combo = QComboBox()
            gestation_combo.addItems(["<", "<=", "=", ">=", ">"])
            gestation_combo.setCurrentText(">=")
            gestation_combo.setStyleSheet(form_styles)
            gestation_combo.setFixedWidth(70)

            gestation_spin = QSpinBox()
            gestation_spin.setRange(0, 300)
            gestation_spin.setValue(0)
            gestation_spin.setSuffix("å¤©")
            gestation_spin.setStyleSheet(form_styles)

            gestation_layout.addWidget(gestation_combo)
            gestation_layout.addWidget(gestation_spin)
            gestation_layout.addStretch()

            gestation_widget = QWidget()
            gestation_widget.setLayout(gestation_layout)
            config_layout.addRow("åœ¨èƒå¤©æ•°:", gestation_widget)
            widget.gestation_combo = gestation_combo
            widget.gestation_spin = gestation_spin

        # å…¬å…±é…ç½®é¡¹
        # å‘ç—…æ¬¡æ•°æ¡ä»¶ï¼ˆæ²»ç–—é»˜è®¤<=2ï¼Œå…¶ä»–é»˜è®¤>=2ï¼‰
        mastitis_layout = QHBoxLayout()
        mastitis_combo = QComboBox()
        mastitis_combo.addItems(["<", "<=", "=", ">=", ">"])
        # æ²»ç–—çš„å‘ç—…æ¬¡æ•°é»˜è®¤ <=2ï¼Œå…¶ä»–é»˜è®¤ >=2
        if method_key == 'treatment':
            mastitis_combo.setCurrentText("<=")
        else:
            mastitis_combo.setCurrentText(">=")
        mastitis_combo.setStyleSheet(form_styles)
        mastitis_combo.setFixedWidth(70)

        mastitis_spin = QSpinBox()
        mastitis_spin.setRange(0, 10)
        mastitis_spin.setValue(2)
        mastitis_spin.setSuffix("æ¬¡")
        mastitis_spin.setStyleSheet(form_styles)

        mastitis_layout.addWidget(mastitis_combo)
        mastitis_layout.addWidget(mastitis_spin)
        mastitis_layout.addStretch()

        mastitis_widget = QWidget()
        mastitis_widget.setLayout(mastitis_layout)
        config_layout.addRow("å‘ç—…æ¬¡æ•°:", mastitis_widget)
        widget.mastitis_combo = mastitis_combo
        widget.mastitis_spin = mastitis_spin

        # æ³Œä¹³å¤©æ•°æ¡ä»¶ï¼ˆçä¹³åŒº/æå‰å¹²å¥¶/æ²»ç–—é»˜è®¤>=0ï¼Œæ·˜æ±°/ç¦é…éš”ç¦»é»˜è®¤>=200ï¼‰
        lactation_layout = QHBoxLayout()
        lactation_combo = QComboBox()
        lactation_combo.addItems(["<", "<=", "=", ">=", ">"])
        lactation_combo.setCurrentText(">=")
        lactation_combo.setStyleSheet(form_styles)
        lactation_combo.setFixedWidth(70)

        lactation_spin = QSpinBox()
        lactation_spin.setRange(0, 500)
        # çä¹³åŒº/æå‰å¹²å¥¶/æ²»ç–—çš„æ³Œä¹³å¤©æ•°é»˜è®¤ >=0ï¼Œæ·˜æ±°/ç¦é…éš”ç¦»é»˜è®¤ >=200
        if method_key in ['blind_quarter', 'early_dry', 'treatment']:
            lactation_spin.setValue(0)
        else:
            lactation_spin.setValue(200)
        lactation_spin.setSuffix("å¤©")
        lactation_spin.setStyleSheet(form_styles)

        lactation_layout.addWidget(lactation_combo)
        lactation_layout.addWidget(lactation_spin)
        lactation_layout.addStretch()

        lactation_widget = QWidget()
        lactation_widget.setLayout(lactation_layout)
        config_layout.addRow("æ³Œä¹³å¤©æ•°:", lactation_widget)
        widget.lactation_combo = lactation_combo
        widget.lactation_spin = lactation_spin

        # ç¹æ®–çŠ¶æ€å¤šé€‰
        breeding_status_label = QLabel("ç¹æ®–çŠ¶æ€:")
        breeding_status_widget = QWidget()
        breeding_status_layout = QGridLayout(breeding_status_widget)
        breeding_status_layout.setContentsMargins(0, 0, 0, 0)

        # æ ¹æ®å¤„ç½®åŠæ³•ç±»å‹è®¾ç½®ä¸åŒçš„é»˜è®¤ç¹æ®–çŠ¶æ€
        # æ·˜æ±°/ç¦é…éš”ç¦»ï¼šäº§çŠŠã€ç¦é…ã€å¯é…ã€å·²é…ã€äº§åæœªé…ã€åˆæ£€ç©ºæ€€ã€å‘æƒ…æœªé…ã€æµäº§æœªé…ã€å·²é…æœªæ£€
        # çä¹³åŒº/æå‰å¹²å¥¶/æ²»ç–—ï¼šåˆæ£€å­•ã€å¤æ£€å­•ã€å¹²å¥¶ã€å¦Šå¨ 
        if method_key in ['cull', 'isolate']:
            default_statuses = ['äº§çŠŠ', 'ç¦é…', 'å¯é…', 'å·²é…', 'äº§åæœªé…', 'åˆæ£€ç©ºæ€€', 'å‘æƒ…æœªé…', 'æµäº§æœªé…', 'å·²é…æœªæ£€']
        else:
            default_statuses = ['åˆæ£€å­•', 'å¤æ£€å­•', 'å¹²å¥¶', 'å¦Šå¨ ']

        widget.breeding_checkboxes = {}

        for i, status in enumerate(default_statuses):
            cb = QCheckBox(status)
            cb.setChecked(True)  # é»˜è®¤å…¨é€‰
            widget.breeding_checkboxes[status] = cb

            # è®¡ç®—è¡Œåˆ—ä½ç½®ï¼ˆæ¯è¡Œ3ä¸ªï¼‰
            row = i // 3
            col = i % 3
            breeding_status_layout.addWidget(cb, row, col)

        config_layout.addRow(breeding_status_label, breeding_status_widget)
        
        # å¯ç”¨/ç¦ç”¨é…ç½®åŒºåŸŸ
        enabled_cb.toggled.connect(lambda checked: config_widget.setVisible(checked))
        config_widget.setVisible(True)  # é»˜è®¤æ˜¾ç¤º
        
        layout.addWidget(config_widget)
        
        # å­˜å‚¨å¼•ç”¨
        widget.enabled_cb = enabled_cb
        widget.config_widget = config_widget
        
        return widget
    
    def update_mastitis_screen_button_state(self):
        """æ›´æ–°ç­›æŸ¥æŒ‰é’®çŠ¶æ€"""
        # æ£€æŸ¥ç³»ç»Ÿé€‰æ‹©å’Œæ–‡ä»¶ä¸Šä¼ çŠ¶æ€
        system_selected = self.current_mastitis_system is not None
        
        if not system_selected:
            self.mastitis_screen_btn.setEnabled(False)
            self.mastitis_status_label.setText("è¯·é€‰æ‹©æ•°æ®ç®¡ç†ç³»ç»Ÿ")
            return
        
        # æ£€æŸ¥æ–‡ä»¶ä¸Šä¼ çŠ¶æ€
        all_files_uploaded = True
        missing_files = []
        
        for file_key, widget in self.mastitis_file_uploads.items():
            if not hasattr(widget, 'file_path') or widget.file_path is None:
                all_files_uploaded = False
                missing_files.append(file_key)
        
        if not all_files_uploaded:
            self.mastitis_screen_btn.setEnabled(False)
            self.mastitis_status_label.setText(f"è¯·ä¸Šä¼ ç¼ºå¤±çš„æ–‡ä»¶: {', '.join(missing_files)}")
            return
        
        # æ£€æŸ¥DHIæ•°æ®æ˜¯å¦å·²ä¸Šä¼ 
        dhi_data_available = hasattr(self, 'data_list') and self.data_list
        
        if not dhi_data_available:
            self.mastitis_screen_btn.setEnabled(False)
            self.mastitis_status_label.setText("è¯·å…ˆåœ¨åŸºç¡€æ•°æ®æ ‡ç­¾é¡µä¸Šä¼ DHIæŠ¥å‘Š")
            return
        
        # æ‰€æœ‰æ¡ä»¶æ»¡è¶³
        self.mastitis_screen_btn.setEnabled(True)
        self.mastitis_status_label.setText("å‡†å¤‡å°±ç»ªï¼Œå¯ä»¥å¼€å§‹ç­›æŸ¥")
    
    def start_mastitis_screening(self):
        """å¼€å§‹æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥"""
        try:
            # æ¸…ç©ºå³ä¾§å¤„ç†è¿‡ç¨‹é¢æ¿å¹¶åˆ‡æ¢åˆ°è¯¥æ ‡ç­¾é¡µ
            self.process_log_widget.clear()
            self.tab_widget.setCurrentWidget(self.process_log_widget)
            
            # æ˜¾ç¤ºå¼€å§‹ä¿¡æ¯
            start_message = f"""
ğŸ¥ æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥å¼€å§‹
â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”

âš™ï¸ ç³»ç»Ÿç±»å‹: {self.current_mastitis_system}
ğŸ“… å¼€å§‹æ—¶é—´: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

ğŸ”„ æ­£åœ¨å¤„ç†æ•°æ®æ–‡ä»¶...
"""
            self.process_log_widget.append(start_message)
            
            # åˆ›å»ºè¿›åº¦å¯¹è¯æ¡†
            self.mastitis_progress_dialog = SmoothProgressDialog(
                "æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥",
                "å–æ¶ˆ",
                0, 100,
                self
            )
            self.mastitis_progress_dialog.setWindowTitle("æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥è¿›åº¦")
            self.mastitis_progress_dialog.show()
            
            # æ›´æ–°è¿›åº¦
            self.mastitis_progress_dialog.setLabelText("æ­¥éª¤ 1/8: æ”¶é›†æ–‡ä»¶è·¯å¾„...")
            self.mastitis_progress_dialog.setValue(5)
            QApplication.processEvents()
            
            # æ”¶é›†æ–‡ä»¶è·¯å¾„å’Œå­—æ®µæ˜ å°„
            file_paths = {}
            field_mappings = {}
            
            for file_key, widget in self.mastitis_file_uploads.items():
                file_paths[file_key] = widget.file_path
                
                # å¦‚æœæ˜¯è‡ªå®šä¹‰ç³»ç»Ÿï¼Œæ”¶é›†å­—æ®µæ˜ å°„
                if hasattr(widget, 'mapping_inputs'):
                    field_mappings[file_key] = {}
                    for field, input_widget in widget.mapping_inputs.items():
                        column_name = input_widget.text().strip()
                        if column_name:
                            field_mappings[file_key][field] = column_name
            
            # å¤„ç†ç³»ç»Ÿæ–‡ä»¶
            self.mastitis_progress_dialog.setValue(10)
            self.mastitis_progress_dialog.setLabelText("æ­¥éª¤ 2/8: å¤„ç†ç³»ç»Ÿæ–‡ä»¶...")
            self.process_log_widget.append("ğŸ“‚ å¼€å§‹å¤„ç†ç³»ç»Ÿæ–‡ä»¶...")
            QApplication.processEvents()
            
            success, message, processed_data = self.data_processor.process_mastitis_system_files(
                self.current_mastitis_system, file_paths, field_mappings
            )
            
            if not success:
                error_msg = f"âŒ æ–‡ä»¶å¤„ç†å¤±è´¥: {message}"
                self.process_log_widget.append(error_msg)
                QMessageBox.warning(self, "æ–‡ä»¶å¤„ç†å¤±è´¥", message)
                self.mastitis_progress_dialog.close()
                return
            
            self.process_log_widget.append(f"âœ… ç³»ç»Ÿæ–‡ä»¶å¤„ç†æˆåŠŸ: {message}")
            
            # ä¿å­˜ç‰›ç¾¤åŸºç¡€ä¿¡æ¯åˆ°ä¸»çª—å£ï¼Œä¾›ç›‘æµ‹åŠŸèƒ½ä½¿ç”¨
            self.cattle_basic_info = processed_data['cattle_info']
            self.current_system = self.current_mastitis_system
            print(f"ğŸ” ç‰›ç¾¤åŸºç¡€ä¿¡æ¯å·²ä¿å­˜åˆ°ä¸»çª—å£: {len(self.cattle_basic_info)}å¤´ç‰›")
            print(f"ğŸ” ç³»ç»Ÿç±»å‹å·²ä¿å­˜: {self.current_system}")
            
            # æ›´æ–°éšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹çš„æ•°æ®çŠ¶æ€æ˜¾ç¤º
            if hasattr(self, 'update_monitoring_data_status'):
                self.update_monitoring_data_status()
            
            self.mastitis_progress_dialog.setValue(30)
            self.mastitis_progress_dialog.setLabelText("æ­¥éª¤ 3/8: è®¡ç®—æœ€è¿‘7å¤©å¥¶é‡...")
            self.process_log_widget.append("ğŸ§® æ­£åœ¨è®¡ç®—å…³é”®æŒ‡æ ‡...")
            QApplication.processEvents()
            
            # è®¡ç®—æœ€è¿‘7å¤©å¹³å‡å¥¶é‡ï¼ˆä»…ä¼Šèµ·ç‰›ç³»ç»Ÿéœ€è¦ï¼‰
            if self.current_mastitis_system == 'yiqiniu':
                self.process_log_widget.append("ğŸ¥› è®¡ç®—æœ€è¿‘7å¤©å¹³å‡å¥¶é‡...")
                milk_yield_df = self.data_processor.calculate_recent_7day_avg_yield(processed_data['milk_yield'])
                # åˆå¹¶åˆ°ç‰›ç¾¤ä¿¡æ¯ä¸­
                processed_data['cattle_info'] = processed_data['cattle_info'].merge(
                    milk_yield_df, on='ear_tag', how='left'
                )
                self.process_log_widget.append(f"âœ… å®Œæˆ{len(milk_yield_df)}å¤´ç‰›çš„å¥¶é‡è®¡ç®—")
            
            self.mastitis_progress_dialog.setValue(50)
            self.mastitis_progress_dialog.setLabelText("æ­¥éª¤ 4/8: ç»Ÿè®¡ä¹³æˆ¿ç‚å‘ç—…...")
            QApplication.processEvents()
            
            # è®¡ç®—ä¹³æˆ¿ç‚å‘ç—…æ¬¡æ•°
            self.process_log_widget.append("ğŸ¦  è®¡ç®—ä¹³æˆ¿ç‚å‘ç—…æ¬¡æ•°...")
            mastitis_count_df = self.data_processor.calculate_mastitis_count_per_lactation(
                processed_data['cattle_info'], processed_data['disease']
            )
            
            # åˆå¹¶åˆ°ç‰›ç¾¤ä¿¡æ¯ä¸­
            processed_data['cattle_info'] = processed_data['cattle_info'].merge(
                mastitis_count_df, on='ear_tag', how='left'
            )
            
            affected_cows = len(mastitis_count_df[mastitis_count_df['mastitis_count'] > 0])
            total_cases = mastitis_count_df['mastitis_count'].sum()
            self.process_log_widget.append(f"âœ… å‘ç—…ç»Ÿè®¡å®Œæˆ: {affected_cows}å¤´ç‰›å‘ç—…ï¼Œå…±{total_cases}æ¬¡")
            
            self.mastitis_progress_dialog.setValue(70)
            self.mastitis_progress_dialog.setLabelText("æ­¥éª¤ 6/8: è¯†åˆ«æ…¢æ€§æ„ŸæŸ“ç‰›...")
            QApplication.processEvents()
            self.process_log_widget.append("ğŸ”¬ è¯†åˆ«æ…¢æ€§æ„ŸæŸ“ç‰›...")
            
            # æ”¶é›†é€‰ä¸­çš„æœˆä»½
            selected_months = [month for month, cb in self.chronic_month_checkboxes.items() if cb.isChecked()]
            scc_operator = self.scc_threshold_combo.currentText()
            scc_threshold = self.scc_threshold_spin.value()
            
            if not selected_months:
                error_msg = "âŒ è¯·è‡³å°‘é€‰æ‹©ä¸€ä¸ªæœˆä»½è¿›è¡Œæ…¢æ€§æ„ŸæŸ“ç‰›è¯†åˆ«"
                self.process_log_widget.append(error_msg)
                self.show_warning("æœˆä»½é€‰æ‹©é”™è¯¯", "è¯·è‡³å°‘é€‰æ‹©ä¸€ä¸ªæœˆä»½è¿›è¡Œæ…¢æ€§æ„ŸæŸ“ç‰›è¯†åˆ«")
                self.mastitis_progress_dialog.close()
                return
            
            self.process_log_widget.append(f"ğŸ—“ï¸ æ£€æŸ¥æœˆä»½: {', '.join(selected_months)}")
            self.process_log_widget.append(f"ğŸ”¢ ä½“ç»†èƒæ•°æ¡ä»¶: {scc_operator} {scc_threshold}ä¸‡/ml")
            
            # è¯†åˆ«æ…¢æ€§æ„ŸæŸ“ç‰›
            chronic_mastitis_df = self.data_processor.identify_chronic_mastitis_cows(
                self.data_list,
                selected_months,
                scc_threshold,
                scc_operator
            )
            
            chronic_count = len(chronic_mastitis_df[chronic_mastitis_df['chronic_mastitis']])
            self.process_log_widget.append(f"âœ… æ…¢æ€§æ„ŸæŸ“è¯†åˆ«å®Œæˆ: {chronic_count}å¤´ç‰›è¢«è¯†åˆ«ä¸ºæ…¢æ€§æ„ŸæŸ“")
            
            # å°†æ…¢æ€§æ„ŸæŸ“ç»“æœåˆå¹¶åˆ°åŸºç¡€æ•°æ®ä¸­
            if not chronic_mastitis_df.empty:
                # ç¡®å®šåˆå¹¶å­—æ®µ
                cattle_info_columns = processed_data['cattle_info'].columns
                chronic_columns = chronic_mastitis_df.columns
                
                if 'management_id' in cattle_info_columns and 'management_id' in chronic_columns:
                    merge_key = 'management_id'
                elif 'ear_tag' in cattle_info_columns and 'ear_tag' in chronic_columns:
                    merge_key = 'ear_tag'
                else:
                    # å¦‚æœæ²¡æœ‰ç›´æ¥åŒ¹é…çš„å­—æ®µï¼Œå°è¯•åˆ›å»ºåŒ¹é…å­—æ®µ
                    if 'ear_tag' in cattle_info_columns and 'management_id' in chronic_columns:
                        # ç‰›ç¾¤ä¿¡æ¯ç”¨ear_tagï¼Œæ…¢æ€§æ„ŸæŸ“ç»“æœç”¨management_idï¼Œå°è¯•åŒ¹é…
                        merge_key = 'ear_tag'  # ä½¿ç”¨ear_tagä½œä¸ºä¸»é”®
                        chronic_mastitis_df['ear_tag'] = chronic_mastitis_df['management_id']
                    else:
                        self.process_log_widget.append("âŒ æ— æ³•æ‰¾åˆ°åˆé€‚çš„å­—æ®µåˆå¹¶æ…¢æ€§æ„ŸæŸ“ç»“æœ")
                        processed_data['cattle_info']['chronic_mastitis'] = False
                        merge_key = None
                
                if merge_key:
                    processed_data['cattle_info'] = processed_data['cattle_info'].merge(
                        chronic_mastitis_df, 
                        left_on=merge_key, 
                        right_on=merge_key, 
                        how='left'
                    )
                    # å¡«å……ç¼ºå¤±å€¼ä¸ºFalseï¼ˆéæ…¢æ€§æ„ŸæŸ“ï¼‰
                    processed_data['cattle_info']['chronic_mastitis'] = processed_data['cattle_info']['chronic_mastitis'].fillna(False)
                    self.process_log_widget.append(f"âœ… æ…¢æ€§æ„ŸæŸ“ç»“æœå·²ä½¿ç”¨{merge_key}å­—æ®µåˆå¹¶åˆ°åŸºç¡€æ•°æ®ä¸­")
            else:
                # å¦‚æœæ²¡æœ‰æ…¢æ€§æ„ŸæŸ“ç‰›ï¼Œæ‰€æœ‰ç‰›çš„chronic_mastitiséƒ½è®¾ä¸ºFalse
                processed_data['cattle_info']['chronic_mastitis'] = False
                self.process_log_widget.append("â„¹ï¸ æœªå‘ç°æ…¢æ€§æ„ŸæŸ“ç‰›ï¼Œæ‰€æœ‰ç‰›çš„chronic_mastitisè®¾ä¸ºFalse")
            
            self.mastitis_progress_dialog.setValue(85)
            self.mastitis_progress_dialog.setLabelText("æ­¥éª¤ 7/8: åº”ç”¨å¤„ç½®åŠæ³•...")
            self.process_log_widget.append("âš–ï¸ åº”ç”¨å¤„ç½®åŠæ³•åˆ¤æ–­...")
            QApplication.processEvents()
            
            # æ”¶é›†å¤„ç½®åŠæ³•é…ç½®
            treatment_config = self.build_treatment_config()
            enabled_treatments = [k for k, v in treatment_config.items() if v.get('enabled', False)]
            self.process_log_widget.append(f"ğŸ“‹ å¯ç”¨çš„å¤„ç½®åŠæ³•: {', '.join(enabled_treatments)}")
            
            # åº”ç”¨å¤„ç½®åŠæ³•åˆ¤æ–­ï¼ˆåªå¯¹æ…¢æ€§æ„ŸæŸ“ç‰›è¿›è¡Œåˆ¤æ–­ï¼‰
            final_results = self.data_processor.apply_treatment_decisions(
                processed_data['cattle_info'], treatment_config
            )
            
            self.mastitis_progress_dialog.setValue(95)
            self.mastitis_progress_dialog.setLabelText("æ­¥éª¤ 8/8: ç”Ÿæˆç­›æŸ¥æŠ¥å‘Š...")
            self.process_log_widget.append("ğŸ“Š ç”Ÿæˆç­›æŸ¥æŠ¥å‘Š...")
            QApplication.processEvents()
            
            # ç”Ÿæˆç­›æŸ¥æŠ¥å‘Š
            screening_report = self.data_processor.create_mastitis_screening_report(
                final_results, 
                selected_months, 
                self.data_list
            )
            
            self.mastitis_progress_dialog.setValue(100)
            self.mastitis_progress_dialog.setLabelText("ç­›æŸ¥å®Œæˆï¼")
            # å»¶è¿Ÿå…³é—­è¿›åº¦å¯¹è¯æ¡†
            QTimer.singleShot(2000, lambda: self.mastitis_progress_dialog.close())
            
            completion_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            if not screening_report.empty:
                self.mastitis_screening_results = screening_report
                self.mastitis_export_btn.setEnabled(True)
                result_message = f"âœ… ç­›æŸ¥å®Œæˆï¼å‘ç°{len(screening_report)}å¤´ç‰›éœ€è¦å¤„ç½®"
                self.mastitis_status_label.setText(result_message)
                
                self.process_log_widget.append(f"""
{result_message}
ğŸ“… å®Œæˆæ—¶é—´: {completion_time}
ğŸ“Š ç­›æŸ¥ç»“æœå·²æ˜¾ç¤ºåœ¨å³ä¾§"ç­›é€‰ç»“æœ"æ ‡ç­¾é¡µ

â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”
ğŸ¯ æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥ä»»åŠ¡å®Œæˆ
""")
                
                # æ˜¾ç¤ºç»“æœåˆ°å³ä¾§ç­›é€‰ç»“æœè¡¨æ ¼
                self.display_mastitis_results_in_table(screening_report)
            else:
                no_result_message = "âœ… ç­›æŸ¥å®Œæˆï¼Œæœªå‘ç°éœ€è¦å¤„ç½®çš„ç‰›åª"
                self.mastitis_status_label.setText(no_result_message)
                self.process_log_widget.append(f"""
{no_result_message}
ğŸ“… å®Œæˆæ—¶é—´: {completion_time}

â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”
ğŸ¯ æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥ä»»åŠ¡å®Œæˆ
""")
            
        except Exception as e:
            error_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            error_message = f"""
âŒ ç­›æŸ¥è¿‡ç¨‹ä¸­å‡ºç°é”™è¯¯
ğŸ“… é”™è¯¯æ—¶é—´: {error_time}
ğŸ” é”™è¯¯è¯¦æƒ…: {str(e)}

â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”
âŒ æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥ä»»åŠ¡å¤±è´¥
"""
            self.process_log_widget.append(error_message)
            self.mastitis_progress_dialog.close()
            QMessageBox.critical(self, "ç­›æŸ¥å¤±è´¥", f"ç­›æŸ¥è¿‡ç¨‹ä¸­å‡ºç°é”™è¯¯ï¼š{str(e)}")
            self.mastitis_status_label.setText("ç­›æŸ¥å¤±è´¥")
    
    def display_mastitis_results_in_table(self, results_df):
        """å°†æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥ç»“æœæ˜¾ç¤ºåˆ°æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥ç»“æœè¡¨æ ¼"""
        try:
            # åˆ‡æ¢åˆ°ç­›é€‰ç»“æœæ ‡ç­¾é¡µï¼Œç„¶ååˆ‡æ¢åˆ°æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥å­æ ‡ç­¾é¡µ
            self.tab_widget.setCurrentWidget(self.result_widget)
            # åˆ‡æ¢åˆ°æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥å­æ ‡ç­¾é¡µ
            for i in range(self.result_sub_tabs.count()):
                if "æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥" in self.result_sub_tabs.tabText(i):
                    self.result_sub_tabs.setCurrentIndex(i)
                    break
            
            # è®¾ç½®è¡¨æ ¼è¡Œåˆ—æ•°
            self.mastitis_screening_table.setRowCount(len(results_df))
            self.mastitis_screening_table.setColumnCount(len(results_df.columns))
            
            # è®¾ç½®è¡¨å¤´
            self.mastitis_screening_table.setHorizontalHeaderLabels(results_df.columns.tolist())
            
            # å¡«å……æ•°æ®
            for i in range(len(results_df)):
                for j, value in enumerate(results_df.iloc[i]):
                    item = QTableWidgetItem(str(value) if pd.notna(value) else "")
                    
                    # ä¸ºä¸åŒçš„å¤„ç½®åŠæ³•è®¾ç½®ä¸åŒçš„èƒŒæ™¯è‰²
                    if j == results_df.columns.get_loc('æ¨èå¤„ç½®åŠæ³•') if 'æ¨èå¤„ç½®åŠæ³•' in results_df.columns else -1:
                        if 'æ·˜æ±°' in str(value):
                            item.setBackground(QColor(255, 235, 238))  # æ·¡çº¢è‰²
                        elif 'ç¦é…éš”ç¦»' in str(value):
                            item.setBackground(QColor(255, 243, 205))  # æ·¡æ©™è‰²
                        elif 'çä¹³åŒº' in str(value):
                            item.setBackground(QColor(217, 237, 247))  # æ·¡è“è‰²
                        elif 'æå‰å¹²å¥¶' in str(value):
                            item.setBackground(QColor(230, 247, 236))  # æ·¡ç»¿è‰²
                        elif 'æ²»ç–—' in str(value):
                            item.setBackground(QColor(248, 249, 250))  # æ·¡ç°è‰²
                    
                    self.mastitis_screening_table.setItem(i, j, item)
            
            # è°ƒæ•´åˆ—å®½
            self.mastitis_screening_table.resizeColumnsToContents()
            
            # é™åˆ¶åˆ—å®½æœ€å¤§å€¼
            for col in range(self.mastitis_screening_table.columnCount()):
                if self.mastitis_screening_table.columnWidth(col) > 200:
                    self.mastitis_screening_table.setColumnWidth(col, 200)
            
            # å¯ç”¨æ’åº
            self.mastitis_screening_table.setSortingEnabled(True)
            
            # åœ¨å¤„ç†è¿‡ç¨‹ä¸­æ·»åŠ ç»“æœè¯´æ˜
            result_summary = f"""
ğŸ“Š ç»“æœå·²æ˜¾ç¤ºåœ¨æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥ç»“æœè¡¨æ ¼ä¸­
â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”

ğŸ¯ ç­›æŸ¥ç»“æœæ‘˜è¦:
â€¢ æ€»è®¡: {len(results_df)} å¤´ç‰›éœ€è¦å¤„ç½®
â€¢ æ•°æ®åˆ—: {len(results_df.columns)} ä¸ªå­—æ®µ
â€¢ è¡¨æ ¼æ”¯æŒç‚¹å‡»åˆ—å¤´æ’åº

ğŸ’¡ å¤„ç½®åŠæ³•é¢œè‰²è¯´æ˜:
ğŸ”´ æ·˜æ±° - æ·¡çº¢è‰²èƒŒæ™¯
ğŸŸ  ç¦é…éš”ç¦» - æ·¡æ©™è‰²èƒŒæ™¯  
ğŸ”µ çä¹³åŒº - æ·¡è“è‰²èƒŒæ™¯
ğŸŸ¢ æå‰å¹²å¥¶ - æ·¡ç»¿è‰²èƒŒæ™¯
âšª æ²»ç–— - æ·¡ç°è‰²èƒŒæ™¯

ğŸ“ æŸ¥çœ‹ç»“æœï¼šç‚¹å‡»å³ä¾§"ç­›æŸ¥ç»“æœ"æ ‡ç­¾é¡µä¸­çš„"æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥"å­æ ‡ç­¾
"""
            self.process_log_widget.append(result_summary)
            
        except Exception as e:
            error_msg = f"âŒ æ˜¾ç¤ºç­›æŸ¥ç»“æœæ—¶å‡ºé”™: {str(e)}"
            self.process_log_widget.append(error_msg)
            print(f"æ˜¾ç¤ºç­›æŸ¥ç»“æœæ—¶å‡ºé”™: {e}")
    
    def build_treatment_config(self) -> Dict[str, Any]:
        """æ„å»ºå¤„ç½®åŠæ³•é…ç½®"""
        config = {}

        for method_key, widget in self.treatment_configs.items():
            if widget.enabled_cb.isChecked():
                method_config = {
                    'enabled': True,
                    'mastitis_operator': widget.mastitis_combo.currentText(),
                    'mastitis_value': widget.mastitis_spin.value(),
                    'lactation_operator': widget.lactation_combo.currentText(),
                    'lactation_value': widget.lactation_spin.value(),
                    'breeding_status': [status for status, cb in widget.breeding_checkboxes.items() if cb.isChecked()]
                }

                # æ·»åŠ äº§å¥¶é‡é…ç½®ï¼ˆæ·˜æ±°ã€ç¦é…éš”ç¦»ã€çä¹³åŒºã€æå‰å¹²å¥¶ã€æ²»ç–—éƒ½æœ‰ï¼‰
                if hasattr(widget, 'yield_combo'):
                    method_config['yield_operator'] = widget.yield_combo.currentText()
                    method_config['yield_value'] = widget.yield_spin.value()

                # æ·»åŠ åœ¨èƒå¤©æ•°é…ç½®ï¼ˆçä¹³åŒºã€æå‰å¹²å¥¶ã€æ²»ç–—æœ‰ï¼‰
                if hasattr(widget, 'gestation_combo'):
                    method_config['gestation_operator'] = widget.gestation_combo.currentText()
                    method_config['gestation_value'] = widget.gestation_spin.value()

                config[method_key] = method_config
            else:
                config[method_key] = {'enabled': False}

        return config
    
    def show_mastitis_results_preview(self, results_df):
        """æ˜¾ç¤ºç­›æŸ¥ç»“æœé¢„è§ˆ"""
        if results_df.empty:
            return
        
        # åˆ›å»ºé¢„è§ˆå¯¹è¯æ¡†
        dialog = QDialog(self)
        dialog.setWindowTitle("æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥ç»“æœé¢„è§ˆ")
        dialog.resize(1000, 600)
        
        layout = QVBoxLayout(dialog)
        
        # ç»Ÿè®¡ä¿¡æ¯
        stats_label = QLabel(f"å…±å‘ç° {len(results_df)} å¤´ç‰›éœ€è¦å¤„ç½®")
        stats_label.setStyleSheet("font-weight: bold; font-size: 16px; padding: 10px;")
        layout.addWidget(stats_label)
        
        # ç»“æœè¡¨æ ¼
        table = QTableWidget()
        table.setRowCount(min(len(results_df), 100))  # æœ€å¤šæ˜¾ç¤º100è¡Œ
        table.setColumnCount(len(results_df.columns))
        table.setHorizontalHeaderLabels(results_df.columns.tolist())
        
        # å¡«å……æ•°æ®
        for i in range(min(len(results_df), 100)):
            for j, value in enumerate(results_df.iloc[i]):
                item = QTableWidgetItem(str(value) if pd.notna(value) else "")
                table.setItem(i, j, item)
        
        # è‡ªé€‚åº”åˆ—å®½
        table.resizeColumnsToContents()
        layout.addWidget(table)
        
        # æŒ‰é’®
        button_layout = QHBoxLayout()
        close_btn = QPushButton("å…³é—­")
        close_btn.clicked.connect(dialog.accept)
        export_btn = QPushButton("å¯¼å‡ºå®Œæ•´ç»“æœ")
        export_btn.clicked.connect(lambda: [dialog.accept(), self.export_mastitis_results()])
        
        button_layout.addStretch()
        button_layout.addWidget(close_btn)
        button_layout.addWidget(export_btn)
        layout.addLayout(button_layout)
        
        dialog.exec()
    
    def export_mastitis_results(self):
        """å¯¼å‡ºæ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥ç»“æœ"""
        if self.mastitis_screening_results is None or self.mastitis_screening_results.empty:
            QMessageBox.warning(self, "å¯¼å‡ºå¤±è´¥", "æ²¡æœ‰ç­›æŸ¥ç»“æœå¯ä»¥å¯¼å‡º")
            return
        
        # é€‰æ‹©ä¿å­˜è·¯å¾„
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥ç»“æœ_{timestamp}.xlsx"
        
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getSaveFileName(
            self, "ä¿å­˜ç­›æŸ¥ç»“æœ", default_filename, 
            "Excelæ–‡ä»¶ (*.xlsx);;æ‰€æœ‰æ–‡ä»¶ (*)"
        )
        
        if file_path:
            try:
                success = self.data_processor.export_mastitis_screening_results(
                    self.mastitis_screening_results, file_path
                )
                
                if success:
                    self.show_export_success_dialog("ç­›æŸ¥ç»“æœå·²ä¿å­˜åˆ°ï¼š", file_path)
                else:
                    QMessageBox.warning(self, "å¯¼å‡ºå¤±è´¥", "å¯¼å‡ºè¿‡ç¨‹ä¸­å‡ºç°é”™è¯¯")
                    
            except Exception as e:
                QMessageBox.critical(self, "å¯¼å‡ºå¤±è´¥", f"å¯¼å‡ºæ—¶å‡ºç°é”™è¯¯ï¼š{str(e)}")

    def show_export_success_dialog(self, message: str, file_path: str):
        """æ˜¾ç¤ºå¯¼å‡ºæˆåŠŸå¯¹è¯æ¡†ï¼ŒåŒ…å«æ‰“å¼€æ–‡ä»¶å’Œæ‰“å¼€æ–‡ä»¶å¤¹æŒ‰é’®"""
        dialog = QDialog(self)
        dialog.setWindowTitle("å¯¼å‡ºæˆåŠŸ")
        dialog.setFixedSize(500, 200)
        dialog.setWindowIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_MessageBoxInformation))
        
        # ä¸»å¸ƒå±€
        layout = QVBoxLayout(dialog)
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)
        
        # æˆåŠŸå›¾æ ‡å’Œæ¶ˆæ¯
        message_layout = QHBoxLayout()
        
        # æˆåŠŸå›¾æ ‡
        icon_label = QLabel()
        icon_pixmap = self.style().standardIcon(QStyle.StandardPixmap.SP_MessageBoxInformation).pixmap(48, 48)
        icon_label.setPixmap(icon_pixmap)
        icon_label.setAlignment(Qt.AlignmentFlag.AlignTop)
        
        # æ¶ˆæ¯æ–‡æœ¬
        message_label = QLabel(f"{message}\n{file_path}")
        message_label.setWordWrap(True)
        message_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        message_label.setStyleSheet("font-size: 14px; color: #333333;")
        
        message_layout.addWidget(icon_label)
        message_layout.addWidget(message_label)
        message_layout.addStretch()
        
        layout.addLayout(message_layout)
        
        # æŒ‰é’®å¸ƒå±€
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        
        # æ‰“å¼€æ–‡ä»¶æŒ‰é’®
        open_file_btn = QPushButton("ğŸ“„ æ‰“å¼€æ–‡ä»¶")
        open_file_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 6px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
        """)
        open_file_btn.clicked.connect(lambda: self.open_file(file_path))
        
        # æ‰“å¼€æ–‡ä»¶å¤¹æŒ‰é’®
        open_folder_btn = QPushButton("ğŸ“ æ‰“å¼€æ–‡ä»¶å¤¹")
        open_folder_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 6px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:pressed {
                background-color: #1565C0;
            }
        """)
        open_folder_btn.clicked.connect(lambda: self.open_file_folder(file_path))
        
        # ç¡®å®šæŒ‰é’®
        ok_btn = QPushButton("ç¡®å®š")
        ok_btn.setStyleSheet("""
            QPushButton {
                background-color: #757575;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 6px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #616161;
            }
            QPushButton:pressed {
                background-color: #424242;
            }
        """)
        ok_btn.clicked.connect(dialog.accept)
        
        # æ·»åŠ æŒ‰é’®åˆ°å¸ƒå±€
        button_layout.addWidget(open_file_btn)
        button_layout.addWidget(open_folder_btn)
        button_layout.addStretch()
        button_layout.addWidget(ok_btn)
        
        layout.addLayout(button_layout)
        
        # è®¾ç½®å¯¹è¯æ¡†æ ·å¼
        dialog.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 8px;
            }
            QLabel {
                background-color: transparent;
            }
        """)
        
        # æ˜¾ç¤ºå¯¹è¯æ¡†
        dialog.exec()
    
    def start_mastitis_monitoring(self):
        """å¯åŠ¨éšå½¢ä¹³æˆ¿ç‚æœˆåº¦ç›‘æµ‹åˆ†æ"""
        try:
            print(f"\nğŸ” [è¯¦ç»†è°ƒè¯•] å¼€å§‹éšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹åˆ†æ...")
            
            # è¯¦ç»†æ£€æŸ¥ç³»ç»ŸçŠ¶æ€
            print(f"ğŸ” [è¯¦ç»†è°ƒè¯•] æ£€æŸ¥DHIæ•°æ®å¯ç”¨æ€§...")
            print(f"   hasattr(self, 'data_list'): {hasattr(self, 'data_list')}")
            if hasattr(self, 'data_list'):
                print(f"   self.data_list is not None: {self.data_list is not None}")
                if self.data_list:
                    print(f"   DHIæ•°æ®æ–‡ä»¶æ•°é‡: {len(self.data_list)}")
                    for i, item in enumerate(self.data_list):
                        print(f"     æ–‡ä»¶{i+1}: {item.get('filename', 'Unknown')} - æ•°æ®è¡Œæ•°: {len(item['data']) if item.get('data') is not None else 0}")
            
            print(f"ğŸ” [è¯¦ç»†è°ƒè¯•] æ£€æŸ¥ç‰›ç¾¤åŸºç¡€ä¿¡æ¯...")
            print(f"   hasattr(self, 'cattle_basic_info'): {hasattr(self, 'cattle_basic_info')}")
            if hasattr(self, 'cattle_basic_info'):
                print(f"   self.cattle_basic_info is not None: {self.cattle_basic_info is not None}")
                if self.cattle_basic_info is not None:
                    print(f"   ç‰›ç¾¤åŸºç¡€ä¿¡æ¯æ•°é‡: {len(self.cattle_basic_info)}")
                    print(f"   ç‰›ç¾¤æ•°æ®åˆ—å: {list(self.cattle_basic_info.columns)}")
                    print(f"   ç³»ç»Ÿç±»å‹: {getattr(self, 'current_system', 'Unknown')}")
                    # æ˜¾ç¤ºå‰å‡ å¤´ç‰›çš„ä¿¡æ¯
                    if len(self.cattle_basic_info) > 0:
                        print(f"   å‰3å¤´ç‰›ç¤ºä¾‹:")
                        for i in range(min(3, len(self.cattle_basic_info))):
                            cow_data = self.cattle_basic_info.iloc[i]
                            print(f"     ç‰›{i+1}: è€³å·={cow_data.get('ear_tag', 'N/A')}, åœ¨èƒå¤©æ•°={cow_data.get('gestation_days', 'N/A')}")
            
            print(f"ğŸ” [è¯¦ç»†è°ƒè¯•] æ£€æŸ¥ç›‘æµ‹è®¡ç®—å™¨...")
            print(f"   hasattr(self, 'mastitis_monitoring_calculator'): {hasattr(self, 'mastitis_monitoring_calculator')}")
            if hasattr(self, 'mastitis_monitoring_calculator'):
                print(f"   è®¡ç®—å™¨å¯¹è±¡: {self.mastitis_monitoring_calculator}")
                if self.mastitis_monitoring_calculator:
                    print(f"   è®¡ç®—å™¨ä¸­æ˜¯å¦æœ‰ç‰›ç¾¤æ•°æ®: {hasattr(self.mastitis_monitoring_calculator, 'cattle_basic_info')}")
                    if hasattr(self.mastitis_monitoring_calculator, 'cattle_basic_info'):
                        print(f"   è®¡ç®—å™¨ä¸­ç‰›ç¾¤æ•°æ®: {self.mastitis_monitoring_calculator.cattle_basic_info is not None}")
            
            # æ£€æŸ¥DHIæ•°æ®æ˜¯å¦å¯ç”¨
            if not hasattr(self, 'data_list') or not self.data_list:
                self.show_warning("è­¦å‘Š", "è¯·å…ˆåœ¨'DHIåŸºç¡€ç­›é€‰'æ ‡ç­¾é¡µä¸­ä¸Šä¼ DHIæ•°æ®")
                return
            
            # è·å–é˜ˆå€¼è®¾ç½®
            scc_threshold = self.monitoring_scc_threshold.value()
            
            # ä»ç›‘æµ‹è®¡ç®—æ¨¡å—å¯¼å…¥
            from mastitis_monitoring import MastitisMonitoringCalculator
            
            # åˆ›å»ºç›‘æµ‹è®¡ç®—å™¨
            self.mastitis_monitoring_calculator = MastitisMonitoringCalculator(scc_threshold=scc_threshold)
            
            # å‡†å¤‡DHIæ•°æ®
            dhi_data_list = []
            for item in self.data_list:
                if item['data'] is not None and not item['data'].empty:
                    dhi_data_list.append(item['data'])
            
            if len(dhi_data_list) == 0:
                QMessageBox.warning(self, "è­¦å‘Š", "æ²¡æœ‰å¯ç”¨çš„DHIæ•°æ®è¿›è¡Œåˆ†æ")
                return
            
            # åŠ è½½DHIæ•°æ®
            load_result = self.mastitis_monitoring_calculator.load_dhi_data(dhi_data_list)
            
            if not load_result['success']:
                QMessageBox.critical(self, "é”™è¯¯", f"DHIæ•°æ®åŠ è½½å¤±è´¥: {load_result.get('error', 'æœªçŸ¥é”™è¯¯')}")
                return
            
            # è‡ªåŠ¨åŠ è½½æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥ä¸­çš„ç‰›ç¾¤åŸºç¡€ä¿¡æ¯
            print(f"\nğŸ” æ£€æŸ¥ç‰›ç¾¤åŸºç¡€ä¿¡æ¯...")
            print(f"   hasattr(self, 'cattle_basic_info'): {hasattr(self, 'cattle_basic_info')}")
            if hasattr(self, 'cattle_basic_info'):
                print(f"   self.cattle_basic_info is not None: {self.cattle_basic_info is not None}")
                if self.cattle_basic_info is not None:
                    print(f"   ç‰›ç¾¤åŸºç¡€ä¿¡æ¯æ•°é‡: {len(self.cattle_basic_info)}")
                    print(f"   ç³»ç»Ÿç±»å‹: {getattr(self, 'current_system', 'Unknown')}")
            
            if hasattr(self, 'cattle_basic_info') and self.cattle_basic_info is not None:
                print(f"   âœ… å‘ç°æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥ä¸­çš„ç‰›ç¾¤æ•°æ®ï¼Œè‡ªåŠ¨åŠ è½½åˆ°ç›‘æµ‹è®¡ç®—å™¨...")
                print(f"   ç‰›ç¾¤æ•°æ®è¯¦æƒ…: {len(self.cattle_basic_info)}å¤´ç‰›, ç³»ç»Ÿç±»å‹: {getattr(self, 'current_system', 'Unknown')}")
                print(f"   ç‰›ç¾¤æ•°æ®åˆ—: {list(self.cattle_basic_info.columns)}")
                
                cattle_result = self.mastitis_monitoring_calculator.load_cattle_basic_info(
                    self.cattle_basic_info, self.current_system)
                    
                print(f"   åŠ è½½ç»“æœ: {cattle_result}")
                
                if not cattle_result['success']:
                    print(f"   âŒ ç‰›ç¾¤åŸºç¡€ä¿¡æ¯åŠ è½½å¤±è´¥: {cattle_result.get('error')}")
                    self.show_warning("æç¤º", f"ç‰›ç¾¤åŸºç¡€ä¿¡æ¯åŠ è½½å¤±è´¥: {cattle_result.get('error', 'æœªçŸ¥é”™è¯¯')}\nå°†æ— æ³•è®¡ç®—å¹²å¥¶å‰æµè¡Œç‡")
                else:
                    print(f"   âœ… ç‰›ç¾¤åŸºç¡€ä¿¡æ¯åŠ è½½æˆåŠŸï¼Œå¯è®¡ç®—å¹²å¥¶å‰æµè¡Œç‡")
                    print(f"   åŠ è½½è¯¦æƒ…: {cattle_result.get('message', 'æ— è¯¦æƒ…')}")
                    # æ›´æ–°çŠ¶æ€æ˜¾ç¤º
                    self.update_monitoring_data_status()
            else:
                print(f"   âŒ è·³è¿‡ç‰›ç¾¤åŸºç¡€ä¿¡æ¯åŠ è½½ï¼šæ•°æ®ä¸å­˜åœ¨")
                print(f"   ğŸ’¡ æç¤ºï¼šå¦‚éœ€è®¡ç®—å¹²å¥¶å‰æµè¡Œç‡ï¼Œè¯·å…ˆåˆ°'æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥'ä¸­ä¸Šä¼ ç‰›ç¾¤åŸºç¡€ä¿¡æ¯")
            
            # å®šä¹‰å¼‚æ­¥è®¡ç®—ä»»åŠ¡
            def calculate_task(progress_callback=None, status_callback=None, check_cancelled=None):
                if status_callback:
                    status_callback("æ­£åœ¨è®¡ç®—å„é¡¹æŒ‡æ ‡...")
                
                if progress_callback:
                    progress_callback(30)
                
                results = self.mastitis_monitoring_calculator.calculate_all_indicators()
                
                if progress_callback:
                    progress_callback(100)
                    
                return results
            
            # æ‰§è¡Œè®¡ç®—
            self.start_monitoring_btn.setText("è®¡ç®—ä¸­...")
            self.start_monitoring_btn.setEnabled(False)
            
            # ä½¿ç”¨å¼‚æ­¥è¿›åº¦ç®¡ç†å™¨æ‰§è¡Œè®¡ç®—
            progress_manager = AsyncProgressManager(self)
            results = progress_manager.execute_with_progress(
                calculate_task,
                title="éšæ€§ä¹³æˆ¿ç‚ç›‘æµ‹åˆ†æ",
                cancel_text="å–æ¶ˆ",
                total_steps=100
            )
            
            if not results['success']:
                QMessageBox.critical(self, "é”™è¯¯", f"æŒ‡æ ‡è®¡ç®—å¤±è´¥: {results.get('error', 'æœªçŸ¥é”™è¯¯')}")
                self.start_monitoring_btn.setText("å¼€å§‹åˆ†æ")
                self.start_monitoring_btn.setEnabled(True)
                return
            
            # ä¿å­˜ç»“æœ
            self.mastitis_monitoring_results = results
            
            # æ˜¾ç¤ºç»“æœ
            self.display_mastitis_monitoring_results(results)
            
            # å¯ç”¨å¯¼å‡ºæŒ‰é’®
            self.export_monitoring_btn.setEnabled(True)
            
            # é‡ç½®æŒ‰é’®
            self.start_monitoring_btn.setText("é‡æ–°åˆ†æ")
            self.start_monitoring_btn.setEnabled(True)
            
            QMessageBox.information(self, "å®Œæˆ", f"éšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹åˆ†æå®Œæˆï¼\nåˆ†æäº†{results['month_count']}ä¸ªæœˆä»½çš„æ•°æ®")
            
        except Exception as e:
            QMessageBox.critical(self, "é”™è¯¯", f"åˆ†æè¿‡ç¨‹ä¸­å‘ç”Ÿé”™è¯¯: {str(e)}")
            logger.error(f"éšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹åˆ†æå¤±è´¥: {e}")
            self.start_monitoring_btn.setText("å¼€å§‹åˆ†æ")
            self.start_monitoring_btn.setEnabled(True)
    
    def display_mastitis_monitoring_results(self, results):
        """æ˜¾ç¤ºéšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹ç»“æœ"""
        try:
            # æ›´æ–°è¡¨æ ¼
            self.update_monitoring_table(results)
            
            # æ›´æ–°å›¾è¡¨
            self.update_monitoring_chart(results)
            
        except Exception as e:
            logger.error(f"æ˜¾ç¤ºç›‘æµ‹ç»“æœå¤±è´¥: {e}")
            QMessageBox.warning(self, "è­¦å‘Š", f"æ˜¾ç¤ºç»“æœå¤±è´¥: {str(e)}")
    
    def update_monitoring_table(self, results):
        """æ›´æ–°ç›‘æµ‹ç»“æœè¡¨æ ¼"""
        try:
            months = results['months']
            indicators = results['indicators']
            
            # å®šä¹‰è¡¨æ ¼åˆ—
            columns = [
                'æœˆä»½', 'å½“æœˆæµè¡Œç‡(%)', 'æ–°å‘æ„ŸæŸ“ç‡(%)', 'æ…¢æ€§æ„ŸæŸ“ç‡(%)', 
                'æ…¢æ€§æ„ŸæŸ“ç‰›å æ¯”(%)', 'å¤´èƒé¦–æµ‹æµè¡Œç‡(%)', 'ç»äº§é¦–æµ‹æµè¡Œç‡(%)', 'å¹²å¥¶å‰æµè¡Œç‡(%)'
            ]
            
            self.mastitis_monitoring_table.setColumnCount(len(columns))
            self.mastitis_monitoring_table.setHorizontalHeaderLabels(columns)
            self.mastitis_monitoring_table.setRowCount(len(months))
            
            for row, month in enumerate(months):
                month_data = indicators.get(month, {})
                
                # æœˆä»½
                self.mastitis_monitoring_table.setItem(row, 0, QTableWidgetItem(month))
                
                # å½“æœˆæµè¡Œç‡
                cp = month_data.get('current_prevalence', {})
                cp_value = f"{cp['value']:.1f}" if cp.get('value') is not None else "N/A"
                cp_item = QTableWidgetItem(cp_value)
                if cp.get('value') is not None:
                    cp_item.setToolTip(cp.get('formula', ''))
                self.mastitis_monitoring_table.setItem(row, 1, cp_item)
                
                # æ–°å‘æ„ŸæŸ“ç‡
                nir = month_data.get('new_infection_rate', {})
                nir_value = f"{nir['value']:.1f}" if nir.get('value') is not None else "N/A"
                nir_item = QTableWidgetItem(nir_value)
                if nir.get('value') is not None:
                    tooltip = nir.get('formula', '')
                    if nir.get('warning'):
                        tooltip += f"\nâš ï¸ {nir['warning']}"
                    nir_item.setToolTip(tooltip)
                self.mastitis_monitoring_table.setItem(row, 2, nir_item)
                
                # æ…¢æ€§æ„ŸæŸ“ç‡
                cir = month_data.get('chronic_infection_rate', {})
                cir_value = f"{cir['value']:.1f}" if cir.get('value') is not None else "N/A"
                cir_item = QTableWidgetItem(cir_value)
                if cir.get('value') is not None:
                    tooltip = cir.get('formula', '')
                    if cir.get('warning'):
                        tooltip += f"\nâš ï¸ {cir['warning']}"
                    cir_item.setToolTip(tooltip)
                self.mastitis_monitoring_table.setItem(row, 3, cir_item)
                
                # æ…¢æ€§æ„ŸæŸ“ç‰›å æ¯”
                cip = month_data.get('chronic_infection_proportion', {})
                cip_value = f"{cip['value']:.1f}" if cip.get('value') is not None else "N/A"
                cip_item = QTableWidgetItem(cip_value)
                if cip.get('value') is not None:
                    tooltip = cip.get('formula', '')
                    if cip.get('warning'):
                        tooltip += f"\nâš ï¸ {cip['warning']}"
                    cip_item.setToolTip(tooltip)
                self.mastitis_monitoring_table.setItem(row, 4, cip_item)
                
                # å¤´èƒé¦–æµ‹æµè¡Œç‡
                ftp = month_data.get('first_test_prevalence', {})
                primi_value = "N/A"
                if ftp and 'primiparous' in ftp:
                    primi_data = ftp['primiparous']
                    primi_value = f"{primi_data['value']:.1f}" if primi_data.get('value') is not None else "N/A"
                primi_item = QTableWidgetItem(primi_value)
                if ftp and 'primiparous' in ftp:
                    primi_item.setToolTip(ftp['primiparous'].get('formula', ''))
                self.mastitis_monitoring_table.setItem(row, 5, primi_item)
                
                # ç»äº§é¦–æµ‹æµè¡Œç‡
                multi_value = "N/A"
                if ftp and 'multiparous' in ftp:
                    multi_data = ftp['multiparous']
                    multi_value = f"{multi_data['value']:.1f}" if multi_data.get('value') is not None else "N/A"
                multi_item = QTableWidgetItem(multi_value)
                if ftp and 'multiparous' in ftp:
                    multi_item.setToolTip(ftp['multiparous'].get('formula', ''))
                self.mastitis_monitoring_table.setItem(row, 6, multi_item)
                
                # å¹²å¥¶å‰æµè¡Œç‡ï¼ˆåªåœ¨æœ€æ–°æœˆä»½æ˜¾ç¤ºï¼‰
                pdp = month_data.get('pre_dry_prevalence', {})
                is_latest_month = (row == len(months) - 1)  # åˆ¤æ–­æ˜¯å¦ä¸ºæœ€æ–°æœˆä»½
                
                # è°ƒè¯•è¾“å‡º
                print(f"ğŸ” å¹²å¥¶å‰æµè¡Œç‡è°ƒè¯• - æœˆä»½: {month}")
                print(f"   æ˜¯å¦æœ€æ–°æœˆä»½: {is_latest_month}")
                print(f"   å¹²å¥¶å‰æµè¡Œç‡æ•°æ®: {pdp}")
                print(f"   æ•°å€¼: {pdp.get('value')}")
                print(f"   è¯Šæ–­: {pdp.get('diagnosis')}")
                
                if is_latest_month and pdp.get('value') is not None:
                    # æœ€æ–°æœˆä»½ä¸”æœ‰æ•°å€¼
                    pdp_value = f"{pdp['value']:.1f}"
                    pdp_item = QTableWidgetItem(pdp_value)
                    
                    # è®¾ç½®è¯¦ç»†çš„å·¥å…·æç¤ºï¼ŒåŒ…å«è¯Šæ–­ä¿¡æ¯
                    if pdp.get('formula'):
                        # å°†HTMLæ ‡ç­¾è½¬æ¢ä¸ºçº¯æ–‡æœ¬ç”¨äºå·¥å…·æç¤º
                        tooltip_text = pdp.get('formula', '').replace('<br/>', '\n').replace('ã€€', '  ')
                        # ç§»é™¤HTMLæ ‡ç­¾
                        import re
                        tooltip_text = re.sub(r'<[^>]+>', '', tooltip_text)
                        pdp_item.setToolTip(tooltip_text)
                    
                    # è®¾ç½®æˆåŠŸè®¡ç®—çš„é¢œè‰²
                    pdp_item.setBackground(QColor('#e8f5e8'))  # æµ…ç»¿è‰²
                    
                elif is_latest_month and pdp.get('formula'):
                    # æœ€æ–°æœˆä»½ä½†è®¡ç®—å¤±è´¥ï¼Œæ˜¾ç¤ºå…·ä½“é”™è¯¯
                    print(f"   ğŸ’¡ å¹²å¥¶å‰æµè¡Œç‡è®¡ç®—å¤±è´¥ï¼Œæ˜¾ç¤ºN/A")
                    pdp_value = "N/A"
                    pdp_item = QTableWidgetItem(pdp_value)
                    
                    # è®¾ç½®è¯¦ç»†çš„å·¥å…·æç¤ºï¼ŒåŒ…å«è¯Šæ–­ä¿¡æ¯
                    tooltip_text = pdp.get('formula', '').replace('<br/>', '\n').replace('ã€€', '  ')
                    # ç§»é™¤HTMLæ ‡ç­¾
                    import re
                    tooltip_text = re.sub(r'<[^>]+>', '', tooltip_text)
                    pdp_item.setToolTip(tooltip_text)
                    
                    # æ ¹æ®è¯Šæ–­ç»“æœè®¾ç½®ä¸åŒçš„é¢œè‰²
                    diagnosis = pdp.get('diagnosis', '')
                    if diagnosis in ['ç¼ºå°‘ç‰›ç¾¤åŸºç¡€ä¿¡æ¯', 'æ•°æ®æ— æ³•åŒ¹é…']:
                        pdp_item.setBackground(QColor('#ffebee'))  # æµ…çº¢è‰²
                        pdp_item.setForeground(QColor('black'))  # é»‘è‰²å­—ä½“
                    elif diagnosis in ['ç¼ºå°‘åœ¨èƒå¤©æ•°å­—æ®µ', 'åŒ¹é…ç‰›åªæ— åœ¨èƒå¤©æ•°æ•°æ®', 'æ— ç¬¦åˆå¹²å¥¶å‰æ¡ä»¶çš„ç‰›åª']:
                        pdp_item.setBackground(QColor('#fff3e0'))  # æµ…æ©™è‰²
                        pdp_item.setForeground(QColor('black'))  # é»‘è‰²å­—ä½“
                else:
                    # éæœ€æ–°æœˆä»½ï¼Œæ˜¾ç¤º"-"
                    print(f"   ğŸ’¡ æ˜¾ç¤º'-'ï¼ŒåŸå› : éæœ€æ–°æœˆä»½æˆ–æ— æ•°æ®")
                    pdp_value = "-"
                    pdp_item = QTableWidgetItem(pdp_value)
                    pdp_item.setToolTip("å¹²å¥¶å‰æµè¡Œç‡åªåœ¨æœ€æ–°æœˆä»½è®¡ç®—")
                    pdp_item.setForeground(QColor('black'))  # é»‘è‰²å­—ä½“
                
                self.mastitis_monitoring_table.setItem(row, 7, pdp_item)
            
            # è°ƒæ•´åˆ—å®½
            self.mastitis_monitoring_table.resizeColumnsToContents()
            
        except Exception as e:
            logger.error(f"æ›´æ–°ç›‘æµ‹è¡¨æ ¼å¤±è´¥: {e}")
            raise
    
    def update_monitoring_chart(self, results):
        """æ›´æ–°ç›‘æµ‹è¶‹åŠ¿å›¾è¡¨"""
        try:
            import pyqtgraph as pg
            
            # æ¸…é™¤ç°æœ‰å›¾è¡¨
            self.mastitis_monitoring_plot.clear()
            
            months = results['months']
            indicators = results['indicators']
            
            if len(months) == 0:
                return
            
            # å‡†å¤‡æ•°æ®
            x_values = list(range(len(months)))
            x_labels = months
            
            # å®šä¹‰çº¿æ¡æ ·å¼
            line_styles = [
                {'color': '#e74c3c', 'width': 2, 'style': None},  # å½“æœˆæµè¡Œç‡ - çº¢è‰²
                {'color': '#f39c12', 'width': 2, 'style': None},  # æ–°å‘æ„ŸæŸ“ç‡ - æ©™è‰²  
                {'color': '#9b59b6', 'width': 2, 'style': None},  # æ…¢æ€§æ„ŸæŸ“ç‡ - ç´«è‰²
                {'color': '#3498db', 'width': 2, 'style': None},  # æ…¢æ€§æ„ŸæŸ“ç‰›å æ¯” - è“è‰²
                {'color': '#27ae60', 'width': 2, 'style': None},  # é¦–æµ‹æµè¡Œç‡ - ç»¿è‰²
            ]
            
            line_index = 0
            
            # ç»˜åˆ¶å„æŒ‡æ ‡çº¿æ¡ï¼ˆé»˜è®¤æ˜¾ç¤ºæ‰€æœ‰æŒ‡æ ‡ï¼‰
            # 1. å½“æœˆæµè¡Œç‡
            y_values = []
            for month in months:
                cp = indicators[month].get('current_prevalence', {})
                y_values.append(cp.get('value', None))
            
            valid_points = [(x, y) for x, y in zip(x_values, y_values) if y is not None]
            if valid_points:
                x_valid, y_valid = zip(*valid_points)
                self.mastitis_monitoring_plot.plot(
                    x_valid, y_valid, 
                    pen=pg.mkPen(color=line_styles[line_index]['color'], width=line_styles[line_index]['width']),
                    symbol='o', symbolSize=8, symbolBrush=line_styles[line_index]['color'],
                    name='å½“æœˆæµè¡Œç‡'
                )
            line_index += 1
            
            # 2. æ–°å‘æ„ŸæŸ“ç‡
            y_values = []
            for month in months:
                nir = indicators[month].get('new_infection_rate', {})
                y_values.append(nir.get('value', None))
            
            valid_points = [(x, y) for x, y in zip(x_values, y_values) if y is not None]
            if valid_points:
                x_valid, y_valid = zip(*valid_points)
                self.mastitis_monitoring_plot.plot(
                    x_valid, y_valid,
                    pen=pg.mkPen(color=line_styles[line_index]['color'], width=line_styles[line_index]['width']),
                    symbol='s', symbolSize=8, symbolBrush=line_styles[line_index]['color'],
                    name='æ–°å‘æ„ŸæŸ“ç‡'
                )
            line_index += 1
            
            # 3. æ…¢æ€§æ„ŸæŸ“ç‡
            y_values = []
            for month in months:
                cir = indicators[month].get('chronic_infection_rate', {})
                y_values.append(cir.get('value', None))
            
            valid_points = [(x, y) for x, y in zip(x_values, y_values) if y is not None]
            if valid_points:
                x_valid, y_valid = zip(*valid_points)
                self.mastitis_monitoring_plot.plot(
                    x_valid, y_valid,
                    pen=pg.mkPen(color=line_styles[line_index]['color'], width=line_styles[line_index]['width']),
                    symbol='t', symbolSize=8, symbolBrush=line_styles[line_index]['color'],
                    name='æ…¢æ€§æ„ŸæŸ“ç‡'
                )
            line_index += 1
            
            # 4. æ…¢æ€§æ„ŸæŸ“ç‰›å æ¯”
            y_values = []
            for month in months:
                cip = indicators[month].get('chronic_infection_proportion', {})
                y_values.append(cip.get('value', None))
            
            valid_points = [(x, y) for x, y in zip(x_values, y_values) if y is not None]
            if valid_points:
                x_valid, y_valid = zip(*valid_points)
                self.mastitis_monitoring_plot.plot(
                    x_valid, y_valid,
                    pen=pg.mkPen(color=line_styles[line_index]['color'], width=line_styles[line_index]['width']),
                    symbol='d', symbolSize=8, symbolBrush=line_styles[line_index]['color'],
                    name='æ…¢æ€§æ„ŸæŸ“ç‰›å æ¯”'
                )
            line_index += 1
            
            # 5. é¦–æµ‹æµè¡Œç‡ï¼ˆå¤´èƒï¼‰
            y_values = []
            for month in months:
                ftp = indicators[month].get('first_test_prevalence', {})
                if ftp and 'primiparous' in ftp:
                    y_values.append(ftp['primiparous'].get('value', None))
                else:
                    y_values.append(None)
            
            valid_points = [(x, y) for x, y in zip(x_values, y_values) if y is not None]
            if valid_points:
                x_valid, y_valid = zip(*valid_points)
                self.mastitis_monitoring_plot.plot(
                    x_valid, y_valid,
                    pen=pg.mkPen(color=line_styles[line_index]['color'], width=line_styles[line_index]['width'], style=pg.QtCore.Qt.PenStyle.DashLine),
                    symbol='+', symbolSize=10, symbolBrush=line_styles[line_index]['color'],
                    name='å¤´èƒé¦–æµ‹æµè¡Œç‡'
                )
            
            # è®¾ç½®Xè½´æ ‡ç­¾
            x_axis = self.mastitis_monitoring_plot.getAxis('bottom')
            x_axis.setTicks([[(i, month) for i, month in enumerate(months)]])
            
            # è®¾ç½®Yè½´èŒƒå›´
            self.mastitis_monitoring_plot.setYRange(0, 100)
            
        except Exception as e:
            logger.error(f"æ›´æ–°ç›‘æµ‹å›¾è¡¨å¤±è´¥: {e}")
            raise
    
    def export_monitoring_results(self):
        """å¯¼å‡ºéšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹ç»“æœåˆ°Excel"""
        try:
            if not self.mastitis_monitoring_results:
                QMessageBox.warning(self, "è­¦å‘Š", "æ²¡æœ‰å¯å¯¼å‡ºçš„ç›‘æµ‹ç»“æœ")
                return
            
            # é€‰æ‹©ä¿å­˜è·¯å¾„
            file_path, _ = QFileDialog.getSaveFileName(
                self, "å¯¼å‡ºç›‘æµ‹ç»“æœ", 
                f"éšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹ç»“æœ_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excelæ–‡ä»¶ (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # å‡†å¤‡å¯¼å‡ºæ•°æ®
            results = self.mastitis_monitoring_results
            months = results['months']
            indicators = results['indicators']
            
            # åˆ›å»ºDataFrame
            export_data = []
            for month in months:
                month_data = indicators.get(month, {})
                
                row = {
                    'æœˆä»½': month,
                    'å½“æœˆæµè¡Œç‡(%)': self._get_indicator_value(month_data, 'current_prevalence'),
                    'å½“æœˆæµè¡Œç‡_å…¬å¼': self._get_indicator_formula(month_data, 'current_prevalence'),
                    'æ–°å‘æ„ŸæŸ“ç‡(%)': self._get_indicator_value(month_data, 'new_infection_rate'),
                    'æ–°å‘æ„ŸæŸ“ç‡_å…¬å¼': self._get_indicator_formula(month_data, 'new_infection_rate'),
                    'æ…¢æ€§æ„ŸæŸ“ç‡(%)': self._get_indicator_value(month_data, 'chronic_infection_rate'),
                    'æ…¢æ€§æ„ŸæŸ“ç‡_å…¬å¼': self._get_indicator_formula(month_data, 'chronic_infection_rate'),
                    'æ…¢æ€§æ„ŸæŸ“ç‰›å æ¯”(%)': self._get_indicator_value(month_data, 'chronic_infection_proportion'),
                    'æ…¢æ€§æ„ŸæŸ“ç‰›å æ¯”_å…¬å¼': self._get_indicator_formula(month_data, 'chronic_infection_proportion'),
                    'å¤´èƒé¦–æµ‹æµè¡Œç‡(%)': self._get_first_test_value(month_data, 'primiparous'),
                    'å¤´èƒé¦–æµ‹æµè¡Œç‡_å…¬å¼': self._get_first_test_formula(month_data, 'primiparous'),
                    'ç»äº§é¦–æµ‹æµè¡Œç‡(%)': self._get_first_test_value(month_data, 'multiparous'),
                    'ç»äº§é¦–æµ‹æµè¡Œç‡_å…¬å¼': self._get_first_test_formula(month_data, 'multiparous'),
                    'å¹²å¥¶å‰æµè¡Œç‡(%)': self._get_indicator_value(month_data, 'pre_dry_prevalence'),
                    'å¹²å¥¶å‰æµè¡Œç‡_å…¬å¼': self._get_indicator_formula(month_data, 'pre_dry_prevalence'),
                }
                export_data.append(row)
            
            df = pd.DataFrame(export_data)
            
            # ä¿å­˜åˆ°Excel
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='ç›‘æµ‹ç»“æœ', index=False)
                
                # æ·»åŠ æ±‡æ€»ä¿¡æ¯sheet
                summary_data = {
                    'é¡¹ç›®': ['åˆ†ææ—¥æœŸ', 'ä½“ç»†èƒé˜ˆå€¼', 'åˆ†ææœˆä»½æ•°', 'æ—¥æœŸèŒƒå›´', 'æœˆä»½è¿ç»­æ€§'],
                    'å€¼': [
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        f"{results['scc_threshold']}ä¸‡/ml",
                        results['month_count'],
                        f"{months[0]} è‡³ {months[-1]}" if months else "æ— ",
                        "è¿ç»­" if results['continuity_check']['is_continuous'] else f"ä¸è¿ç»­ï¼Œç¼ºå¤±ï¼š{', '.join(results['continuity_check']['missing_months'])}"
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='åˆ†ææ±‡æ€»', index=False)
            
            self.show_export_success_dialog(f"éšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹ç»“æœå·²æˆåŠŸå¯¼å‡ºï¼", file_path)
            
        except Exception as e:
            QMessageBox.critical(self, "é”™è¯¯", f"å¯¼å‡ºå¤±è´¥: {str(e)}")
            logger.error(f"å¯¼å‡ºç›‘æµ‹ç»“æœå¤±è´¥: {e}")
    
    def _get_indicator_value(self, month_data, indicator_name):
        """è·å–æŒ‡æ ‡æ•°å€¼"""
        indicator = month_data.get(indicator_name, {})
        value = indicator.get('value')
        return f"{value:.1f}" if value is not None else "N/A"
    
    def _get_indicator_formula(self, month_data, indicator_name):
        """è·å–æŒ‡æ ‡å…¬å¼"""
        indicator = month_data.get(indicator_name, {})
        return indicator.get('formula', '')
    
    def _get_first_test_value(self, month_data, test_type):
        """è·å–é¦–æµ‹æµè¡Œç‡æ•°å€¼"""
        ftp = month_data.get('first_test_prevalence', {})
        if ftp and test_type in ftp:
            value = ftp[test_type].get('value')
            return f"{value:.1f}" if value is not None else "N/A"
        return "N/A"
    
    def _get_first_test_formula(self, month_data, test_type):
        """è·å–é¦–æµ‹æµè¡Œç‡å…¬å¼"""
        ftp = month_data.get('first_test_prevalence', {})
        if ftp and test_type in ftp:
            return ftp[test_type].get('formula', '')
        return ""
    
    def toggle_monitoring_formula_visibility(self):
        """åˆ‡æ¢éšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹å…¬å¼è¯´æ˜çš„æ˜¾ç¤º/éšè—çŠ¶æ€"""
        if self.monitoring_formula_widget.isVisible():
            self.monitoring_formula_widget.setVisible(False)
            self.formula_toggle_btn.setText("â–¼ å±•å¼€å…¬å¼è¯¦æƒ…")
        else:
            self.monitoring_formula_widget.setVisible(True)
            self.formula_toggle_btn.setText("â–² æ”¶èµ·å…¬å¼è¯¦æƒ…")
    
    def update_monitoring_threshold(self):
        """æ›´æ–°éšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹çš„ä½“ç»†èƒé˜ˆå€¼"""
        if hasattr(self, 'mastitis_monitoring_calculator') and self.mastitis_monitoring_calculator:
            new_threshold = self.monitoring_scc_threshold.value()
            self.mastitis_monitoring_calculator.set_scc_threshold(new_threshold)
            logger.info(f"ä½“ç»†èƒé˜ˆå€¼å·²æ›´æ–°ä¸º: {new_threshold} ä¸‡/ml")
    
    def update_monitoring_display(self):
        """æ›´æ–°éšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹æ˜¾ç¤ºï¼ˆé‡æ–°è®¡ç®—å¹¶æ˜¾ç¤ºç»“æœï¼‰"""
        if hasattr(self, 'mastitis_monitoring_results') and self.mastitis_monitoring_results:
            # é‡æ–°è®¡ç®—æŒ‡æ ‡
            results = self.mastitis_monitoring_calculator.calculate_all_indicators()
            if results['success']:
                self.mastitis_monitoring_results = results
                self.display_mastitis_monitoring_results(results)
            else:
                QMessageBox.warning(self, "æ›´æ–°å¤±è´¥", f"é‡æ–°è®¡ç®—å¤±è´¥: {results.get('error', 'æœªçŸ¥é”™è¯¯')}")
    
    def upload_dhi_for_monitoring(self):
        """ä¸ºéšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹ä¸Šä¼ DHIæ•°æ®"""
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "é€‰æ‹©DHIæŠ¥å‘Šæ–‡ä»¶ï¼ˆéšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹ï¼‰",
            "",
            "æ”¯æŒçš„æ–‡ä»¶ (*.zip *.xlsx *.xls);;ZIPæ–‡ä»¶ (*.zip);;Excelæ–‡ä»¶ (*.xlsx *.xls)"
        )
        
        if not files:
            return
        
        try:
            # ä½¿ç”¨æ–°çš„æµç•…è¿›åº¦æ¡
            progress_dialog = SmoothProgressDialog("æ­£åœ¨å¤„ç†DHIæ–‡ä»¶...", "å–æ¶ˆ", 0, len(files), self)
            progress_dialog.show()
            
            # å¤„ç†DHIæ–‡ä»¶
            from data_processor import DataProcessor
            processor = DataProcessor()
            
            all_dhi_data = []
            success_files = []
            failed_files = []
            
            for i, file_path in enumerate(files):
                if progress_dialog.wasCanceled():
                    break
                
                filename = os.path.basename(file_path)
                progress_dialog.setLabelText(f"å¤„ç†: {filename}")
                progress_dialog.setValue(i)
                QApplication.processEvents()
                
                try:
                    # å¤„ç†å•ä¸ªDHIæ–‡ä»¶
                    success, message, df = processor.process_file(file_path, filename)
                    
                    if success and df is not None and not df.empty:
                        all_dhi_data.append(df)
                        success_files.append(filename)
                    else:
                        failed_files.append(f"{filename}: {message}")
                        
                except Exception as e:
                    failed_files.append(f"{filename}: {str(e)}")
            
            progress_dialog.setValue(len(files))
            progress_dialog.close()
            
            if all_dhi_data:
                # åˆå§‹åŒ–ç›‘æµ‹è®¡ç®—å™¨
                try:
                    from mastitis_monitoring import MastitisMonitoringCalculator
                    
                    if not hasattr(self, 'mastitis_monitoring_calculator'):
                        self.mastitis_monitoring_calculator = MastitisMonitoringCalculator()
                    
                    # åŠ è½½DHIæ•°æ®
                    load_result = self.mastitis_monitoring_calculator.load_dhi_data(all_dhi_data)
                    
                    if load_result['success']:
                        # æ›´æ–°ç›‘æµ‹çŠ¶æ€æ˜¾ç¤º
                        self.update_monitoring_data_status()
                        
                        # æ›´æ–°çŠ¶æ€ä¿¡æ¯
                        self.monitoring_status_label.setText(f"âœ… DHIæ•°æ®å·²ä¸Šä¼ ï¼ŒåŒ…å« {load_result['month_count']} ä¸ªæœˆä»½çš„æ•°æ®")
                        
                        QMessageBox.information(
                            self, 
                            "ä¸Šä¼ æˆåŠŸ", 
                            f"DHIæ•°æ®ä¸Šä¼ æˆåŠŸï¼\n\n"
                            f"æˆåŠŸå¤„ç†: {len(success_files)} ä¸ªæ–‡ä»¶\n"
                            f"æ•°æ®æœˆä»½: {load_result['month_count']} ä¸ªæœˆ\n"
                            f"æ—¶é—´èŒƒå›´: {load_result['date_range']['start']} - {load_result['date_range']['end']}"
                        )
                        
                    else:
                        QMessageBox.warning(self, "åŠ è½½å¤±è´¥", f"DHIæ•°æ®åŠ è½½å¤±è´¥: {load_result.get('error', 'æœªçŸ¥é”™è¯¯')}")
                        
                except ImportError:
                    QMessageBox.critical(self, "æ¨¡å—ç¼ºå¤±", "ç¼ºå°‘éšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹æ¨¡å—ï¼Œè¯·æ£€æŸ¥å®‰è£…ã€‚")
                except Exception as e:
                    QMessageBox.critical(self, "å¤„ç†å¤±è´¥", f"DHIæ•°æ®å¤„ç†å¤±è´¥: {str(e)}")
                    
            else:
                error_msg = "æ²¡æœ‰æˆåŠŸå¤„ç†çš„DHIæ–‡ä»¶ã€‚\n\nå¤±è´¥åŸå› :\n" + "\n".join(failed_files[:5])
                if len(failed_files) > 5:
                    error_msg += f"\n... è¿˜æœ‰ {len(failed_files) - 5} ä¸ªæ–‡ä»¶å¤±è´¥"
                QMessageBox.warning(self, "å¤„ç†å¤±è´¥", error_msg)
                
        except Exception as e:
            QMessageBox.critical(self, "ä¸Šä¼ å¤±è´¥", f"DHIæ–‡ä»¶ä¸Šä¼ å¤±è´¥: {str(e)}")
    
    def upload_cattle_info_for_monitoring(self):
        """ä¸ºéšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹ä¸Šä¼ ç‰›ç¾¤åŸºç¡€ä¿¡æ¯"""
        system_type_map = {
            "ä¼Šèµ·ç‰›": "yiqiniu",
            "æ…§ç‰§äº‘": "huimuyun", 
            "å…¶ä»–": "custom"
        }
        
        system_type = system_type_map.get(self.monitoring_system_combo.currentText(), "custom")
        
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "é€‰æ‹©ç‰›ç¾¤åŸºç¡€ä¿¡æ¯æ–‡ä»¶",
            "",
            "Excelæ–‡ä»¶ (*.xlsx *.xls)"
        )
        
        if not file_path:
            return
        
        try:
            # å¤„ç†ç‰›ç¾¤åŸºç¡€ä¿¡æ¯æ–‡ä»¶
            from data_processor import DataProcessor
            processor = DataProcessor()
            
            if system_type == "custom":
                # è‡ªå®šä¹‰ç³»ç»Ÿï¼šå°è¯•ç›´æ¥è¯»å–å¹¶è¯†åˆ«å­—æ®µ
                try:
                    cattle_df = pd.read_excel(file_path)
                    
                    # çµæ´»åŒ¹é…è€³å·å­—æ®µ
                    ear_tag_field = None
                    for field in ['è€³å·', 'ear_tag', 'ç‰›å·', 'cow_id']:
                        if field in cattle_df.columns:
                            ear_tag_field = field
                            break
                    
                    # çµæ´»åŒ¹é…åœ¨èƒå¤©æ•°å­—æ®µ
                    pregnancy_field = None
                    for field in ['åœ¨èƒå¤©æ•°', 'æ€€å­•å¤©æ•°', 'gestation_days', 'pregnancy_days']:
                        if field in cattle_df.columns:
                            pregnancy_field = field
                            break
                    
                    if not ear_tag_field or not pregnancy_field:
                        missing_fields = []
                        if not ear_tag_field:
                            missing_fields.append('è€³å·ï¼ˆæˆ–ear_tag/ç‰›å·ï¼‰')
                        if not pregnancy_field:
                            missing_fields.append('åœ¨èƒå¤©æ•°ï¼ˆæˆ–æ€€å­•å¤©æ•°/gestation_daysï¼‰')
                        
                        QMessageBox.warning(
                            self,
                            "å­—æ®µç¼ºå¤±",
                            f"ç‰›ç¾¤åŸºç¡€ä¿¡æ¯æ–‡ä»¶ç¼ºå°‘å¿…è¦å­—æ®µï¼š\n{', '.join(missing_fields)}\n\n"
                            f"å½“å‰æ–‡ä»¶åŒ…å«å­—æ®µï¼š\n{', '.join(cattle_df.columns[:10])}..."
                        )
                        return
                    
                    # æ ‡å‡†åŒ–æ•°æ®
                    result_df = pd.DataFrame()
                    result_df['ear_tag'] = cattle_df[ear_tag_field].astype(str).str.lstrip('0').replace('', '0')
                    result_df['gestation_days'] = pd.to_numeric(cattle_df[pregnancy_field], errors='coerce')
                    
                    # æ·»åŠ å…¶ä»–å¯ç”¨å­—æ®µ
                    optional_fields = ['èƒæ¬¡', 'æ³Œä¹³å¤©æ•°', 'ç¹è‚²çŠ¶æ€', 'æœ€è¿‘äº§çŠŠæ—¥æœŸ']
                    for field in optional_fields:
                        if field in cattle_df.columns:
                            if field == 'èƒæ¬¡':
                                result_df['parity'] = pd.to_numeric(cattle_df[field], errors='coerce')
                            elif field == 'æ³Œä¹³å¤©æ•°':
                                result_df['lactation_days'] = pd.to_numeric(cattle_df[field], errors='coerce')
                            elif field == 'ç¹è‚²çŠ¶æ€':
                                result_df['breeding_status'] = cattle_df[field].astype(str)
                            elif field == 'æœ€è¿‘äº§çŠŠæ—¥æœŸ':
                                result_df['last_calving_date'] = pd.to_datetime(cattle_df[field], errors='coerce')
                    
                    # æ¸…ç†æ•°æ®
                    result_df = result_df.dropna(subset=['ear_tag'])
                    result_df = result_df[result_df['ear_tag'] != 'nan']
                    
                    processed_data = {'cattle_info': result_df}
                    success = True
                    message = f"æˆåŠŸå¤„ç†{len(result_df)}æ¡ç‰›ç¾¤åŸºç¡€ä¿¡æ¯"
                    
                except Exception as e:
                    QMessageBox.critical(
                        self,
                        "æ–‡ä»¶å¤„ç†å¤±è´¥",
                        f"æ— æ³•å¤„ç†è‡ªå®šä¹‰æ ¼å¼çš„ç‰›ç¾¤åŸºç¡€ä¿¡æ¯æ–‡ä»¶ï¼š\n{str(e)}\n\n"
                        "è¯·ç¡®ä¿æ–‡ä»¶åŒ…å«'è€³å·'å’Œ'åœ¨èƒå¤©æ•°'å­—æ®µï¼ˆæˆ–ç›¸åº”çš„è‹±æ–‡å­—æ®µåï¼‰"
                    )
                    return
            else:
                # ä½¿ç”¨æ…¢æ€§ä¹³æˆ¿ç‚ç­›æŸ¥çš„æ–‡ä»¶å¤„ç†æ–¹æ³•
                success, message, processed_data = processor.process_mastitis_system_files(
                    system_type, 
                    {'cattle_info': file_path}
                )
            
            if success and 'cattle_info' in processed_data:
                cattle_df = processed_data['cattle_info']
                
                # åˆå§‹åŒ–ç›‘æµ‹è®¡ç®—å™¨ï¼ˆå¦‚æœè¿˜æ²¡æœ‰ï¼‰
                try:
                    from mastitis_monitoring import MastitisMonitoringCalculator
                    
                    if not hasattr(self, 'mastitis_monitoring_calculator'):
                        self.mastitis_monitoring_calculator = MastitisMonitoringCalculator()
                    
                    # åŠ è½½ç‰›ç¾¤åŸºç¡€ä¿¡æ¯
                    load_result = self.mastitis_monitoring_calculator.load_cattle_basic_info(cattle_df, system_type)
                    
                    if load_result['success']:
                        # åŒæ—¶ä¿å­˜åˆ°ä¸»çª—å£ï¼Œç¡®ä¿å…¶ä»–æ¨¡å—ä¹Ÿèƒ½ä½¿ç”¨
                        self.cattle_basic_info = cattle_df
                        self.current_system = system_type
                        
                        # æ›´æ–°çŠ¶æ€æ˜¾ç¤º
                        status_text = f"âœ… å·²ä¸Šä¼ ç‰›ç¾¤åŸºç¡€ä¿¡æ¯\n"
                        status_text += f"ğŸ“Š ç³»ç»Ÿç±»å‹: {self.monitoring_system_combo.currentText()}\n"
                        status_text += f"ğŸ„ ç‰›åªæ•°é‡: {load_result['cattle_count']} å¤´\n"
                        if load_result.get('pregnancy_field'):
                            status_text += f"ğŸ¤° åœ¨èƒå¤©æ•°å­—æ®µ: {load_result['pregnancy_field']}\n"
                        
                        # ç»Ÿè®¡åœ¨èƒå¤©æ•°>180å¤©çš„ç‰›åª
                        pregnancy_field = load_result.get('pregnancy_field')
                        if pregnancy_field and pregnancy_field in cattle_df.columns:
                            pregnancy_data = cattle_df[pregnancy_field].dropna()
                            if len(pregnancy_data) > 0:
                                over_180_count = (pregnancy_data > 180).sum()
                                status_text += f"ğŸ¯ å¹²å¥¶å‰ç‰›åª(>180å¤©): {over_180_count} å¤´"
                                
                        self.cattle_info_label.setText(status_text)
                        self.cattle_info_label.setStyleSheet("color: #27ae60; font-size: 10px; padding: 5px; background-color: #f8f9fa; border: 1px solid #27ae60; border-radius: 3px;")
                        
                        # æ›´æ–°åˆ†ææŒ‰é’®çŠ¶æ€
                        self.check_monitoring_analysis_ready()
                        
                        QMessageBox.information(
                            self,
                            "ä¸Šä¼ æˆåŠŸ",
                            f"ğŸ‰ ç‰›ç¾¤åŸºç¡€ä¿¡æ¯ä¸Šä¼ æˆåŠŸï¼\n\n"
                            f"ğŸ“Š ç³»ç»Ÿç±»å‹: {self.monitoring_system_combo.currentText()}\n"
                            f"ğŸ„ æ•°æ®é‡: {load_result['cattle_count']} å¤´ç‰›\n"
                            f"ğŸ¤° åœ¨èƒå¤©æ•°å­—æ®µ: {load_result.get('pregnancy_field', 'æœªè¯†åˆ«')}\n"
                            f"ğŸ’¡ ç°åœ¨å¯ä»¥è¿›è¡Œéšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹åˆ†æï¼Œå°†åŒ…å«å¹²å¥¶å‰æµè¡Œç‡è®¡ç®—"
                        )
                        
                    else:
                        QMessageBox.warning(self, "åŠ è½½å¤±è´¥", f"ç‰›ç¾¤åŸºç¡€ä¿¡æ¯åŠ è½½å¤±è´¥: {load_result.get('error', 'æœªçŸ¥é”™è¯¯')}")
                        
                except ImportError:
                    QMessageBox.critical(self, "æ¨¡å—ç¼ºå¤±", "ç¼ºå°‘éšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹æ¨¡å—ï¼Œè¯·æ£€æŸ¥å®‰è£…ã€‚")
                except Exception as e:
                    QMessageBox.critical(self, "å¤„ç†å¤±è´¥", f"ç‰›ç¾¤åŸºç¡€ä¿¡æ¯å¤„ç†å¤±è´¥: {str(e)}")
                    
            else:
                QMessageBox.warning(self, "å¤„ç†å¤±è´¥", f"ç‰›ç¾¤åŸºç¡€ä¿¡æ¯å¤„ç†å¤±è´¥: {message}")
                
        except Exception as e:
            QMessageBox.critical(self, "ä¸Šä¼ å¤±è´¥", f"ç‰›ç¾¤åŸºç¡€ä¿¡æ¯ä¸Šä¼ å¤±è´¥: {str(e)}")
    
    def check_monitoring_analysis_ready(self):
        """æ£€æŸ¥éšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹åˆ†ææ˜¯å¦å‡†å¤‡å°±ç»ª"""
        if hasattr(self, 'mastitis_monitoring_calculator') and self.mastitis_monitoring_calculator:
            # æ£€æŸ¥æ˜¯å¦æœ‰DHIæ•°æ®
            has_dhi_data = hasattr(self.mastitis_monitoring_calculator, 'monthly_data') and \
                          self.mastitis_monitoring_calculator.monthly_data
            
            # å¯ç”¨åˆ†ææŒ‰é’®
            if has_dhi_data:
                self.analyze_monitoring_btn.setEnabled(True)
                logger.info("éšå½¢ä¹³æˆ¿ç‚ç›‘æµ‹åˆ†æå·²å‡†å¤‡å°±ç»ª")
            else:
                self.analyze_monitoring_btn.setEnabled(False)
        else:
            self.analyze_monitoring_btn.setEnabled(False)
    
    def start_heartbeat(self):
        """å¯åŠ¨å¿ƒè·³æœºåˆ¶"""
        if self.heartbeat_timer:
            self.heartbeat_timer.stop()
        
        self.heartbeat_timer = QTimer()
        self.heartbeat_timer.timeout.connect(self.send_heartbeat)
        self.heartbeat_timer.start(90000)  # 90ç§’
        
        # ç«‹å³å‘é€ä¸€æ¬¡å¿ƒè·³
        self.send_heartbeat()
    
    def send_heartbeat(self):
        """å‘é€å¿ƒè·³"""
        if self.auth_service:
            success = self.auth_service.heartbeat()
            if not success:
                # ä¼šè¯å¤±æ•ˆï¼Œéœ€è¦é‡æ–°ç™»å½•
                self.heartbeat_timer.stop()
                QMessageBox.warning(
                    self,
                    "ä¼šè¯å¤±æ•ˆ",
                    "æ‚¨çš„ç™»å½•ä¼šè¯å·²å¤±æ•ˆï¼Œè¯·é‡æ–°ç™»å½•ã€‚"
                )
                self.logout()
    
    def logout(self):
        """æ³¨é”€"""
        reply = QMessageBox.question(
            self,
            "ç¡®è®¤æ³¨é”€",
            "ç¡®å®šè¦é€€å‡ºç™»å½•å—ï¼Ÿ",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            # åœæ­¢å¿ƒè·³
            if self.heartbeat_timer:
                self.heartbeat_timer.stop()
            
            # è°ƒç”¨æ³¨é”€æ¥å£
            if self.auth_service:
                self.auth_service.logout()
            
            # å…³é—­ä¸»çª—å£
            self.close()
            
            # é‡æ–°æ˜¾ç¤ºç™»å½•å¯¹è¯æ¡†
            login_dialog = LoginDialog(None, self.auth_service)
            if login_dialog.exec() == QDialog.DialogCode.Accepted:
                # ç™»å½•æˆåŠŸï¼Œåˆ›å»ºæ–°çš„ä¸»çª—å£
                new_window = MainWindow(
                    username=login_dialog.get_username(),
                    auth_service=self.auth_service
                )
                new_window.showMaximized()
            else:
                # ç”¨æˆ·å–æ¶ˆç™»å½•ï¼Œé€€å‡ºåº”ç”¨
                QApplication.quit()
    
    def closeEvent(self, event):
        """çª—å£å…³é—­äº‹ä»¶"""
        # åœæ­¢å¿ƒè·³
        if self.heartbeat_timer:
            self.heartbeat_timer.stop()
        
        # æ³¨é”€
        if self.auth_service:
            self.auth_service.logout()
        
        event.accept()


class DHIDesktopApp:
    """DHIæ¡Œé¢åº”ç”¨ç¨‹åº"""
    
    def __init__(self):
        self.app = None
        self.window = None
        self.auth_service = None
        self.username = None
    
    def run(self):
        """è¿è¡Œåº”ç”¨ç¨‹åº"""
        try:
            # å¯ç”¨é«˜DPIæ”¯æŒ
            QApplication.setHighDpiScaleFactorRoundingPolicy(Qt.HighDpiScaleFactorRoundingPolicy.PassThrough)
            
            # åˆ›å»ºQApplication
            self.app = QApplication(sys.argv)
            self.app.setApplicationName("DHIç­›æŸ¥åŠ©æ‰‹")
            self.app.setApplicationVersion("2.0.0")
            self.app.setOrganizationName("DHI")
            self.app.setOrganizationDomain("dhi.com")
            self.app.setStyle('Fusion')  # ä½¿ç”¨ç°ä»£æ ·å¼
            
            # è®¾ç½®åº”ç”¨ç¨‹åºå›¾æ ‡
            try:
                if os.path.exists("whg3r-qi1nv-001.ico"):
                    self.app.setWindowIcon(QIcon("whg3r-qi1nv-001.ico"))
            except:
                pass
            
            # åˆ›å»ºç®€åŒ–çš„è®¤è¯æœåŠ¡ï¼ˆç›´æ¥è¿æ¥é˜¿é‡Œäº‘æ•°æ®åº“ï¼‰
            print("æ­£åœ¨è¿æ¥è®¤è¯æœåŠ¡...")
            self.auth_service = SimpleAuthService()
            
            # æ£€æŸ¥æ•°æ®åº“è¿æ¥
            if not self.auth_service.check_server_health():
                QMessageBox.critical(
                    None,
                    "æ•°æ®åº“è¿æ¥å¤±è´¥",
                    "æ— æ³•è¿æ¥åˆ°æ•°æ®åº“ã€‚\nè¯·æ£€æŸ¥ç½‘ç»œè¿æ¥åé‡è¯•ã€‚"
                )
                return 0
            
            # æ˜¾ç¤ºç™»å½•å¯¹è¯æ¡†
            login_dialog = LoginDialog(None, self.auth_service)
            if login_dialog.exec() == QDialog.DialogCode.Accepted:
                self.username = login_dialog.get_username()
            else:
                # ç”¨æˆ·å–æ¶ˆç™»å½•
                return 0
            
            # åˆ›å»ºä¸»çª—å£ï¼Œä¼ å…¥ç”¨æˆ·åå’Œè®¤è¯æœåŠ¡
            self.window = MainWindow(username=self.username, auth_service=self.auth_service)
            self.window.showMaximized()  # è‡ªåŠ¨æœ€å¤§åŒ–æ˜¾ç¤º
            
            # è¿è¡Œäº‹ä»¶å¾ªç¯
            return self.app.exec()
            
        except Exception as e:
            print(f"åº”ç”¨ç¨‹åºå¯åŠ¨å¤±è´¥: {e}")
            if self.app:
                QMessageBox.critical(
                    None,
                    "å¯åŠ¨å¤±è´¥", 
                    f"åº”ç”¨ç¨‹åºå¯åŠ¨å¤±è´¥:\n{e}"
                )
            return 1


def main():
    """ä¸»å‡½æ•°"""
    try:
        app = DHIDesktopApp()
        return app.run()
    except Exception as e:
        print(f"ç¨‹åºå¯åŠ¨å¤±è´¥: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())